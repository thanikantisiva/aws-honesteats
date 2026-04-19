#!/usr/bin/env python3
"""
Enrich Kohinoor Spice menu Excel with subcategory, veg/non-veg, description,
unique Unsplash images and generate API payloads for import.

Usage:
  python3 scripts/enrich_and_import_kohinoor.py            # enrich + dry run
  python3 scripts/enrich_and_import_kohinoor.py --apply     # enrich + insert
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RESTAURANT_ID = "RES-1776403494998-2791"
XLSX_PATH = Path("/Users/user/Downloads/Kohinoor_Spice_Zomato_vs_YumDude.xlsx")

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get("HONESTEATS_RETOOL_BYPASS", "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc")
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

# ── ALL items are Veg (Kohinoor is a pure-veg restaurant) ────────────────────────

# ── Subcategory mapping ──────────────────────────────────────────────────────────
SUBCATEGORY: dict[str, str] = {
    # Icecream Scoops
    "American Dry Nuts": "Premium Scoops", "Anjeer Badam": "Premium Scoops",
    "Belgium Dark Chocolate": "Premium Scoops", "Black Current": "Premium Scoops",
    "Black Forest": "Premium Scoops", "Choco Chips": "Premium Scoops",
    "Kaju Kismiss": "Premium Scoops", "Caramel Nuts Scoop": "Premium Scoops",
    "Rajbhog Scoop": "Premium Scoops", "Italian Delight Scoop": "Premium Scoops",
    "Badam Scoop": "Classic Scoops", "Butterscotch Scoop": "Classic Scoops",
    "Chocolate Scoop": "Classic Scoops", "Mango Scoop": "Classic Scoops",
    "Pineapple Scoop": "Classic Scoops", "Pista Scoop": "Classic Scoops",
    "Strawberry Scoop": "Classic Scoops", "Tooti Frooti Scoop": "Classic Scoops",
    "Vanilla Scoop": "Classic Scoops",
    # Natural Fruit Based
    "Alphonso Mango": "Fruit Ice Cream", "Banana Caramel": "Fruit Ice Cream",
    "Chatpata Guava": "Fruit Ice Cream", "Fruit Ninja": "Fruit Ice Cream",
    "Jack Fruit": "Fruit Ice Cream", "Kesar Pista": "Fruit Ice Cream",
    "Musk Melon": "Fruit Ice Cream", "Sapota": "Fruit Ice Cream",
    "Sithaphal": "Fruit Ice Cream", "Strawberry": "Fruit Ice Cream",
    "Tender Coconut": "Fruit Ice Cream",
    # Starters
    "Babycorn Chilli": "Baby Corn", "Babycorn Manchurian": "Baby Corn",
    "Gobi Chilli": "Gobi", "Gobi Manchurian": "Gobi",
    "Mushroom Chilli": "Mushroom", "Mushroom Machurian": "Mushroom",
    "Paneer 555": "Paneer", "Paneer Basket": "Paneer",
    "Paneer Chilli": "Paneer", "Paneer Majestic": "Paneer",
    "Paneer Manchurian": "Paneer", "Veg Allwin": "Mixed Veg",
    "Veg Basket": "Mixed Veg", "Veg Bullet": "Mixed Veg",
    "Veg Gulzar": "Mixed Veg", "Veg Machurian": "Mixed Veg",
    "Masala French Fries": "Fries", "Salted French Fries": "Fries",
    # Sundaes
    "Badam Pista Deluxe Sundaes": "Sundaes", "Deluxe Icecream Sundaes": "Sundaes",
    "Evening Special Sundaes": "Sundaes", "Fruit Salad Sundaes": "Sundaes",
    "Holiday Special Sundaes": "Sundaes", "Honey Moon Sundaes": "Sundaes",
    "Kova Special Sundaes": "Sundaes", "Mango Deluxe Sundaes": "Sundaes",
    "Mughal Sundaes": "Sundaes", "Pista Deluxe Sundaes": "Sundaes",
    "Sunday Special Sundaes": "Sundaes", "Tropical Boat Sundaes": "Sundaes",
    # Milkshakes
    "Alphonso Mango Milkshake": "Fruit Milkshake", "Banana Caramel Milkshake": "Fruit Milkshake",
    "Fruit Ninja Milkshake": "Fruit Milkshake", "Jack Fruit Milkshake": "Fruit Milkshake",
    "Kesar Pista Milkshake": "Fruit Milkshake", "Mango Milkshake": "Fruit Milkshake",
    "Musk Melon Milkshake": "Fruit Milkshake", "Pine Apple Milkshake": "Fruit Milkshake",
    "Sapota Milkshake": "Fruit Milkshake", "Sithaphal Milkshake": "Fruit Milkshake",
    "Strawberry Milkshake": "Fruit Milkshake", "Tender Coconut Milkshake": "Fruit Milkshake",
    "Badam Milk Milkshake": "Classic Milkshake", "Badam Milkshake": "Classic Milkshake",
    "Butterscotch Milkshake": "Classic Milkshake", "Caramel Nuts Milkshake": "Classic Milkshake",
    "Chocolate Milkshake": "Classic Milkshake", "Italian Delight Milkshake": "Classic Milkshake",
    "Pista Milkshake": "Classic Milkshake", "Rajbhog Milkshake": "Classic Milkshake",
    "Tooti Frooti Milkshake": "Classic Milkshake", "Vanilla Milkshake": "Classic Milkshake",
    # Noodles
    "Gobi Noodles": "Noodles", "Mushroom Noodles": "Noodles",
    "Paneer Noodles": "Noodles", "Veg Noodles": "Noodles",
    # Fried Rice
    "Chinese Fried Rice": "Fried Rice", "Gobi Fried Rice": "Fried Rice",
    "Mushroom Fried Rice": "Fried Rice", "Paneer Fried Rice": "Fried Rice",
    "Schezwan Fried Rice": "Fried Rice", "Singapore Fried Rice": "Fried Rice",
    "Tomato Fried Rice": "Fried Rice", "Veg Fried Rice": "Fried Rice",
    "Curd Rice": "Plain Rice", "Ghee Rice": "Plain Rice",
    "Jeera Rice": "Plain Rice", "Kaju Rice": "Plain Rice",
    "Kashmir Pulao": "Plain Rice",
    # Biryanis
    "Kajui Biryani": "Veg Biryani", "Mushroom Biryani": "Veg Biryani",
    "Paneer Biryani": "Veg Biryani", "Veg Biryani": "Veg Biryani",
    # Rotis
    "Butter Phulka": "Roti", "Butter Rumali Roti": "Roti",
    "Butter Wheat Phulka": "Roti", "Phulka": "Roti",
    "Rumali Roti": "Roti", "Wheat Phulka": "Roti",
    # Curries
    "Kadai Paneer Masala": "Paneer Curry", "Paneer Butter Masala": "Paneer Curry",
    "Paneer Mushroom Masala": "Paneer Curry", "Paneer Tikka Masala": "Paneer Curry",
    "Methi Chaman": "Paneer Curry", "Palak Paneer": "Paneer Curry",
    "Kaju Paneer Masala": "Kaju Curry", "Kaju Mushroom Masala": "Kaju Curry",
    "Kaju Masala": "Kaju Curry", "Kaju Tomato Curry": "Kaju Curry",
    "Kadai Vegetable": "Veg Curry", "Kofta Curry": "Veg Curry",
    "Mushroom Masala": "Veg Curry", "Veg Doboja": "Veg Curry",
    "Veg Mix Curry": "Veg Curry",
    # Cold Pressed Juices
    "Pineapple": "Cold Pressed Juice", "Watermelon": "Cold Pressed Juice",
    # Pizza
    "Chocolate Pizza": "Sweet Pizza", "Corn Delight Pizza": "Veg Pizza",
    "Double Cheese Pizza": "Cheese Pizza", "Mushroom Pizza": "Veg Pizza",
    "Paneer Delight": "Paneer Pizza", "Pepper Corn Veg": "Veg Pizza",
    "Pineapple Pizza": "Veg Pizza", "Spicy Veg Pizza": "Veg Pizza",
    "Veg Delight Pizza": "Veg Pizza", "Veg Pizza": "Veg Pizza",
    "Veg Supreme Pizza": "Veg Pizza",
    # Burgers
    "Aloo Tikki": "Veg Burger", "Cheese Burger": "Veg Burger",
    "Paneer Burger": "Veg Burger", "Veg Burger": "Veg Burger",
    # Sandwich
    "Chocolate": "Sweet Sandwich", "Corn Sandwich": "Veg Sandwich",
    "Paneer": "Veg Sandwich", "Pineapple": "Veg Sandwich",
    "Veg Cheese Sandwich": "Veg Sandwich", "Veg Sandwich": "Veg Sandwich",
}

# ── Descriptions ─────────────────────────────────────────────────────────────────
DESC: dict[str, str] = {
    "American Dry Nuts": "Premium ice cream loaded with almonds, cashews and pistachios",
    "Anjeer Badam": "Rich anjeer and badam flavored ice cream scoop",
    "Badam Scoop": "Creamy badam (almond) flavored ice cream",
    "Belgium Dark Chocolate": "Intense Belgian dark chocolate ice cream",
    "Black Current": "Tangy black currant flavored ice cream",
    "Black Forest": "Classic black forest ice cream with chocolate and cherry",
    "Butterscotch Scoop": "Smooth butterscotch ice cream with crunchy bits",
    "Caramel Nuts Scoop": "Rich caramel ice cream studded with roasted nuts",
    "Choco Chips": "Chocolate ice cream loaded with choco chips",
    "Chocolate Scoop": "Classic rich chocolate ice cream",
    "Italian Delight Scoop": "Premium Italian-style ice cream with mixed flavors",
    "Kaju Kismiss": "Cashew and raisin flavored premium ice cream",
    "Mango Scoop": "Fresh mango flavored ice cream",
    "Pineapple Scoop": "Refreshing pineapple ice cream",
    "Pista Scoop": "Premium pistachio flavored ice cream",
    "Rajbhog Scoop": "Traditional rajbhog flavored ice cream with saffron",
    "Strawberry Scoop": "Fresh strawberry flavored ice cream",
    "Tooti Frooti Scoop": "Colorful tutti frutti ice cream with candied fruits",
    "Vanilla Scoop": "Classic vanilla bean ice cream",
    "Alphonso Mango": "Natural alphonso mango ice cream made with real fruit",
    "Banana Caramel": "Banana ice cream with caramel swirl",
    "Chatpata Guava": "Tangy guava ice cream with a chatpata twist",
    "Fruit Ninja": "Mixed fruit ice cream medley",
    "Jack Fruit": "Exotic jackfruit flavored natural ice cream",
    "Kesar Pista": "Saffron and pistachio natural ice cream",
    "Musk Melon": "Refreshing musk melon natural ice cream",
    "Sapota": "Natural sapota (chikoo) flavored ice cream",
    "Sithaphal": "Seasonal custard apple natural ice cream",
    "Strawberry": "Natural strawberry ice cream made with real berries",
    "Tender Coconut": "Fresh tender coconut flavored natural ice cream",
    "Badam Pista Deluxe Sundaes": "Deluxe sundae with badam and pista ice cream layers",
    "Deluxe Icecream Sundaes": "Classic deluxe ice cream sundae with toppings",
    "Evening Special Sundaes": "Special evening treat sundae with assorted flavors",
    "Fruit Salad Sundaes": "Fresh fruit salad topped with ice cream and cream",
    "Holiday Special Sundaes": "Festive holiday special ice cream sundae",
    "Honey Moon Sundaes": "Romantic honey-drizzled ice cream sundae",
    "Kova Special Sundaes": "Traditional kova (milk sweet) based sundae",
    "Mango Deluxe Sundaes": "Rich mango ice cream sundae with mango pulp",
    "Mughal Sundaes": "Royal Mughal-inspired ice cream sundae with dry fruits",
    "Pista Deluxe Sundaes": "Pistachio ice cream sundae with crushed pistachios",
    "Sunday Special Sundaes": "Weekly special sundae with chef's choice toppings",
    "Tropical Boat Sundaes": "Grand tropical fruit boat sundae — serves to share",
    "Alphonso Mango Milkshake": "Thick alphonso mango milkshake made with real pulp",
    "Badam Milk Milkshake": "Traditional badam milk served chilled",
    "Badam Milkshake": "Rich almond milkshake with badam paste",
    "Banana Caramel Milkshake": "Banana milkshake with caramel drizzle",
    "Butterscotch Milkshake": "Creamy butterscotch milkshake",
    "Caramel Nuts Milkshake": "Caramel milkshake with crushed nuts",
    "Chocolate Milkshake": "Rich chocolate milkshake",
    "Fruit Ninja Milkshake": "Mixed fruit milkshake blend",
    "Italian Delight Milkshake": "Italian-style premium milkshake",
    "Jack Fruit Milkshake": "Exotic jackfruit milkshake",
    "Kesar Pista Milkshake": "Saffron and pistachio milkshake",
    "Mango Milkshake": "Classic mango milkshake",
    "Musk Melon Milkshake": "Refreshing musk melon milkshake",
    "Pine Apple Milkshake": "Tangy pineapple milkshake",
    "Pista Milkshake": "Pistachio milkshake with crushed pista",
    "Rajbhog Milkshake": "Saffron-infused rajbhog milkshake",
    "Sapota Milkshake": "Thick sapota (chikoo) milkshake",
    "Sithaphal Milkshake": "Seasonal custard apple milkshake",
    "Strawberry Milkshake": "Fresh strawberry milkshake",
    "Tender Coconut Milkshake": "Tender coconut milkshake with coconut pieces",
    "Tooti Frooti Milkshake": "Colorful tutti frutti milkshake",
    "Vanilla Milkshake": "Classic vanilla milkshake",
    "Babycorn Chilli": "Crispy baby corn tossed in spicy chilli sauce",
    "Babycorn Manchurian": "Baby corn in tangy Indo-Chinese Manchurian sauce",
    "Gobi Chilli": "Cauliflower florets in spicy chilli sauce",
    "Gobi Manchurian": "Crispy cauliflower in tangy Manchurian sauce",
    "Masala French Fries": "Golden fries seasoned with spicy masala",
    "Mushroom Chilli": "Button mushrooms in fiery chilli sauce",
    "Mushroom Machurian": "Mushrooms in tangy Manchurian sauce",
    "Paneer 555": "Crispy paneer in signature 555 spice blend",
    "Paneer Basket": "Assorted paneer starters served in an edible basket",
    "Paneer Chilli": "Paneer cubes tossed in spicy chilli sauce",
    "Paneer Majestic": "Crispy paneer tossed with curry leaves and spices",
    "Paneer Manchurian": "Paneer balls in Indo-Chinese Manchurian sauce",
    "Salted French Fries": "Classic golden salted french fries",
    "Veg Allwin": "Mixed vegetable starter with chef's special seasoning",
    "Veg Basket": "Assorted vegetable fritters in a crispy basket",
    "Veg Bullet": "Spiced vegetable bullets, deep-fried crispy",
    "Veg Gulzar": "Mixed veg starter tossed in tangy Gulzar masala",
    "Veg Machurian": "Mixed vegetable balls in Manchurian sauce",
    "Gobi Noodles": "Stir-fried noodles with cauliflower and vegetables",
    "Mushroom Noodles": "Noodles tossed with mushrooms and veggies",
    "Paneer Noodles": "Noodles with soft paneer cubes and vegetables",
    "Veg Noodles": "Classic vegetable stir-fried noodles",
    "Chinese Fried Rice": "Indo-Chinese style vegetable fried rice",
    "Curd Rice": "Classic South Indian curd rice, cool and comforting",
    "Ghee Rice": "Aromatic basmati rice cooked in pure ghee",
    "Gobi Fried Rice": "Fried rice with cauliflower and vegetables",
    "Jeera Rice": "Fragrant basmati rice tempered with cumin seeds",
    "Kaju Rice": "Ghee rice topped with roasted cashew nuts",
    "Kashmir Pulao": "Mildly sweet Kashmiri pulao with dry fruits",
    "Mushroom Fried Rice": "Fried rice with button mushrooms",
    "Paneer Fried Rice": "Fried rice with soft paneer cubes",
    "Schezwan Fried Rice": "Spicy Schezwan-style vegetable fried rice",
    "Singapore Fried Rice": "Singapore-style fried rice with mixed veggies",
    "Tomato Fried Rice": "Tangy tomato-flavored fried rice",
    "Veg Fried Rice": "Classic vegetable fried rice",
    "Kajui Biryani": "Fragrant biryani loaded with roasted cashew nuts",
    "Mushroom Biryani": "Aromatic biryani with button mushrooms",
    "Paneer Biryani": "Dum biryani with soft paneer cubes and spices",
    "Veg Biryani": "Mixed vegetable dum biryani with aromatic spices",
    "Butter Phulka": "Soft phulka brushed with butter",
    "Butter Rumali Roti": "Thin rumali roti with butter finish",
    "Butter Wheat Phulka": "Whole wheat phulka with butter",
    "Phulka": "Soft puffed Indian bread",
    "Rumali Roti": "Paper-thin handkerchief roti",
    "Wheat Phulka": "Whole wheat puffed bread",
    "Kadai Paneer Masala": "Paneer cooked kadai-style with bell peppers",
    "Kadai Vegetable": "Mixed vegetables in kadai with capsicum",
    "Kaju Masala": "Rich cashew nut curry in creamy gravy",
    "Kaju Mushroom Masala": "Cashews and mushrooms in rich masala gravy",
    "Kaju Paneer Masala": "Cashews and paneer in creamy gravy",
    "Kaju Tomato Curry": "Cashew nuts in tangy tomato-based curry",
    "Kofta Curry": "Soft vegetable kofta balls in aromatic gravy",
    "Methi Chaman": "Fresh fenugreek and paneer in Kashmiri-style gravy",
    "Mushroom Masala": "Button mushrooms in rich masala gravy",
    "Palak Paneer": "Paneer cubes in smooth spinach gravy",
    "Paneer Butter Masala": "Paneer in creamy tomato-butter gravy — a classic",
    "Paneer Mushroom Masala": "Paneer and mushrooms in rich gravy",
    "Paneer Tikka Masala": "Grilled paneer tikka in smoky masala gravy",
    "Veg Doboja": "Mixed vegetables in Kohinoor's signature Doboja gravy",
    "Veg Mix Curry": "Mixed vegetables in mild and flavorful curry",
    "Pineapple": "Fresh cold-pressed pineapple juice",
    "Watermelon": "Refreshing cold-pressed watermelon juice",
    "Chocolate Pizza": "Sweet pizza with chocolate spread and toppings",
    "Corn Delight Pizza": "Pizza topped with sweet corn and cheese",
    "Double Cheese Pizza": "Extra cheesy pizza with double mozzarella",
    "Mushroom Pizza": "Pizza with sautéed mushrooms and cheese",
    "Paneer Delight": "Pizza loaded with paneer cubes and veggies",
    "Pepper Corn Veg": "Pizza with peppercorn seasoning and vegetables",
    "Pineapple Pizza": "Pizza with pineapple chunks and cheese",
    "Spicy Veg Pizza": "Pizza with spicy vegetables and jalapeños",
    "Veg Delight Pizza": "Pizza loaded with assorted vegetables",
    "Veg Pizza": "Classic vegetable pizza with mozzarella",
    "Veg Supreme Pizza": "Premium pizza with all veggies and extra cheese",
    "Aloo Tikki": "Crispy potato tikki burger with chutney",
    "Cheese Burger": "Veg burger with melted cheese slice",
    "Paneer Burger": "Paneer patty burger with fresh veggies",
    "Veg Burger": "Classic vegetable patty burger",
    "Chocolate": "Chocolate spread sandwich",
    "Corn Sandwich": "Grilled sandwich with sweet corn filling",
    "Paneer": "Grilled paneer sandwich with veggies",
    "Veg Cheese Sandwich": "Grilled veg sandwich with melted cheese",
    "Veg Sandwich": "Classic grilled vegetable sandwich",
}

# ── Unique Unsplash images ───────────────────────────────────────────────────────
_POOLS = {
    "icecream": [
        "https://images.unsplash.com/photo-1497034825429-c343d7c6a68f?w=800&q=80",
        "https://images.unsplash.com/photo-1501443762994-82bd5dace89a?w=800&q=80",
        "https://images.unsplash.com/photo-1560008581-09826d1de69e?w=800&q=80",
        "https://images.unsplash.com/photo-1488900128323-21503983a07e?w=800&q=80",
        "https://images.unsplash.com/photo-1570197571499-166b36435e9f?w=800&q=80",
        "https://images.unsplash.com/photo-1629385701021-fcd568a743e8?w=800&q=80",
        "https://images.unsplash.com/photo-1576506295286-5cda18df43e7?w=800&q=80",
        "https://images.unsplash.com/photo-1633933358116-a27b902fad35?w=800&q=80",
        "https://images.unsplash.com/photo-1514849302-984523450cf4?w=800&q=80",
        "https://images.unsplash.com/photo-1563805042-7684c019e1cb?w=800&q=80",
        "https://images.unsplash.com/photo-1551024506-0bccd828d307?w=800&q=80",
        "https://images.unsplash.com/photo-1580915411954-282cb1b0d780?w=800&q=80",
        "https://images.unsplash.com/photo-1579954115545-a95591f28bfc?w=800&q=80",
        "https://images.unsplash.com/photo-1557142046-c704a3adf364?w=800&q=80",
        "https://images.unsplash.com/photo-1505394033641-40c6ad1178d7?w=800&q=80",
        "https://images.unsplash.com/photo-1516559828984-fb3b99548b21?w=800&q=80",
        "https://images.unsplash.com/photo-1621303837174-89787a7d4729?w=800&q=80",
        "https://images.unsplash.com/photo-1587563871167-1ee9c731aefb?w=800&q=80",
        "https://images.unsplash.com/photo-1612203985729-70726954388c?w=800&q=80",
    ],
    "fruit_icecream": [
        "https://images.unsplash.com/photo-1615478503562-ec2d8aa0e24e?w=800&q=80",
        "https://images.unsplash.com/photo-1567206563064-6f60f40a2b57?w=800&q=80",
        "https://images.unsplash.com/photo-1543255006-d6395b6f1171?w=800&q=80",
        "https://images.unsplash.com/photo-1595348020949-87cdfbb44174?w=800&q=80",
        "https://images.unsplash.com/photo-1600002423562-975eabb78d5a?w=800&q=80",
        "https://images.unsplash.com/photo-1632170684742-9c8b38c1aeab?w=800&q=80",
        "https://images.unsplash.com/photo-1628607189631-96e9e8a3cedc?w=800&q=80",
        "https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=800&q=80",
        "https://images.unsplash.com/photo-1622483767028-3f66f32aef97?w=800&q=80",
        "https://images.unsplash.com/photo-1659432873335-3b5a6d7f1a4f?w=800&q=80",
        "https://images.unsplash.com/photo-1625869767142-1fb8faf7e8d9?w=800&q=80",
    ],
    "sundae": [
        "https://images.unsplash.com/photo-1563805042-7684c019e1cb?w=800&q=80",
        "https://images.unsplash.com/photo-1514849302-984523450cf4?w=800&q=80",
        "https://images.unsplash.com/photo-1488900128323-21503983a07e?w=800&q=80",
        "https://images.unsplash.com/photo-1580915411954-282cb1b0d780?w=800&q=80",
        "https://images.unsplash.com/photo-1579954115545-a95591f28bfc?w=800&q=80",
        "https://images.unsplash.com/photo-1557142046-c704a3adf364?w=800&q=80",
        "https://images.unsplash.com/photo-1551024506-0bccd828d307?w=800&q=80",
        "https://images.unsplash.com/photo-1505394033641-40c6ad1178d7?w=800&q=80",
        "https://images.unsplash.com/photo-1516559828984-fb3b99548b21?w=800&q=80",
        "https://images.unsplash.com/photo-1621303837174-89787a7d4729?w=800&q=80",
        "https://images.unsplash.com/photo-1587563871167-1ee9c731aefb?w=800&q=80",
        "https://images.unsplash.com/photo-1612203985729-70726954388c?w=800&q=80",
    ],
    "milkshake": [
        "https://images.unsplash.com/photo-1572490122747-3968b75cc699?w=800&q=80",
        "https://images.unsplash.com/photo-1541658016709-82535e94bc69?w=800&q=80",
        "https://images.unsplash.com/photo-1553787499-6f9133860278?w=800&q=80",
        "https://images.unsplash.com/photo-1568901839119-631418a3910d?w=800&q=80",
        "https://images.unsplash.com/photo-1619158401201-8fa932695178?w=800&q=80",
        "https://images.unsplash.com/photo-1579954115545-a95591f28bfc?w=800&q=80",
        "https://images.unsplash.com/photo-1626082927389-6cd097cdc6ec?w=800&q=80",
        "https://images.unsplash.com/photo-1615478503562-ec2d8aa0e24e?w=800&q=80",
        "https://images.unsplash.com/photo-1600002423562-975eabb78d5a?w=800&q=80",
        "https://images.unsplash.com/photo-1543255006-d6395b6f1171?w=800&q=80",
        "https://images.unsplash.com/photo-1595348020949-87cdfbb44174?w=800&q=80",
        "https://images.unsplash.com/photo-1632170684742-9c8b38c1aeab?w=800&q=80",
        "https://images.unsplash.com/photo-1628607189631-96e9e8a3cedc?w=800&q=80",
        "https://images.unsplash.com/photo-1622483767028-3f66f32aef97?w=800&q=80",
        "https://images.unsplash.com/photo-1659432873335-3b5a6d7f1a4f?w=800&q=80",
        "https://images.unsplash.com/photo-1625869767142-1fb8faf7e8d9?w=800&q=80",
        "https://images.unsplash.com/photo-1567206563064-6f60f40a2b57?w=800&q=80",
        "https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=800&q=80",
        "https://images.unsplash.com/photo-1497034825429-c343d7c6a68f?w=800&q=80",
        "https://images.unsplash.com/photo-1501443762994-82bd5dace89a?w=800&q=80",
        "https://images.unsplash.com/photo-1560008581-09826d1de69e?w=800&q=80",
        "https://images.unsplash.com/photo-1570197571499-166b36435e9f?w=800&q=80",
    ],
    "starter": [
        "https://images.unsplash.com/photo-1601050690597-df0568f70950?w=800&q=80",
        "https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=800&q=80",
        "https://images.unsplash.com/photo-1540189549336-e6e99c3679fe?w=800&q=80",
        "https://images.unsplash.com/photo-1567337710282-00832b415979?w=800&q=80",
        "https://images.unsplash.com/photo-1585032226651-759b368d7246?w=800&q=80",
        "https://images.unsplash.com/photo-1572715376701-98568319fd0b?w=800&q=80",
        "https://images.unsplash.com/photo-1559847844-5315695dadae?w=800&q=80",
        "https://images.unsplash.com/photo-1606574977732-e8e5f1f46c23?w=800&q=80",
        "https://images.unsplash.com/photo-1564834724105-918b73d1b8e0?w=800&q=80",
        "https://images.unsplash.com/photo-1512621776951-a57141f2eefd?w=800&q=80",
        "https://images.unsplash.com/photo-1574484284002-952d92456975?w=800&q=80",
        "https://images.unsplash.com/photo-1567620832903-9fc6debc209f?w=800&q=80",
        "https://images.unsplash.com/photo-1606728035253-49e8a23146de?w=800&q=80",
        "https://images.unsplash.com/photo-1599487488170-d11ec9c172f0?w=800&q=80",
        "https://images.unsplash.com/photo-1565299624946-b28f40a0ae38?w=800&q=80",
        "https://images.unsplash.com/photo-1573080496219-bb080dd4f877?w=800&q=80",
        "https://images.unsplash.com/photo-1585937421612-70a008356fbe?w=800&q=80",
        "https://images.unsplash.com/photo-1580217593608-61931cefc821?w=800&q=80",
    ],
    "noodles": [
        "https://images.unsplash.com/photo-1569718212165-3a8278d5f624?w=800&q=80",
        "https://images.unsplash.com/photo-1612929633738-8fe44f7ec841?w=800&q=80",
        "https://images.unsplash.com/photo-1585032226651-759b368d7246?w=800&q=80",
        "https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=800&q=80",
    ],
    "rice": [
        "https://images.unsplash.com/photo-1596560548464-f010549b84d7?w=800&q=80",
        "https://images.unsplash.com/photo-1516714435131-44d6b64dc6a2?w=800&q=80",
        "https://images.unsplash.com/photo-1536304993881-460587633ee1?w=800&q=80",
        "https://images.unsplash.com/photo-1512058564366-18510be2db19?w=800&q=80",
        "https://images.unsplash.com/photo-1645696301019-35adcc552067?w=800&q=80",
        "https://images.unsplash.com/photo-1603133872878-684f208fb84b?w=800&q=80",
        "https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=800&q=80",
        "https://images.unsplash.com/photo-1563379091339-03b21ab4a4f4?w=800&q=80",
        "https://images.unsplash.com/photo-1589302168068-964664d93dc0?w=800&q=80",
        "https://images.unsplash.com/photo-1631515243349-e0cb75fb8d3a?w=800&q=80",
        "https://images.unsplash.com/photo-1596797038530-2c107229654b?w=800&q=80",
        "https://images.unsplash.com/photo-1585937421612-70a008356fbe?w=800&q=80",
        "https://images.unsplash.com/photo-1574484284002-952d92456975?w=800&q=80",
    ],
    "biryani": [
        "https://images.unsplash.com/photo-1563379091339-03b21ab4a4f4?w=800&q=80",
        "https://images.unsplash.com/photo-1589302168068-964664d93dc0?w=800&q=80",
        "https://images.unsplash.com/photo-1631515243349-e0cb75fb8d3a?w=800&q=80",
        "https://images.unsplash.com/photo-1642821373181-16a5bc9f5801?w=800&q=80",
    ],
    "roti": [
        "https://images.unsplash.com/photo-1565557623262-b51c2513a641?w=800&q=80",
        "https://images.unsplash.com/photo-1600326145359-3a44909d1a39?w=800&q=80",
        "https://images.unsplash.com/photo-1574071318508-1cdbab80d002?w=800&q=80",
        "https://images.unsplash.com/photo-1586444248879-bc604bc77212?w=800&q=80",
        "https://images.unsplash.com/photo-1567620832903-9fc6debc209f?w=800&q=80",
        "https://images.unsplash.com/photo-1555939594-58d7cb561ad1?w=800&q=80",
    ],
    "curry": [
        "https://images.unsplash.com/photo-1565557623262-b51c2513a641?w=800&q=80",
        "https://images.unsplash.com/photo-1585937421612-70a008356fbe?w=800&q=80",
        "https://images.unsplash.com/photo-1631515243349-e0cb75fb8d3a?w=800&q=80",
        "https://images.unsplash.com/photo-1574484284002-952d92456975?w=800&q=80",
        "https://images.unsplash.com/photo-1455619452474-d2be8b1e70cd?w=800&q=80",
        "https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=800&q=80",
        "https://images.unsplash.com/photo-1596797038530-2c107229654b?w=800&q=80",
        "https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=800&q=80",
        "https://images.unsplash.com/photo-1567337710282-00832b415979?w=800&q=80",
        "https://images.unsplash.com/photo-1612929633738-8fe44f7ec841?w=800&q=80",
        "https://images.unsplash.com/photo-1576402187878-974f70c890a5?w=800&q=80",
        "https://images.unsplash.com/photo-1574653853027-5382a3d23a15?w=800&q=80",
        "https://images.unsplash.com/photo-1628294895950-9805252327bc?w=800&q=80",
        "https://images.unsplash.com/photo-1551881192-5e377f1b2142?w=800&q=80",
        "https://images.unsplash.com/photo-1573080496219-bb080dd4f877?w=800&q=80",
    ],
    "juice": [
        "https://images.unsplash.com/photo-1534353473418-4cfa6c56fd38?w=800&q=80",
        "https://images.unsplash.com/photo-1600271886742-f049cd451bba?w=800&q=80",
    ],
    "pizza": [
        "https://images.unsplash.com/photo-1565299624946-b28f40a0ae38?w=800&q=80",
        "https://images.unsplash.com/photo-1574071318508-1cdbab80d002?w=800&q=80",
        "https://images.unsplash.com/photo-1513104890138-7c749659a591?w=800&q=80",
        "https://images.unsplash.com/photo-1604382354936-07c5d9983bd3?w=800&q=80",
        "https://images.unsplash.com/photo-1588315029754-2dd089d39a1a?w=800&q=80",
        "https://images.unsplash.com/photo-1571407970349-bc81e7e96d47?w=800&q=80",
        "https://images.unsplash.com/photo-1593560708920-61dd98c46a4e?w=800&q=80",
        "https://images.unsplash.com/photo-1595854341625-f33ee10dbf94?w=800&q=80",
        "https://images.unsplash.com/photo-1528137871618-79d2761e3fd5?w=800&q=80",
        "https://images.unsplash.com/photo-1600628421060-939639517883?w=800&q=80",
        "https://images.unsplash.com/photo-1576458088443-04a19bb13da6?w=800&q=80",
    ],
    "burger": [
        "https://images.unsplash.com/photo-1568901346375-23c9450c58cd?w=800&q=80",
        "https://images.unsplash.com/photo-1550547660-d9450f859349?w=800&q=80",
        "https://images.unsplash.com/photo-1571091718767-18b5b1457add?w=800&q=80",
        "https://images.unsplash.com/photo-1586190848861-99aa4a171e90?w=800&q=80",
    ],
    "sandwich": [
        "https://images.unsplash.com/photo-1528735602780-2552fd46c7af?w=800&q=80",
        "https://images.unsplash.com/photo-1553909489-cd47e0907980?w=800&q=80",
        "https://images.unsplash.com/photo-1509722747041-616f39b57569?w=800&q=80",
        "https://images.unsplash.com/photo-1528736235302-52922df5c122?w=800&q=80",
        "https://images.unsplash.com/photo-1539252554453-80ab65ce3586?w=800&q=80",
        "https://images.unsplash.com/photo-1481070555726-e2fe8357725c?w=800&q=80",
    ],
}

_used: set[str] = set()

def _pick(cat: str) -> str:
    cat_lower = cat.lower()
    if "scoop" in cat_lower or cat_lower in ("icecream scoops",):
        pool = _POOLS["icecream"]
    elif "fruit" in cat_lower or cat_lower == "natural fruit based":
        pool = _POOLS["fruit_icecream"]
    elif "sundae" in cat_lower:
        pool = _POOLS["sundae"]
    elif "milkshake" in cat_lower:
        pool = _POOLS["milkshake"]
    elif "starter" in cat_lower:
        pool = _POOLS["starter"]
    elif "noodle" in cat_lower:
        pool = _POOLS["noodles"]
    elif "biryani" in cat_lower:
        pool = _POOLS["biryani"]
    elif "rice" in cat_lower or "pulao" in cat_lower:
        pool = _POOLS["rice"]
    elif "roti" in cat_lower or cat_lower == "rotis":
        pool = _POOLS["roti"]
    elif "curry" in cat_lower or "masala" in cat_lower or cat_lower == "curries":
        pool = _POOLS["curry"]
    elif "juice" in cat_lower:
        pool = _POOLS["juice"]
    elif "pizza" in cat_lower:
        pool = _POOLS["pizza"]
    elif "burger" in cat_lower:
        pool = _POOLS["burger"]
    elif "sandwich" in cat_lower:
        pool = _POOLS["sandwich"]
    else:
        pool = _POOLS["starter"]
    for url in pool:
        if url not in _used:
            _used.add(url)
            return url
    return pool[0]


def enrich():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    for ci in range(1, 5):
        c = ws.cell(row=1, column=ci)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center")

    for ci, h in [(5, "Description"), (6, "Subcategory"), (8, "Veg/Non-Veg"), (9, "Image URL")]:
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center")

    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=2).value
        cat = ws.cell(row=r, column=1).value
        if not item:
            continue
        item = str(item).strip()
        cat = str(cat).strip() if cat else ""

        ws.cell(row=r, column=5, value=DESC.get(item, f"Delicious {item}"))
        ws.cell(row=r, column=6, value=SUBCATEGORY.get(item, cat))
        ws.cell(row=r, column=8, value="Veg")  # Pure veg restaurant
        c = ws.cell(row=r, column=9, value=_pick(cat))
        c.font = Font(color="0563C1", underline="single", size=9)

    for col in range(1, 10):
        mx = max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row+1, 200))), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 3, 55)

    wb.save(XLSX_PATH)
    print(f"Enriched {XLSX_PATH.name}")


def build_payloads() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    payloads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat, name, price, hike = row[0], row[1], row[2], row[3]
        desc, subcat, veg_raw, img = row[4], row[5], row[7], row[8]
        if not name or price is None:
            continue
        hike_val = 0
        if hike is not None:
            try:
                h = float(hike)
                hike_val = round(h * 100, 1) if h < 1 else round(h, 1)
            except (ValueError, TypeError):
                hike_val = 0
        payloads.append({
            "name": str(name).strip(),
            "restaurantPrice": float(price),
            "hikePercentage": hike_val,
            "category": str(cat).strip() if cat else None,
            "subCategory": str(subcat).strip() if subcat else None,
            "isVeg": True,
            "isAvailable": True,
            "description": str(desc).strip() if desc else None,
            "image": [str(img).strip()] if img else [],
        })
    wb.close()
    return payloads


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--bearer-token", default=os.environ.get("HONESTEATS_BEARER_TOKEN"))
    parser.add_argument("--delay", type=float, default=0.3)
    args = parser.parse_args()

    enrich()
    payloads = build_payloads()
    print(f"\n{len(payloads)} items | Restaurant: {RESTAURANT_ID}\n")

    menu_url = f"{args.api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"

    for i, p in enumerate(payloads, 1):
        print(f"[{i:3d}] {p['name']}")
        print(f"      ₹{p['restaurantPrice']} + {p['hikePercentage']}% | {p['category']} > {p['subCategory']}")
        if not args.apply:
            print(f"      {json.dumps(p, indent=None, ensure_ascii=False)}")
        print()

    if not args.apply:
        print("=" * 60)
        print("DRY RUN — no HTTP calls.")
        print(f"To insert: python3 {sys.argv[0]} --apply")
        print("=" * 60)
        return 0

    s = requests.Session()
    s.headers["Content-Type"] = "application/json"
    if args.bearer_token:
        s.headers["Authorization"] = f"Bearer {args.bearer_token}"
    else:
        s.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS

    ok = err = 0
    for i, p in enumerate(payloads, 1):
        if args.delay > 0:
            time.sleep(args.delay)
        try:
            r = s.post(menu_url, json=p, timeout=60)
        except requests.RequestException as ex:
            err += 1; print(f"  ✗ [{i}] {p['name']}: {ex}", file=sys.stderr); continue
        if r.status_code == 201:
            ok += 1; print(f"  ✓ [{i}] {p['name']} created")
        else:
            err += 1; print(f"  ✗ [{i}] {p['name']}: HTTP {r.status_code} {r.text[:200]}", file=sys.stderr)
            if r.status_code == 401:
                print("\nUnauthorized.", file=sys.stderr); return 1

    print(f"\nDone: {ok} created, {err} failed.")
    return 0 if err == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
