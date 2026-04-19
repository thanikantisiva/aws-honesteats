#!/usr/bin/env python3
"""
Add pricing columns to Surabhi_DineIn_vs_Zomato.xlsx

Zomato commission 36% -> R_z = Z * 0.64
Our commission 22% -> restaurant P * 0.78, platform P * 0.22

P = (0.40 * Z + 0.60 * R_z) / 0.85, rounded to nearest ₹1

No Zomato price (blank / - / None): P = round(D * 1.29), Our Hike = 29% fixed.

Expected columns A–E:
  Category | Subcategory | Item | Dine-In (₹) | Zomato (₹)
"""
from __future__ import annotations

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

XLSX = "/Users/user/Downloads/Surabhi_DineIn_vs_Zomato.xlsx"

COL_D = 4  # Dine-In
COL_Z = 5  # Zomato

REST_ZOMATO = 0.64
REST_US = 0.78
OUR_COMM = 0.22
FIXED_HIKE_NO_Z = 0.29

HEADERS = [
    "Zomato Hike %",
    "Restaurant gets from Zomato (R_z)",
    "Our Recommended Price (P)",
    "Our Hike on Dine-in %",
    "Restaurant gets from Us",
    "Our Revenue (per order)",
    "Customer saves vs Zomato (₹)",
    "Customer saves vs Zomato %",
    "Restaurant gains vs Zomato (₹)",
    "Feasible?",
]


def parse_z(val) -> float | None:
    if val is None:
        return None
    if isinstance(val, str) and val.strip() in ("-", "", "None", "N/A", "n/a"):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def main() -> None:
    wb = load_workbook(XLSX)
    ws = wb.active

    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="2E86AB")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    gray = PatternFill("solid", fgColor="E7E6E6")
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    start_col = 6
    for i, h in enumerate(HEADERS):
        c = start_col + i
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = brd

    feasible_yes = feasible_no = na = 0

    for r in range(2, ws.max_row + 1):
        d_val = ws.cell(row=r, column=COL_D).value
        z_val = ws.cell(row=r, column=COL_Z).value
        if d_val is None:
            continue
        try:
            D = float(d_val)
        except (TypeError, ValueError):
            continue
        if D <= 0:
            continue

        Z = parse_z(z_val)

        if Z is not None:
            z_hike = (Z - D) / D * 100
            R_z = Z * REST_ZOMATO
            P = round((0.40 * Z + 0.60 * R_z) / 0.85)
            our_hike = (P - D) / D * 100
            rest_us = P * REST_US
            our_rev = P * OUR_COMM
            cust_save = Z - P
            cust_save_pct = (Z - P) / Z * 100 if Z > 0 else 0.0
            rest_gain = rest_us - R_z
            feasible = P < Z and rest_us > R_z
            feas_str = "Yes" if feasible else "No"
            if feasible:
                feasible_yes += 1
            else:
                feasible_no += 1

            row_vals = [
                round(z_hike, 1),
                round(R_z, 2),
                float(P),
                round(our_hike, 1),
                round(rest_us, 2),
                round(our_rev, 2),
                round(cust_save, 2),
                round(cust_save_pct, 1),
                round(rest_gain, 2),
                feas_str,
            ]
            feas_fill = green if feasible else red
            feas_color = "006100" if feasible else "9C0006"
        else:
            P = round(D * (1 + FIXED_HIKE_NO_Z))
            rest_us = P * REST_US
            our_rev = P * OUR_COMM
            row_vals = [
                "-",
                "-",
                float(P),
                29.0,
                round(rest_us, 2),
                round(our_rev, 2),
                "-",
                "-",
                "-",
                "N/A",
            ]
            feas_fill = gray
            feas_color = "000000"
            na += 1

        for i, val in enumerate(row_vals):
            c = start_col + i
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = brd
            cell.alignment = right if i != 9 else center
            if i == 9:
                cell.fill = feas_fill
                cell.font = Font(bold=True, color=feas_color)

    for col_letter, w in [
        ("F", 14),
        ("G", 26),
        ("H", 22),
        ("I", 18),
        ("J", 22),
        ("K", 20),
        ("L", 24),
        ("M", 22),
        ("N", 26),
        ("O", 12),
    ]:
        ws.column_dimensions[col_letter].width = w

    wb.save(XLSX)
    print(f"Saved: {XLSX}")
    print(f"With Zomato: Feasible Yes={feasible_yes}, No={feasible_no}")
    print(f"No Zomato (29% hike): {na}")


if __name__ == "__main__":
    main()
