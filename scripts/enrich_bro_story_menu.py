#!/usr/bin/env python3
"""
Enrich Bro Story menu Excel and emit menu-item JSON; optional POST to API.

Reads:  ~/Downloads/Bro story menu_category_subcategory_item_price_inr.xlsx
        Cols A–F: Category | Subcategory | Item | Description | Price (₹) | Our Hike on Dine-in %

Writes / refreshes:
  D — Description (only when currently blank / '-' / missing)
  G — (blank spacer)
  H — isVeg (TRUE / FALSE)
  I — image (single Unsplash URL per row; no repeats)

Writes JSON → ~/Downloads/Bro_Story_Menu.menu_items.json
Restaurant ID: RES-1776586021827-3708

Usage:
  python3 scripts/enrich_bro_story_menu.py                     # enrich + JSON only (default, dry-run)
  python3 scripts/enrich_bro_story_menu.py --apply             # enrich + POST to API
  python3 scripts/enrich_bro_story_menu.py --from-json --apply # POST from existing JSON
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

XLSX_PATH = Path.home() / "Downloads" / "Bro story menu_category_subcategory_item_price_inr.xlsx"
JSON_PATH = Path.home() / "Downloads" / "Bro_Story_Menu.menu_items.json"
RESTAURANT_ID = "RES-1776586021827-3708"

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get(
    "HONESTEATS_RETOOL_BYPASS",
    "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc",
)
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

COL_CAT, COL_SUB, COL_NAME, COL_DESC, COL_PRICE, COL_HIKE = 1, 2, 3, 4, 5, 6
COL_GAP, COL_VEG, COL_IMG = 7, 8, 9

_BRAND = "Bro Story"

# ── Unsplash photo-ID pools (stable IDs; unique across pools enforced below) ──
_POOLS: dict[str, list[str]] = {
    "fried_chicken": [
        "1562967914-608f82629710", "1626082927389-6cd097cdc6ec",
        "1513185041617-8ab03f83d6c5", "1606755962773-d324e0a13086",
        "1604908176997-125f25cc6f3d", "1598515214211-89d3c73ae83b",
        "1587593810167-a84920ea0781", "1608039755401-742074f0548d",
        "1562967916-eb82221dfb92", "1619221882220-947b3d3c8861",
    ],
    "chicken_wings": [
        "1527477396000-e27163b481c2", "1608039755401-742074f0548d",
        "1567620832903-9fc6debc209f", "1624726175512-19b9baf9fbd1",
        "1610057099431-d73a1c9d2f2f",
    ],
    "chicken_popcorn": [
        "1562967914-608f82629710", "1632778149955-e80f8ceca2e8",
        "1565680018093-ebb6e3062e7b", "1559339352-11d035aa65de",
    ],
    "chicken_strips": [
        "1606755962773-d324e0a13086", "1604908176997-125f25cc6f3d",
        "1553621042-f6e147245754",
    ],
    "nashville_hot": [
        "1562967915-6d45f4b8f3c8", "1619221882220-947b3d3c8861",
        "1610057099431-d73a1c9d2f2f",
    ],
    "burger_veg": [
        "1571091718767-18b5b1457add", "1586190848861-99aa4a171e90",
        "1565299585323-38d6b0865b47", "1520072959219-c595dc870360",
        "1606131731446-5568d87113aa",
    ],
    "burger_nonveg": [
        "1568901346375-23c9450c58cd", "1550317138-10000687a72b",
        "1551782450-a2132b4ba21d", "1571091655789-405127d2c2e1",
        "1561758033-d89a9ad46330", "1553979459-d2229ba7433b",
    ],
    "sandwich_veg": [
        "1539252554453-80ab65ce3586", "1528735602780-2552fd46c7af",
        "1509722747041-616f39b57569", "1521390188846-e2a3a97453a0",
        "1519864600395-3404e40a0eda",
    ],
    "sandwich_nonveg": [
        "1528735602780-2552fd46c7af", "1521390188846-e2a3a97453a0",
        "1504113888036-c19ddbbe3e10", "1567234669003-dce7a7a88821",
    ],
    "wrap_veg": [
        "1600891964092-4316c288032e", "1626700051175-f5d1f5e9c07a",
        "1565299507177-b0ac66763828",
    ],
    "wrap_nonveg": [
        "1563379091339-03b21ab4a4f4", "1626700051175-f5d1f5e9c07a",
        "1626700051175-5f7de4c7a776", "1609167830220-7164aa360951",
    ],
    "fries": [
        "1541592106381-b31e9677c0e5", "1576107232684-1279f390859f",
        "1573080496219-bb080dd4f877", "1585109649139-366815a0d713",
        "1630384060421-cb20d0e0649d",
    ],
    "dips": [
        "1607013251379-e6eecfffe234", "1572441713132-c542fc4fe282",
        "1608500218890-c4f9077f79a1", "1528735602780-2552fd46c7af",
        "1546548970-71785318a17b",
    ],
    "mojito": [
        "1536935338788-846bb9981813", "1541658016709-82535e94bc69",
        "1572490122747-3968b75cc699", "1534353473418-4cfa6c56fd38",
        "1625772299848-391b6a87d7b3", "1600271886742-f049cd451bba",
        "1553787499-6f9133860278",
    ],
    "waffles": [
        "1562376552-0d160a2f238d", "1504754524776-8f4f37790ca0",
        "1551024601-bec78aea704b", "1558030006-450675393462",
        "1464195244916-ccda7a04b9c8",
    ],
    "thickshake": [
        "1568644396922-5c3bfae12521", "1572490122747-3968b75cc699",
        "1551024739-22082dc40f9f", "1625772299848-391b6a87d7b3",
        "1553787499-6f9133860278", "1517959105821-eaf2591984ca",
    ],
    "milkshake": [
        "1572490122747-3968b75cc699", "1551024601-bec78aea704b",
        "1568644396922-5c3bfae12521", "1534353473418-4cfa6c56fd38",
        "1541658016709-82535e94bc69",
    ],
    "combo_veg": [
        "1565299507177-b0ac66763828", "1571091718767-18b5b1457add",
        "1586190848861-99aa4a171e90",
    ],
    "combo_nonveg": [
        "1571091655789-405127d2c2e1", "1568901346375-23c9450c58cd",
        "1550317138-10000687a72b", "1513104890138-7c749659a591",
    ],
}


def _photo_url(pid: str, sig: int | None = None) -> str:
    base = f"https://images.unsplash.com/photo-{pid}?auto=format&fit=crop&w=800&q=80"
    return base if sig is None else f"{base}&sig={sig}"


_RX = re.compile(r"^\d{6,}-[0-9a-f]+$")
for _k, _lst in list(_POOLS.items()):
    _POOLS[_k] = [p for p in _lst if _RX.match(p)]

_seen_ids: set[str] = set()
_GLOBAL: list[str] = []
for _pn in list(_POOLS.keys()):
    for _pid in _POOLS[_pn]:
        if _pid not in _seen_ids:
            _seen_ids.add(_pid)
            _GLOBAL.append(_pid)


class _Picker:
    """Deduplicates photo IDs across rows; falls back to signed variants if exhausted."""

    def __init__(self) -> None:
        self.used: set[str] = set()
        self._sig = 0

    def _themed(self, key: str) -> str | None:
        for pid in _POOLS.get(key, []):
            if pid not in self.used:
                self.used.add(pid)
                return _photo_url(pid)
        return None

    def _global(self) -> str | None:
        for pid in _GLOBAL:
            if pid not in self.used:
                self.used.add(pid)
                return _photo_url(pid)
        return None

    def _recycled(self) -> str:
        self._sig += 1
        pid = _GLOBAL[self._sig % len(_GLOBAL)]
        return _photo_url(pid, sig=self._sig)

    def pick(self, key: str) -> str:
        return self._themed(key) or self._global() or self._recycled()


# ── Veg / non-veg ────────────────────────────────────────────────────────────
_NV_KW = (
    "chicken", "ckn", "egg", "mutton", "fish", "prawn", "meat",
    "non veg", "non-veg", "nonveg", "nashville", "indo american",
    "madmax", "juicy",
)
_VEG_OVERRIDE = ("paneer", "veg ", "veg_", "mojito")


def _is_nonveg(cat: str, sub: str, name: str, desc: str = "") -> bool:
    cl = cat.lower()
    nl = name.lower()
    if cl.startswith("veg ") or "veg wrap" in nl or "bro veg" in nl:
        return False
    if cl == "mojitos" or cl == "waffles" or cl == "thickshakes" or cl == "milkshakes":
        return False
    if cl == "dips":
        return False
    base_text = f"{cat} {sub} {name}".lower()
    if any(k in base_text for k in _NV_KW):
        return True
    # Only consult description for ambiguous cases (combos / meals).
    if desc and ("combo" in cl or "meal" in nl or "meal" in cl):
        if any(k in desc.lower() for k in _NV_KW):
            return True
    return False


# ── Pool selection ────────────────────────────────────────────────────────────
def select_pool(cat: str, sub: str, name: str) -> str:
    cl, sl, nl = cat.lower(), (sub or "").lower(), name.lower()
    is_nv = _is_nonveg(cat, sub, name)

    if "combo" in cl:
        return "combo_nonveg" if is_nv else "combo_veg"
    if "waffle" in cl:
        return "waffles"
    if "thickshake" in cl:
        return "thickshake"
    if "milkshake" in cl:
        return "milkshake"
    if "mojito" in cl or "mojito" in nl:
        return "mojito"
    if "fries" in cl or "fries" in nl:
        return "fries"
    if "dip" in cl or "mayo" in nl or "sauce" in nl or "dip" in nl or "cheese" in nl and cat.lower() == "dips":
        return "dips"
    if "wrap" in cl or "wrap" in nl:
        return "wrap_nonveg" if is_nv else "wrap_veg"
    if "sandwich" in cl:
        return "sandwich_nonveg" if is_nv else "sandwich_veg"
    if "burger" in cl:
        return "burger_nonveg" if is_nv else "burger_veg"
    if "nashville" in cl or "nashville" in nl:
        return "nashville_hot"
    if "popcorn" in nl:
        return "chicken_popcorn"
    if "wing" in nl:
        return "chicken_wings"
    if "strip" in nl:
        return "chicken_strips"
    if "chicken" in cl or "chicken" in nl or is_nv:
        return "fried_chicken"
    return "burger_veg"


# ── Description templates (used only when col D is missing / blank / '-') ─────
def build_description(cat: str, sub: str, name: str, is_veg: bool) -> str:
    veg = "vegetarian" if is_veg else "non-vegetarian"
    cl = cat.lower()
    sl = (sub or "").lower()
    nl = name.lower()

    if "fried chicken" in cl:
        if "wing" in nl:
            return f"{name} — crispy, juicy fried chicken wings seasoned in our signature mix. {_BRAND}."
        if "strip" in nl:
            return f"{name} — boneless fried chicken strips with a crunchy coating. {_BRAND}."
        if "popcorn" in nl:
            return f"{name} — bite-size popcorn chicken, fried golden and crispy. {_BRAND}."
        return f"{name} — classic crispy fried chicken, juicy inside and crunchy outside. {_BRAND}."

    if "indo american" in cl:
        return f"{name} — chicken dipped in spiced butter and coated with bold Indo-American seasoning. {_BRAND}."

    if "nashville" in cl:
        return f"{name} — fiery Nashville-style hot chicken coated in signature chilli oil. {_BRAND}."

    if "wrap" in cl:
        return f"{name} — soft tortilla wrap loaded with fresh fillings and house sauces. {_BRAND}."

    if "burger" in cl:
        cheese = "with melty cheese" if "with cheese" in sl else "without cheese" if "without cheese" in sl else ""
        return f"{name} — hand-assembled burger {cheese} with mayo, onion, and {_BRAND}'s signature sauces.".replace("  ", " ")

    if "sandwich" in cl:
        cheese = "with cheese" if "with cheese" in sl else "without cheese" if "without cheese" in sl else ""
        return f"{name} — toasted sandwich {cheese}, packed with flavourful filling and mayo. {_BRAND}.".replace("  ", " ")

    if "fries" in cl:
        if "peri" in nl:
            return f"{name} — crispy fries tossed in peri-peri seasoning for a spicy kick. {_BRAND}."
        if "loaded" in nl:
            return f"{name} — fries loaded with chicken, cheese, and house sauces. {_BRAND}."
        return f"{name} — golden, freshly fried salted potato fries. {_BRAND}."

    if "dip" in cl or cat.lower() == "dips":
        return f"{name} — creamy, flavour-packed dip to pair with fries, burgers and chicken. {_BRAND}."

    if "mojito" in cl:
        return f"{name} — refreshing mojito mocktail with mint, lime, and a fruity twist. {_BRAND}."

    if "waffle" in cl:
        return f"{name} — crisp golden waffle loaded with indulgent toppings. {_BRAND}."

    if "thickshake" in cl:
        return f"{name} — extra-thick blended shake packed with ice cream and toppings. {_BRAND}."

    if "milkshake" in cl:
        return f"{name} — smooth, creamy milkshake in your favourite flavour. {_BRAND}."

    if "combo" in cl:
        return f"{name} — value combo bringing together our signature favourites. {_BRAND}."

    return f"{name} — a {veg} favourite from the {_BRAND} kitchen."


def _has_description(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return bool(s) and s != "-"


def parse_num(val) -> float | int:
    if val is None or val == "" or val == "-":
        return 0
    if isinstance(val, (int, float)):
        return int(val) if float(val) == int(val) else round(float(val), 2)
    try:
        x = float(str(val).strip().rstrip("%"))
        return int(x) if x == int(x) else round(x, 2)
    except ValueError:
        return 0


# ── Main enrichment ──────────────────────────────────────────────────────────
def enrich_workbook() -> list[dict]:
    if not XLSX_PATH.is_file():
        raise SystemExit(f"Missing input workbook: {XLSX_PATH}")

    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    hdr_fill = PatternFill("solid", fgColor="2E86AB")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col, label in (
        (COL_DESC, "Description"),
        (COL_VEG, "isVeg"),
        (COL_IMG, "image"),
    ):
        c = ws.cell(1, col, label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    picker = _Picker()
    payloads: list[dict] = []

    last_cat = ""
    last_sub = ""
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, COL_NAME).value
        if name is None or str(name).strip() == "":
            continue

        cat_raw = ws.cell(r, COL_CAT).value
        sub_raw = ws.cell(r, COL_SUB).value
        cat = str(cat_raw).strip() if cat_raw not in (None, "") else last_cat
        sub = str(sub_raw).strip() if sub_raw not in (None, "") else last_sub
        if cat_raw not in (None, ""):
            last_cat = cat
            last_sub = ""  # new category resets subcat
        if sub_raw not in (None, ""):
            last_sub = sub
        if sub == "-":
            sub = ""

        name_s = str(name).strip()
        price = ws.cell(r, COL_PRICE).value
        hike = ws.cell(r, COL_HIKE).value

        desc_cell = ws.cell(r, COL_DESC).value
        desc_existing = str(desc_cell).strip() if _has_description(desc_cell) else ""
        is_veg = not _is_nonveg(cat, sub, name_s, desc_existing)

        if desc_existing:
            desc = desc_existing
        else:
            desc = build_description(cat, sub, name_s, is_veg)
            ws.cell(r, COL_DESC, desc).alignment = wrap

        img_url = picker.pick(select_pool(cat, sub, name_s))

        ws.cell(r, COL_GAP, None)
        ws.cell(r, COL_VEG, bool(is_veg)).alignment = center
        ws.cell(r, COL_IMG, img_url).alignment = wrap

        payloads.append(
            {
                "name": name_s,
                "restaurantPrice": parse_num(price),
                "hikePercentage": parse_num(hike),
                "category": cat,
                "subCategory": sub,
                "isVeg": bool(is_veg),
                "isAvailable": True,
                "description": desc,
                "image": [img_url],
                "restaurantId": RESTAURANT_ID,
            }
        )

    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 64

    wb.save(XLSX_PATH)
    JSON_PATH.write_text(
        json.dumps(payloads, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    print(f"Enriched rows: {len(payloads)}")
    print(f"Saved workbook: {XLSX_PATH}")
    print(f"Saved JSON:     {JSON_PATH}")
    print(f"Unique images:  {len({p['image'][0] for p in payloads})}")
    return payloads


# ── Optional API import (disabled by default) ─────────────────────────────────
def menu_item_api_body(row: dict) -> dict:
    cat = (row.get("category") or "").strip() or None
    sub = (row.get("subCategory") or "").strip() or None
    desc = row.get("description")
    if isinstance(desc, str):
        desc = desc.strip() or None
    return {
        "name": row["name"],
        "restaurantPrice": float(row["restaurantPrice"]),
        "hikePercentage": float(row["hikePercentage"]),
        "category": cat,
        "subCategory": sub,
        "isVeg": row.get("isVeg", True),
        "isAvailable": row.get("isAvailable", True),
        "description": desc,
        "image": row.get("image") or [],
    }


def post_menu_items(
    payloads: list[dict],
    *,
    api_url: str,
    bearer_token: str | None,
    delay_sec: float,
) -> int:
    menu_url = f"{api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"
    session = requests.Session()
    session.headers["Content-Type"] = "application/json"
    if bearer_token:
        session.headers["Authorization"] = f"Bearer {bearer_token}"
    else:
        session.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS

    ok = err = 0
    for i, raw in enumerate(payloads, 1):
        body = menu_item_api_body(raw)
        if delay_sec > 0:
            time.sleep(delay_sec)
        try:
            r = session.post(menu_url, json=body, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"  x [{i}] {body['name']}: {ex}", file=sys.stderr)
            continue
        if r.status_code == 201:
            ok += 1
            print(f"  ok [{i}] {body['name']}")
        else:
            err += 1
            print(
                f"  x [{i}] {body['name']}: HTTP {r.status_code} {r.text[:300]}",
                file=sys.stderr,
            )
            if r.status_code == 401:
                print(
                    "\nUnauthorized: set HONESTEATS_RETOOL_BYPASS or HONESTEATS_BEARER_TOKEN.",
                    file=sys.stderr,
                )
                return 1

    print(f"\nPOST done: {ok} created, {err} failed. URL base: {api_url.rstrip('/')}")
    return 0 if err == 0 else 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich Bro Story menu + optional API import")
    parser.add_argument("--apply", action="store_true", help="POST each item to the API")
    parser.add_argument("--from-json", action="store_true", help="Load payloads from JSON instead of re-enriching")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--bearer-token", default=os.environ.get("HONESTEATS_BEARER_TOKEN"))
    parser.add_argument("--delay", type=float, default=0.3)
    args = parser.parse_args()

    if args.from_json:
        if not JSON_PATH.is_file():
            print(f"Missing {JSON_PATH}", file=sys.stderr)
            return 1
        payloads = json.loads(JSON_PATH.read_text(encoding="utf-8"))
        print(f"Loaded {len(payloads)} items from {JSON_PATH.name}")
    else:
        payloads = enrich_workbook()

    if not args.apply:
        print("No POST (pass --apply to import).")
        return 0

    return post_menu_items(
        payloads,
        api_url=args.api_url,
        bearer_token=args.bearer_token,
        delay_sec=args.delay,
    )


if __name__ == "__main__":
    raise SystemExit(main())
