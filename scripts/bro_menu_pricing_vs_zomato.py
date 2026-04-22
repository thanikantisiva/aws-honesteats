"""
Populate pricing comparison columns on Bro story menu xlsx.

Input: ~/Downloads/Bro story menu_category_subcategory_item_price_inr.xlsx
Columns E = Dine-in (D), F = Zomato (Z).

Zomato commission 39% → restaurant gets 61% of Z.
Our commission 25% → restaurant gets 75% of our listed price P.

With Z:
  Zomato Hike % = (Z - D) / D * 100
  R_z = Z * 0.61
  P = round((0.40 * Z + 0.60 * R_z) / 0.85)
  (Constraints checked: P < Z and 0.75 * P > R_z)

Without Z:
  P = round(D * 1.29)  # ~29% hike on dine-in
  R_z / Zomato columns left blank; customer savings vs Zomato blank.

Run: python3 scripts/bro_menu_pricing_vs_zomato.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path.home() / "Downloads" / "Bro story menu_category_subcategory_item_price_inr.xlsx"

NEW_HEADERS = [
    "Zomato Hike %",
    "Restaurant gets from Zomato R_z (₹)",
    "Our Recommended Price P (₹)",
    "Our Hike on Dine-in %",
    "Restaurant gets from Us (₹)",
    "Our Revenue (per order) (₹)",
    "Customer saves vs Zomato (₹)",
    "Customer saves vs Zomato %",
    "Restaurant gains vs Zomato (₹)",
]


def _num(x) -> float | None:
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip()
        if not s or s == "-":
            return None
        if s.startswith("="):
            return None
        try:
            return float(s)
        except ValueError:
            return None
    if isinstance(x, (int, float)):
        return float(x)
    return None


def _zomato_price(ws, r: int, D: float) -> float | None:
    """Column F: numeric Z, or formula '=En*1.32' (Zomato = 32% markup on dine-in)."""
    v = ws.cell(r, 6).value
    z = _num(v)
    if z is not None and z > 0:
        return z
    if isinstance(v, str):
        s = v.strip().replace(" ", "")
        if s.upper().startswith("=E") and "*1.32" in s:
            return D * 1.32
    return None


def compute_p_with_z(Z: float, D: float) -> tuple[int, float]:
    R_z = Z * 0.61
    P = round((0.40 * Z + 0.60 * R_z) / 0.85)
    if not (P < Z and 0.75 * P > R_z):
        for delta in (0, -1, 1, -2, 2, -3, 3):
            cand = P + delta
            if cand > 0 and cand < Z and 0.75 * cand > R_z:
                P = cand
                break
        else:
            raise ValueError(f"No P satisfies constraints for Z={Z}, D={D}")
    return P, R_z


def main() -> None:
    wb = load_workbook(SRC)
    ws = wb.active
    start_col = 7  # G

    # Write headers
    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for i, title in enumerate(NEW_HEADERS):
        c = start_col + i
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    n_z = 0
    n_no_z = 0
    for r in range(2, ws.max_row + 1):
        D = _num(ws.cell(r, 5).value)
        if D is None or D <= 0:
            continue

        Z = _zomato_price(ws, r, D)

        c0 = start_col
        if Z is not None and Z > 0:
            n_z += 1
            R_z = Z * 0.61
            P, _ = compute_p_with_z(Z, D)
            z_hike = (Z - D) / D * 100
            us_hike = (P - D) / D * 100
            rest_us = 0.75 * P
            rev = 0.25 * P
            save = Z - P
            save_pct = (Z - P) / Z * 100
            rest_gain = rest_us - R_z

            ws.cell(r, c0 + 0, value=round(z_hike, 2))
            ws.cell(r, c0 + 1, value=round(R_z, 2))
            ws.cell(r, c0 + 2, value=P)
            ws.cell(r, c0 + 3, value=round(us_hike, 2))
            ws.cell(r, c0 + 4, value=round(rest_us, 2))
            ws.cell(r, c0 + 5, value=round(rev, 2))
            ws.cell(r, c0 + 6, value=round(save, 2))
            ws.cell(r, c0 + 7, value=round(save_pct, 2))
            ws.cell(r, c0 + 8, value=round(rest_gain, 2))
        else:
            n_no_z += 1
            P = round(D * 1.29)
            us_hike = 29.0  # fixed display per spec
            rest_us = 0.75 * P
            rev = 0.25 * P

            ws.cell(r, c0 + 0, value=None)
            ws.cell(r, c0 + 1, value=None)
            ws.cell(r, c0 + 2, value=P)
            ws.cell(r, c0 + 3, value=us_hike)
            ws.cell(r, c0 + 4, value=round(rest_us, 2))
            ws.cell(r, c0 + 5, value=round(rev, 2))
            ws.cell(r, c0 + 6, value=None)
            ws.cell(r, c0 + 7, value=None)
            ws.cell(r, c0 + 8, value=None)

    for i in range(len(NEW_HEADERS)):
        ws.column_dimensions[get_column_letter(start_col + i)].width = 22 if i != 2 else 14

    wb.save(SRC)
    print(f"Wrote {len(NEW_HEADERS)} columns to {SRC}")
    print(f"Rows with Z: {n_z}, without Z: {n_no_z}")


if __name__ == "__main__":
    main()
