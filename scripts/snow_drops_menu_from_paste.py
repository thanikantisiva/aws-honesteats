#!/usr/bin/env python3
"""One-off: build Snow Drops style menu xlsx from embedded TSV rows."""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path.home() / "Downloads" / "Snow_Drops_Menu.xlsx"

# Tab-separated: Item, Option1, Price1, Option2, Price2 (use "" for empty, "na" as literal from menu)
ROWS_RAW = r"""Belgium Chocolate Shake	Milk shake	180	Thick shake	240
Berry Blast Shake	Milk shake	180	Thick shake	240
Black Current Oreo Shake	Milk shake	180	Thick shake	240
Blue Berry Shake	Milk shake	170	Thick shake	230
Brownie Oreo Shake	Milk shake	180	Thick shake	240
Candy Delight Shake	Milk shake	180	Thick shake	240
Candy Shake Shake	Milk shake	180	Thick shake	240
Caramel Coffee Shake	Milk shake	180	Thick shake	240
Choco Blast Shake	Milk shake	180	Thick shake	240
Chocolate Brownie Shake	Milk shake	180	Thick shake	240
Chocolate Shake	Milk shake	180	Thick shake	240
Classic Cold Coffee Shake	Milk shake	180	Thick shake	240
Coffee Brownie Shake	Milk shake	180	Thick shake	240
Ferrerocher Shake	Milk shake	299	na	na
Fiber Blast Shake	Milk shake	180	Thick shake	240
Five Star Shake	Milk shake	180	Thick shake	240
Gems Shake	Milk shake	180	Thick shake	240
Green Apple Shake	Milk shake	180	Thick shake	240
Kit Kat Shake	Milk shake	180	Thick shake	240
Lichi Shake	Milk shake	180	Thick shake	240
Mango Shake	Milk shake	180	Thick shake	240
Nutella Brownie Shake	Milk shake	180	Thick shake	240
Nutella Chocolate Shake	Milk shake	180	Thick shake	240
Orea Nutella Shake	Milk shake	180	Thick shake	240
Oreo Caramel Shake	Milk shake	180	Thick shake	240
Oreo Coffee Shake	Milk shake	180	Thick shake	240
Oreo Shake	Milk shake	180	Thick shake	240
Peanut Butter Oreo Shake	Milk shake	180	Thick shake	240
Peanut Butter Shake	Milk shake	180	Thick shake	240
Pine Apple Shake	Milk shake	180	Thick shake	240
Rose Milk Shake	Milk shake	180	Thick shake	240
Sitaphal Shake	Milk shake	230	Thick shake	299
Snickers Shake	Milk shake	180	Thick shake	240
Strawberry Chocolate Shake	Milk shake	180	Thick shake	240
Strawberry Kiwi Shake	Milk shake	180	Thick shake	240
Strawberry Shake	Milk shake	180	Thick shake	240
Vanilla Shake	Milk shake	180	Thick shake	240
American Nuts Scoops		80	na	
Anjeer Badam Scoops		80	na	
Belgium Dark Chocolate Scoops		80	na	
Black Current Scoops		80	na	
Caramel Brownie Scoops		80	na	
Caramel Nuts Scoops		80	na	
Choco Chips Scoops		90	na	
Chocolate Scoops		80	na	
Dry Fruits Scoops		80	na	
Fig And Honey		80	na	
Kesar Pista Scoops		80	na	
Litchi Scoops		80	na	
Mango Scoops		80	na	
Musk Melon Scoops		80	na	
Sitaphal Scoops		90	na	
Strawberry Scoops		70	na	
Tender Coconut Scoops		90	na	
Totty Fruity		70	na	
Vanilla Scoops		70	na	
Best In Dark Chocolate	small	180	large	240
Chocolate Over Load	small	180	large	240
Chocolava @250		299	na	
Ferrero Rocher @250		299	na	
Nutella Choco Brownie	small	180	large	240
Oreo Cookies Crunch	small	180	large	240
Black Current Nuts	small	180	large	240
Candy Delight	small	180	large	240
Dry Fruit Delight	small	180	large	240
Nutty Caramel Brownie	small	180	large	240
Nutty Overload	small	180	large	240
Peanut Butter	small	180	large	240
Brownie With Fudge		150	na	
Brownie With Vanilla		199	na	
Death By Chocolate		240	na	
American Nuts Family Packs		320	na	
Belgium Dark Chocolate Family Packs		399	na	
Caramel Brownie Family Packs		399	na	
Caramel Nuts Family Packs		399	na	
Chocolate Family Packs		280	na	
Dry Fruit Delight Family Packs		399	na	
Dry Fruit Family Packs		399	na	
Honey Almond Family Packs		399	na	
Honeymoon Delight Family Packs		399	na	
Mango Family Packs		290	na	
Sitaphal Family Packs		290	na	
Strawberry Family Packs		230	na	
Vanilla Family Packs		230	na	
Blue Angel Mocktail		79	na	
Green Apple Mocktail		79	na	
Strawberry Mocktail		79	na	
Virgin Mojito		79	na	
Cheesy French Fries		120	na	
Peri-peri French Fries		100	na	
Regular French Fries		80	na	
Crispy		120	na	
Paneer Popcorn		150	na	
Veg Fingers		150	na	
Veg Nuggets		120	na	
Veg Smilies		100	na	
Chicken Fingers		170	na	
Chicken Leg Piece (1pc)		75	na	
Chicken Lolly Pop (1pc)		45	na	
Chicken Nuggets		150	na	
Chicken Popcorn		150	na	
Chicken Tandoori Burger		99	na	
Chicken Tikka Burger		99	na	
Classic Chicken Burger		89	na	
Snow Drops Special Chicken Burger		110	na	
Classic Veg Burger		69	na	
Fried Paneer Burger		99	na	
Snow Drops Special Veg Burger		110	na	
Tandoori Paneer Burger		99	na	
Chicken Cheese Sandwich		89	na	
Chicken Tandoori Sandwich		99	na	
Chicken Tikka Sandwich		99	na	
Snow Drops Special Chicken Sandwich		110	na	
Corn Sandwich		89	na	
Snow Drops Special Veg Sandwich		99	na	
Tandoori Paneer Sandwich		89	na	
Veg Club Sandwich		79	na	
Double Cheese Pizza		159	na	
Mushroom Veg Pizza		159	na	
Peri-peri Veg Pizza		179	na	
Plain Cheese Pizza		99	na	
Schezwan Veg And Corn Pizza		159	na	
Snow Drops Special Veg Pizza		199	na	
Sweet Corn Pizza		159	na	
Tandoori Paneer Pizza		179	na	
Veg Bbq Pizza		179	na	
Veg Blast Pizza		179	na	
Veg Supreme Pizza		179	na	
Chicken Bbq Pizza		199	na	
Chicken Blast Pizza		159	na	
Chicken Cheese Pizza		149	na	
Chicken Double Cheese Pizza		159	na	
Chicken Peri-peri Pizza		199	na	
Chicken Tandoori Pizza		159	na	
Chicken Tikka Pizza		199	na	
Mexican Chicken Pizza		199	na	
Snow Drops Special Chicken Pizza		210	na	
2 Chicken Leg Piece		169	na	
4 Chicken Leg Piece		329	na	
6 Chicken Leg Piece		489	na	
2 Piece Chicken Lollypop		80	na	
4 Piece Chicken Lollypop		140	na	
6 Piece Chicken Lollypop		200	na	
8 Piece Chicken Lollypop		260	na	
Large Bucket (3 legpiece + 3 lollypop + 4 wings)		359	na	
Medium Bucket (2 legpiece + 2 lollypop + 2 wings)		249	na	
Small Bucket (1 legpiece + 2 lollypop + 2 wings)		178	na	
Snow Drops Special Bucket (4 legpiece + 4 lollypop + 4 wings)		479	na	
combos				
any chicken burger + fries + milkshake		300	na	
any chicken burger + fries + mocktail		250	na	
any chicken burger + fries + thickshake		340	na	
any veg burger + fries + milkshake		270	na	
any veg burger + fries + mocktail		230	na	
any veg burger + fries + thickshake		330	na	
any chicken sandwich + fries + milkshake		289	na	
any chicken sandwich + fries + mocktail		249	na	
any chicken sandwich + fries + thickshake		320	na	
any veg sandwich + fries + milkshake		269	na	
any veg sandwich + fries + mocktale		229	na	
any veg sandwich + fries + thickshake		339	na	
Any veg pizza + milkshake		299	na	
Any veg pizza + mojito		259	na	
Any veg pizza + thickshake		349	na	
Any chicken pizza + milkshake		300	na	
Any chicken pizza + mocktail		269	na	
Any chicken pizza + thickshake		359	na	
2 Chicken Leg Piece + fries + milkshake		299	na	
2 Chicken Leg Piece + fries + mocktail		279	na	
2 Chicken lollypop + fries + milkshake		249	na	
2 Chicken lollypop + fries + mocktail		199	na	
4 Chicken Leg Piece + fries + milkshake		449	na	
4 Chicken Leg Piece + fries + mocktail		429	na	
4 Chicken lollypop + fries + milkshake		299	na	
4 Chicken lollypop + fries + mocktail		249	na"""


def _parse_price(s: str):
    s = (s or "").strip()
    if not s or s.lower() == "na":
        return None
    try:
        return int(s)
    except ValueError:
        return s


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Menu"
    headers = [
        "Item",
        "Customization Option 1",
        "Price (₹)",
        "Customization Option 2",
        "Price (₹)",
    ]
    ws.append(headers)

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    n = 0
    for line in ROWS_RAW.strip().splitlines():
        parts = line.split("\t")
        while len(parts) < 5:
            parts.append("")
        item, o1, p1, o2, p2 = parts[0], parts[1], parts[2], parts[3], parts[4]
        item = item.strip()
        if not item:
            continue
        row = [
            item,
            o1.strip() or None,
            _parse_price(p1),
            (o2.strip() or None) if o2.strip().lower() != "na" else None,
            _parse_price(p2),
        ]
        ws.append(row)
        n += 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
        if row[2].value is not None and isinstance(row[2].value, int):
            row[2].alignment = Alignment(horizontal="right")
        if row[4].value is not None and isinstance(row[4].value, int):
            row[4].alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [42, 22, 12, 22, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(OUT)
    print(f"rows: {n}")


if __name__ == "__main__":
    main()
