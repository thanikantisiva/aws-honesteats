#!/usr/bin/env python3
"""
Read Snow_Drops_Menu.xlsx (Item, Option1, Price1, Option2, Price2), expand each
sellable variant into its own row (e.g. two shake rows), then add Category and
SubCategory.

Output: ~/Downloads/Snow_Drops_Menu_expanded.xlsx

  A: Category
  B: SubCategory
  C: Item Name   (expanded, e.g. "Belgium Chocolate Shake - Milk Shake")
  D: Price (₹)
  E: addOnOptions (JSON array) — only for combos whose name starts with
     "any " / "Any " (pick-one mains). Each option is a real menu row from
     Burgers, Sandwiches, or Pizza (extraPrice 0 = included swap; edit if needed).

Does not call any API.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_IN = Path.home() / "Downloads" / "Snow_Drops_Menu.xlsx"
DEFAULT_OUT = Path.home() / "Downloads" / "Snow_Drops_Menu_expanded.xlsx"


def _clean(v) -> str:
    return str(v).strip() if v is not None else ""


def _title_words(s: str) -> str:
    s = _clean(s)
    if not s:
        return ""
    return " ".join(w[:1].upper() + w[1:].lower() for w in s.split())


def _is_na_opt(s: str) -> bool:
    t = s.strip().lower()
    return not t or t == "na"


def _as_int_price(v) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        if v != v:  # NaN
            return None
        return int(v) if v == int(v) else int(round(v))
    s = _clean(v).replace("₹", "").replace(",", "")
    if not s or s.lower() == "na":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _expand_row(item: str, o1, p1, o2, p2) -> list[tuple[str, int]]:
    """Return list of (display_name, price)."""
    item = _clean(item)
    if not item or item.lower() == "combos":
        return []

    opt1 = _clean(o1)
    opt2 = _clean(o2)
    pr1 = _as_int_price(p1)
    pr2 = _as_int_price(p2)
    na1 = _is_na_opt(opt1)
    na2 = _is_na_opt(opt2)

    out: list[tuple[str, int]] = []
    if not na1 and pr1 is not None and not na2 and pr2 is not None:
        out.append((f"{item} - {_title_words(opt1)}", pr1))
        out.append((f"{item} - {_title_words(opt2)}", pr2))
        return out
    if not na1 and pr1 is not None:
        out.append((f"{item} - {_title_words(opt1)}", pr1))
        return out
    if not na2 and pr2 is not None:
        out.append((f"{item} - {_title_words(opt2)}", pr2))
        return out
    if pr1 is not None:
        out.append((item, pr1))
        return out
    return []


SCOOP_STYLE_NAMES = frozenset(
    {
        "fig and honey",
        "totty fruity",
    }
)


def _classify(name: str) -> tuple[str, str]:
    s = name.lower()

    if "+ fries +" in s or "+fries+" in s.replace(" ", ""):
        if "pizza" in s:
            if "chicken" in s:
                return ("Combos", "Chicken Pizza Combos")
            return ("Combos", "Veg Pizza Combos")
        if "sandwich" in s:
            if "chicken" in s:
                return ("Combos", "Chicken Sandwich Combos")
            return ("Combos", "Veg Sandwich Combos")
        if "burger" in s:
            if "chicken" in s:
                return ("Combos", "Chicken Burger Combos")
            return ("Combos", "Veg Burger Combos")
        if "lollypop" in s or "lolly" in s:
            return ("Combos", "Chicken Lollipop Combos")
        if "leg piece" in s:
            return ("Combos", "Chicken Leg Combos")
        return ("Combos", "Value Combos")

    if "any veg pizza" in s or ("veg pizza" in s and "milkshake" in s.replace(" ", "")):
        return ("Combos", "Veg Pizza Combos")
    if "any chicken pizza" in s or ("chicken pizza" in s and "milkshake" in s.replace(" ", "")):
        return ("Combos", "Chicken Pizza Combos")

    if "family packs" in s:
        return ("Desserts", "Family Packs")
    if " scoops" in s or s.endswith("scoops") or s in SCOOP_STYLE_NAMES:
        return ("Desserts", "Ice Cream Scoops")
    if "death by chocolate" in s or "brownie with" in s:
        return ("Desserts", "Brownies")
    if "chocolava" in s or ("ferrero" in s and "@" in name):
        return ("Desserts", "Premium Singles")
    if re.search(r" - (small|large)\b", s):
        return ("Desserts", "Sundaes")

    if "mocktail" in s or "mojito" in s or s.startswith("virgin mojito"):
        return ("Beverages", "Mocktails & Mojitos")

    if "fries" in s:
        return ("Snacks", "Fries")

    if "popcorn" in s or " nuggets" in s or s.endswith("nuggets") or "fingers" in s or "smilies" in s:
        return ("Snacks", "Starters")
    if s == "crispy" or s.startswith("crispy"):
        return ("Snacks", "Starters")
    if "paneer popcorn" in s:
        return ("Snacks", "Starters")

    if "burger" in s:
        if "veg" in s or "paneer" in s:
            return ("Burgers", "Veg Burgers")
        return ("Burgers", "Chicken Burgers")

    if "sandwich" in s:
        if "veg" in s or "corn" in s or "paneer" in s or "club" in s:
            return ("Sandwiches", "Veg Sandwiches")
        return ("Sandwiches", "Chicken Sandwiches")

    if "pizza" in s:
        veg_markers = (
            "veg", "corn", "mushroom", "peri-peri veg", "plain cheese",
            "schezwan", "sweet corn", "tandoori paneer", "veg bbq", "veg blast",
            "veg supreme", "double cheese",
        )
        if any(m in s for m in veg_markers) and "chicken" not in s:
            return ("Pizza", "Veg Pizza")
        return ("Pizza", "Chicken Pizza")

    if "bucket" in s:
        return ("Fried Chicken", "Buckets")
    if "lollypop" in s or "lolly pop" in s:
        return ("Fried Chicken", "Lollipops")
    if "leg piece" in s or "legpiece" in s.replace(" ", ""):
        return ("Fried Chicken", "Chicken Legs")

    if "shake" in s:
        return ("Beverages", "Shakes")

    return ("Menu", "Other")


def _combo_addon_pool_key(item_name: str) -> str | None:
    """Which burger / sandwich / pizza pool applies for 'any …' combo lines."""
    s = item_name.strip().lower()
    if s.startswith("any veg pizza"):
        return "veg_pizza"
    if s.startswith("any chicken pizza"):
        return "chicken_pizza"
    if s.startswith("any chicken burger"):
        return "chicken_burger"
    if s.startswith("any veg burger"):
        return "veg_burger"
    if s.startswith("any chicken sandwich"):
        return "chicken_sandwich"
    if s.startswith("any veg sandwich"):
        return "veg_sandwich"
    return None


def _option_id(seed: str, label: str, idx: int) -> str:
    h = hashlib.md5(f"{seed}|{label}|{idx}".encode()).hexdigest()[:12]
    return f"sd_{h}"


def _build_addon_pools(
    rows: list[tuple[str, str, str, int]],
) -> dict[str, list[tuple[str, int]]]:
    """Collect (name, price) from non-combo menu rows by pool key."""
    pools: dict[str, list[tuple[str, int]]] = {
        "veg_burger": [],
        "chicken_burger": [],
        "veg_sandwich": [],
        "chicken_sandwich": [],
        "veg_pizza": [],
        "chicken_pizza": [],
    }
    for cat, sub, name, price in rows:
        if cat == "Combos":
            continue
        if cat == "Burgers" and sub == "Veg Burgers":
            pools["veg_burger"].append((name, price))
        elif cat == "Burgers" and sub == "Chicken Burgers":
            pools["chicken_burger"].append((name, price))
        elif cat == "Sandwiches" and sub == "Veg Sandwiches":
            pools["veg_sandwich"].append((name, price))
        elif cat == "Sandwiches" and sub == "Chicken Sandwiches":
            pools["chicken_sandwich"].append((name, price))
        elif cat == "Pizza" and sub == "Veg Pizza":
            pools["veg_pizza"].append((name, price))
        elif cat == "Pizza" and sub == "Chicken Pizza":
            pools["chicken_pizza"].append((name, price))
    return pools


def _addons_json_for_combo(
    combo_name: str,
    pools: dict[str, list[tuple[str, int]]],
) -> str | None:
    key = _combo_addon_pool_key(combo_name)
    if not key:
        return None
    choices = pools.get(key) or []
    if not choices:
        return json.dumps([], ensure_ascii=False)
    seed = combo_name.strip()
    opts = [
        {
            "name": n,
            "optionId": _option_id(seed, n, i),
            "extraPrice": 0,
        }
        for i, (n, _p) in enumerate(choices)
    ]
    return json.dumps(opts, ensure_ascii=False)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in-xlsx", type=Path, default=DEFAULT_IN)
    ap.add_argument("--out-xlsx", type=Path, default=DEFAULT_OUT)
    args = ap.parse_args()

    if not args.in_xlsx.is_file():
        print(f"Not found: {args.in_xlsx}", file=sys.stderr)
        return 1

    wb_in = openpyxl.load_workbook(args.in_xlsx, read_only=True, data_only=True)
    ws_in = wb_in[wb_in.sheetnames[0]]
    rows_out: list[tuple[str, str, str, int]] = []

    for r in range(2, (ws_in.max_row or 1) + 1):
        item = ws_in.cell(r, 1).value
        o1 = ws_in.cell(r, 2).value
        p1 = ws_in.cell(r, 3).value
        o2 = ws_in.cell(r, 4).value
        p2 = ws_in.cell(r, 5).value
        for disp, price in _expand_row(item, o1, p1, o2, p2):
            cat, sub = _classify(disp)
            rows_out.append((cat, sub, disp, price))

    wb_in.close()

    pools = _build_addon_pools(rows_out)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menu"
    headers = ["Category", "SubCategory", "Item Name", "Price (₹)", "addOnOptions"]
    ws.append(headers)
    for cat, sub, name, price in rows_out:
        addons = _addons_json_for_combo(name, pools) if cat == "Combos" else None
        ws.append([cat, sub, name, price, addons])

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
        row[3].alignment = Alignment(horizontal="right")
        if row[4].value:
            row[4].alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate([18, 22, 48, 12, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    args.out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out_xlsx)
    wb.close()

    print(args.out_xlsx)
    print(f"rows: {len(rows_out)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
