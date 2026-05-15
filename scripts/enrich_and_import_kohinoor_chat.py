#!/usr/bin/env python3
"""
Enrich Kohinoor Chat menu Excel by adding `Is Veg` (col G) and `Description`
(col I), and POST each row to the menu API.

Source columns (input — already in user's spec):
    A = Item             (Name, title-cased)
    B = Dine-in (D)      (restaurantPrice)
    C = Our Hike on D %  (hikePercentage)

Added columns:
    G = Is Veg           ("Veg" / "Non-Veg")
    I = Description

If a row has no hike% in column C, a fixed 35% hike is applied.

Payload (per row):
    {
      "name":             A (Title-cased),
      "restaurantPrice":  B,
      "description":      I,
      "hikePercentage":   C  (or 35 if blank),
      "category":         "Chat",
      "subCategory":      "Chat",
      "isVeg":            true/false (from G),
      "isAvailable":      true
    }

Usage:
  python3 scripts/enrich_and_import_kohinoor_chat.py             # enrich + dry run
  python3 scripts/enrich_and_import_kohinoor_chat.py --apply     # enrich + POST
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RESTAURANT_ID = "RES-1776403494998-2791"
XLSX_PATH = Path("/Users/user/Downloads/Kohinoor chat Items 1.xlsx")

CATEGORY = "Chat"
SUB_CATEGORY = "Chat"
DEFAULT_HIKE_PCT = 35.0  # fallback if column C is empty

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get(
    "HONESTEATS_RETOOL_BYPASS", "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc"
)
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

# ── Descriptions (key = title-cased item name) ───────────────────────────────────
DESC: dict[str, str] = {
    "Pani Poori": "Crispy hollow puris filled with spiced tangy water, mashed potato and chickpeas",
    "Bhel Puri": "Mumbai-style puffed rice chaat tossed with sev, onions, tomatoes and tangy chutneys",
    "Butter Pav Bhaji": "Spicy mashed vegetable curry topped with butter, served with toasted pav",
    "Cheese Pav Bhaji": "Classic pav bhaji topped with grated cheese and butter, served with pav",
    "Cutlet": "Spiced potato and vegetable cutlet, deep-fried to a crisp golden brown",
    "Dahi Papdi": "Crispy papdi topped with whipped curd, sweet-and-spicy chutneys, sev and pomegranate",
    "Dahi Puri": "Mini puris stuffed with potato, curd, tangy chutneys and a sprinkle of sev",
    "Ghee Cutlet": "Rich ghee-roasted vegetable cutlet with aromatic Indian spices",
    "Kaju Cutlet": "Crispy cutlet loaded with cashews for a delicate nutty crunch",
    "Masala Puri": "South Indian street-style chaat with crispy puris in spicy peas masala",
    "Mixing Pav Bhaji": "Special blend pav bhaji with extra veggies, butter and a kick of spices",
    "Mushroom Cutlet": "Tender mushroom and potato cutlet, golden fried with savory spices",
    "Paneer Cutlet": "Soft paneer and potato cutlet with a crispy spiced outer crust",
    "Papdi Cutlet": "Papdi-style cutlet served with tangy tamarind and green chutneys",
    "Pav Bhaji": "Mumbai's iconic spicy mashed vegetable curry served with buttered pav",
    "Samosa Cutlet": "Samosa-stuffed cutlet, deep-fried to a crisp golden brown",
    "Sev Puri": "Crispy puris topped with potato, onion, tangy chutneys and a generous topping of sev",
}

# All Kohinoor chaat items are vegetarian.  Add any non-veg overrides here.
NON_VEG_ITEMS: set[str] = set()


def _title_case(name: str) -> str:
    return " ".join(w.capitalize() for w in str(name).strip().split())


def _to_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def enrich() -> None:
    """Add headers/values for `Is Veg` (G) and `Description` (I) in-place."""
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    # Existing headers (A, B, C) — style only, don't overwrite if user already set them
    for col in (1, 2, 3):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center")

    # Add Is Veg (G=7) and Description (I=9) headers
    for col, header in [(7, "Is Veg"), (9, "Description")]:
        c = ws.cell(row=1, column=col, value=header)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center")

    rows_processed = 0
    for r in range(2, ws.max_row + 1):
        raw_name = ws.cell(row=r, column=1).value
        if not raw_name:
            continue
        name = _title_case(raw_name)
        # Persist the canonical title-cased name back into column A
        ws.cell(row=r, column=1, value=name)

        is_veg_str = "Non-Veg" if name in NON_VEG_ITEMS else "Veg"
        description = DESC.get(name, f"Delicious {name}")

        ws.cell(row=r, column=7, value=is_veg_str)
        ws.cell(row=r, column=9, value=description)
        rows_processed += 1

    # Tidy column widths
    for col in range(1, 10):
        widths = [
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, ws.max_row + 1)
        ]
        ws.column_dimensions[get_column_letter(col)].width = min(max(widths or [10]) + 3, 60)

    wb.save(XLSX_PATH)
    print(f"Enriched {XLSX_PATH.name} → {rows_processed} rows (added G=Is Veg, I=Description)")


def build_payloads() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    payloads: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = _title_case(row[0])
        restaurant_price = row[1] if len(row) > 1 else None
        hike_raw = row[2] if len(row) > 2 else None
        is_veg_str = row[6] if len(row) > 6 else "Veg"
        description = row[8] if len(row) > 8 else None

        if restaurant_price is None:
            continue

        hike_pct = _to_float(hike_raw, DEFAULT_HIKE_PCT)
        if hike_raw in (None, ""):
            hike_pct = DEFAULT_HIKE_PCT

        is_veg = True if str(is_veg_str or "").strip().lower() == "veg" else False

        payloads.append({
            "name": name,
            "restaurantPrice": float(restaurant_price),
            "description": str(description).strip() if description else None,
            "hikePercentage": round(hike_pct, 2),
            "category": CATEGORY,
            "subCategory": SUB_CATEGORY,
            "isVeg": is_veg,
            "isAvailable": True,
        })
    wb.close()
    return payloads


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true", help="POST to API (default: dry run)")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--bearer-token", default=os.environ.get("HONESTEATS_BEARER_TOKEN"))
    parser.add_argument("--delay", type=float, default=0.3)
    parser.add_argument("--skip-enrich", action="store_true", help="Skip Excel enrichment step")
    args = parser.parse_args()

    if not args.skip_enrich:
        enrich()
    payloads = build_payloads()

    print(f"\n{len(payloads)} items | Restaurant: {RESTAURANT_ID}\n")
    for i, p in enumerate(payloads, 1):
        print(
            f"[{i:3d}] {p['name']:<22}  ₹{p['restaurantPrice']:>5} + {p['hikePercentage']:>5}%  "
            f"| {p['category']} > {p['subCategory']} | Veg={p['isVeg']}"
        )
        if not args.apply:
            print(f"      {json.dumps(p, ensure_ascii=False)}")

    menu_url = f"{args.api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"

    if not args.apply:
        print("\n" + "=" * 60)
        print("DRY RUN — no HTTP calls.")
        print(f"To insert: python3 {sys.argv[0]} --apply")
        print("=" * 60)
        return 0

    s = requests.Session()
    s.headers["Content-Type"] = "application/json"
    if args.bearer_token:
        s.headers["Authorization"] = f"Bearer {args.bearer_token}"
    else:
        s.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS

    ok = err = 0
    print(f"\nPOST {menu_url}")
    for i, p in enumerate(payloads, 1):
        if args.delay > 0:
            time.sleep(args.delay)
        try:
            r = s.post(menu_url, json=p, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"  ✗ [{i}] {p['name']}: {ex}", file=sys.stderr)
            continue
        if r.status_code == 201:
            ok += 1
            print(f"  ✓ [{i}] {p['name']} created")
        else:
            err += 1
            print(f"  ✗ [{i}] {p['name']}: HTTP {r.status_code} {r.text[:200]}", file=sys.stderr)
            if r.status_code == 401:
                print("\nUnauthorized — provide --bearer-token or HONESTEATS_BEARER_TOKEN.", file=sys.stderr)
                return 1

    print(f"\nDone: {ok} created, {err} failed.")
    return 0 if err == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
