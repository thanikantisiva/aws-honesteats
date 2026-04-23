#!/usr/bin/env python3
"""
Insert SBK (Sree Bheema's Kitchen) menu items from the enriched pricing Excel.

Reads: ~/Downloads/SBK_menu_insert_bodies.json
   or: ~/Downloads/SBK_Pricing_Analysis copy.xlsx  (fallback — rebuilds bodies on the fly)

POST {base}/api/v1/restaurants/{restaurantId}/menu

Usage:
  # Dry run (default) — prints what would be sent, no HTTP calls
  python3 scripts/insert_sbk_menu.py

  # Actually insert
  python3 scripts/insert_sbk_menu.py --apply

  # Custom API / auth
  python3 scripts/insert_sbk_menu.py --apply \
    --api-url https://api.yumdude.com \
    --bearer-token <JWT>

Env:
  HONESTEATS_API_URL           — API host (default https://api.yumdude.com)
  HONESTEATS_RETOOL_BYPASS     — x-retool-header value
  HONESTEATS_BEARER_TOKEN      — JWT alternative
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

import requests

_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))
from menu_item_images import resolve_menu_item_image

RESTAURANT_ID = "RES-1776938607201-3458"

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get(
    "HONESTEATS_RETOOL_BYPASS",
    "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc",
)
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

JSON_PATH = Path.home() / "Downloads" / "SBK_menu_insert_bodies.json"
XLSX_PATH = Path.home() / "Downloads" / "SBK_Pricing_Analysis copy.xlsx"


def _normalize_api_base(url: str) -> str:
    u = url.rstrip("/")
    if u.endswith("/api/v1"):
        u = u[: -len("/api/v1")]
    return u


def load_items_from_json(path: Path) -> list[dict]:
    with open(path) as f:
        data = json.load(f)
    return data["items"]


def load_items_from_xlsx(path: Path) -> list[dict]:
    """Fallback: read directly from the enriched Excel (cols A–H)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 8:
            continue
        cat, _sub_old, name, dinein, our_price, sub_cat, is_veg_str, desc = row[:8]
        if not name:
            continue
        dinein = float(dinein) if dinein else 0
        our_price = float(our_price) if our_price else 0
        hike = round((our_price - dinein) / dinein * 100, 2) if dinein > 0 else 0
        is_veg = str(is_veg_str).strip().lower() == "veg" if is_veg_str else None
        items.append({
            "name": str(name).strip(),
            "restaurantPrice": dinein,
            "hikePercentage": hike,
            "category": str(cat).strip() if cat else "",
            "subCategory": str(sub_cat).strip() if sub_cat else "",
            "isVeg": is_veg,
            "isAvailable": True,
            "description": str(desc).strip() if desc else "",
        })
    wb.close()
    return items


def build_session(args: argparse.Namespace) -> requests.Session:
    s = requests.Session()
    s.headers["Content-Type"] = "application/json"
    if args.no_retool_bypass:
        bypass = None
    elif (args.retool_bypass or "").strip():
        bypass = args.retool_bypass.strip()
    else:
        bypass = DEFAULT_RETOOL_BYPASS.strip() or None
    if bypass:
        s.headers[RETOOL_BYPASS_HEADER] = bypass
    token = (args.bearer_token or "").strip()
    if token:
        s.headers["Authorization"] = f"Bearer {token}"
    return s


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Insert SBK menu items via API")
    p.add_argument(
        "--restaurant-id",
        default=RESTAURANT_ID,
        help=f"Restaurant ID (default: {RESTAURANT_ID})",
    )
    p.add_argument(
        "--api-url",
        default=DEFAULT_API_URL,
        help="API base URL",
    )
    p.add_argument("--no-retool-bypass", action="store_true")
    p.add_argument("--retool-bypass", default="")
    p.add_argument(
        "--bearer-token",
        default=os.environ.get("HONESTEATS_BEARER_TOKEN", ""),
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        default=True,
        help="Parse and print only (default)",
    )
    p.add_argument("--apply", action="store_true", help="POST each item to the API")
    p.add_argument(
        "--delay-sec",
        type=float,
        default=float(os.environ.get("HONESTEATS_REQUEST_DELAY_SEC", "0.05")),
    )
    p.add_argument("--no-images", action="store_true", help="Skip stock image resolution")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    if args.apply:
        args.dry_run = False

    if JSON_PATH.is_file():
        items = load_items_from_json(JSON_PATH)
        print(f"Loaded {len(items)} items from {JSON_PATH.name}")
    elif XLSX_PATH.is_file():
        items = load_items_from_xlsx(XLSX_PATH)
        print(f"Loaded {len(items)} items from {XLSX_PATH.name}")
    else:
        print(f"Neither {JSON_PATH} nor {XLSX_PATH} found.", file=sys.stderr)
        return 1

    base = _normalize_api_base(args.api_url)
    menu_url = f"{base}/api/v1/restaurants/{args.restaurant_id}/menu"

    print(f"API:        POST {menu_url}")
    print(f"Restaurant: {args.restaurant_id}")
    print(f"Items:      {len(items)}")
    print(f"Mode:       {'DRY-RUN' if args.dry_run else 'APPLY'}")
    print()

    veg = sum(1 for it in items if it.get("isVeg") is True)
    nv = sum(1 for it in items if it.get("isVeg") is False)
    print(f"Veg: {veg}  |  Non-Veg: {nv}  |  Unknown: {len(items) - veg - nv}")

    for i, it in enumerate(items[:5], 1):
        img = ""
        if not args.no_images:
            resolved = resolve_menu_item_image(
                it["name"], it["category"], it.get("subCategory", ""), None
            )
            img = f" | img={resolved[:56]}…"
        print(
            f"  {i}. {it['name']} | {it['category']} / {it.get('subCategory', '')} "
            f"| ₹{it['restaurantPrice']} +{it['hikePercentage']}% "
            f"| {'V' if it.get('isVeg') else 'NV'}{img}"
        )
    if len(items) > 5:
        print(f"  ... +{len(items) - 5} more")

    if args.dry_run:
        print("\nDry run — no HTTP calls. Pass --apply to insert via API.")
        return 0

    session = build_session(args)
    has_bypass = RETOOL_BYPASS_HEADER in session.headers
    has_bearer = "Authorization" in session.headers
    if not has_bypass and not has_bearer:
        print(
            "Error: no auth configured. Set HONESTEATS_RETOOL_BYPASS or "
            "pass --bearer-token.",
            file=sys.stderr,
        )
        return 1

    ok = err = 0
    for it in items:
        payload = {
            "name": it["name"],
            "restaurantPrice": it["restaurantPrice"],
            "hikePercentage": it["hikePercentage"],
            "category": it["category"],
            "subCategory": it.get("subCategory") or None,
            "isVeg": it.get("isVeg"),
            "isAvailable": True,
            "description": it.get("description") or None,
        }
        if not args.no_images:
            payload["image"] = resolve_menu_item_image(
                it["name"], it["category"], it.get("subCategory", ""), None
            )

        if args.delay_sec > 0:
            time.sleep(args.delay_sec)

        try:
            r = session.post(menu_url, json=payload, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"ERROR {it['name']!r}: {ex}", file=sys.stderr)
            continue

        if r.status_code == 201:
            ok += 1
            if ok % 25 == 0:
                print(f"  ... {ok}/{len(items)} created")
        else:
            err += 1
            print(f"ERROR {it['name']!r}: HTTP {r.status_code} {r.text[:300]}", file=sys.stderr)
            if r.status_code == 401:
                print("\nUnauthorized — check auth config.", file=sys.stderr)
                return 1

    print(f"\nDone: {ok} created, {err} failed.")
    return 0 if err == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
