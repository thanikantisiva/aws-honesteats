#!/usr/bin/env python3
"""
Compare Aatithyam Pure Veg: Zomato PDF vs dine-in Excel.

Inputs:
  Zomato: /Users/user/Downloads/Aatithyam Pure Veg, Nandyal Locality order online - Zomato.pdf
  Dine-in: /Users/user/Downloads/Aathithyam_Dinein.xlsx

Output:
  /Users/user/Downloads/Aatithyam_Zomato_vs_DineIn.xlsx
  Columns: Category | subcategory | Item | Dine-In (₹) | Zomato (₹)
"""
from __future__ import annotations
import re
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from rapidfuzz import fuzz

ZOMATO_PDF = Path(
    "/Users/user/Downloads/Aatithyam Pure Veg, Nandyal Locality order online - Zomato.pdf"
)
DINEIN_XLSX = Path("/Users/user/Downloads/Aathithyam_Dinein.xlsx")
OUT_XLSX = Path("/Users/user/Downloads/Aatithyam_Zomato_vs_DineIn.xlsx")

ZOMATO_SECTIONS = {
    "Soups",
    "Indian Appetizers",
    "Tandoor Appetizers",
    "Chinese Appetizers",
    "Indian Main Course",
    "Pulaos",
    "Accompaniments",
    "Breads",
    "Aatithyam Specials",
    "Deserts",
}

SKIP_SUBSTR = (
    "order online",
    "zomato.com",
    "nandyal locality",
    "search for restaurant",
    "home /",
    "direction",
    "delivery only",
    "share reviews",
    "dining ratings",
    "north indian",
    "saneeva nagar",
    "sanjeewa nagar",
    "+91",
    "currently closed",
    "live tracking",
    "search within menu",
    "related to aatithyam",
    "restaurants in nandyal",
    "frequent searches",
    "lic. no.",
    "your order",
    "subtotal",
    "download the app",
    "terms of service",
    "cookie policy",
    "privacy policy",
    "about zomato",
    "india english",
)

PRICE_RE = re.compile(r"^₹\s*([\d,]+(?:\.\d+)?)$")
SIDEBAR_RE = re.compile(r".*\(\d+\)\s*$")


def is_noise_line(s: str) -> bool:
    t = s.strip()
    if not t:
        return True
    low = t.lower()
    if t.startswith("Page ") and "of" in t:
        return True
    if low.startswith("https://"):
        return True
    for x in SKIP_SUBSTR:
        if x in low:
            return True
    if re.match(r"^aatithyam pure veg\b", low) and len(t) < 80:
        return True
    return False


def is_description(s: str) -> bool:
    low = s.lower()
    if "read more" in low:
        return True
    if len(s) > 95:
        return True
    if s.strip().endswith("..."):
        return True
    # Full sentence descriptions (not short dish names like "Crispy Babycorn")
    if len(s) > 55 and s[0].islower():
        return True
    if len(s) > 55 and re.match(
        r"^(a |an |the |yellow|soft |tangy|mildly|hung |pepper|classic|thick |"
        r"aromatic|mixed |assorted|broccoli|spinach)",
        s,
        re.I,
    ):
        return True
    return False


def is_item_candidate(s: str) -> bool:
    t = s.strip()
    if len(t) < 2 or len(t) > 62:
        return False
    if t in ZOMATO_SECTIONS:
        return False
    if SIDEBAR_RE.match(t):
        return False
    if PRICE_RE.match(t):
        return False
    if is_description(t):
        return False
    if t.lower() in {"chilli", "honey chilli", "manchuria", "salt & pepper"}:
        return True
    return True


def parse_zomato_pdf() -> list[tuple[str, str, float]]:
    """Return [(zomato_section, item_name, price), ...]."""
    lines: list[str] = []
    with pdfplumber.open(ZOMATO_PDF) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for line in text.splitlines():
                s = line.strip()
                if is_noise_line(s):
                    continue
                lines.append(s)

    items: list[tuple[str, str, float]] = []
    current_section = ""
    pending_item: str | None = None

    for s in lines:
        if s in ZOMATO_SECTIONS:
            current_section = s
            pending_item = None
            continue
        if SIDEBAR_RE.match(s):
            continue
        m = PRICE_RE.match(s)
        if m:
            if pending_item:
                price = float(m.group(1).replace(",", ""))
                items.append((current_section, pending_item, price))
            pending_item = None
            continue
        if is_description(s):
            continue
        if is_item_candidate(s):
            pending_item = s

    return items


def norm(s: str) -> str:
    s = s.lower()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for a, b in [
        ("jee", "jeedi"),
        ("jeedipappu", "jeedipappu"),
        ("manchuria", "manchurian"),
        ("schezwan", "schezwan"),
        ("kulcha", "kulcha"),
    ]:
        s = s.replace(a, b)
    return s


# Dine-in item name -> Zomato menu item name (as parsed from PDF)
MANUAL: dict[str, str] = {
    "Dahi ke Kebab": "Aatithyam Special Dahi Ke Kebab",
    "Veg Manchuria (Wet)": "Manchuria",
    "Veg Manchuria (Dry)": "Manchuria",
    "Baby Corn Manchuria (Wet)": "Manchuria",
    "Baby Corn Manchuria (Dry)": "Manchuria",
    "Mushroom Manchuria (Wet)": "Manchuria",
    "Mushroom Manchuria (Dry)": "Manchuria",
    "Paneer Manchuria (Wet)": "Manchuria",
    "Paneer Manchuria (Dry)": "Manchuria",
    "Veg Chilli (Wet)": "Chilli",
    "Veg Chilli (Dry)": "Chilli",
    "Mushroom Chilli (Wet)": "Chilli",
    "Mushroom Chilli (Dry)": "Chilli",
    "Baby Corn Chilli (Wet)": "Chilli",
    "Baby Corn Chilli (Dry)": "Chilli",
    "Paneer Chilli (Wet)": "Chilli",
    "Paneer Chilli (Dry)": "Chilli",
    "Veg Salt & Pepper": "Crispy Salt & Pepper",
    "Baby Corn Salt & Pepper": "Crispy Salt & Pepper",
    "Mushroom Salt & Pepper": "Crispy Salt & Pepper",
    "Paneer Salt & Pepper": "Crispy Salt & Pepper",
    "Veg Honey Chilli": "Honey Chilli",
    "Mushroom Honey Chilli": "Honey Chilli",
    "Baby Corn Honey Chilli": "Honey Chilli",
    "Paneer Honey Chilli": "Honey Chilli",
    "Veg Kung Pao": "Honey Chilli",
    "Mushroom Kung Pao": "Honey Chilli",
    "Baby Corn Kung Pao": "Honey Chilli",
    "Paneer Kung Pao": "Honey Chilli",
    "Tomato Jeedipappu Mushroom": "Tomato Jeedipappu Mushroom Curry",
    "Tomato Jeedipappu Paneer": "Tomato Jeedipappu Paneer Curry",
    "Gongura Paneer Pulao": "Chitimutyala Gongura Paneer Pulao",
    "Gongura Mushroom Pulao": "Chitimutyala Gongura Mushroom Pulao",
    "Dal Thadka": "Dal Tadka",
}


STOPWORDS = {
    "veg", "paneer", "mushroom", "baby", "corn", "mix", "fried", "rice",
    "dry", "wet", "chilli", "honey", "salt", "pepper", "curry", "soup",
    "the", "and", "with", "naan", "kulcha", "garlic", "butter",
}


def best_zomato_price(dine_item: str, zlist: list[tuple[str, str, float]]) -> float | None:
    alias = MANUAL.get(dine_item)
    if alias:
        an = norm(alias)
        for _sec, zname, price in zlist:
            if norm(zname) == an:
                return price

    dn = norm(dine_item)
    d_toks = set(dn.split())
    best_score = 0
    best_price: float | None = None
    best_name: str | None = None
    for _sec, zname, price in zlist:
        zn = norm(zname)
        score = fuzz.token_sort_ratio(dn, zn)
        if score > best_score:
            best_score = score
            best_price = price
            best_name = zname
    if best_score < 82 or best_name is None:
        return None
    z_toks = set(norm(best_name).split())
    common = d_toks & z_toks
    distinctive = [w for w in common if w not in STOPWORDS and len(w) >= 3]
    if distinctive or best_score >= 93:
        return best_price
    return None


def load_dinein() -> list[tuple[str, str, str, int]]:
    wb = load_workbook(DINEIN_XLSX, read_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(row=r, column=1).value or ""
        sub = ws.cell(row=r, column=2).value or ""
        item = ws.cell(row=r, column=3).value or ""
        price = ws.cell(row=r, column=5).value
        if item and price is not None:
            rows.append((str(cat), str(sub), str(item), int(price)))
    wb.close()
    return rows


def main() -> None:
    zlist = parse_zomato_pdf()
    dine = load_dinein()

    wb = Workbook()
    ws = wb.active
    ws.title = "Zomato vs Dine-In"

    hdr = ["Category", "subcategory", "Item", "Dine-In (₹)", "Zomato (₹)"]
    ws.append(hdr)

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2E86AB")
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_al = Alignment(horizontal="right", vertical="center")

    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = brd

    matched = 0
    for cat, sub, item, dprice in dine:
        zprice = best_zomato_price(item, zlist)
        zcell = zprice if zprice is not None else "-"
        if zprice is not None:
            matched += 1
        ws.append([cat, sub, item, dprice, zcell])
        r = ws.max_row
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = brd
        ws.cell(row=r, column=4).alignment = right_al
        ws.cell(row=r, column=5).alignment = right_al

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A2"

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)

    print(f"Zomato items parsed: {len(zlist)}")
    print(f"Dine-in rows:       {len(dine)}")
    print(f"Matched prices:     {matched}")
    print(f"Saved: {OUT_XLSX}")


if __name__ == "__main__":
    main()
