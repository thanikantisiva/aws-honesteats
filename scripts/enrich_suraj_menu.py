#!/usr/bin/env python3
"""
Enrich Suraj Restaurant menu Excel and emit menu-item JSON; optional POST to API.

Reads:  /Users/user/Downloads/Suraj_Restaurant_Menu.xlsx
        Cols A–E: Category | Subcategory | Item Name | Price | Hike

Adds / refreshes columns:
  F — Description
  G — (blank spacer)
  H — isVeg (TRUE/FALSE)
  I — image (single Unsplash URL per row; no repeats)

Writes JSON array → ~/Downloads/Suraj_Restaurant_Menu.menu_items.json

Usage:
  python3 scripts/enrich_suraj_menu.py                    # enrich + JSON only
  python3 scripts/enrich_suraj_menu.py --apply             # enrich + POST
  python3 scripts/enrich_suraj_menu.py --from-json --apply # POST from existing JSON
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

XLSX_PATH = Path("/Users/user/Downloads/Suraj_Restaurant_Menu.xlsx")
JSON_PATH = Path("/Users/user/Downloads/Suraj_Restaurant_Menu.menu_items.json")
RESTAURANT_ID = "RES-1776582609750-1046"

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get(
    "HONESTEATS_RETOOL_BYPASS",
    "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc",
)
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

COL_CAT, COL_SUB, COL_NAME, COL_PRICE, COL_HIKE = 1, 2, 3, 4, 5
COL_DESC, COL_GAP, COL_VEG, COL_IMG = 6, 7, 8, 9

# ── Unsplash photo-ID pools by theme ─────────────────────────────────────────
_POOLS: dict[str, list[str]] = {
    "soup": [
        "1547592166-23ac45744acd", "1603105037880-880cd4f5b2e6",
        "1476718406336-bb5a9690ee2a", "1588566565463-180a5b2090d2",
        "1613844237701-8f3664fc2eff", "1594756202469-9ff9799b2e4e",
        "1604152135912-04a022e23696", "1607330289024-1535c6b4e1c1",
        "1583608205776-bfd35f0d9f83", "1509358271058-acd22cc93898",
        "1617093727343-374698b1b08d", "1597227129956-93bad7e18d08",
        "1603894584373-5ac82b2ae328",
    ],
    "starters_veg": [
        "1601050690597-df0568f70950", "1546069901-ba9599a7e63c",
        "1540189549336-e6e99c3679fe", "1567337710282-00832b415979",
        "1585032226651-759b368d7246", "1572715376701-98568319fd0b",
        "1559847844-5315695dadae", "1606574977732-e8e5f1f46c23",
        "1564834724105-918b73d1b8e0", "1512621776951-a57141f2eefd",
        "1574484284002-952d92456975", "1567620832903-9fc6debc209f",
        "1606728035253-49e8a23146de", "1599487488170-d11ec9c172f0",
        "1565299624946-b28f40a0ae38", "1573080496219-bb080dd4f877",
        "1585937421612-70a008356fbe", "1580217593608-61931cefc821",
        "1614398751058-bca239de00ca", "1603360946369-dc9bb6258143",
        "1565557623262-b51c2513a641", "1606755962773-d324e0a13086",
        "1612874742237-6526221588e3", "1625937286520-3ef7955f3813",
        "1576402187878-974f70c890a5", "1455619452474-d2be8b1e70cd",
        "1596797038530-2c107229654b", "1604908176997-125f25cc6f3d",
        "1612929633738-8fe44f7ec841", "1631452180519-c014fe946bc7",
    ],
    "starters_nonveg": [
        "1626082927389-6cd097cdc6ec", "1562967914-608f82629710",
        "1608039755401-742074f0548d", "1610057099431-d73a1c9d2f2f",
        "1598515214211-89d3c73ae83b", "1587593810167-a84920ea0781",
        "1619221882220-947b3d3c8861", "1632778149955-e80f8ceca2e8",
        "1624726175512-19b9baf9fbd1", "1559339352-11d035aa65de",
        "1553621042-f6e147245754", "1565680018093-ebb6e3062e7b",
    ],
    "seafood": [
        "1615141982883-c7ad0e69fd62", "1504674900247-0877df9cc836",
        "1535140728325-a4d3707eee61", "1510130113581-4ae76c0f6e7f",
        "1559039448-9b03d2e3c18e", "1519708227418-c8fd9a32b7a2",
        "1498654896293-37aacf113fd9", "1580476262798-bddd9f4b7369",
        "1551504734-5ee1c4a1479b", "1569058242567-93de6f36f8e6",
    ],
    "tandoori": [
        "1601050690117-94f5f6fa8bd7", "1628294895950-9805252327bc",
        "1567188040759-fb8a883dc6d8", "1574653853027-5382a3d23a15",
        "1551881192-5e377f1b2142", "1545247181-516773cae754",
        "1606491956689-2ea866880049", "1609501676725-7186f017a4b7",
        "1617692855027-33b14f061079", "1633321702518-7fecdafb94d5",
        "1618449840665-9ed506d73a34", "1642821373181-16a5bc9f5801",
    ],
    "bread": [
        "1600326145359-3a44909d1a39", "1574071318508-1cdbab80d002",
        "1586444248879-bc604bc77212", "1555939594-58d7cb561ad1",
        "1555507036-ab1f4038024a", "1573675542321-f51b18e6e759",
        "1515024014929-c2ba2c4da1d0", "1519864600395-3404e40a0eda",
        "1509722747041-616f39b57569", "1528736235302-52922df5c122",
        "1605888969139-42cca4308aa2",
    ],
    "curry_veg": [
        "1631515243349-e0cb75fb8d3a", "1505253758473-96b7015fcd40",
        "1607116667573-1c7d73636ba0", "1599043513900-ed6fe01d3833",
        "1563379091339-03b21ab4a4f4", "1627662168223-7df99068099a",
        "1567188040759-fb8a883dc6d8",
    ],
    "curry_nonveg": [
        "1610057099431-d73a1c9d2f2f", "1598515214211-89d3c73ae83b",
        "1587593810167-a84920ea0781", "1619221882220-947b3d3c8861",
        "1632778149955-e80f8ceca2e8", "1624726175512-19b9baf9fbd1",
    ],
    "biryani": [
        "1589302168068-964664d93dc0", "1633321702518-7fecdafb94d5",
        "1612874742237-6526221588e3", "1625937286520-3ef7955f3813",
        "1599043513900-ed6fe01d3833", "1628294895950-9805252327bc",
        "1607116667573-1c7d73636ba0", "1574653853027-5382a3d23a15",
        "1603894584373-5ac82b2ae328", "1551881192-5e377f1b2142",
        "1505253758473-96b7015fcd40", "1627662168223-7df99068099a",
        "1545247181-516773cae754", "1609501676725-7186f017a4b7",
        "1618449840665-9ed506d73a34", "1617692855027-33b14f061079",
    ],
    "fried_rice": [
        "1596560548464-f010549b84d7", "1516714435131-44d6b64dc6a2",
        "1536304993881-460587633ee1", "1512058564366-18510be2db19",
        "1645696301019-35adcc552067", "1603133872878-684f208fb84b",
    ],
    "noodles": [
        "1569058242567-93de6f36f8e6", "1585032226651-759b368d7246",
        "1572715376701-98568319fd0b", "1559847844-5315695dadae",
    ],
    "drinks": [
        "1534353473418-4cfa6c56fd38", "1600271886742-f049cd451bba",
        "1572490122747-3968b75cc699", "1541658016709-82535e94bc69",
        "1553787499-6f9133860278", "1568901839119-631418a3910d",
        "1619158401201-8fa932695178", "1579954115545-a95591f28bfc",
        "1497034825429-c343d7c6a68f", "1501443762994-82bd5dace89a",
        "1560008581-09826d1de69e", "1570197571499-166b36435e9f",
        "1629385701021-fcd568a743e8", "1576506295286-5cda18df43e7",
        "1633933358116-a27b902fad35", "1514849302-984523450cf4",
        "1563805042-7684c019e1cb", "1551024506-0bccd828d307",
        "1580915411954-282cb1b0d780", "1557142046-c704a3adf364",
    ],
    "dessert": [
        "1571091718767-18b5b1457add", "1606312619070-d48b4c652a52",
        "1551024506-0bccd828d307", "1488477181946-6428a0291777",
    ],
    "raita": [
        "1568901346375-23c9450c58cd", "1550547660-d9450f859349",
        "1586190848861-99aa4a171e90",
    ],
}


def _photo_url(pid: str, sig: int | None = None) -> str:
    base = f"https://images.unsplash.com/photo-{pid}?auto=format&fit=crop&w=800&q=80"
    return base if sig is None else f"{base}&sig={sig}"


_RX = re.compile(r"^\d{6,}-[0-9a-f]+$")
for _k, _lst in list(_POOLS.items()):
    _POOLS[_k] = [p for p in _lst if _RX.match(p)]

_seen_ids: set[str] = set()
_GLOBAL: list[str] = []
for _pn in list(_POOLS.keys()):
    for _pid in _POOLS[_pn]:
        if _pid not in _seen_ids:
            _seen_ids.add(_pid)
            _GLOBAL.append(_pid)


class _Picker:
    def __init__(self):
        self.used: set[str] = set()
        self._sig = 0

    def _themed(self, key: str) -> str | None:
        for pid in _POOLS.get(key, []):
            if pid not in self.used:
                self.used.add(pid)
                return _photo_url(pid)
        return None

    def _global(self) -> str | None:
        for pid in _GLOBAL:
            if pid not in self.used:
                self.used.add(pid)
                return _photo_url(pid)
        return None

    def _recycled(self) -> str:
        self._sig += 1
        pid = _GLOBAL[self._sig % len(_GLOBAL)]
        return _photo_url(pid, sig=self._sig)

    def pick(self, key: str) -> str:
        return self._themed(key) or self._global() or self._recycled()


# ── Pool selection ────────────────────────────────────────────────────────────
def select_pool(cat: str, sub: str, name: str) -> str:
    cl, sl, nl = cat.lower(), sub.lower(), name.lower()

    if "soup" in cl:
        return "soup"
    if "raita" in sl:
        return "raita"
    if "dessert" in cl or "dessert" in sl:
        return "dessert"
    if "beverage" in cl or "drink" in sl or "milkshake" in sl:
        return "drinks"
    if "bread" in cl or any(k in nl for k in ("naan", "roti", "paratha", "kulcha", "phulka")):
        return "bread"
    if "biryani" in sl:
        return "biryani"
    if "noodle" in sl:
        return "noodles"
    if "rice" in sl:
        return "fried_rice"
    if "tandoori" in sl or any(k in nl for k in ("tandoori", "tikka", "kebab", "kabab")):
        return "tandoori"
    if any(k in nl for k in ("fish", "prawn", "jingha", "apollo fish")):
        return "seafood"

    is_nv = _is_nonveg(cat, sub, name)
    if "starter" in cl:
        return "starters_nonveg" if is_nv else "starters_veg"
    if "main course" in cl or "curry" in cl:
        return "curry_nonveg" if is_nv else "curry_veg"

    return "starters_veg"


# ── Veg / non-veg ────────────────────────────────────────────────────────────
_NV_KW = (
    "non veg", "non-veg", "nonveg", "chicken", "mutton", "fish", "prawn",
    "prawns", "jingha", "egg", "lamb", "beef", "drumstick", "tangdi",
    "murgh", "kheema", "apollo",
)


def _is_nonveg(cat: str, sub: str, name: str) -> bool:
    text = f"{cat} {sub} {name}".lower()
    return any(k in text for k in _NV_KW)


# ── Description ───────────────────────────────────────────────────────────────
_HOUSE = "Suraj Restaurant"


def build_description(cat: str, sub: str, name: str, is_veg: bool) -> str:
    veg = "vegetarian" if is_veg else "non-vegetarian"
    cl = cat.lower()

    if "soup" in cl:
        return f"{name} — aromatic {veg} soup simmered with select spices. Served piping-hot at {_HOUSE}."
    if "raita" in sub.lower():
        return f"{name} — cooling curd-based accompaniment, ideal with biryanis and kebabs. {_HOUSE}."
    if "starter" in cl and "tandoori" in sub.lower():
        return f"{name} — marinated and char-grilled in the tandoor for a smoky, tender finish. {_HOUSE}."
    if "starter" in cl and "chinese" in sub.lower():
        return f"{name} — Indo-Chinese style {veg} appetiser with bold sauces and crisp textures. {_HOUSE}."
    if "starter" in cl and "indian" in sub.lower():
        return f"{name} — traditional Indian {veg} appetiser, spiced and crisped to order. {_HOUSE}."
    if "starter" in cl:
        return f"{name} — {veg} starter with a bold spice profile. {_HOUSE}."
    if "main course" in cl and "chinese" in cl:
        return f"{name} — Indo-Chinese gravy preparation, wok-tossed with aromatic sauces. {_HOUSE}."
    if "main course" in cl:
        return f"{name} — slow-cooked {veg} curry from {_HOUSE}'s main-course kitchen."
    if "biryani" in sub.lower():
        return f"{name} — dum-cooked basmati layered with spices; served with raita and salan. {_HOUSE}."
    if "rice" in sub.lower():
        return f"{name} — fragrant rice preparation, perfectly seasoned. {_HOUSE}."
    if "noodle" in sub.lower():
        return f"{name} — stir-fried noodles with vegetables and signature sauces. {_HOUSE}."
    if "bread" in cl or "bread" in sub.lower():
        return f"{name} — freshly baked in the tandoor; soft and ideal with curries. {_HOUSE}."
    if "dessert" in cl:
        return f"{name} — a sweet finish to your meal at {_HOUSE}."
    if "beverage" in cl or "drink" in sub.lower():
        return f"{name} — refreshing drink from {_HOUSE}'s beverages menu."
    if "milkshake" in sub.lower():
        return f"{name} — thick, creamy milkshake blended fresh. {_HOUSE}."
    return f"{name} — a {veg} dish from {_HOUSE}."


# ── Parsing helpers ───────────────────────────────────────────────────────────
def parse_num(val) -> float | int:
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return int(val) if float(val) == int(val) else float(val)
    try:
        x = float(str(val).strip().rstrip("%"))
        return int(x) if x == int(x) else x
    except ValueError:
        return 0


# ── Main ──────────────────────────────────────────────────────────────────────
def enrich_workbook() -> list[dict]:
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    hdr_fill = PatternFill("solid", fgColor="2E86AB")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col, label in ((COL_DESC, "Description"), (COL_VEG, "isVeg"), (COL_IMG, "image")):
        c = ws.cell(1, col, label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    picker = _Picker()
    payloads: list[dict] = []
    seen_urls: set[str] = set()

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, COL_NAME).value
        if name is None or str(name).strip() == "":
            continue
        cat = str(ws.cell(r, COL_CAT).value or "").strip()
        sub = str(ws.cell(r, COL_SUB).value or "").strip()
        price = ws.cell(r, COL_PRICE).value
        hike = ws.cell(r, COL_HIKE).value
        name_s = str(name).strip()

        is_veg = not _is_nonveg(cat, sub, name_s)
        desc = build_description(cat, sub, name_s, is_veg)

        pool_key = select_pool(cat, sub, name_s)
        img_url = picker.pick(pool_key)
        if img_url in seen_urls:
            img_url = picker._recycled()
        seen_urls.add(img_url)

        ws.cell(r, COL_DESC, desc).alignment = wrap
        ws.cell(r, COL_GAP, None)
        ws.cell(r, COL_VEG, bool(is_veg)).alignment = center
        ws.cell(r, COL_IMG, img_url).alignment = wrap

        payloads.append(
            {
                "name": name_s,
                "restaurantPrice": parse_num(price),
                "hikePercentage": parse_num(hike),
                "category": cat,
                "subCategory": sub,
                "isVeg": bool(is_veg),
                "isAvailable": True,
                "description": desc,
                "image": [img_url],
                "restaurantId": RESTAURANT_ID,
            }
        )

    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 64

    wb.save(XLSX_PATH)
    JSON_PATH.write_text(json.dumps(payloads, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Enriched rows: {len(payloads)}")
    print(f"Saved workbook: {XLSX_PATH}")
    print(f"Saved JSON: {JSON_PATH}")
    print(f"Unique image URLs: {len(seen_urls)}")
    return payloads


def menu_item_api_body(row: dict) -> dict:
    cat = (row.get("category") or "").strip() or None
    sub = (row.get("subCategory") or "").strip() or None
    desc = row.get("description")
    if isinstance(desc, str):
        desc = desc.strip() or None
    return {
        "name": row["name"],
        "restaurantPrice": float(row["restaurantPrice"]),
        "hikePercentage": float(row["hikePercentage"]),
        "category": cat,
        "subCategory": sub,
        "isVeg": row.get("isVeg", True),
        "isAvailable": row.get("isAvailable", True),
        "description": desc,
        "image": row.get("image") or [],
    }


def post_menu_items(
    payloads: list[dict],
    *,
    api_url: str,
    bearer_token: str | None,
    delay_sec: float,
) -> int:
    menu_url = f"{api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"
    session = requests.Session()
    session.headers["Content-Type"] = "application/json"
    if bearer_token:
        session.headers["Authorization"] = f"Bearer {bearer_token}"
    else:
        session.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS

    ok = err = 0
    for i, raw in enumerate(payloads, 1):
        body = menu_item_api_body(raw)
        if delay_sec > 0:
            time.sleep(delay_sec)
        try:
            r = session.post(menu_url, json=body, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"  ✗ [{i}] {body['name']}: {ex}", file=sys.stderr)
            continue
        if r.status_code == 201:
            ok += 1
            print(f"  ✓ [{i}] {body['name']}")
        else:
            err += 1
            print(
                f"  ✗ [{i}] {body['name']}: HTTP {r.status_code} {r.text[:300]}",
                file=sys.stderr,
            )
            if r.status_code == 401:
                print("\nUnauthorized: set HONESTEATS_RETOOL_BYPASS or HONESTEATS_BEARER_TOKEN.", file=sys.stderr)
                return 1

    print(f"\nPOST done: {ok} created, {err} failed. URL base: {api_url.rstrip('/')}")
    return 0 if err == 0 else 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich Suraj menu + optional API import")
    parser.add_argument("--apply", action="store_true", help="POST each item to the API")
    parser.add_argument("--from-json", action="store_true", help="Load payloads from JSON instead of re-enriching")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--bearer-token", default=os.environ.get("HONESTEATS_BEARER_TOKEN"))
    parser.add_argument("--delay", type=float, default=0.3)
    args = parser.parse_args()

    if args.from_json:
        if not JSON_PATH.is_file():
            print(f"Missing {JSON_PATH}", file=sys.stderr)
            return 1
        payloads = json.loads(JSON_PATH.read_text(encoding="utf-8"))
        print(f"Loaded {len(payloads)} items from {JSON_PATH.name}")
    else:
        payloads = enrich_workbook()

    if not args.apply:
        print("No POST (pass --apply to import).")
        return 0

    return post_menu_items(
        payloads,
        api_url=args.api_url,
        bearer_token=args.bearer_token,
        delay_sec=args.delay,
    )


if __name__ == "__main__":
    raise SystemExit(main())
