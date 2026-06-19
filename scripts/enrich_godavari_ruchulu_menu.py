#!/usr/bin/env python3
"""
Read 'godavari ruchulu.xlsx' (columns A-D: Category, Subcategory, Item Name,
Price), enrich each row with a Veg/Non-Veg flag in column E and a hand-written
description in column I, write both back into the workbook, and emit JSON bodies
for the menu items. Does NOT call any API / insert anything.

Body shape (matches the example provided by the user):
  {
    "restaurantId": "RES-1781888349188-7788",
    "name": <Item from column C, capitalised>,
    "restaurantPrice": <number from column D>,
    "description": <generated description, also persisted to column I>,
    "hikePercentage": 0,
    "category": <capitalised column A>,
    "subCategory": <capitalised column B>,
    "isVeg": <bool derived from column E>,
    "isAvailable": true
  }

Default paths can be overridden with --input / --output.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

RESTAURANT_ID_DEFAULT = "RES-1781888349188-7788"
HIKE_PERCENTAGE = 0
VEG_COLUMN = 5          # column E
VEG_HEADER = "Veg/Non-Veg"
DESCRIPTION_COLUMN = 9  # column I
DESCRIPTION_HEADER = "Description"


# True = Veg, False = Non-Veg. Keyed by the capitalised item name (see enrich_label).
ITEM_VEG: dict[str, bool] = {
    "Veg Meals": True,
    "Non-Veg Meals (1 Egg, Curry)": False,
    "Meals With Fish Curry": False,
    "Meals With Thalakaya Curry": False,
    "Meals With Mutton Curry": False,
    "Konaseema Chicken Curry": False,
    "Chicken Masala": False,
    "Rayalaseema Chicken Curry": False,
    "Punjabi Chicken": False,
    "Butter Chicken": False,
    "Afghani Chicken": False,
    "Kadai Chicken": False,
    "Prawns Masala": False,
    "Fish Masala": False,
    "Egg Masala": False,
    "Chicken Fry Pulao": False,
    "Chicken Joint Pulao": False,
    "Chicken Ghee Roast Pulao": False,
    "Spl. Chicken Pulao (Boneless)": False,
    "Chicken Wings Pulao": False,
    "Royyala Pulao": False,
    "Panneer Fry Pulao": True,
    "Mushroom Fry Pulao": True,
    "Rajugari Kodi Pulao": False,
    "Kaju Fry Pulao": True,
    "Pachimirchi Panneer Pulao": True,
    "Chicken Dum Biryani": False,
    "Chicken Roast Biryani": False,
    "Chicken Spl. Biryani (Boneless)": False,
    "Chicken 65 Biryani": False,
    "Panneer Biryani": True,
    "Mushroom Biryani": True,
    "Kadai Veg": True,
    "Panneer Butter Masala": True,
    "Mushroom Masala": True,
    "Kaju Masala": True,
    "Kaju Panner Masala": True,
    "Kaju Mashroom Masala": True,
    "Methi Chman": True,
    "Pulka": True,
    "Chapathi": True,
    "Jonna Rotte": True,
    "French Fries": True,
    "Crispy Corn": True,
    "Chilli Panner": True,
    "Panner 65": True,
    "Mushroom 65": True,
    "Chilli Mushroom": True,
    "Dragon Panner": True,
    "Kaju Fry": True,
    "Baby Corn Munchuria": True,
    "Pachimirchi Panner Fry": True,
    "Jeedipappu Puttagudugula Pakodi": True,
    "Konaseema Puttagodugula Vepudu": True,
    "Konaseema Panneer Vepudu": True,
    "Omlet": False,
    "Egg Chilli Roast": False,
    "Egg Burji": False,
    "Chicken 65": False,
    "Chicken Lollipop (6p)": False,
    "Pepper Chicken": False,
    "Lemon Chicken": False,
    "Garlic Chicken": False,
    "Kaju Chicken": False,
    "Chilli Chicken": False,
    "Chicken Wings": False,
    "Veg Fried Rice": True,
    "Egg Fried Rice": False,
    "Chicken Fried Rice": False,
    "Panner Fried Rice": True,
    "Mushroom Fried Rice": True,
    "Kaju Fried Rice": True,
    "Konaseema Kodi Roast": False,
    "Guntur Karam Kodi Roast": False,
    "Nellore Chicken Vepudu": False,
    "Rayalaseema Kodi Wings": False,
    "Pachimirchi Kodi Roast": False,
    "Bangla Kodi Chips": False,
    "Garlic Prawns": False,
    "Apollo Fish": False,
    "Chilli Fish": False,
    "Chilli Royyala Vepudu": False,
}


ITEM_DESCRIPTIONS: dict[str, str] = {
    "Veg Meals": (
        "Traditional vegetarian meal with rice, sambar, rasam, assorted veg curries, curd, papad and pickle."
    ),
    "Non-Veg Meals (1 Egg, Curry)": (
        "Hearty rice meal with rasam and curd, served with a boiled egg and a non-veg curry."
    ),
    "Meals With Fish Curry": (
        "Full rice meal served with tangy Andhra-style fish curry, sambar, rasam, curd and papad."
    ),
    "Meals With Thalakaya Curry": (
        "Rice meal paired with traditional thalakaya (goat head) curry, sambar, rasam, curd and papad."
    ),
    "Meals With Mutton Curry": (
        "Rice meal served with spicy slow-cooked mutton curry, sambar, rasam, curd and papad."
    ),
    "Konaseema Chicken Curry": (
        "Coastal Konaseema-style chicken curry simmered in freshly ground spices and coconut."
    ),
    "Chicken Masala": (
        "Tender chicken cooked in a rich onion-tomato masala with aromatic spices."
    ),
    "Rayalaseema Chicken Curry": (
        "Fiery Rayalaseema-style chicken curry loaded with red chillies and bold spices."
    ),
    "Punjabi Chicken": (
        "North Indian-style chicken in a creamy, mildly spiced onion-tomato gravy."
    ),
    "Butter Chicken": (
        "Soft chicken pieces in a velvety buttery tomato gravy with a hint of cream."
    ),
    "Afghani Chicken": (
        "Mildly spiced chicken in a rich, creamy white cashew-and-yogurt gravy."
    ),
    "Kadai Chicken": (
        "Chicken tossed with bell peppers and onions in a freshly ground kadai masala."
    ),
    "Prawns Masala": (
        "Juicy prawns cooked in a spicy onion-tomato masala with coastal spices."
    ),
    "Fish Masala": (
        "Fish simmered in a tangy, spicy Andhra-style masala gravy."
    ),
    "Egg Masala": (
        "Boiled eggs in a spicy onion-tomato gravy tempered with aromatic spices."
    ),
    "Chicken Fry Pulao": (
        "Fragrant pulao layered with spicy chicken fry and caramelised onions."
    ),
    "Chicken Joint Pulao": (
        "Flavourful pulao cooked with bone-in chicken joints and whole spices."
    ),
    "Chicken Ghee Roast Pulao": (
        "Aromatic pulao served with rich, spicy chicken ghee roast."
    ),
    "Spl. Chicken Pulao (Boneless)": (
        "Special pulao loaded with tender boneless chicken and house spices."
    ),
    "Chicken Wings Pulao": (
        "Fragrant pulao paired with crispy spiced chicken wings."
    ),
    "Royyala Pulao": (
        "Coastal-style pulao cooked with juicy prawns (royyala) and aromatic spices."
    ),
    "Panneer Fry Pulao": (
        "Mildly spiced pulao tossed with golden pan-fried paneer cubes."
    ),
    "Mushroom Fry Pulao": (
        "Aromatic pulao with spiced sauteed mushrooms and fresh herbs."
    ),
    "Rajugari Kodi Pulao": (
        "Royal-style country chicken pulao slow-cooked with traditional spices."
    ),
    "Kaju Fry Pulao": (
        "Fragrant pulao tossed with crunchy fried cashews and mild spices."
    ),
    "Pachimirchi Panneer Pulao": (
        "Green-chilli flavoured pulao with soft paneer and a fresh, spicy kick."
    ),
    "Chicken Dum Biryani": (
        "Long-grain rice and marinated chicken slow-cooked on dum with aromatic spices."
    ),
    "Chicken Roast Biryani": (
        "Spicy roasted chicken layered with fragrant biryani rice and fried onions."
    ),
    "Chicken Spl. Biryani (Boneless)": (
        "Special boneless chicken biryani dum-cooked with rich spices and herbs."
    ),
    "Chicken 65 Biryani": (
        "Fragrant biryani topped with crispy, spicy chicken 65."
    ),
    "Panneer Biryani": (
        "Aromatic dum biryani layered with spiced paneer and fried onions."
    ),
    "Mushroom Biryani": (
        "Fragrant biryani cooked with spiced mushrooms and aromatic herbs."
    ),
    "Kadai Veg": (
        "Mixed vegetables tossed with capsicum and onions in a freshly ground kadai masala."
    ),
    "Panneer Butter Masala": (
        "Soft paneer cubes in a rich, creamy buttery tomato gravy."
    ),
    "Mushroom Masala": (
        "Mushrooms simmered in a spicy onion-tomato masala gravy."
    ),
    "Kaju Masala": (
        "Cashews cooked in a creamy, mildly spiced onion-tomato gravy."
    ),
    "Kaju Panner Masala": (
        "Cashews and paneer in a rich, creamy tomato-onion gravy."
    ),
    "Kaju Mashroom Masala": (
        "Cashews and mushrooms in a creamy, mildly spiced masala gravy."
    ),
    "Methi Chman": (
        "Soft paneer and fresh fenugreek leaves in a creamy, lightly spiced gravy."
    ),
    "Pulka": (
        "Soft, fluffy unleavened wheat flatbread cooked on an open flame."
    ),
    "Chapathi": (
        "Soft whole-wheat flatbread, lightly roasted and ideal with any curry."
    ),
    "Jonna Rotte": (
        "Traditional sorghum (jowar) flatbread, wholesome and gluten-free."
    ),
    "French Fries": (
        "Crispy golden potato fries lightly salted to perfection."
    ),
    "Crispy Corn": (
        "Crunchy fried sweet corn tossed with spices and herbs."
    ),
    "Chilli Panner": (
        "Paneer cubes tossed with onions, capsicum and a spicy Indo-Chinese chilli sauce."
    ),
    "Panner 65": (
        "Crispy fried paneer tossed in a spicy, tangy 65 masala."
    ),
    "Mushroom 65": (
        "Crispy fried mushrooms tossed in a spicy, tangy 65 masala."
    ),
    "Chilli Mushroom": (
        "Mushrooms tossed with onions, capsicum and a spicy Indo-Chinese chilli sauce."
    ),
    "Dragon Panner": (
        "Crispy paneer tossed in a fiery, sweet-and-spicy dragon sauce with cashews."
    ),
    "Kaju Fry": (
        "Crunchy cashews lightly fried and tossed with spices."
    ),
    "Baby Corn Munchuria": (
        "Crispy baby corn tossed in a tangy, spicy Manchurian sauce."
    ),
    "Pachimirchi Panner Fry": (
        "Paneer stir-fried with green chillies for a fresh, spicy kick."
    ),
    "Jeedipappu Puttagudugula Pakodi": (
        "Crispy fritters of cashews and mushrooms, fried to a golden crunch."
    ),
    "Konaseema Puttagodugula Vepudu": (
        "Konaseema-style mushroom fry tempered with curry leaves and spices."
    ),
    "Konaseema Panneer Vepudu": (
        "Konaseema-style paneer fry tossed with onions, curry leaves and spices."
    ),
    "Omlet": (
        "Fluffy egg omelette cooked with onions, green chillies and spices."
    ),
    "Egg Chilli Roast": (
        "Boiled eggs roasted with onions, capsicum and a spicy chilli masala."
    ),
    "Egg Burji": (
        "Scrambled eggs cooked with onions, tomatoes, green chillies and spices."
    ),
    "Chicken 65": (
        "Crispy deep-fried chicken tossed in a spicy, tangy 65 masala with curry leaves."
    ),
    "Chicken Lollipop (6p)": (
        "Six spicy, crispy frenched chicken drumettes - a juicy, finger-licking starter."
    ),
    "Pepper Chicken": (
        "Chicken tossed with crushed black pepper, curry leaves and spices."
    ),
    "Lemon Chicken": (
        "Crispy chicken tossed in a tangy lemon-pepper glaze."
    ),
    "Garlic Chicken": (
        "Chicken tossed in a punchy garlic sauce with a crisp finish."
    ),
    "Kaju Chicken": (
        "Chicken stir-fried with crunchy cashews and aromatic spices."
    ),
    "Chilli Chicken": (
        "Crispy chicken tossed with onions, capsicum and a spicy Indo-Chinese chilli sauce."
    ),
    "Chicken Wings": (
        "Crispy spiced chicken wings, juicy inside with a flavour-packed coating."
    ),
    "Veg Fried Rice": (
        "Wok-tossed rice with crunchy mixed vegetables and Indo-Chinese sauces."
    ),
    "Egg Fried Rice": (
        "Wok-tossed rice with scrambled egg, vegetables and Indo-Chinese sauces."
    ),
    "Chicken Fried Rice": (
        "Wok-tossed rice with spiced chicken, vegetables and Indo-Chinese sauces."
    ),
    "Panner Fried Rice": (
        "Wok-tossed rice with golden paneer cubes and crunchy vegetables."
    ),
    "Mushroom Fried Rice": (
        "Wok-tossed rice with sauteed mushrooms and crunchy vegetables."
    ),
    "Kaju Fried Rice": (
        "Wok-tossed rice with crunchy cashews and mixed vegetables."
    ),
    "Konaseema Kodi Roast": (
        "Konaseema-style country chicken roast cooked with bold coastal spices."
    ),
    "Guntur Karam Kodi Roast": (
        "Fiery Guntur-chilli chicken roast with a bold, spicy kick."
    ),
    "Nellore Chicken Vepudu": (
        "Nellore-style chicken fry tempered with curry leaves and aromatic spices."
    ),
    "Rayalaseema Kodi Wings": (
        "Spicy Rayalaseema-style chicken wings packed with red-chilli heat."
    ),
    "Pachimirchi Kodi Roast": (
        "Country chicken roast with green chillies for a fresh, fiery kick."
    ),
    "Bangla Kodi Chips": (
        "Crispy thin chicken chips fried to a crunchy, spicy finish."
    ),
    "Garlic Prawns": (
        "Juicy prawns tossed in a punchy garlic sauce with a crisp finish."
    ),
    "Apollo Fish": (
        "Crispy boneless fish tossed in a tangy, spicy Apollo-style masala."
    ),
    "Chilli Fish": (
        "Crispy fish tossed with onions, capsicum and a spicy Indo-Chinese chilli sauce."
    ),
    "Chilli Royyala Vepudu": (
        "Prawns stir-fried with green chillies, onions and curry leaves for a spicy coastal kick."
    ),
}


def _num(v: Any) -> float | int:
    """Parse a numeric cell; use int when value is mathematically whole."""
    if v is None or (isinstance(v, str) and not str(v).strip()):
        return 0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and v == int(v):
            return int(v)
        return v
    s = str(v).strip().replace(",", "")
    try:
        x = float(s)
        return int(x) if x == int(x) else x
    except ValueError:
        return 0


def enrich_label(value: Any) -> str:
    """Trim, normalize whitespace, then capitalize each word start (after space, /, -, or opening bracket)."""
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        s = str(value).strip()
        if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
            s = s[:-2]
        return s
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    s = " ".join(s.split()).lower()
    return re.sub(r"(^|[\s(/\-])([a-zà-ÿ])", lambda m: m.group(1) + m.group(2).upper(), s)


def _parse_is_veg(value: Any) -> bool | None:
    """Read an existing Veg/Non-Veg cell; None when blank/unrecognised."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    s = str(value).strip().lower()
    if s in {"veg", "vegetarian", "true", "v", "yes", "1"}:
        return True
    if s in {"non-veg", "nonveg", "non veg", "false", "nv", "no", "0"}:
        return False
    return None


def infer_is_veg(name: str) -> bool:
    """Veg flag from the embedded map; default to non-veg if an unknown item appears."""
    return ITEM_VEG.get(name, False)


def generic_description(name: str, category: str, sub_category: str) -> str:
    """Last-resort fallback when an item is not in ITEM_DESCRIPTIONS."""
    bits = [b for b in (sub_category, category) if b]
    suffix = f" ({' · '.join(bits)})" if bits else ""
    return f"{name}{suffix} - a freshly prepared dish from Godavari Ruchulu."


def describe(name: str, category: str, sub_category: str) -> str:
    return ITEM_DESCRIPTIONS.get(name) or generic_description(name, category, sub_category)


def row_to_body(
    name: str,
    price: Any,
    description: str,
    category: str,
    sub_category: str,
    is_veg: bool,
    restaurant_id: str,
) -> dict[str, Any]:
    return {
        "restaurantId": restaurant_id,
        "name": name,
        "restaurantPrice": _num(price),
        "description": description,
        "hikePercentage": HIKE_PERCENTAGE,
        "category": category,
        "subCategory": sub_category,
        "isVeg": is_veg,
        "isAvailable": True,
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--input",
        type=Path,
        default=Path("/Users/user/Downloads/godavari ruchulu.xlsx"),
        help="Path to the Godavari Ruchulu pricing workbook",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=Path("/Users/user/Downloads/godavari_ruchulu_menu_payloads.json"),
        help="Where to write the JSON array of menu bodies",
    )
    p.add_argument(
        "--restaurant-id",
        default=RESTAURANT_ID_DEFAULT,
        help="Restaurant id stamped on each payload",
    )
    p.add_argument(
        "--no-write-xlsx",
        action="store_true",
        help="Skip writing Veg/Non-Veg (col E) and descriptions (col I) back into the workbook",
    )
    args = p.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        return 1

    if not args.input.is_file():
        print(f"Input not found: {args.input}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active

    if ws.cell(row=1, column=VEG_COLUMN).value in (None, ""):
        ws.cell(row=1, column=VEG_COLUMN).value = VEG_HEADER
    if ws.cell(row=1, column=DESCRIPTION_COLUMN).value in (None, ""):
        ws.cell(row=1, column=DESCRIPTION_COLUMN).value = DESCRIPTION_HEADER

    bodies: list[dict[str, Any]] = []
    missing_desc: list[str] = []
    missing_veg: list[str] = []

    for row_idx in range(2, ws.max_row + 1):
        raw_cat = ws.cell(row=row_idx, column=1).value
        raw_sub = ws.cell(row=row_idx, column=2).value
        raw_name = ws.cell(row=row_idx, column=3).value
        raw_price = ws.cell(row=row_idx, column=4).value

        if str(raw_name or "").strip() == "":
            continue

        name = enrich_label(raw_name)
        category = enrich_label(raw_cat)
        sub_category = enrich_label(raw_sub)

        # Veg flag: respect an existing column-E value, else fall back to the embedded map.
        is_veg = _parse_is_veg(ws.cell(row=row_idx, column=VEG_COLUMN).value)
        if is_veg is None:
            if name not in ITEM_VEG:
                missing_veg.append(name)
            is_veg = infer_is_veg(name)

        description = describe(name, category, sub_category)
        if name not in ITEM_DESCRIPTIONS:
            missing_desc.append(name)

        ws.cell(row=row_idx, column=VEG_COLUMN).value = "Veg" if is_veg else "Non-Veg"
        ws.cell(row=row_idx, column=DESCRIPTION_COLUMN).value = description

        bodies.append(
            row_to_body(
                name=name,
                price=raw_price,
                description=description,
                category=category,
                sub_category=sub_category,
                is_veg=is_veg,
                restaurant_id=args.restaurant_id,
            )
        )

    if not args.no_write_xlsx:
        wb.save(args.input)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(bodies, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    veg = sum(1 for b in bodies if b["isVeg"])
    print(f"Wrote {len(bodies)} bodies to {args.output} (Veg: {veg}, Non-Veg: {len(bodies) - veg})")
    if not args.no_write_xlsx:
        print(f"Wrote Veg/Non-Veg to column {VEG_COLUMN} and descriptions to column {DESCRIPTION_COLUMN} of {args.input}")
    if missing_veg:
        print(f"WARNING: {len(missing_veg)} item(s) not in ITEM_VEG (defaulted to Non-Veg): {missing_veg}", file=sys.stderr)
    if missing_desc:
        print(f"WARNING: {len(missing_desc)} item(s) used the generic fallback description: {missing_desc}", file=sys.stderr)
    if bodies:
        print("Sample (first item):\n", json.dumps(bodies[0], indent=2, ensure_ascii=False))
    print("No API calls were made. Nothing was inserted.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
