#!/usr/bin/env python3
"""
Read 'Shree Anna Millet house — Pricing.xlsx' (columns A–E: Category, Subcategory,
Item, Dine-in Price, Hike), enrich each row with a hand-written description,
write that description back into column I of the workbook, and emit JSON bodies
for the menu items. Does NOT call any API.

Body shape (matches the example provided by the user):
  {
    "restaurantId": "RES-1779292271565-1476",
    "name": <Item from column C>,
    "restaurantPrice": <number from column D>,
    "description": <generated description, also persisted to column I>,
    "hikePercentage": 35,
    "category": <Title-cased column A>,
    "subCategory": <Title-cased column B>,
    "isVeg": true,
    "isAvailable": true
  }

Default paths can be overridden with --input / --output.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

RESTAURANT_ID_DEFAULT = "RES-1779292271565-1476"
HIKE_PERCENTAGE = 35
DESCRIPTION_COLUMN = 9
DESCRIPTION_HEADER = "Description"


ITEM_DESCRIPTIONS: dict[str, str] = {
    "Millet Plain Dosa": (
        "Crisp, golden millet dosa made from fermented millet batter, served with sambar and coconut chutney."
    ),
    "Millet Karam Dosa": (
        "Millet dosa smeared with fiery red karam chutney for a spicy kick, paired with sambar."
    ),
    "Millet Masala Dosa": (
        "Classic millet dosa wrapped around a lightly spiced potato masala filling, served with sambar and chutneys."
    ),
    "Millet Onion Dosa": (
        "Crispy millet dosa topped with finely chopped onions and green chilies, served with sambar."
    ),
    "Millet Ghee Karam Dosa": (
        "Millet dosa roasted in pure ghee and brushed with spicy karam chutney."
    ),
    "Millet Onion Uttapam": (
        "Thick, soft millet pancake topped with onions, tomatoes and green chilies, griddled to a golden finish."
    ),
    "Millet Upma Dosa": (
        "Soft upma layered inside a crisp millet dosa — a hearty signature combination."
    ),
    "Millet Special Dosa": (
        "House-special millet dosa loaded with masala, onions and a generous drizzle of ghee."
    ),
    "Millet Paneer Dosa": (
        "Millet dosa filled with spiced paneer crumble, served with sambar and chutneys."
    ),
    "Millet Palak Paneer Dosa": (
        "Millet dosa stuffed with paneer cooked in a fresh spinach (palak) gravy."
    ),
    "Millet Thattu Idly": (
        "Soft, fluffy thattu-style millet idlies served with sambar and chutney."
    ),
    "Millet Single Idly": (
        "One soft, steam-cooked millet idly served with sambar and chutney."
    ),
    "Millet Idly (2)": (
        "Two soft millet idlies, light and easy to digest, served with sambar and chutney."
    ),
    "Millet Idly (Full)": (
        "Full plate of soft millet idlies served with sambar and an assortment of chutneys."
    ),
    "Millet Sambar Idly": (
        "Soft millet idlies dunked in piping-hot, tangy sambar topped with ghee."
    ),
    "Millet Tawa Idly": (
        "Millet idlies tossed on a tawa with spices, curry leaves and ghee for a smoky finish."
    ),
    "Millet Paneer Tawa Idly": (
        "Tawa-fried millet idlies tossed with crumbled paneer, onions and spices."
    ),
    "Millet Chitti Idly": (
        "Bite-sized millet chitti idlies served in tangy sambar — perfect for sharing."
    ),
    "Single Vada": (
        "One crisp medu vada with a fluffy center, served with sambar and chutney."
    ),
    "Vada (Full)": (
        "Full plate of crisp medu vadas with fluffy centers, served with sambar and chutney."
    ),
    "Sambar Vada": (
        "Crisp medu vada soaked in tangy, ghee-tempered sambar."
    ),
    "Ragi Idly Vada": (
        "Soft ragi idly paired with a crisp medu vada, served with sambar and chutney."
    ),
    "Jowar Idly Vada": (
        "Soft jowar millet idly paired with a crisp medu vada, served with sambar and chutney."
    ),
    "Multi Millet Idly Vada": (
        "Multi-millet idly served alongside a crisp medu vada, with sambar and chutney."
    ),
    "Tawa Vada": (
        "Medu vada tossed on a tawa with spices, onions and curry leaves for a crisp, savory finish."
    ),
    "Perugu Vada": (
        "Cool, creamy yogurt-soaked vadas tempered with mustard and curry leaves (dahi vada style)."
    ),
    "Perugu Vada (1)": (
        "Single yogurt-soaked vada tempered with mustard and curry leaves."
    ),
    "Millet Thali": (
        "Wholesome millet meal with rice, dal, sambar, curry, papad, curd and pickle on a single plate."
    ),
    "Multi Millet Kichidi": (
        "Comforting one-pot khichdi made with a blend of millets, lentils and mild spices."
    ),
    "Millet Curd Rice": (
        "Cooling millet curd rice tempered with mustard, curry leaves and ginger."
    ),
    "Millet Bisi Bele Bath": (
        "Classic Karnataka-style spiced rice with lentils and vegetables, made wholesome with millet."
    ),
    "Single Upma": (
        "Steamy semolina upma tempered with mustard, curry leaves and cashews."
    ),
    "Jowar Upma": (
        "Light, savory upma made with jowar millet, tempered with mustard and curry leaves."
    ),
    "Foxtail Upma": (
        "Soft foxtail millet upma tempered with vegetables, mustard and curry leaves."
    ),
    "Kodo Upma": (
        "Hearty kodo millet upma tempered with curry leaves, mustard and a hint of ginger."
    ),
    "Single Pongal": (
        "Soft, peppery pongal with rice, moong dal, ghee, cashews and crushed black pepper."
    ),
    "Pongal (Foxtail)": (
        "Traditional pongal made with foxtail millet, ghee, cashews and crushed pepper."
    ),
    "Single Kichidi": (
        "Single-portion millet khichdi with lentils and mild spices — light and comforting."
    ),
    "Single Curd Rice": (
        "Single portion of cooling curd rice tempered with mustard and curry leaves."
    ),
    "Curd Rice (Little)": (
        "Light, refreshing millet curd rice finished with a mustard and curry-leaf tempering."
    ),
    "Jonna Sangati": (
        "Jowar sangati — soft Andhra-style millet dumplings, served with curry."
    ),
    "Jonna Sangati (Full)": (
        "Full plate of jowar sangati served with a side curry — a traditional Rayalaseema meal."
    ),
    "Single Puri": (
        "One puffed, golden-fried puri served with potato curry."
    ),
    "Millet Puri": (
        "Golden puffed millet puris served with a spiced potato curry."
    ),
    "Single Chapathi": (
        "Single soft whole-wheat chapathi served warm with curry."
    ),
    "Millet Chapathi": (
        "Soft millet chapathi made with wholesome millet flour, served with curry."
    ),
    "Rayachoor Jonna Roti (2)": (
        "Two Rayachoor-style jowar rotis served with a flavourful curry."
    ),
    "Rayachoor Jonna Roti (1)": (
        "One traditional Rayachoor-style jowar roti served with curry."
    ),
    "1 Jonna Roti Without Curry": (
        "One Rayachoor-style jowar (millet) roti, served plain without curry."
    ),
    "Ambali (Spiced)": (
        "Spiced fermented millet drink (ambali) — a cooling, gut-friendly traditional beverage."
    ),
    "Ashwagandha Health Mix": (
        "Nutritious health mix with ashwagandha and millets, blended in warm milk."
    ),
    "Bellam Coffee": (
        "South Indian filter coffee sweetened naturally with jaggery (bellam) instead of sugar."
    ),
    "Bellam Badam Palu": (
        "Warm almond milk sweetened with jaggery — a rich, traditional bedtime drink."
    ),
    "Desi Cow Buttermilk": (
        "Cool, frothy buttermilk from desi-cow milk, tempered with ginger and curry leaves."
    ),
}


def _num(v: Any) -> float | int:
    """Parse a numeric cell; use int when value is mathematically whole."""
    if v is None or (isinstance(v, str) and not str(v).strip()):
        return 0
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and v == int(v):
            return int(v)
        return v
    s = str(v).strip().replace(",", "")
    try:
        x = float(s)
        return int(x) if x == int(x) else x
    except ValueError:
        return 0


def enrich_label(value: Any) -> str:
    """Trim, normalize whitespace, then capitalize each word start (after space, /, -, or opening bracket)."""
    if value is None:
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        s = str(value).strip()
        if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
            s = s[:-2]
        return s
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""
    s = " ".join(s.split()).lower()
    return re.sub(r"(^|[\s(/\-])([a-zà-ÿ])", lambda m: m.group(1) + m.group(2).upper(), s)


def generic_description(name: str, category: str, sub_category: str) -> str:
    """Last-resort fallback when an item is not in ITEM_DESCRIPTIONS."""
    bits = [b for b in (sub_category, category) if b]
    suffix = f" ({' · '.join(bits)})" if bits else ""
    return f"{name}{suffix} — a wholesome millet-based preparation from Shree Anna Millet House."


def describe(name: str, category: str, sub_category: str) -> str:
    return ITEM_DESCRIPTIONS.get(name) or generic_description(name, category, sub_category)


def row_to_body(
    name: str,
    price: Any,
    description: str,
    category: str,
    sub_category: str,
    restaurant_id: str,
) -> dict[str, Any]:
    return {
        "restaurantId": restaurant_id,
        "name": name,
        "restaurantPrice": _num(price),
        "description": description,
        "hikePercentage": HIKE_PERCENTAGE,
        "category": category,
        "subCategory": sub_category,
        "isVeg": True,
        "isAvailable": True,
    }


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--input",
        type=Path,
        default=Path("/Users/user/Downloads/Shree Anna Millet house — Pricing.xlsx"),
        help="Path to the Shree Anna Millet House pricing workbook",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=Path("/Users/user/Downloads/Shree_Anna_Millet_House_menu_payloads.json"),
        help="Where to write the JSON array of menu bodies",
    )
    p.add_argument(
        "--restaurant-id",
        default=RESTAURANT_ID_DEFAULT,
        help="Restaurant id stamped on each payload",
    )
    p.add_argument(
        "--no-write-xlsx",
        action="store_true",
        help="Skip writing descriptions back into column I of the workbook",
    )
    args = p.parse_args()

    try:
        import openpyxl
    except ImportError:
        print("Install openpyxl: pip install openpyxl", file=sys.stderr)
        return 1

    if not args.input.is_file():
        print(f"Input not found: {args.input}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(args.input, data_only=True)
    ws = wb.active

    if ws.cell(row=1, column=DESCRIPTION_COLUMN).value in (None, ""):
        ws.cell(row=1, column=DESCRIPTION_COLUMN).value = DESCRIPTION_HEADER

    bodies: list[dict[str, Any]] = []
    missing_desc: list[str] = []

    for row_idx in range(2, ws.max_row + 1):
        raw_cat = ws.cell(row=row_idx, column=1).value
        raw_sub = ws.cell(row=row_idx, column=2).value
        raw_name = ws.cell(row=row_idx, column=3).value
        raw_price = ws.cell(row=row_idx, column=4).value

        if raw_name is None and raw_cat is None and raw_sub is None:
            continue
        if str(raw_name or "").strip() == "":
            continue

        name = enrich_label(raw_name)
        category = enrich_label(raw_cat)
        sub_category = enrich_label(raw_sub)
        description = describe(name, category, sub_category)

        if name not in ITEM_DESCRIPTIONS:
            missing_desc.append(name)

        ws.cell(row=row_idx, column=DESCRIPTION_COLUMN).value = description

        bodies.append(
            row_to_body(
                name=name,
                price=raw_price,
                description=description,
                category=category,
                sub_category=sub_category,
                restaurant_id=args.restaurant_id,
            )
        )

    if not args.no_write_xlsx:
        wb.save(args.input)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(bodies, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Wrote {len(bodies)} bodies to {args.output}")
    if not args.no_write_xlsx:
        print(f"Wrote descriptions to column {DESCRIPTION_COLUMN} of {args.input}")
    if missing_desc:
        print(
            f"WARNING: {len(missing_desc)} item(s) used the generic fallback description: "
            f"{missing_desc}",
            file=sys.stderr,
        )
    if bodies:
        print("Sample (first item):\n", json.dumps(bodies[0], indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
