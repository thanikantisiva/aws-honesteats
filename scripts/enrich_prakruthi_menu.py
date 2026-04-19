#!/usr/bin/env python3
"""
Enrich Prakruthi Millet Foods menu Excel and emit menu-item JSON; optional POST to API.

Expected sheet columns A–E:
  Category | Sub-Category | Item Name | Price (₹) | Our Hike on Dine-in %

Adds / refreshes:
  F — Description
  G — (reserved, left blank)
  H — isVeg (TRUE — all items veg)
  I — image (single Unsplash URL per row; no repeats within the sheet)

Writes JSON array of objects (same field order; restaurantId last for import):
  name, restaurantPrice, hikePercentage, category, subCategory, isVeg,
  isAvailable, description, image, restaurantId

Usage:
  python3 scripts/enrich_prakruthi_menu.py                    # enrich + JSON only
  python3 scripts/enrich_prakruthi_menu.py --apply             # enrich + POST each item
  python3 scripts/enrich_prakruthi_menu.py --from-json --apply # POST from JSON (no Excel write)

Prod: set HONESTEATS_API_URL (default https://api.yumdude.com) and either
  HONESTEATS_RETOOL_BYPASS (x-retool-header) or HONESTEATS_BEARER_TOKEN.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

XLSX_PATH = Path("/Users/user/Downloads/Prakruthi_Millet_Foods_Menu.xlsx")
JSON_PATH = Path("/Users/user/Downloads/Prakruthi_Millet_Foods_Menu.menu_items.json")
RESTAURANT_ID = "RES-1776527535909-6436"

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get(
    "HONESTEATS_RETOOL_BYPASS",
    "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc",
)
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

COL_CAT, COL_SUB, COL_NAME, COL_PRICE, COL_HIKE = 1, 2, 3, 4, 5
COL_DESC, COL_GAP, COL_VEG, COL_IMG = 6, 7, 8, 9

# 40+ distinct Unsplash photo IDs (food / millet / Indian / drinks) — no duplicates.
_UNSPLASH_IDS = [
    "1497034825429-c343d7c6a68f",
    "1501443762994-82bd5dace89a",
    "1560008581-09826d1de69e",
    "1488900128323-21503983a07e",
    "1570197571499-166b36435e9f",
    "1629385701021-fcd568a743e8",
    "1576506295286-5cda18df43e7",
    "1633933358116-a27b902fad35",
    "1514849302-984523450cf4",
    "1563805042-7684c019e1cb",
    "1551024506-0bccd828d307",
    "1580915411954-282cb1b0d780",
    "1579954115545-a95591f28bfc",
    "1557142046-c704a3adf364",
    "1505394033641-40c6ad1178d7",
    "1516559828984-fb3b99548b21",
    "1621303837174-89787a7d4729",
    "1587563871167-1ee9c731aefb",
    "1612203985729-70726954388c",
    "1615478503562-ec2d8aa0e24e",
    "1567206563064-6f60f40a2b57",
    "1543255006-d6395b6f1171",
    "1595348020949-87cdfbb44174",
    "1600002423562-975eabb78d5a",
    "1632170684742-9c8b38c1aeab",
    "1628607189631-96e9e8a3cedc",
    "1546069901-ba9599a7e63c",
    "1622483767028-3f66f32aef97",
    "1625869767142-1fb8faf7e8d9",
    "1572490122747-3968b75cc699",
    "1541658016709-82535e94bc69",
    "1553787499-6f9133860278",
    "1568901839119-631418a3910d",
    "1619158401201-8fa932695178",
    "1626082927389-6cd097cdc6ec",
    "1601050690597-df0568f70950",
    "1540189549336-e6e99c3679fe",
    "1567337710282-00832b415979",
    "1585032226651-759b368d7246",
    "1572715376701-98568319fd0b",
    "1559847844-5315695dadae",
    "1606574977732-e8e5f1f46c23",
    "1564834724105-918b73d1b8e0",
    "1512621776951-a57141f2eefd",
    "1574484284002-952d92456975",
    "1567620832903-9fc6debc209f",
    "1606728035253-49e8a23146de",
    "1599487488170-d11ec9c172f0",
    "1565299624946-b28f40a0ae38",
    "1573080496219-bb080dd4f877",
    "1585937421612-70a008356fbe",
    "1523677013480-48917ea9fd5f",
    "1497534446932-c925b458314e",
    "1511920170033-f8396924c348",
    "1556679343-c73006c196f3",
    "1504753794599-61249cd1bd47",
]

UNSPLASH_IDS = [pid for pid in _UNSPLASH_IDS if re.match(r"^[0-9a-zA-Z-]+$", pid)]


def unsplash_url(photo_id: str) -> str:
    return (
        f"https://images.unsplash.com/photo-{photo_id}"
        "?auto=format&fit=crop&w=800&q=80"
    )


def clean_display_name(raw: str | None) -> str:
    if not raw:
        return ""
    s = str(raw)
    s = re.sub(r"\s*\([^)]*\)\s*", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def build_description(category: str | None, subcategory: str | None, item_name: str | None) -> str:
    cat = (category or "").strip()
    sub = (subcategory or "").strip()
    base = clean_display_name(item_name) or (item_name or "Menu item").strip()
    focus = "Prakruthi Millet Foods focuses on whole millets, traditional grains, and natural jaggery."

    if "Hot Drinks" in cat or "hot drink" in (sub or "").lower():
        return (
            f"{base} — a warming drink from our hot drinks section, made with millet-based "
            f"ferments or jaggery-sweetened brews. {focus}"
        )
    if "Healthy Juices" in cat or "juice" in base.lower():
        return (
            f"{base} — fresh, vegetable- and fruit-forward juice from our healthy juices menu. {focus}"
        )
    if "Afternoon" in cat or "Meals" in sub and "Afternoon" in cat:
        if "payasam" in base.lower():
            return f"{base} — a millet-based sweet payasam, gently cooked for afternoon service. {focus}"
        if "sangati" in base.lower() or "roti" in base.lower():
            return f"{base} — hearty millet meal or flatbread, ideal with curries and sides. {focus}"
        return f"{base} — a wholesome afternoon plate built around millets and organic grains. {focus}"
    if "Evening" in cat:
        if "idli" in base.lower():
            return f"{base} — soft steamed millet idlis, fermented for flavour and digestibility. {focus}"
        if "khichdi" in base.lower() or "bath" in base.lower() or "pongali" in base.lower():
            return f"{base} — one-pot millet comfort food, spiced and finished for evening dining. {focus}"
        return f"{base} — evening millet specialities from our kitchen. {focus}"
    if "Dosa" in sub or "dosa" in base.lower() or "uthappam" in base.lower():
        return (
            f"{base} — crisp or soft millet dosa from the tawa, with classic Andhra-style toppings as listed. {focus}"
        )
    return f"{base} — millet-forward dish from {sub or cat or 'our menu'}. {focus}"


def parse_hike(val) -> float | int:
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return int(val) if float(val) == int(val) else float(val)
    try:
        x = float(str(val).strip().rstrip("%"))
        return int(x) if x == int(x) else x
    except ValueError:
        return 0


def parse_price(val) -> float | int:
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val) if float(val) == int(val) else float(val)
    try:
        x = float(str(val).strip())
        return int(x) if x == int(x) else x
    except ValueError:
        return 0


def enrich_workbook() -> list[dict]:
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    hdr_fill = PatternFill("solid", fgColor="2E86AB")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    ws.cell(1, COL_DESC, "Description").font = hdr_font
    ws.cell(1, COL_DESC).fill = hdr_fill
    ws.cell(1, COL_DESC).alignment = center
    ws.cell(1, COL_GAP, "").alignment = center
    ws.cell(1, COL_VEG, "isVeg").font = hdr_font
    ws.cell(1, COL_VEG).fill = hdr_fill
    ws.cell(1, COL_VEG).alignment = center
    ws.cell(1, COL_IMG, "image").font = hdr_font
    ws.cell(1, COL_IMG).fill = hdr_fill
    ws.cell(1, COL_IMG).alignment = center

    urls = [unsplash_url(pid) for pid in UNSPLASH_IDS]
    seen: set[str] = set()
    unique_urls: list[str] = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            unique_urls.append(u)

    payloads: list[dict] = []
    idx_img = 0

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, COL_NAME).value
        if name is None or str(name).strip() == "":
            continue
        cat = ws.cell(r, COL_CAT).value
        sub = ws.cell(r, COL_SUB).value
        price = ws.cell(r, COL_PRICE).value
        hike = ws.cell(r, COL_HIKE).value

        desc = build_description(
            str(cat) if cat else "",
            str(sub) if sub else "",
            str(name),
        )
        if idx_img >= len(unique_urls):
            raise RuntimeError(
                f"Ran out of unique Unsplash URLs (need more IDs for row {r}). "
                f"Extend UNSPLASH_IDS in {__file__}."
            )
        img_url = unique_urls[idx_img]
        idx_img += 1

        ws.cell(r, COL_DESC, desc).alignment = wrap
        ws.cell(r, COL_GAP, None)
        ws.cell(r, COL_VEG, True)
        ws.cell(r, COL_VEG).alignment = center
        ws.cell(r, COL_IMG, img_url).alignment = wrap

        hp = parse_hike(hike)
        rp = parse_price(price)
        payloads.append(
            {
                "name": str(name).strip(),
                "restaurantPrice": rp,
                "hikePercentage": hp,
                "category": str(cat).strip() if cat else "",
                "subCategory": str(sub).strip() if sub else "",
                "isVeg": True,
                "isAvailable": True,
                "description": desc,
                "image": [img_url],
                "restaurantId": RESTAURANT_ID,
            }
        )

    ws.column_dimensions["F"].width = 56
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 64

    wb.save(XLSX_PATH)
    JSON_PATH.write_text(json.dumps(payloads, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Enriched rows: {len(payloads)}")
    print(f"Saved workbook: {XLSX_PATH}")
    print(f"Saved JSON: {JSON_PATH}")
    return payloads


def menu_item_api_body(row: dict) -> dict:
    """POST body for /api/v1/restaurants/{id}/menu (no restaurantId)."""
    cat = (row.get("category") or "").strip() or None
    sub = (row.get("subCategory") or "").strip() or None
    desc = row.get("description")
    if isinstance(desc, str):
        desc = desc.strip() or None
    img = row.get("image") or []
    return {
        "name": row["name"],
        "restaurantPrice": float(row["restaurantPrice"]),
        "hikePercentage": float(row["hikePercentage"]),
        "category": cat,
        "subCategory": sub,
        "isVeg": row.get("isVeg", True),
        "isAvailable": row.get("isAvailable", True),
        "description": desc,
        "image": img,
    }


def post_menu_items(
    payloads: list[dict],
    *,
    api_url: str,
    bearer_token: str | None,
    delay_sec: float,
) -> int:
    menu_url = f"{api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"
    session = requests.Session()
    session.headers["Content-Type"] = "application/json"
    if bearer_token:
        session.headers["Authorization"] = f"Bearer {bearer_token}"
    else:
        session.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS

    ok = err = 0
    for i, raw in enumerate(payloads, 1):
        body = menu_item_api_body(raw)
        if delay_sec > 0:
            time.sleep(delay_sec)
        try:
            r = session.post(menu_url, json=body, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"  ✗ [{i}] {body['name']}: {ex}", file=sys.stderr)
            continue
        if r.status_code == 201:
            ok += 1
            print(f"  ✓ [{i}] {body['name']}")
        else:
            err += 1
            print(
                f"  ✗ [{i}] {body['name']}: HTTP {r.status_code} {r.text[:300]}",
                file=sys.stderr,
            )
            if r.status_code == 401:
                print(
                    "\nUnauthorized: set HONESTEATS_RETOOL_BYPASS (prod secret) or "
                    "HONESTEATS_BEARER_TOKEN.",
                    file=sys.stderr,
                )
                return 1

    print(f"\nPOST done: {ok} created, {err} failed. URL base: {api_url.rstrip('/')}")
    return 0 if err == 0 else 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich Prakruthi menu Excel + optional API import")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="POST each item to HONESTEATS_API_URL (default prod yumdude)",
    )
    parser.add_argument(
        "--from-json",
        action="store_true",
        help="Load payloads from JSON_PATH instead of re-enriching Excel",
    )
    parser.add_argument("--api-url", default=DEFAULT_API_URL, help="API base URL")
    parser.add_argument(
        "--bearer-token",
        default=os.environ.get("HONESTEATS_BEARER_TOKEN"),
        help="JWT (optional; else x-retool-header from HONESTEATS_RETOOL_BYPASS)",
    )
    parser.add_argument("--delay", type=float, default=0.3, help="Seconds between POSTs")
    args = parser.parse_args()

    if args.from_json:
        if not JSON_PATH.is_file():
            print(f"Missing {JSON_PATH}", file=sys.stderr)
            return 1
        payloads = json.loads(JSON_PATH.read_text(encoding="utf-8"))
        print(f"Loaded {len(payloads)} items from {JSON_PATH.name}")
    else:
        payloads = enrich_workbook()

    if not args.apply:
        print("No POST (pass --apply to import).")
        return 0

    return post_menu_items(
        payloads,
        api_url=args.api_url,
        bearer_token=args.bearer_token,
        delay_sec=args.delay,
    )


if __name__ == "__main__":
    raise SystemExit(main())
