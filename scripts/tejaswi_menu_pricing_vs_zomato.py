"""
Populate pricing comparison columns on Tejaswi dine-in vs Zomato xlsx.

Input: ~/Downloads/Tejaswi_DineIn_vs_Zomato.xlsx
Columns A=Category, B=Subcategory, C=Item, D=Dine-in (D), E=Zomato (Z).

Zomato commission 36% → restaurant gets 64% of Z.
Our commission 17% → restaurant gets 83% of our listed price P.

With Z:
  Zomato Hike % = (Z - D) / D * 100
  R_z = Z * 0.64
  P = round((0.40 * Z + 0.60 * R_z) / 0.85)
  Constraints: P < Z AND 0.83 * P > R_z

Without Z (Zomato cell is "-" / empty):
  P = round(D * 1.20)          # fixed 20% hike on dine-in
  Zomato-related columns left blank.

Run: python3 scripts/tejaswi_menu_pricing_vs_zomato.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SRC = Path.home() / "Downloads" / "Tejaswi_DineIn_vs_Zomato.xlsx"

NEW_HEADERS = [
    "Zomato Hike %",
    "Restaurant gets from Zomato R_z (\u20b9)",
    "Our Recommended Price P (\u20b9)",
    "Our Hike on Dine-in %",
    "Restaurant gets from Us (\u20b9)",
    "Our Revenue (per order) (\u20b9)",
    "Customer saves vs Zomato (\u20b9)",
    "Customer saves vs Zomato %",
    "Restaurant gains vs Zomato (\u20b9)",
]

ZOMATO_RATIO = 0.64   # restaurant share on Zomato
OUR_RATIO = 0.83      # restaurant share on our app
NO_Z_HIKE_PCT = 20.0  # fixed hike when Zomato price absent


def _num(x) -> float | None:
    if x is None:
        return None
    if isinstance(x, str):
        s = x.strip()
        if not s or s in {"-", "—"}:
            return None
        try:
            return float(s.replace(",", ""))
        except ValueError:
            return None
    if isinstance(x, (int, float)):
        return float(x)
    return None


def compute_p_with_z(Z: float) -> tuple[int, float]:
    R_z = Z * ZOMATO_RATIO
    P = round((0.40 * Z + 0.60 * R_z) / 0.85)
    # Constraint satisfaction: P < Z AND OUR_RATIO * P > R_z
    if not (P < Z and OUR_RATIO * P > R_z):
        for delta in (0, -1, 1, -2, 2, -3, 3, -4, 4, -5, 5):
            cand = P + delta
            if cand > 0 and cand < Z and OUR_RATIO * cand > R_z:
                P = cand
                break
        else:
            raise ValueError(f"No P satisfies constraints for Z={Z}")
    return P, R_z


def _find_columns(ws) -> tuple[int, int]:
    """Return (dine_col, zom_col) by detecting headers; defaults to D, E."""
    dine_col = zom_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").lower()
        if dine_col is None and "dine" in h:
            dine_col = c
        if zom_col is None and "zomato" in h and "hike" not in h and "r_z" not in h:
            zom_col = c
    return dine_col or 4, zom_col or 5


def main() -> None:
    wb = load_workbook(SRC)
    ws = wb.active
    dine_col, zom_col = _find_columns(ws)
    start_col = max(ws.max_column + 1, 6)  # append after the last populated column

    hdr_fill = PatternFill("solid", fgColor="2E75B6")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for i, title in enumerate(NEW_HEADERS):
        c = start_col + i
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    n_z = 0
    n_no_z = 0
    for r in range(2, ws.max_row + 1):
        D = _num(ws.cell(r, dine_col).value)
        if D is None or D <= 0:
            continue
        Z = _num(ws.cell(r, zom_col).value)

        c0 = start_col
        if Z is not None and Z > 0:
            n_z += 1
            P, R_z = compute_p_with_z(Z)
            z_hike = (Z - D) / D * 100
            us_hike = (P - D) / D * 100
            rest_us = OUR_RATIO * P
            rev = (1 - OUR_RATIO) * P
            save = Z - P
            save_pct = (Z - P) / Z * 100
            rest_gain = rest_us - R_z

            ws.cell(r, c0 + 0, value=round(z_hike, 2))
            ws.cell(r, c0 + 1, value=round(R_z, 2))
            ws.cell(r, c0 + 2, value=P)
            ws.cell(r, c0 + 3, value=round(us_hike, 2))
            ws.cell(r, c0 + 4, value=round(rest_us, 2))
            ws.cell(r, c0 + 5, value=round(rev, 2))
            ws.cell(r, c0 + 6, value=round(save, 2))
            ws.cell(r, c0 + 7, value=round(save_pct, 2))
            ws.cell(r, c0 + 8, value=round(rest_gain, 2))
        else:
            n_no_z += 1
            P = round(D * (1 + NO_Z_HIKE_PCT / 100))
            rest_us = OUR_RATIO * P
            rev = (1 - OUR_RATIO) * P

            ws.cell(r, c0 + 0, value=None)
            ws.cell(r, c0 + 1, value=None)
            ws.cell(r, c0 + 2, value=P)
            ws.cell(r, c0 + 3, value=NO_Z_HIKE_PCT)
            ws.cell(r, c0 + 4, value=round(rest_us, 2))
            ws.cell(r, c0 + 5, value=round(rev, 2))
            ws.cell(r, c0 + 6, value=None)
            ws.cell(r, c0 + 7, value=None)
            ws.cell(r, c0 + 8, value=None)

    for i in range(len(NEW_HEADERS)):
        ws.column_dimensions[get_column_letter(start_col + i)].width = 22 if i != 2 else 16

    wb.save(SRC)
    print(f"Wrote {len(NEW_HEADERS)} columns to {SRC}")
    print(f"Rows with Zomato price: {n_z}")
    print(f"Rows without Zomato price (flat {NO_Z_HIKE_PCT:g}% hike): {n_no_z}")


if __name__ == "__main__":
    main()
