#!/usr/bin/env python3
"""
Enrich Maharaja menu Excel with subcategory, veg/non-veg, description, Unsplash images
and generate API payloads for import.

Reads:  /Users/user/Downloads/Maharaja_Zomato_vs_DineIn_Comparison upload.xlsx
Writes: same file (enriched with cols E–I)
Also:   dry-run prints payloads; pass --apply to POST to API.

Usage:
  python3 scripts/enrich_and_import_maharaja.py            # enrich + dry run
  python3 scripts/enrich_and_import_maharaja.py --apply     # enrich + insert
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment

RESTAURANT_ID = "RES-1776354042494-9358"
XLSX_PATH = Path("/Users/user/Downloads/Maharaja_Zomato_vs_DineIn_Comparison upload.xlsx")

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get("HONESTEATS_RETOOL_BYPASS", "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc")
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

# ─── Subcategory mapping ────────────────────────────────────────────────────────
SUBCATEGORY_MAP: dict[str, str] = {
    # Soups
    "Veg Hot & Sour Soup": "Veg Soups",
    "Tomato Soup": "Veg Soups",
    "Veg Lemon Coriander Soup": "Veg Soups",
    "Sweet Corn Soup": "Veg Soups",
    "Veg Cantonese Soup": "Veg Soups",
    "Baby Corn Soup": "Veg Soups",
    "Mutton Hot N Sour Soup": "Non-Veg Soups",
    "Chicken Hot N Sour Soup": "Non-Veg Soups",
    "Chicken Clear Soup": "Non-Veg Soups",
    "Chicken Cantonese Soup": "Non-Veg Soups",
    "Chicken Manchow Soup": "Non-Veg Soups",
    "Chicken Schezwan Soup": "Non-Veg Soups",
    # Starters - Chicken
    "Chicken Lollipop [6 Pieces]": "Chicken Starters",
    "Chicken Gulzar": "Chicken Starters",
    "Chicken Majestic": "Chicken Starters",
    "Chilli Chicken": "Chicken Starters",
    "Chicken 555": "Chicken Starters",
    "Chicken Pepper": "Chicken Starters",
    "Chicken Wings": "Chicken Starters",
    "Chicken Roast": "Chicken Starters",
    "Chicken Rayalaseema Dry": "Chicken Starters",
    "Chicken Shangri La [8 Pieces]": "Chicken Starters",
    "Chicken Gold Coin": "Chicken Starters",
    "Chicken Keema Balls [10 Pieces]": "Chicken Starters",
    "Chicken Bullets": "Chicken Starters",
    "Chicken Finger": "Chicken Starters",
    "Chicken Drumstick [3 Pieces]": "Chicken Starters",
    "Chicken Lemon": "Chicken Starters",
    "Murgh Musallam": "Chicken Starters",
    "Pudina Chicken": "Chicken Starters",
    "Hot Chicken": "Chicken Starters",
    "Chicken Sukka": "Chicken Starters",
    # Starters - Tandoori/Tikka
    "Chicken Tandoori": "Tandoori & Tikka",
    "Tangdi Kabab": "Tandoori & Tikka",
    "Chicken Tikka": "Tandoori & Tikka",
    "Chicken Kalmi Kabab": "Tandoori & Tikka",
    "Tandoori Lemon Chilli Prawns": "Tandoori & Tikka",
    "Fish Hariyali Tikka": "Tandoori & Tikka",
    "Mutton Kabab": "Tandoori & Tikka",
    "Tandoori Pomfret": "Tandoori & Tikka",
    "Lasooni Fish Tikka": "Tandoori & Tikka",
    # Starters - Mutton
    "Mutton Gulzar": "Mutton Starters",
    "Mutton Kaju Dry": "Mutton Starters",
    "Mutton Roast": "Mutton Starters",
    "Mutton Liver Dry": "Mutton Starters",
    "Mutton Methi": "Mutton Starters",
    "Mutton Pepper Dry": "Mutton Starters",
    "Mutton Rayalaseema Dry": "Mutton Starters",
    "Mutton Fry": "Mutton Starters",
    "Mutton Keema Gupta": "Mutton Starters",
    "Mutton Chilli": "Mutton Starters",
    "Mutton 65": "Mutton Starters",
    # Starters - Egg
    "Cauliflower Egg Fry": "Egg Starters",
    "Egg Pepper Fry": "Egg Starters",
    "Egg Chilli": "Egg Starters",
    "Egg 65": "Egg Starters",
    "Egg Garlic": "Egg Starters",
    "Egg Manchurian": "Egg Starters",
    # Starters - Fish/Seafood
    "Fish Roast": "Seafood Starters",
    "Fish Bengali Dry": "Seafood Starters",
    "Apollo Fish Dry": "Seafood Starters",
    "Lemon Pepper Fish Fry": "Seafood Starters",
    "Fish 65": "Seafood Starters",
    "Fish 555": "Seafood Starters",
    "Fish Chilli": "Seafood Starters",
    "Pomfret Tandoori [1 Piece]": "Seafood Starters",
    "White Pomfret [1 Piece]": "Seafood Starters",
    "Bhutan Fish Dry": "Seafood Starters",
    # Starters - Prawns
    "Prawns Gulzar": "Prawns Starters",
    "Prawns Loose": "Prawns Starters",
    "Prawns Red Chilli": "Prawns Starters",
    "Prawns 65": "Prawns Starters",
    "Prawns 888": "Prawns Starters",
    "Prawns Korea": "Prawns Starters",
    "Prawns Apollo": "Prawns Starters",
    "Maharaja Prawns Dry": "Prawns Starters",
    # Starters - Kamju
    "Kamju Oil Roast": "Kamju Starters",
    "Kamju": "Kamju Starters",
    "Kamju Gulzar": "Kamju Starters",
    # Starters - Veg
    "Aloo Gulzar": "Veg Starters",
    "Gobi 65": "Veg Starters",
    "Baby Corn Manchurian": "Veg Starters",
    "Gobhi Manchurian": "Veg Starters",
    "Paneer 65": "Veg Starters",
    "Gobi Chilli": "Veg Starters",
    "Veg Manchurian": "Veg Starters",
    "Mushroom Manchurian": "Veg Starters",
    "Mushroom Chilli": "Veg Starters",
    "Aloo 65": "Veg Starters",
    "Aloo Manchurian": "Veg Starters",
    # Main Course - Veg
    "Kaju Curry": "Veg Main Course",
    "Vegetable Kadai": "Veg Main Course",
    "Paneer Malai Kofta": "Veg Main Course",
    "Baby Corn Curry": "Veg Main Course",
    "Veg Kolhapuri": "Veg Main Course",
    "Tomato Curry": "Veg Main Course",
    "Spice Vegetable Kofta": "Veg Main Course",
    "Capsicum Masala": "Veg Main Course",
    "Aloo Palak": "Veg Main Course",
    "Plain Palak": "Veg Main Course",
    "Cauliflower Gravy": "Veg Main Course",
    # Main Course - Egg
    "Egg Bhurji": "Egg Main Course",
    "Egg Curry": "Egg Main Course",
    "Egg Rayalseema Curry": "Egg Main Course",
    "Egg Pataka": "Egg Main Course",
    "Egg Mobile": "Egg Main Course",
    "Egg Tomato": "Egg Main Course",
    "Egg Saagwala": "Egg Main Course",
    # Main Course - Chicken
    "Boneless Chicken Rayalaseema": "Chicken Main Course",
    "Butter Chicken": "Chicken Main Course",
    "Chicken Curry": "Chicken Main Course",
    "Chicken Mughlai": "Chicken Main Course",
    "Chicken Kadai": "Chicken Main Course",
    "Chicken Kolhapuri": "Chicken Main Course",
    "Boneless Chicken Curry": "Chicken Main Course",
    "Chicken Masala Curry": "Chicken Main Course",
    "Chicken Methi": "Chicken Main Course",
    "Chicken Nawabi": "Chicken Main Course",
    "Pepper Chicken Curry": "Chicken Main Course",
    "Chicken Bhaji": "Chicken Main Course",
    "Boneless Chicken Bhaji": "Chicken Main Course",
    "Chicken Hyderabadi": "Chicken Main Course",
    "Chicken Malai": "Chicken Main Course",
    "Chicken Chandrakala": "Chicken Main Course",
    "Chicken Kakatiya": "Chicken Main Course",
    "Chicken Rayalaseema": "Chicken Main Course",
    "Chicken Kashmiri": "Chicken Main Course",
    "Boneless Chicken Kashmiri": "Chicken Main Course",
    "Chicken Cashew Nut": "Chicken Main Course",
    "Chicken Dum Ka Murgi": "Chicken Main Course",
    "Natu Kodi Iguru": "Chicken Main Course",
    "Ulava Charu Chicken": "Chicken Main Course",
    "Curry Leaf Chicken": "Chicken Main Course",
    "Malabar Chicken Curry": "Chicken Main Course",
    "Chicken Kolhapuri Curry": "Chicken Main Course",
    # Main Course - Kamju
    "Kamju Masala": "Kamju Main Course",
    "Kamju Golden Curry": "Kamju Main Course",
    # Main Course - Mutton
    "Mutton Keema Curry": "Mutton Main Course",
    "Mutton Rayalaseema": "Mutton Main Course",
    "Mutton Masala": "Mutton Main Course",
    "Mutton Liver Curry": "Mutton Main Course",
    "Mutton Pepper Red Chilli Curry": "Mutton Main Course",
    "Mutton Kadai": "Mutton Main Course",
    "Mutton Bhaji": "Mutton Main Course",
    "Mutton Mughlai": "Mutton Main Course",
    "Mutton Rogan Josh": "Mutton Main Course",
    "Mutton Kandari": "Mutton Main Course",
    "Mutton Basanthi": "Mutton Main Course",
    "Mutton Saagwala": "Mutton Main Course",
    "Hariyali Mutton Curry": "Mutton Main Course",
    # Main Course - Seafood
    "Prawns Malai Curry": "Seafood Main Course",
    "Prawns Andhra": "Seafood Main Course",
    "Fish Curry": "Seafood Main Course",
    "Fish Andhra Curry": "Seafood Main Course",
    "Fish Bengali": "Seafood Main Course",
    "Red Hot Chilli Fish": "Seafood Main Course",
    "Prawns Kadai": "Seafood Main Course",
    "Prawns Golden Curry": "Seafood Main Course",
    "Prawns Bhaji": "Seafood Main Course",
    "Prawns Red Chilli Curry": "Seafood Main Course",
    # Breads
    "Kerala Paratha": "Paratha",
    "Ghee Paratha": "Paratha",
    "Gobi Paratha": "Paratha",
    "Lachha Paratha": "Paratha",
    "Methi Paratha": "Paratha",
    "Tandoori Roti": "Roti & Naan",
    "Butter Roti": "Roti & Naan",
    "Plain Naan": "Roti & Naan",
    "Butter Naan": "Roti & Naan",
    "Garlic Naan": "Roti & Naan",
    "Plain Kulcha": "Kulcha",
    "Butter Kulcha": "Kulcha",
    "Chicken Keema Naan": "Stuffed Naan",
    "Mutton Keema Naan": "Stuffed Naan",
    # Rice and Biryani - Rice
    "Special Curd Rice": "Rice",
    "Curd Rice": "Rice",
    "Tomato Rice": "Rice",
    "Lemon Rice": "Rice",
    "Jeera Rice": "Rice",
    "Steamed Rice": "Rice",
    "Ghee Rice": "Rice",
    "Coconut Rice": "Rice",
    # Rice and Biryani - Veg Biryani
    "Paneer Kaju Biryani": "Veg Biryani",
    "Kaju Biryani": "Veg Biryani",
    "Paneer Biryani": "Veg Biryani",
    "Veg Biryani": "Veg Biryani",
    # Rice and Biryani - Non-Veg Biryani
    "Chicken Lollipop Biryani [4 Pieces]": "Non-Veg Biryani",
    "Chicken Hyderabadi Dum Biryani": "Non-Veg Biryani",
    "Hyderabadi Mutton Ghee Dum Biryani": "Non-Veg Biryani",
    "Chicken Special Biryani": "Non-Veg Biryani",
    "Mutton Fry Biryani": "Non-Veg Biryani",
    "Chicken Mughlai Biryani": "Non-Veg Biryani",
    "Chicken Roast Biryani": "Non-Veg Biryani",
    "Mutton Mughlai Biryani": "Non-Veg Biryani",
    "Egg Biryani": "Non-Veg Biryani",
    "Chicken Ghee Pot Dum Biryani": "Non-Veg Biryani",
    "Mutton Keema Fry Biryani": "Non-Veg Biryani",
    "Fish Biryani": "Non-Veg Biryani",
    "Chicken Leg Fry Biryani": "Non-Veg Biryani",
    "Chicken Keema Fry Biryani": "Non-Veg Biryani",
    "Kamju Biryani": "Non-Veg Biryani",
    "Mutton Ghee Pot Dum Biryani": "Non-Veg Biryani",
    "Prawns Ghee Pot Dum Biryani": "Non-Veg Biryani",
    # Fried Rice
    "Veg Fried Rice": "Veg Fried Rice",
    "Veg Schezwan Fried Rice": "Veg Fried Rice",
    "Paneer Fried Rice": "Veg Fried Rice",
    "Chicken Fried Rice": "Non-Veg Fried Rice",
    "Egg Spice Fried Rice": "Non-Veg Fried Rice",
    "Mutton Fried Rice": "Non-Veg Fried Rice",
    "Fish Fried Rice": "Non-Veg Fried Rice",
    "Prawns Fried Rice": "Non-Veg Fried Rice",
    # Snacks
    "Prawns Pakoda": "Snacks",
    "Paneer Kulcha": "Snacks",
    "Mutton Keema Bullets": "Snacks",
    "Mutton Keema Cutlet": "Snacks",
    "Fish Finger": "Snacks",
    # Mutton Methi appears in both Starters and Main Course
    "Mutton Methi": "Mutton Starters",
}

# ─── Veg/Non-Veg ────────────────────────────────────────────────────────────────
VEG_ITEMS = {
    "Veg Hot & Sour Soup", "Tomato Soup", "Veg Lemon Coriander Soup", "Sweet Corn Soup",
    "Veg Cantonese Soup", "Baby Corn Soup", "Aloo Gulzar", "Gobi 65", "Baby Corn Manchurian",
    "Gobhi Manchurian", "Paneer 65", "Gobi Chilli", "Veg Manchurian", "Mushroom Manchurian",
    "Mushroom Chilli", "Aloo 65", "Aloo Manchurian", "Kaju Curry", "Vegetable Kadai",
    "Paneer Malai Kofta", "Baby Corn Curry", "Veg Kolhapuri", "Tomato Curry",
    "Spice Vegetable Kofta", "Capsicum Masala", "Aloo Palak", "Plain Palak",
    "Cauliflower Gravy", "Kerala Paratha", "Ghee Paratha", "Tandoori Roti", "Butter Roti",
    "Plain Naan", "Butter Naan", "Garlic Naan", "Gobi Paratha", "Lachha Paratha",
    "Methi Paratha", "Plain Kulcha", "Butter Kulcha", "Paneer Kulcha",
    "Special Curd Rice", "Curd Rice", "Tomato Rice", "Lemon Rice", "Jeera Rice",
    "Steamed Rice", "Ghee Rice", "Coconut Rice", "Paneer Kaju Biryani", "Kaju Biryani",
    "Paneer Biryani", "Veg Biryani", "Veg Fried Rice", "Veg Schezwan Fried Rice",
    "Paneer Fried Rice",
}

# ─── Descriptions ────────────────────────────────────────────────────────────────
DESCRIPTIONS: dict[str, str] = {
    "Veg Hot & Sour Soup": "A tangy and spicy vegetable soup with a perfect balance of flavors",
    "Tomato Soup": "Classic creamy tomato soup, rich and comforting",
    "Veg Lemon Coriander Soup": "Light and refreshing soup with lemon zing and fresh coriander",
    "Sweet Corn Soup": "Creamy sweet corn soup with a hint of pepper",
    "Veg Cantonese Soup": "Indo-Chinese style thick vegetable soup",
    "Baby Corn Soup": "Delicate soup with tender baby corn and vegetables",
    "Mutton Hot N Sour Soup": "Spicy and tangy mutton soup with bold flavors",
    "Chicken Hot N Sour Soup": "Classic hot and sour chicken soup with mushrooms and bamboo shoots",
    "Chicken Clear Soup": "Light aromatic chicken broth with herbs and vegetables",
    "Chicken Cantonese Soup": "Rich Cantonese-style chicken soup with vegetables",
    "Chicken Manchow Soup": "Spicy Indo-Chinese chicken manchow soup topped with crispy noodles",
    "Chicken Schezwan Soup": "Fiery Schezwan-flavored chicken soup with bold spices",
    "Chicken Lollipop [6 Pieces]": "6 crispy fried chicken lollipops tossed in tangy sauce",
    "Chicken Gulzar": "Signature spicy chicken starter with aromatic masala",
    "Chicken Majestic": "Crispy fried chicken tossed with curry leaves and spices",
    "Chilli Chicken": "Indo-Chinese classic — crispy chicken in spicy chilli sauce",
    "Chicken 555": "Hyderabadi-style spicy deep-fried chicken with green chillies",
    "Chicken Pepper": "Pepper-crusted chicken with a bold peppery kick",
    "Chicken Wings": "Crispy golden chicken wings with signature seasoning",
    "Chicken Roast": "Succulent roasted chicken with aromatic Indian spices",
    "Chicken Rayalaseema Dry": "Fiery Rayalaseema-style dry chicken with red chillies",
    "Chicken Shangri La [8 Pieces]": "8 pieces of crispy chicken in sweet and spicy Shangri La sauce",
    "Chicken Gold Coin": "Crispy coin-shaped chicken patties with masala seasoning",
    "Chicken Keema Balls [10 Pieces]": "10 spiced minced chicken balls, deep-fried to perfection",
    "Chicken Bullets": "Bite-sized crispy chicken pieces with spicy coating",
    "Chicken Finger": "Golden-fried chicken fingers with a crunchy exterior",
    "Chicken Drumstick [3 Pieces]": "3 juicy marinated chicken drumsticks, deep-fried crispy",
    "Chicken Lemon": "Tangy lemon-glazed chicken with a citrusy twist",
    "Murgh Musallam": "Whole chicken slow-cooked in rich Mughlai gravy — a royal feast",
    "Pudina Chicken": "Fresh mint-infused chicken with green herb marinade",
    "Hot Chicken": "Extra spicy chicken with fiery red chilli marinade",
    "Chicken Sukka": "Dry-roasted chicken with coconut and South Indian spices",
    "Aloo Gulzar": "Crispy potato cubes tossed in tangy masala",
    "Chicken Tandoori": "Classic clay-oven roasted chicken with yogurt-spice marinade",
    "Tangdi Kabab": "Juicy chicken leg kababs marinated in tandoori spices",
    "Chicken Tikka": "Boneless chicken chunks grilled in smoky tikka marinade",
    "Chicken Kalmi Kabab": "Tender chicken drumettes with rich cream and spice marinade",
    "Tandoori Lemon Chilli Prawns": "Tandoor-grilled prawns with lemon and green chilli",
    "Fish Hariyali Tikka": "Fish tikka marinated in green herb and mint paste",
    "Mutton Kabab": "Succulent minced mutton kababs with aromatic spices",
    "Tandoori Pomfret": "Whole pomfret marinated and grilled in tandoor",
    "Lasooni Fish Tikka": "Garlic-marinated fish tikka grilled to perfection",
    "Mutton Gulzar": "Spicy mutton dry starter with Maharaja's signature masala",
    "Mutton Kaju Dry": "Tender mutton pieces with cashew and dry spices",
    "Mutton Roast": "Slow-roasted mutton with deep South Indian flavors",
    "Mutton Liver Dry": "Pan-fried mutton liver with onions and spices",
    "Mutton Methi": "Fenugreek-flavored mutton with aromatic dry masala",
    "Mutton Pepper Dry": "Black pepper-crusted dry mutton with bold flavor",
    "Mutton Rayalaseema Dry": "Fiery Rayalaseema-style dry mutton with red chillies",
    "Mutton Fry": "Crispy deep-fried mutton with traditional spice blend",
    "Mutton Keema Gupta": "Spiced minced mutton prepared in a secret recipe",
    "Cauliflower Egg Fry": "Cauliflower and egg stir-fried with Indian spices",
    "Egg Pepper Fry": "Boiled eggs tossed in black pepper and curry leaf masala",
    "Egg Chilli": "Indo-Chinese egg chilli with bell peppers and soy sauce",
    "Egg 65": "Crispy deep-fried egg fritters with spicy masala coating",
    "Egg Garlic": "Eggs sautéed with garlic and aromatic spices",
    "Egg Manchurian": "Egg balls in tangy Indo-Chinese Manchurian sauce",
    "Fish Roast": "Whole fish pan-roasted with South Indian masala",
    "Fish Bengali Dry": "Bengali-style dry fish with mustard and spices",
    "Apollo Fish Dry": "Hyderabadi-style crispy Apollo fish with spicy batter",
    "Lemon Pepper Fish Fry": "Fish fillets fried with lemon and cracked pepper",
    "Fish 65": "Crispy deep-fried fish with spicy 65 masala",
    "Fish 555": "Spicy fried fish with Hyderabadi 555 seasoning",
    "Fish Chilli": "Indo-Chinese fish chilli with peppers and soy",
    "Pomfret Tandoori [1 Piece]": "Whole pomfret marinated in tandoori spices, clay-oven roasted",
    "White Pomfret [1 Piece]": "Delicate white pomfret pan-fried with mild spices",
    "Bhutan Fish Dry": "Crispy fried fish with Bhutanese-inspired spice blend",
    "Prawns Gulzar": "Spicy prawns tossed in signature Gulzar masala",
    "Prawns Loose": "Tender loose prawns stir-fried with aromatic spices",
    "Prawns Red Chilli": "Prawns in fiery red chilli sauce",
    "Prawns 65": "Crispy deep-fried prawns with spicy 65 masala",
    "Prawns 888": "Signature triple-8 spiced prawns",
    "Prawns Korea": "Korean-inspired spicy prawns with sweet chilli glaze",
    "Prawns Apollo": "Hyderabadi Apollo-style crispy prawns",
    "Maharaja Prawns Dry": "Chef's special dry prawns — Maharaja's signature recipe",
    "Kamju Oil Roast": "Kamju (crab) oil-roasted with South Indian spices",
    "Kamju": "Classic kamju (crab) preparation with coastal flavors",
    "Kamju Gulzar": "Kamju (crab) in spicy Gulzar masala",
    "Gobi 65": "Crispy cauliflower fritters with spicy 65 seasoning",
    "Baby Corn Manchurian": "Baby corn in tangy Manchurian sauce",
    "Gobhi Manchurian": "Cauliflower florets in Indo-Chinese Manchurian gravy",
    "Paneer 65": "Crispy paneer cubes with spicy 65 masala coating",
    "Gobi Chilli": "Cauliflower florets tossed in spicy chilli sauce",
    "Veg Manchurian": "Mixed vegetable balls in tangy Manchurian sauce",
    "Mushroom Manchurian": "Button mushrooms in rich Manchurian sauce",
    "Mushroom Chilli": "Stir-fried mushrooms in spicy chilli sauce",
    "Aloo 65": "Crispy potato cubes with spicy 65 masala",
    "Aloo Manchurian": "Potato balls in tangy Indo-Chinese Manchurian sauce",
    "Mutton Chilli": "Tender mutton in fiery Indo-Chinese chilli sauce",
    "Mutton 65": "Crispy deep-fried mutton with classic 65 spice blend",
    "Kaju Curry": "Rich cashew curry in creamy gravy",
    "Vegetable Kadai": "Mixed vegetables cooked in kadai with bell peppers",
    "Paneer Malai Kofta": "Soft paneer kofta balls in rich creamy gravy",
    "Baby Corn Curry": "Tender baby corn in mildly spiced gravy",
    "Veg Kolhapuri": "Mixed vegetables in fiery Kolhapuri masala",
    "Tomato Curry": "Simple and flavorful tomato-based curry",
    "Spice Vegetable Kofta": "Spiced vegetable balls in aromatic gravy",
    "Capsicum Masala": "Bell peppers cooked in rich masala gravy",
    "Aloo Palak": "Potato and spinach in a smooth green gravy",
    "Plain Palak": "Creamy spinach gravy, simple and nutritious",
    "Cauliflower Gravy": "Cauliflower florets in mild and flavorful gravy",
    "Boneless Chicken Rayalaseema": "Boneless chicken in fiery Rayalaseema-style curry",
    "Butter Chicken": "Creamy tomato-based butter chicken — a North Indian classic",
    "Chicken Curry": "Traditional Indian chicken curry with aromatic spices",
    "Chicken Mughlai": "Rich Mughlai-style chicken in cashew and cream gravy",
    "Chicken Kadai": "Chicken cooked kadai-style with bell peppers and spices",
    "Chicken Kolhapuri": "Spicy Kolhapuri-style chicken with bold masala",
    "Boneless Chicken Curry": "Boneless chicken pieces in traditional curry gravy",
    "Chicken Masala Curry": "Chicken in rich and spicy masala gravy",
    "Chicken Methi": "Chicken cooked with fresh fenugreek leaves",
    "Chicken Nawabi": "Royal Nawabi-style chicken in rich cream gravy",
    "Pepper Chicken Curry": "Black pepper-infused chicken curry",
    "Egg Bhurji": "Indian-style scrambled eggs with onions and spices",
    "Egg Curry": "Boiled eggs simmered in aromatic onion-tomato gravy",
    "Egg Rayalseema Curry": "Eggs in fiery Rayalaseema-style red chilli curry",
    "Egg Pataka": "Spicy firecracker-style egg curry",
    "Egg Mobile": "Unique egg preparation with special house masala",
    "Egg Tomato": "Eggs cooked in tangy tomato-based gravy",
    "Egg Saagwala": "Eggs in creamy spinach gravy",
    "Chicken Bhaji": "Chicken cooked with onions in bhaji-style dry gravy",
    "Boneless Chicken Bhaji": "Boneless chicken in dry bhaji-style preparation",
    "Chicken Hyderabadi": "Rich Hyderabadi-style chicken curry with nuts and spices",
    "Chicken Malai": "Tender chicken in smooth and creamy malai gravy",
    "Chicken Chandrakala": "Chicken in aromatic Chandrakala masala — chef's special",
    "Chicken Kakatiya": "Kakatiya dynasty-inspired spicy chicken curry",
    "Chicken Rayalaseema": "Fiery Rayalaseema-style chicken curry with red chillies",
    "Chicken Kashmiri": "Mildly spiced Kashmiri chicken with aromatic flavors",
    "Boneless Chicken Kashmiri": "Boneless chicken in fragrant Kashmiri-style gravy",
    "Chicken Cashew Nut": "Chicken cooked with roasted cashew nuts in rich gravy",
    "Chicken Dum Ka Murgi": "Whole chicken slow-cooked dum-style with royal spices",
    "Natu Kodi Iguru": "Country chicken iguru — rustic Andhra-style dry chicken",
    "Ulava Charu Chicken": "Traditional Andhra horsegram rasam with chicken",
    "Curry Leaf Chicken": "Chicken infused with aromatic curry leaves and spices",
    "Malabar Chicken Curry": "Coconut-based Malabar-style chicken curry",
    "Chicken Kolhapuri Curry": "Chicken in fiery Kolhapuri-style gravy",
    "Kamju Masala": "Crab cooked in rich masala gravy",
    "Kamju Golden Curry": "Crab in golden turmeric and coconut curry",
    "Mutton Keema Curry": "Minced mutton in aromatic curry gravy",
    "Mutton Rayalaseema": "Fiery Rayalaseema-style mutton curry",
    "Mutton Masala": "Tender mutton in rich and spicy masala gravy",
    "Mutton Liver Curry": "Mutton liver cooked in flavorful curry sauce",
    "Mutton Pepper Red Chilli Curry": "Mutton in pepper and red chilli gravy",
    "Mutton Kadai": "Mutton cooked kadai-style with bell peppers",
    "Mutton Bhaji": "Mutton in dry bhaji-style preparation with onions",
    "Mutton Mughlai": "Rich Mughlai mutton curry with cream and nuts",
    "Mutton Rogan Josh": "Kashmiri-style rogan josh mutton in aromatic gravy",
    "Mutton Kandari": "Mutton in tangy pomegranate-based Kandari gravy",
    "Mutton Basanthi": "Mutton in mild and aromatic Basanthi-style curry",
    "Mutton Saagwala": "Mutton simmered in rich spinach gravy",
    "Hariyali Mutton Curry": "Mutton in vibrant green herb and mint curry",
    "Prawns Malai Curry": "Prawns in luxurious coconut malai curry",
    "Prawns Andhra": "Prawns in spicy Andhra-style masala curry",
    "Fish Curry": "Traditional fish curry in tangy tamarind gravy",
    "Fish Andhra Curry": "Fish in fiery Andhra-style red chilli curry",
    "Fish Bengali": "Bengali-style fish in mustard and turmeric gravy",
    "Red Hot Chilli Fish": "Fish in blazing red hot chilli sauce",
    "Prawns Kadai": "Prawns cooked kadai-style with bell peppers",
    "Prawns Golden Curry": "Prawns in golden turmeric and coconut curry",
    "Prawns Bhaji": "Prawns in dry bhaji-style preparation",
    "Prawns Red Chilli Curry": "Prawns in fiery red chilli curry",
    "Kerala Paratha": "Flaky layered Kerala-style paratha",
    "Ghee Paratha": "Crispy ghee-brushed layered paratha",
    "Tandoori Roti": "Whole wheat roti baked in clay tandoor",
    "Butter Roti": "Soft tandoori roti brushed with butter",
    "Plain Naan": "Soft and fluffy plain naan from tandoor",
    "Butter Naan": "Buttery naan bread baked in clay oven",
    "Garlic Naan": "Aromatic garlic-topped naan from tandoor",
    "Gobi Paratha": "Stuffed paratha with spiced cauliflower filling",
    "Lachha Paratha": "Multi-layered crispy lachha paratha",
    "Methi Paratha": "Paratha flavored with fresh fenugreek leaves",
    "Plain Kulcha": "Soft and fluffy plain kulcha bread",
    "Butter Kulcha": "Buttery kulcha bread baked in tandoor",
    "Chicken Keema Naan": "Naan stuffed with spiced minced chicken",
    "Mutton Keema Naan": "Naan stuffed with spiced minced mutton",
    "Special Curd Rice": "Premium curd rice with tempering and pomegranate",
    "Curd Rice": "Classic South Indian curd rice, cool and comforting",
    "Tomato Rice": "Tangy tomato-flavored rice with South Indian tempering",
    "Lemon Rice": "Zesty lemon rice with peanuts and curry leaves",
    "Jeera Rice": "Fragrant basmati rice tempered with cumin seeds",
    "Steamed Rice": "Plain steamed basmati rice",
    "Ghee Rice": "Aromatic basmati rice cooked in pure ghee",
    "Coconut Rice": "Basmati rice with fresh coconut and curry leaf tempering",
    "Paneer Kaju Biryani": "Rich biryani with paneer and cashew nuts",
    "Kaju Biryani": "Fragrant biryani loaded with roasted cashew nuts",
    "Paneer Biryani": "Aromatic biryani with soft paneer cubes",
    "Veg Biryani": "Mixed vegetable dum biryani with aromatic spices",
    "Chicken Lollipop Biryani [4 Pieces]": "Biryani topped with 4 crispy chicken lollipops",
    "Chicken Hyderabadi Dum Biryani": "Authentic Hyderabadi dum biryani with tender chicken",
    "Hyderabadi Mutton Ghee Dum Biryani": "Royal Hyderabadi mutton biryani slow-cooked in ghee",
    "Chicken Special Biryani": "Chef's special chicken biryani with extra masala",
    "Mutton Fry Biryani": "Biryani served with crispy fried mutton pieces",
    "Chicken Mughlai Biryani": "Rich Mughlai-style chicken biryani with cream and nuts",
    "Chicken Roast Biryani": "Biryani topped with aromatic roasted chicken",
    "Mutton Mughlai Biryani": "Royal Mughlai mutton biryani with rich flavors",
    "Egg Biryani": "Flavorful biryani with boiled eggs and aromatic spices",
    "Chicken Ghee Pot Dum Biryani": "Premium pot biryani slow-cooked in ghee — serves 2-3",
    "Mutton Keema Fry Biryani": "Biryani with spiced fried minced mutton",
    "Fish Biryani": "Aromatic biryani with tender fish pieces",
    "Chicken Leg Fry Biryani": "Biryani topped with crispy fried chicken leg",
    "Chicken Keema Fry Biryani": "Biryani with spiced fried chicken keema",
    "Kamju Biryani": "Unique crab biryani with coastal spices",
    "Mutton Ghee Pot Dum Biryani": "Premium mutton pot biryani in ghee — serves 2-3",
    "Prawns Ghee Pot Dum Biryani": "Premium prawns pot biryani in ghee — serves 2-3",
    "Veg Fried Rice": "Indo-Chinese vegetable fried rice",
    "Veg Schezwan Fried Rice": "Spicy Schezwan-style vegetable fried rice",
    "Paneer Fried Rice": "Fried rice with soft paneer cubes",
    "Chicken Fried Rice": "Classic chicken fried rice with vegetables",
    "Egg Spice Fried Rice": "Spiced egg fried rice with aromatic seasoning",
    "Mutton Fried Rice": "Fried rice with tender mutton pieces",
    "Fish Fried Rice": "Fried rice with flaky fish pieces",
    "Prawns Fried Rice": "Fried rice loaded with juicy prawns",
    "Prawns Pakoda": "Crispy deep-fried prawn fritters",
    "Paneer Kulcha": "Kulcha stuffed with spiced paneer filling",
    "Mutton Keema Bullets": "Spiced minced mutton bullets, deep-fried crispy",
    "Mutton Keema Cutlet": "Pan-fried mutton keema cutlets with crispy coating",
    "Fish Finger": "Golden-fried fish fingers with crunchy breadcrumb coating",
}

# ─── Unsplash image pool (unique per item) ───────────────────────────────────────
# Organized by food type to ensure relevant matches. Each URL used only once.
_UNSPLASH = {
    "veg_soup": [
        "https://images.unsplash.com/photo-1547592166-23ac45744acd?w=800&q=80",
        "https://images.unsplash.com/photo-1603105037880-880cd4f5b2e6?w=800&q=80",
        "https://images.unsplash.com/photo-1476718406336-bb5a9690ee2a?w=800&q=80",
        "https://images.unsplash.com/photo-1588566565463-180a5b2090d2?w=800&q=80",
        "https://images.unsplash.com/photo-1613844237701-8f3664fc2eff?w=800&q=80",
        "https://images.unsplash.com/photo-1594756202469-9ff9799b2e4e?w=800&q=80",
    ],
    "nonveg_soup": [
        "https://images.unsplash.com/photo-1604152135912-04a022e23696?w=800&q=80",
        "https://images.unsplash.com/photo-1607330289024-1535c6b4e1c1?w=800&q=80",
        "https://images.unsplash.com/photo-1583608205776-bfd35f0d9f83?w=800&q=80",
        "https://images.unsplash.com/photo-1509358271058-acd22cc93898?w=800&q=80",
        "https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=800&q=80",
        "https://images.unsplash.com/photo-1597227129956-93bad7e18d08?w=800&q=80",
    ],
    "chicken_starter": [
        "https://images.unsplash.com/photo-1626082927389-6cd097cdc6ec?w=800&q=80",
        "https://images.unsplash.com/photo-1562967914-608f82629710?w=800&q=80",
        "https://images.unsplash.com/photo-1608039755401-742074f0548d?w=800&q=80",
        "https://images.unsplash.com/photo-1567620832903-9fc6debc209f?w=800&q=80",
        "https://images.unsplash.com/photo-1610057099431-d73a1c9d2f2f?w=800&q=80",
        "https://images.unsplash.com/photo-1598515214211-89d3c73ae83b?w=800&q=80",
        "https://images.unsplash.com/photo-1587593810167-a84920ea0781?w=800&q=80",
        "https://images.unsplash.com/photo-1619221882220-947b3d3c8861?w=800&q=80",
        "https://images.unsplash.com/photo-1606728035253-49e8a23146de?w=800&q=80",
        "https://images.unsplash.com/photo-1599487488170-d11ec9c172f0?w=800&q=80",
        "https://images.unsplash.com/photo-1632778149955-e80f8ceca2e8?w=800&q=80",
        "https://images.unsplash.com/photo-1624726175512-19b9baf9fbd1?w=800&q=80",
        "https://images.unsplash.com/photo-1585937421612-70a008356fbe?w=800&q=80",
        "https://images.unsplash.com/photo-1580217593608-61931cefc821?w=800&q=80",
        "https://images.unsplash.com/photo-1614398751058-bca239de00ca?w=800&q=80",
        "https://images.unsplash.com/photo-1603360946369-dc9bb6258143?w=800&q=80",
        "https://images.unsplash.com/photo-1565299624946-b28f40a0ae38?w=800&q=80",
        "https://images.unsplash.com/photo-1606755962773-d324e0a13086?w=800&q=80",
        "https://images.unsplash.com/photo-1612874742237-6526221588e3?w=800&q=80",
        "https://images.unsplash.com/photo-1625937286520-3ef7955f3813?w=800&q=80",
    ],
    "tandoori": [
        "https://images.unsplash.com/photo-1599487488170-d11ec9c172f0?w=800&q=80",
        "https://images.unsplash.com/photo-1610057099431-d73a1c9d2f2f?w=800&q=80",
        "https://images.unsplash.com/photo-1601050690117-94f5f6fa8bd7?w=800&q=80",
        "https://images.unsplash.com/photo-1565557623262-b51c2513a641?w=800&q=80",
        "https://images.unsplash.com/photo-1628294895950-9805252327bc?w=800&q=80",
        "https://images.unsplash.com/photo-1567188040759-fb8a883dc6d8?w=800&q=80",
        "https://images.unsplash.com/photo-1574653853027-5382a3d23a15?w=800&q=80",
        "https://images.unsplash.com/photo-1551881192-5e377f1b2142?w=800&q=80",
        "https://images.unsplash.com/photo-1573080496219-bb080dd4f877?w=800&q=80",
    ],
    "mutton": [
        "https://images.unsplash.com/photo-1545247181-516773cae754?w=800&q=80",
        "https://images.unsplash.com/photo-1574484284002-952d92456975?w=800&q=80",
        "https://images.unsplash.com/photo-1606491956689-2ea866880049?w=800&q=80",
        "https://images.unsplash.com/photo-1633321702518-7fecdafb94d5?w=800&q=80",
        "https://images.unsplash.com/photo-1609501676725-7186f017a4b7?w=800&q=80",
        "https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=800&q=80",
        "https://images.unsplash.com/photo-1617692855027-33b14f061079?w=800&q=80",
        "https://images.unsplash.com/photo-1574653853027-5382a3d23a15?w=800&q=80",
        "https://images.unsplash.com/photo-1642821373181-16a5bc9f5801?w=800&q=80",
        "https://images.unsplash.com/photo-1618449840665-9ed506d73a34?w=800&q=80",
        "https://images.unsplash.com/photo-1603894584373-5ac82b2ae328?w=800&q=80",
    ],
    "egg": [
        "https://images.unsplash.com/photo-1482049016688-2d3e1b311543?w=800&q=80",
        "https://images.unsplash.com/photo-1525351484163-7529414344d8?w=800&q=80",
        "https://images.unsplash.com/photo-1510693206972-df098062cb71?w=800&q=80",
        "https://images.unsplash.com/photo-1598215429751-891f0d5f1f46?w=800&q=80",
        "https://images.unsplash.com/photo-1594901753909-81c64ab91a90?w=800&q=80",
        "https://images.unsplash.com/photo-1555939594-58d7cb561ad1?w=800&q=80",
        "https://images.unsplash.com/photo-1590301157890-4810ed352733?w=800&q=80",
    ],
    "seafood": [
        "https://images.unsplash.com/photo-1606728035253-49e8a23146de?w=800&q=80",
        "https://images.unsplash.com/photo-1615141982883-c7ad0e69fd62?w=800&q=80",
        "https://images.unsplash.com/photo-1504674900247-0877df9cc836?w=800&q=80",
        "https://images.unsplash.com/photo-1535140728325-a4d3707eee61?w=800&q=80",
        "https://images.unsplash.com/photo-1510130113581-4ae76c0f6e7f?w=800&q=80",
        "https://images.unsplash.com/photo-1559039448-9b03d2e3c18e?w=800&q=80",
        "https://images.unsplash.com/photo-1519708227418-c8fd9a32b7a2?w=800&q=80",
        "https://images.unsplash.com/photo-1498654896293-37aacf113fd9?w=800&q=80",
        "https://images.unsplash.com/photo-1580476262798-bddd9f4b7369?w=800&q=80",
        "https://images.unsplash.com/photo-1551504734-5ee1c4a1479b?w=800&q=80",
        "https://images.unsplash.com/photo-1559339352-11d035aa65de?w=800&q=80",
        "https://images.unsplash.com/photo-1553621042-f6e147245754?w=800&q=80",
        "https://images.unsplash.com/photo-1565680018093-ebb6e3062e7b?w=800&q=80",
        "https://images.unsplash.com/photo-1612929633738-8fe44f7ec841?w=800&q=80",
        "https://images.unsplash.com/photo-1569058242567-93de6f36f8e6?w=800&q=80",
        "https://images.unsplash.com/photo-1606731219412-213c1e68ca63?w=800&q=80",
        "https://images.unsplash.com/photo-1617196035154-1e7e6e28b0db?w=800&q=80",
        "https://images.unsplash.com/photo-1610540881815-6e81b56d8ea5?w=800&q=80",
        "https://images.unsplash.com/photo-1615361200098-9e630ec29b4e?w=800&q=80",
    ],
    "veg_starter": [
        "https://images.unsplash.com/photo-1601050690597-df0568f70950?w=800&q=80",
        "https://images.unsplash.com/photo-1546069901-ba9599a7e63c?w=800&q=80",
        "https://images.unsplash.com/photo-1540189549336-e6e99c3679fe?w=800&q=80",
        "https://images.unsplash.com/photo-1567337710282-00832b415979?w=800&q=80",
        "https://images.unsplash.com/photo-1585032226651-759b368d7246?w=800&q=80",
        "https://images.unsplash.com/photo-1572715376701-98568319fd0b?w=800&q=80",
        "https://images.unsplash.com/photo-1559847844-5315695dadae?w=800&q=80",
        "https://images.unsplash.com/photo-1606574977732-e8e5f1f46c23?w=800&q=80",
        "https://images.unsplash.com/photo-1574484284002-952d92456975?w=800&q=80",
        "https://images.unsplash.com/photo-1564834724105-918b73d1b8e0?w=800&q=80",
        "https://images.unsplash.com/photo-1512621776951-a57141f2eefd?w=800&q=80",
    ],
    "veg_curry": [
        "https://images.unsplash.com/photo-1565557623262-b51c2513a641?w=800&q=80",
        "https://images.unsplash.com/photo-1585937421612-70a008356fbe?w=800&q=80",
        "https://images.unsplash.com/photo-1631515243349-e0cb75fb8d3a?w=800&q=80",
        "https://images.unsplash.com/photo-1574484284002-952d92456975?w=800&q=80",
        "https://images.unsplash.com/photo-1455619452474-d2be8b1e70cd?w=800&q=80",
        "https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=800&q=80",
        "https://images.unsplash.com/photo-1596797038530-2c107229654b?w=800&q=80",
        "https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=800&q=80",
        "https://images.unsplash.com/photo-1567337710282-00832b415979?w=800&q=80",
        "https://images.unsplash.com/photo-1612929633738-8fe44f7ec841?w=800&q=80",
        "https://images.unsplash.com/photo-1576402187878-974f70c890a5?w=800&q=80",
    ],
    "chicken_curry": [
        "https://images.unsplash.com/photo-1603894584373-5ac82b2ae328?w=800&q=80",
        "https://images.unsplash.com/photo-1631452180519-c014fe946bc7?w=800&q=80",
        "https://images.unsplash.com/photo-1565557623262-b51c2513a641?w=800&q=80",
        "https://images.unsplash.com/photo-1505253758473-96b7015fcd40?w=800&q=80",
        "https://images.unsplash.com/photo-1574653853027-5382a3d23a15?w=800&q=80",
        "https://images.unsplash.com/photo-1628294895950-9805252327bc?w=800&q=80",
        "https://images.unsplash.com/photo-1607116667573-1c7d73636ba0?w=800&q=80",
        "https://images.unsplash.com/photo-1599043513900-ed6fe01d3833?w=800&q=80",
        "https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=800&q=80",
        "https://images.unsplash.com/photo-1585937421612-70a008356fbe?w=800&q=80",
        "https://images.unsplash.com/photo-1633321702518-7fecdafb94d5?w=800&q=80",
        "https://images.unsplash.com/photo-1596797038530-2c107229654b?w=800&q=80",
        "https://images.unsplash.com/photo-1563379091339-03b21ab4a4f4?w=800&q=80",
        "https://images.unsplash.com/photo-1580217593608-61931cefc821?w=800&q=80",
        "https://images.unsplash.com/photo-1606755962773-d324e0a13086?w=800&q=80",
        "https://images.unsplash.com/photo-1612874742237-6526221588e3?w=800&q=80",
        "https://images.unsplash.com/photo-1625937286520-3ef7955f3813?w=800&q=80",
        "https://images.unsplash.com/photo-1627662168223-7df99068099a?w=800&q=80",
        "https://images.unsplash.com/photo-1631515243349-e0cb75fb8d3a?w=800&q=80",
        "https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=800&q=80",
        "https://images.unsplash.com/photo-1606728035253-49e8a23146de?w=800&q=80",
        "https://images.unsplash.com/photo-1574484284002-952d92456975?w=800&q=80",
        "https://images.unsplash.com/photo-1567188040759-fb8a883dc6d8?w=800&q=80",
        "https://images.unsplash.com/photo-1551881192-5e377f1b2142?w=800&q=80",
        "https://images.unsplash.com/photo-1573080496219-bb080dd4f877?w=800&q=80",
        "https://images.unsplash.com/photo-1614398751058-bca239de00ca?w=800&q=80",
        "https://images.unsplash.com/photo-1608039755401-742074f0548d?w=800&q=80",
    ],
    "bread": [
        "https://images.unsplash.com/photo-1565557623262-b51c2513a641?w=800&q=80",
        "https://images.unsplash.com/photo-1600326145359-3a44909d1a39?w=800&q=80",
        "https://images.unsplash.com/photo-1574071318508-1cdbab80d002?w=800&q=80",
        "https://images.unsplash.com/photo-1586444248879-bc604bc77212?w=800&q=80",
        "https://images.unsplash.com/photo-1567620832903-9fc6debc209f?w=800&q=80",
        "https://images.unsplash.com/photo-1555939594-58d7cb561ad1?w=800&q=80",
        "https://images.unsplash.com/photo-1601050690597-df0568f70950?w=800&q=80",
        "https://images.unsplash.com/photo-1555507036-ab1f4038024a?w=800&q=80",
        "https://images.unsplash.com/photo-1573675542321-f51b18e6e759?w=800&q=80",
        "https://images.unsplash.com/photo-1515024014929-c2ba2c4da1d0?w=800&q=80",
        "https://images.unsplash.com/photo-1519864600395-3404e40a0eda?w=800&q=80",
        "https://images.unsplash.com/photo-1509722747041-616f39b57569?w=800&q=80",
        "https://images.unsplash.com/photo-1528736235302-52922df5c122?w=800&q=80",
        "https://images.unsplash.com/photo-1605888969139-42cca4308aa2?w=800&q=80",
    ],
    "rice": [
        "https://images.unsplash.com/photo-1596560548464-f010549b84d7?w=800&q=80",
        "https://images.unsplash.com/photo-1563379091339-03b21ab4a4f4?w=800&q=80",
        "https://images.unsplash.com/photo-1589302168068-964664d93dc0?w=800&q=80",
        "https://images.unsplash.com/photo-1516714435131-44d6b64dc6a2?w=800&q=80",
        "https://images.unsplash.com/photo-1536304993881-460587633ee1?w=800&q=80",
        "https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=800&q=80",
        "https://images.unsplash.com/photo-1512058564366-18510be2db19?w=800&q=80",
        "https://images.unsplash.com/photo-1645696301019-35adcc552067?w=800&q=80",
    ],
    "biryani": [
        "https://images.unsplash.com/photo-1563379091339-03b21ab4a4f4?w=800&q=80",
        "https://images.unsplash.com/photo-1589302168068-964664d93dc0?w=800&q=80",
        "https://images.unsplash.com/photo-1631515243349-e0cb75fb8d3a?w=800&q=80",
        "https://images.unsplash.com/photo-1642821373181-16a5bc9f5801?w=800&q=80",
        "https://images.unsplash.com/photo-1633321702518-7fecdafb94d5?w=800&q=80",
        "https://images.unsplash.com/photo-1596797038530-2c107229654b?w=800&q=80",
        "https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=800&q=80",
        "https://images.unsplash.com/photo-1574484284002-952d92456975?w=800&q=80",
        "https://images.unsplash.com/photo-1585937421612-70a008356fbe?w=800&q=80",
        "https://images.unsplash.com/photo-1565557623262-b51c2513a641?w=800&q=80",
        "https://images.unsplash.com/photo-1606728035253-49e8a23146de?w=800&q=80",
        "https://images.unsplash.com/photo-1612874742237-6526221588e3?w=800&q=80",
        "https://images.unsplash.com/photo-1625937286520-3ef7955f3813?w=800&q=80",
        "https://images.unsplash.com/photo-1599043513900-ed6fe01d3833?w=800&q=80",
        "https://images.unsplash.com/photo-1628294895950-9805252327bc?w=800&q=80",
        "https://images.unsplash.com/photo-1607116667573-1c7d73636ba0?w=800&q=80",
        "https://images.unsplash.com/photo-1574653853027-5382a3d23a15?w=800&q=80",
        "https://images.unsplash.com/photo-1603894584373-5ac82b2ae328?w=800&q=80",
        "https://images.unsplash.com/photo-1551881192-5e377f1b2142?w=800&q=80",
        "https://images.unsplash.com/photo-1505253758473-96b7015fcd40?w=800&q=80",
        "https://images.unsplash.com/photo-1627662168223-7df99068099a?w=800&q=80",
    ],
    "fried_rice": [
        "https://images.unsplash.com/photo-1603133872878-684f208fb84b?w=800&q=80",
        "https://images.unsplash.com/photo-1512058564366-18510be2db19?w=800&q=80",
        "https://images.unsplash.com/photo-1596560548464-f010549b84d7?w=800&q=80",
        "https://images.unsplash.com/photo-1645696301019-35adcc552067?w=800&q=80",
        "https://images.unsplash.com/photo-1516714435131-44d6b64dc6a2?w=800&q=80",
        "https://images.unsplash.com/photo-1536304993881-460587633ee1?w=800&q=80",
        "https://images.unsplash.com/photo-1617093727343-374698b1b08d?w=800&q=80",
        "https://images.unsplash.com/photo-1604908176997-125f25cc6f3d?w=800&q=80",
    ],
    "snack": [
        "https://images.unsplash.com/photo-1601050690597-df0568f70950?w=800&q=80",
        "https://images.unsplash.com/photo-1585032226651-759b368d7246?w=800&q=80",
        "https://images.unsplash.com/photo-1599487488170-d11ec9c172f0?w=800&q=80",
        "https://images.unsplash.com/photo-1567337710282-00832b415979?w=800&q=80",
        "https://images.unsplash.com/photo-1573080496219-bb080dd4f877?w=800&q=80",
    ],
    "kamju": [
        "https://images.unsplash.com/photo-1559339352-11d035aa65de?w=800&q=80",
        "https://images.unsplash.com/photo-1553621042-f6e147245754?w=800&q=80",
        "https://images.unsplash.com/photo-1565680018093-ebb6e3062e7b?w=800&q=80",
        "https://images.unsplash.com/photo-1612929633738-8fe44f7ec841?w=800&q=80",
        "https://images.unsplash.com/photo-1569058242567-93de6f36f8e6?w=800&q=80",
    ],
}

# Track used URLs globally
_used_urls: set[str] = set()


def _pick_image(item: str, subcat: str) -> str:
    """Pick a unique Unsplash URL for the item based on its subcategory."""
    s = subcat.lower()
    if "veg soup" in s:
        pool = _UNSPLASH["veg_soup"]
    elif "non-veg soup" in s:
        pool = _UNSPLASH["nonveg_soup"]
    elif "tandoori" in s or "tikka" in s:
        pool = _UNSPLASH["tandoori"]
    elif "chicken starter" in s:
        pool = _UNSPLASH["chicken_starter"]
    elif "mutton" in s and "main" in s:
        pool = _UNSPLASH["mutton"]
    elif "mutton" in s:
        pool = _UNSPLASH["mutton"]
    elif "egg" in s and "main" not in s:
        pool = _UNSPLASH["egg"]
    elif "egg" in s:
        pool = _UNSPLASH["egg"]
    elif "seafood" in s or "prawns" in s or "fish" in item.lower() or "prawn" in item.lower():
        pool = _UNSPLASH["seafood"]
    elif "kamju" in s:
        pool = _UNSPLASH["kamju"]
    elif "veg starter" in s:
        pool = _UNSPLASH["veg_starter"]
    elif "veg main" in s:
        pool = _UNSPLASH["veg_curry"]
    elif "chicken main" in s or "chicken" in s:
        pool = _UNSPLASH["chicken_curry"]
    elif "paratha" in s or "roti" in s or "naan" in s or "kulcha" in s or "stuffed" in s:
        pool = _UNSPLASH["bread"]
    elif "rice" in s and "biryani" not in s and "fried" not in s:
        pool = _UNSPLASH["rice"]
    elif "biryani" in s:
        pool = _UNSPLASH["biryani"]
    elif "fried rice" in s:
        pool = _UNSPLASH["fried_rice"]
    elif "snack" in s:
        pool = _UNSPLASH["snack"]
    else:
        pool = _UNSPLASH["chicken_curry"]

    for url in pool:
        if url not in _used_urls:
            _used_urls.add(url)
            return url
    # Fallback: reuse from pool if all exhausted
    return pool[0]


def get_subcategory(item: str, category: str) -> str:
    if item in SUBCATEGORY_MAP:
        return SUBCATEGORY_MAP[item]
    return category


def is_veg(item: str) -> bool:
    return item in VEG_ITEMS


def get_description(item: str) -> str:
    return DESCRIPTIONS.get(item, f"Delicious {item} prepared with authentic spices")


def enrich_excel():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)

    new_cols = {5: "Description", 6: "Subcategory", 7: "Placeholder", 8: "Veg/Non-Veg", 9: "Image URL"}
    # Shift: E=Description, F=Subcategory, G=(unused placeholder removed), H=Veg/Non-Veg, I=Image URL
    # Actually per user mapping: E=Description, F=Subcategory, G=unused, H=Veg/Non-Veg, I=Image URL
    # Let's use exactly E, F, (skip G), H, I

    for ci, hdr in [(5, "Description"), (6, "Subcategory"), (8, "Veg/Non-Veg"), (9, "Image URL")]:
        c = ws.cell(row=1, column=ci, value=hdr)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center")

    # Also style existing headers
    for ci in range(1, 5):
        c = ws.cell(row=1, column=ci)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center")

    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=2).value
        category = ws.cell(row=r, column=1).value
        if not item:
            continue

        item = str(item).strip()
        cat = str(category).strip() if category else ""

        subcat = get_subcategory(item, cat)
        veg = is_veg(item)
        desc = get_description(item)
        img = _pick_image(item, subcat)

        ws.cell(row=r, column=5, value=desc)
        ws.cell(row=r, column=6, value=subcat)
        ws.cell(row=r, column=8, value="Veg" if veg else "Non-Veg")
        c = ws.cell(row=r, column=9, value=img)
        c.font = Font(color="0563C1", underline="single", size=9)

    from openpyxl.utils import get_column_letter
    for col in range(1, 10):
        mx = max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row+1, 250))), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 3, 55)

    wb.save(XLSX_PATH)
    print(f"Enriched {XLSX_PATH.name}: added Description(E), Subcategory(F), Veg/Non-Veg(H), Image(I)")
    return ws.max_row - 1


def build_payloads() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    payloads: list[dict] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        cat = row[0]      # A
        name = row[1]      # B
        price = row[2]     # C
        hike = row[3]      # D
        desc = row[4]      # E
        subcat = row[5]    # F
        veg_raw = row[7]   # H
        img = row[8]       # I

        if not name or price is None:
            continue

        is_veg_val = None
        if veg_raw:
            v = str(veg_raw).strip().lower()
            is_veg_val = v == "veg"

        hike_val = 0
        if hike is not None:
            try:
                h = float(hike)
                hike_val = round(h * 100, 1) if h < 1 else round(h, 1)
            except (ValueError, TypeError):
                hike_val = 0

        payload: dict = {
            "name": str(name).strip(),
            "restaurantPrice": float(price),
            "hikePercentage": hike_val,
            "category": str(cat).strip() if cat else None,
            "subCategory": str(subcat).strip() if subcat else None,
            "isVeg": is_veg_val,
            "isAvailable": True,
            "description": str(desc).strip() if desc else None,
            "image": [str(img).strip()] if img else [],
        }
        payloads.append(payload)

    wb.close()
    return payloads


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich & import Maharaja menu")
    parser.add_argument("--apply", action="store_true", help="POST to API (default is dry run)")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--bearer-token", default=os.environ.get("HONESTEATS_BEARER_TOKEN"))
    parser.add_argument("--delay", type=float, default=0.3)
    args = parser.parse_args()

    # Step 1: Enrich Excel
    enrich_excel()

    # Step 2: Build payloads
    payloads = build_payloads()
    print(f"\nLoaded {len(payloads)} menu items")
    print(f"Restaurant ID: {RESTAURANT_ID}\n")

    menu_url = f"{args.api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"

    for i, p in enumerate(payloads, 1):
        print(f"[{i:3d}] {p['name']}")
        print(f"      ₹{p['restaurantPrice']} + {p['hikePercentage']}% | {p['category']} > {p['subCategory']} | {'Veg' if p['isVeg'] else 'Non-Veg'}")
        if not args.apply:
            print(f"      {json.dumps(p, indent=None, ensure_ascii=False)}")
        print()

    if not args.apply:
        print("=" * 60)
        print("DRY RUN — no HTTP calls.")
        print(f"To insert: python3 {sys.argv[0]} --apply")
        print("=" * 60)
        return 0

    # Insert
    s = requests.Session()
    s.headers["Content-Type"] = "application/json"
    if args.bearer_token:
        s.headers["Authorization"] = f"Bearer {args.bearer_token}"
    else:
        s.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS

    ok = err = 0
    for i, p in enumerate(payloads, 1):
        if args.delay > 0:
            time.sleep(args.delay)
        try:
            r = requests.Session()
            r = s.post(menu_url, json=p, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"  ✗ [{i}] {p['name']}: {ex}", file=sys.stderr)
            continue
        if r.status_code == 201:
            ok += 1
            print(f"  ✓ [{i}] {p['name']} created")
        else:
            err += 1
            print(f"  ✗ [{i}] {p['name']}: HTTP {r.status_code} {r.text[:200]}", file=sys.stderr)
            if r.status_code == 401:
                print("\nUnauthorized.", file=sys.stderr)
                return 1

    print(f"\nDone: {ok} created, {err} failed.")
    return 0 if err == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
