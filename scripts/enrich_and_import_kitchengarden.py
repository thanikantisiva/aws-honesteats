#!/usr/bin/env python3
"""
Enrich Kitchen Garden menu Excel with veg/non-veg, description, unique Unsplash
images and generate API payloads for import.

Columns: A=Category, B=Item, C=Dine-In, D=Hike%, E=Subcategory
Adds:    F=Description, G=(skip), H=Veg/Non-Veg, I=Image URL

Usage:
  python3 scripts/enrich_and_import_kitchengarden.py            # enrich + dry run
  python3 scripts/enrich_and_import_kitchengarden.py --apply     # enrich + insert
"""
from __future__ import annotations
import argparse, json, os, sys, time, re
from pathlib import Path
import openpyxl, requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RESTAURANT_ID = "RES-1776414760552-3958"
XLSX_PATH = Path("/Users/user/Downloads/KitchenGarden_Zomato_vs_DineIn.xlsx")
RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get("HONESTEATS_RETOOL_BYPASS", "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc")
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

# ── Veg items (everything else is Non-Veg) ───────────────────────────────────────
VEG_ITEMS = {
    "Gobi 65","Dragon Paneer","Paneer Majestic","Baby Corn Manchurian","Gobi Manchurian",
    "Paneer 65","Cripsy Corn","Paneer Manchuria","Veg Manchurian","Chilli Gobi",
    "Spicy Crispy Paneer","Mushroom Manchurian","Pepper Mushroom","Chilli Mushroom",
    "Veg Cheese Stick","Babycorn 65","Golden Fried Babycorn","Lemon Veg","Aloo 65",
    "Baby Corn Hot & Pepper","Babycorn & Mushroom Salt & Pepper","Chilli Babycorn",
    "Chilli Paneer","Crispy Mushroom","French Fries","Ginger Gobi","Mushroom 65",
    "Paneer Pepper Fry","Pepper Paneer","Peri Peri French Fries","Schezwan Paneer",
    "Veg Crispy Chilli","Veg Tempura",
    "Veg Manchow Soup","Spicy Corn With Almond Soup","Cream Of Mushroom Soup",
    "Cream Of Spinach Soup","Cream Of Tomato Soup","Cream Of Veg Soup","Veg Noodle Soup",
    "Paneer Butter Masala","Sambar Rice","Spl. Kaju Curry","Kaju Paneer Masala",
    "Mushroom Masala","Kadai Paneer","Dal Fry","Dal Tadka","Kadai Mushroom",
    "Mix Veg Curry","Aloo Gobi Masala","Kadai Veg","Paneer Lajawab","Paneer Tikka Masala",
    "Plain Palak","Sabji Punjabi","Veg Diwani Handi","Veg Jaipuri",
    "Paneer Tikka Kebab","Achari Paneer Tikka","Hariyali Paneer Tikka",
    "Tandoori Babycorn","Tandoori Mushroom","Veg Hara Bhara Kabab",
    "Aloo Paratha","Butter Kulcha","Masala Kulcha","Butter Naan","Butter Roti",
    "Folding Butter Naan","Kashmir Naan","Kulcha","Lachha Paratha",
    "Plain Naan","Plain Paratha","Pudina Paratha","Tandoori Roti",
    "Kaju Fried Rice","Paneer Fried Rice","Mushroom Fried Rice","Jeera Rice",
    "Veg Fried Rice","Mixed Veg Fried Rice","Veg Schezwan Fried Rice",
    "Burnt Garlic Fried Rice","Ghee Rice","Masala Rice","Veg Chopper Fried Rice",
    "Veg Pot Rice","Veg Thai Fried Rice Shanghai","Special Curd Rice","Curd Rice",
    "Dal Khichdi","Veg Pulao","Masala Khichdi",
    "Veg Hakka Noodles","Veg Schezwan Noodles","Pan Fried Noodles",
    "Bell Pepper Noodles","Cantonese Pan Fried Noodles","Veg Shanghai Noodles",
    "Veg American Chopsuey","Veg Chinese Chopsuey",
    "Paneer Biryani","Mushroom Biryani","Biryani Rice","Kaju Biryani","Veg Biryani",
    "Veg Handi Biryani","Veg Family Pack","Veg Jumbo Pack",
    "Ulavacharu Veg Biryani","Veg Patiala Biryani","Tawa Paneer Biryani",
    "Mushroom Tikka Biryani","Ragi Sankati",
    "Pomegranate Juice","Pineapple Juice","Orange Juice","Dryfruits Juice",
    "Watermelon Juice.","Sweet Lime Juice","Apple Juice.","Banana Juice",
    "Mixed Fruit Juice","Seasonal Fruit Juice",
    "Chocolate Milkshake","Mango Milkshake","Strawberry Milkshake","Vanilla Milkshake",
    "Banana Milkshake","Apple Milkshake","Butterscotch Milkshakes","Litchi Milkshake",
    "Oreo Milkshake","Pista Milkshakes",
    "Virgin Mojito","Blue Lagoon","Strawberry Lemonade","Fruit Crush",
    "Blue Hawaiian","Cinderella","Downstreet","Fruit Smarty","Pink Lady",
    "Rainbow Mocktail","Shirley Temple","Spicy Twist","Virgin Pina Colada",
    "Lassi Sweet","Butter Milk","Lassi Salt",
    "Fruit Salad","Arabian Delight","Fruit Salad With Icecream","Apricot Delite","Bengali Sweets",
    "Masala Papad","Roasted Papad","Cup Curd","Pineapple Raita","Veg Raita",
    "Boiled Vegetable Salad","Cucumber Salad","Green Salad","Smoked Vegetable Salad",
}

# ── Descriptions ─────────────────────────────────────────────────────────────────
DESC: dict[str, str] = {
    "Mutton 65": "Crispy deep-fried mutton with classic 65 spice blend",
    "Pepper Mutton": "Black pepper-crusted mutton with bold peppery kick",
    "Apollo Fish": "Hyderabadi-style crispy Apollo fish with spicy batter",
    "Butter Garlic Prawns": "Succulent prawns sautéed in garlic butter",
    "Peri Peri Fish Fingers": "Crispy fish fingers with fiery peri peri seasoning",
    "Peri Peri Prawns": "Prawns tossed in spicy peri peri sauce",
    "Hot Garlic Fish": "Fish in aromatic hot garlic sauce",
    "Golden Fried Prawns": "Crispy golden-battered prawns",
    "Chilli Fried Fish With Red Pepper": "Fried fish with red pepper chilli sauce",
    "Salt & Pepper Prawns": "Prawns with salt and cracked pepper seasoning",
    "Schezwan Fish": "Fish in fiery Schezwan sauce",
    "Gobi 65": "Crispy cauliflower fritters with spicy 65 seasoning",
    "Dragon Paneer": "Paneer in spicy dragon sauce with peppers",
    "Paneer Majestic": "Crispy paneer tossed with curry leaves and spices",
    "Baby Corn Manchurian": "Baby corn in tangy Manchurian sauce",
    "Gobi Manchurian": "Crispy cauliflower in tangy Manchurian sauce",
    "Paneer 65": "Crispy paneer cubes with spicy 65 masala",
    "Cripsy Corn": "Crispy golden fried corn kernels with seasoning",
    "Paneer Manchuria": "Paneer balls in Indo-Chinese Manchurian sauce",
    "Veg Manchurian": "Mixed vegetable balls in Manchurian sauce",
    "Chilli Gobi": "Cauliflower florets tossed in spicy chilli sauce",
    "Spicy Crispy Paneer": "Extra crispy paneer with fiery spice coating",
    "Mushroom Manchurian": "Button mushrooms in rich Manchurian sauce",
    "Pepper Mushroom": "Mushrooms with cracked black pepper seasoning",
    "Chilli Mushroom": "Mushrooms in fiery Indo-Chinese chilli sauce",
    "Veg Cheese Stick": "Crispy cheese sticks with vegetable filling",
    "Babycorn 65": "Crispy baby corn with spicy 65 masala",
    "Golden Fried Babycorn": "Golden-battered crispy baby corn",
    "Lemon Veg": "Mixed vegetables in tangy lemon sauce",
    "Aloo 65": "Crispy potato cubes with spicy 65 masala",
    "Baby Corn Hot & Pepper": "Baby corn with hot pepper seasoning",
    "Babycorn & Mushroom Salt & Pepper": "Baby corn and mushroom with salt and pepper",
    "Chilli Babycorn": "Baby corn tossed in spicy chilli sauce",
    "Chilli Paneer": "Paneer cubes in Indo-Chinese chilli sauce",
    "Crispy Mushroom": "Golden-fried crispy mushrooms",
    "French Fries": "Classic golden salted french fries",
    "Ginger Gobi": "Cauliflower with ginger and aromatic spices",
    "Mushroom 65": "Crispy mushrooms with spicy 65 seasoning",
    "Paneer Pepper Fry": "Paneer stir-fried with black pepper",
    "Pepper Paneer": "Paneer with cracked pepper seasoning",
    "Peri Peri French Fries": "French fries with fiery peri peri seasoning",
    "Schezwan Paneer": "Paneer in spicy Schezwan sauce",
    "Veg Crispy Chilli": "Mixed veggies in crispy chilli preparation",
    "Veg Tempura": "Light and crispy vegetable tempura",
    "Chilli Egg": "Boiled eggs tossed in spicy chilli sauce",
    "Egg 65": "Crispy deep-fried egg fritters with 65 masala",
    "Egg Manchurian": "Egg balls in tangy Manchurian sauce",
    "Ginger Egg Starter": "Eggs sautéed with ginger and spices",
    "Chicken Lollipop": "Crispy fried chicken lollipops in tangy sauce",
    "Chilli Chicken": "Indo-Chinese classic — crispy chicken in chilli sauce",
    "Chicken Majestic": "Crispy chicken tossed with curry leaves and spices",
    "Chicken Drum Stick": "Juicy marinated chicken drumsticks",
    "Chicken 65": "Crispy deep-fried chicken with classic 65 spices",
    "Dragon Chicken": "Chicken in spicy dragon sauce with peppers",
    "Peri Peri Chicken Fingers": "Chicken fingers with peri peri seasoning",
    "Kaju Chicken Pakoda": "Cashew-coated chicken pakoda fritters",
    "Lemon Chicken": "Tangy lemon-glazed chicken",
    "Maharaja Chicken 65 (bone)": "Royal-style bone-in chicken 65",
    "Egg Omlette": "Classic egg omelette",
    "Cheese Omlet": "Fluffy cheese omelette",
    "Chicken Manchurian": "Chicken in tangy Manchurian sauce",
    "Crispy Fried Chicken": "Extra crispy deep-fried chicken",
    "Crispy Fried Lemon Chicken": "Crispy fried chicken with lemon glaze",
    "Ginger Chicken Starter": "Chicken sautéed with ginger and aromatics",
    "Kimchi Kai Chicken": "Korean-inspired kimchi chicken",
    "Masala Omlet": "Spiced masala omelette with onions and chilies",
    "Multi Flavor Chicken": "Chicken with a medley of signature flavors",
    "Roasted Chilli Chicken": "Chicken roasted with fiery red chillies",
    "Veg Manchow Soup": "Spicy Indo-Chinese veg manchow soup with crispy noodles",
    "Chicken Kalimirchi Soup": "Black pepper chicken soup",
    "Spicy Corn With Almond Soup": "Creamy corn soup with toasted almonds",
    "Chicken Hunan Soup": "Spicy Hunan-style chicken soup",
    "Chicken Lemon Coriander Soup": "Chicken soup with lemon and coriander",
    "Chicken Lemon Pepper Soup": "Chicken soup with lemon and pepper",
    "Chicken Noodle Soup": "Classic chicken noodle soup",
    "Cream Of Chicken Soup": "Rich and creamy chicken soup",
    "Cream Of Mushroom Soup": "Velvety cream of mushroom soup",
    "Cream Of Spinach Soup": "Smooth and creamy spinach soup",
    "Cream Of Tomato Soup": "Classic creamy tomato soup",
    "Cream Of Veg Soup": "Creamy mixed vegetable soup",
    "Veg Noodle Soup": "Light vegetable noodle soup",
    "Paneer Butter Masala": "Paneer in creamy tomato-butter gravy",
    "Sambar Rice": "South Indian sambar served with rice",
    "Spl. Kaju Curry": "Special cashew nut curry in rich gravy",
    "Kaju Paneer Masala": "Cashews and paneer in creamy masala",
    "Mushroom Masala": "Button mushrooms in rich masala gravy",
    "Kadai Paneer": "Paneer cooked kadai-style with bell peppers",
    "Dal Fry": "Classic dal fry with aromatic tempering",
    "Dal Tadka": "Yellow dal with sizzling tadka",
    "Kadai Mushroom": "Mushrooms in kadai with capsicum and spices",
    "Mix Veg Curry": "Mixed vegetables in mild curry",
    "Aloo Gobi Masala": "Potato and cauliflower in spiced masala",
    "Kadai Veg": "Mixed vegetables in kadai with peppers",
    "Paneer Lajawab": "Paneer in rich and creamy lajawab gravy",
    "Paneer Tikka Masala": "Grilled paneer tikka in smoky masala gravy",
    "Plain Palak": "Creamy spinach gravy",
    "Sabji Punjabi": "North Indian-style mixed vegetable sabji",
    "Veg Diwani Handi": "Assorted vegetables in handi-style gravy",
    "Veg Jaipuri": "Vegetables in Rajasthani Jaipuri-style curry",
    "Rayalaseema Chicken Curry Boneless": "Fiery Rayalaseema-style boneless chicken",
    "Butter Chicken": "Creamy butter chicken — a North Indian classic",
    "Rayalaseema Chicken Curry Bone": "Rayalaseema-style bone-in chicken curry",
    "Andhra Chicken Curry Boneless": "Spicy Andhra-style boneless chicken",
    "Kadai Chicken": "Chicken kadai with bell peppers and spices",
    "Mughlai Chicken Curry": "Rich Mughlai chicken in cream and nut gravy",
    "Hyderabadi Green Chicken": "Hyderabadi green masala chicken",
    "Chicken Afghani Masala": "Afghani-style creamy chicken masala",
    "Handi Chicken": "Chicken slow-cooked in clay handi",
    "Tawa Chicken Boneless Curry": "Tawa-cooked boneless chicken curry",
    "Chicken Dilkush": "Chef's special chicken dilkush",
    "Kodi Vepudu": "Andhra-style spicy chicken fry",
    "Chicken Begum Bahaar": "Royal Begum Bahaar style chicken",
    "Chicken Chen": "Chef's signature chicken chen preparation",
    "Chicken Lababdar": "Chicken in rich lababdar tomato gravy",
    "Chicken Masala": "Classic spicy chicken masala curry",
    "Chicken Tadka": "Chicken with sizzling tadka seasoning",
    "Dhaba Chicken": "Rustic highway dhaba-style chicken curry",
    "Dum Ka Murgh": "Whole chicken slow-cooked dum-style",
    "Matka Chicken Curry": "Chicken curry served in earthen pot",
    "Raichur Chicken Roast": "Raichur-style roasted chicken",
    "Punjabi Egg Masala": "Eggs in Punjabi-style masala gravy",
    "Ginger Egg Curry": "Eggs in ginger-flavored curry",
    "Egg Do Pyaza": "Eggs in onion-rich do pyaza gravy",
    "Egg Masala": "Boiled eggs in aromatic masala curry",
    "Coriander Mutton Curry": "Mutton in fresh coriander gravy",
    "Mutton Angara": "Charcoal-smoked mutton in fiery angara masala",
    "Mutton Kurma": "Mutton in rich and creamy kurma gravy",
    "Mutton Lal Masala": "Mutton in fiery red lal masala",
    "Mutton Masala": "Tender mutton in rich masala gravy",
    "Mutton Rogan Josh": "Kashmiri-style rogan josh mutton",
    "Mutton Shikari": "Hunter-style mutton with bold spices",
    "Nalli Nihari": "Slow-cooked nalli nihari with marrow",
    "Nihari Mutton Curry": "Traditional Lucknowi nihari mutton",
    "Soufian Mutton Curry": "Sufi-inspired aromatic mutton curry",
    "Prawns Fry": "Crispy masala-fried prawns",
    "Fish Masala": "Fish in rich and spicy masala gravy",
    "Prawns Curry": "Prawns in tangy curry sauce",
    "Tandoori Chicken": "Classic clay-oven roasted tandoori chicken",
    "Chicken Tikka Kebab": "Boneless chicken tikka grilled on skewers",
    "Tangdi Kabab": "Juicy chicken leg kababs in tandoori spices",
    "Alfam Chicken (half)": "Half portion of Arabian-style alfam chicken",
    "Achari Murgh Tikka": "Pickle-marinated chicken tikka",
    "Murgh Afghani Kebab": "Creamy Afghani-style chicken kebab",
    "Tandoori Wings": "Tandoor-roasted spicy chicken wings",
    "Alfam Chicken (full)": "Full Arabian-style alfam chicken",
    "Chicken Dilruba Kebab": "Premium chicken dilruba kebab",
    "Chicken Jojo Kebab": "Signature jojo-style chicken kebab",
    "Chilli Flakes Kabab": "Kebab with crushed chilli flakes",
    "Murgh Beetroot Kebab": "Chicken kebab with beetroot marinade",
    "Murgh Multani Kabab": "Multani-style chicken kebab",
    "Mutton Sangai Kabab": "Premium mutton sangai kebab",
    "Raichur Murgh Tikka": "Raichur-style chicken tikka",
    "Paneer Tikka Kebab": "Grilled paneer tikka on skewers",
    "Achari Paneer Tikka": "Pickle-marinated paneer tikka",
    "Hariyali Paneer Tikka": "Green herb-marinated paneer tikka",
    "Tandoori Babycorn": "Baby corn grilled in tandoor",
    "Tandoori Mushroom": "Mushrooms marinated and tandoor-grilled",
    "Veg Hara Bhara Kabab": "Green vegetable kabab with spinach and peas",
    "Achari Mutton Kebab": "Pickle-marinated mutton kebab",
    "Mutton King Kebab": "Royal king-size mutton kebab",
    "Grilled Chicken": "Perfectly grilled seasoned chicken",
    "Aloo Paratha": "Stuffed potato paratha", "Butter Kulcha": "Buttery kulcha bread",
    "Masala Kulcha": "Spiced masala kulcha", "Butter Naan": "Soft buttery naan",
    "Butter Roti": "Roti brushed with butter", "Chicken Keema Naan": "Naan stuffed with minced chicken",
    "Folding Butter Naan": "Folded layered butter naan", "Kashmir Naan": "Kashmiri naan with dry fruits",
    "Kulcha": "Soft fluffy plain kulcha", "Lachha Paratha": "Multi-layered crispy lachha paratha",
    "Mutton Keema Kulcha": "Kulcha stuffed with spiced mutton keema",
    "Mutton Keema Naan": "Naan stuffed with spiced mutton keema",
    "Plain Naan": "Soft fluffy plain naan", "Plain Paratha": "Classic plain paratha",
    "Pudina Paratha": "Mint-flavored paratha", "Tandoori Roti": "Whole wheat roti from tandoor",
    "Chicken Schezwan Fried Rice": "Spicy Schezwan chicken fried rice",
    "Kaju Fried Rice": "Fried rice with roasted cashews",
    "Chicken Fried Rice": "Classic chicken fried rice",
    "Egg Fried Rice": "Egg fried rice with vegetables",
    "Chicken Mongolian Fried Rice": "Mongolian-style chicken fried rice",
    "Paneer Fried Rice": "Fried rice with paneer cubes",
    "Mushroom Fried Rice": "Fried rice with button mushrooms",
    "Jeera Rice": "Fragrant cumin-tempered rice",
    "Veg Fried Rice": "Classic vegetable fried rice",
    "Mixed Veg Fried Rice": "Fried rice with assorted vegetables",
    "Veg Schezwan Fried Rice": "Spicy Schezwan veg fried rice",
    "Burnt Garlic Fried Rice": "Fried rice with roasted garlic",
    "Chicken Chopper Fried Rice": "Loaded chicken chopper fried rice",
    "Chicken Pot Rice": "Chicken rice cooked in pot",
    "Ghee Rice": "Aromatic basmati rice in pure ghee",
    "Masala Rice": "Spiced masala flavored rice",
    "Veg Chopper Fried Rice": "Loaded veg chopper fried rice",
    "Veg Pot Rice": "Vegetable rice cooked in pot",
    "Veg Thai Fried Rice Shanghai": "Thai-Shanghai style veg fried rice",
    "Special Curd Rice": "Premium curd rice with tempering",
    "Chicken Sambar Rice": "Chicken sambar served with rice",
    "Curd Rice": "Classic South Indian curd rice",
    "Dal Khichdi": "Comforting dal khichdi",
    "Veg Pulao": "Fragrant vegetable pulao",
    "Masala Khichdi": "Spiced masala khichdi",
    "Chicken Garlic Hakka Noodles": "Garlic-flavored chicken hakka noodles",
    "Chicken Shanghai Noodles": "Shanghai-style chicken noodles",
    "Schezwan Chicken Noodles": "Spicy Schezwan chicken noodles",
    "Veg Hakka Noodles": "Classic veg hakka noodles",
    "Veg Schezwan Noodles": "Spicy Schezwan veg noodles",
    "Pan Fried Noodles": "Crispy pan-fried noodles",
    "Bell Pepper Noodles": "Noodles with colorful bell peppers",
    "Cantonese Pan Fried Noodles": "Cantonese-style crispy noodles",
    "Hunan Chicken Noodles": "Spicy Hunan chicken noodles",
    "Veg Shanghai Noodles": "Shanghai-style veg noodles",
    "Chinese Chicken Chopsuey": "Chicken chopsuey with crispy noodles",
    "Veg American Chopsuey": "American-style veg chopsuey",
    "Veg Chinese Chopsuey": "Chinese-style veg chopsuey",
    "Egg Biryani": "Flavorful biryani with boiled eggs",
    "Egg Handi Biryani": "Egg biryani served in clay handi",
    "Paneer Biryani": "Dum biryani with soft paneer cubes",
    "Mushroom Biryani": "Aromatic biryani with mushrooms",
    "Biryani Rice": "Aromatic basmati biryani rice",
    "Kaju Biryani": "Biryani with roasted cashew nuts",
    "Veg Biryani": "Mixed vegetable dum biryani",
    "Veg Handi Biryani": "Veg biryani in clay handi",
    "Veg Family Pack": "Veg biryani family pack — serves 4",
    "Veg Jumbo Pack": "Veg biryani jumbo pack — serves 6",
    "Ulavacharu Veg Biryani": "Horsegram-flavored veg biryani",
    "Veg Patiala Biryani": "Patiala-style rich veg biryani",
    "Tawa Paneer Biryani": "Tawa paneer layered biryani",
    "Chicken Dum Biryani": "Authentic Hyderabadi chicken dum biryani",
    "Chicken Dum Biryani Family Pack": "Chicken dum biryani — serves 4",
    "Special Chicken Biryani": "Chef's special chicken biryani",
    "Chicken Lollipop Biryani": "Biryani topped with chicken lollipops",
    "Chicken Dum Handi Biryani": "Chicken dum biryani in clay handi",
    "Chicken Lollipop Handi Biryani": "Lollipop biryani in clay handi",
    "Chicken Roast Biryani": "Biryani with aromatic roasted chicken",
    "Chicken Dum Biryani Small": "Small portion chicken dum biryani",
    "Chicken Dum Biryani Jumbo Pack": "Chicken biryani jumbo — serves 6-8",
    "Chicken Roast Familypack": "Chicken roast biryani — serves 4",
    "Chicken Roast Biryani Mini": "Mini chicken roast biryani",
    "Chicken Mughlai Biryani": "Rich Mughlai-style chicken biryani",
    "Chicken Roast Handi Biryani": "Chicken roast biryani in handi",
    "Tangdi Biryani": "Biryani with tandoori chicken legs",
    "Ulvacharu Chicken Biryani": "Horsegram chicken biryani",
    "Potlam Biryani": "Biryani wrapped in cloth potlam style",
    "Tandoori Chicken Biryani": "Biryani with tandoori chicken",
    "Natu Kodi Biryani": "Country chicken biryani",
    "Tawa Roast Biryani": "Tawa-roasted chicken biryani",
    "Chicken Bewafa Biryani": "Signature bewafa-style chicken biryani",
    "Chicken Patiala Biryani": "Patiala-style rich chicken biryani",
    "Chicken Tikka Biryani": "Biryani with grilled chicken tikka",
    "Mutton Dum Biryani": "Authentic mutton dum biryani",
    "Mini Mutton Dum Biryani": "Mini portion mutton dum biryani",
    "Nalli Ghost Biryani": "Biryani with mutton nalli (bone marrow)",
    "Mutton Roast Biryani": "Biryani with roasted mutton",
    "Mutton Dum Biryanifamily Pack": "Mutton biryani — serves 4",
    "Mutton Dum Handi Biryani": "Mutton biryani in clay handi",
    "Mutton Kheema Biryani": "Biryani with spiced mutton keema",
    "Mutton Dum Biryani Jumbo Pack": "Mutton biryani jumbo — serves 6-8",
    "Prawns Biryani": "Aromatic biryani with juicy prawns",
    "Fish Biryani": "Biryani with tender fish pieces",
    "Ulavacharu Mutton Biryani": "Horsegram mutton biryani",
    "Ulavacharu Chicken Biryani": "Horsegram chicken biryani",
    "Mushroom Tikka Biryani": "Biryani with grilled mushroom tikka",
    "Natu Kodi Pulusu": "Traditional country chicken pulusu",
    "Mutton Keema Ragi Mudda": "Mutton keema with ragi mudda",
    "Ragi Sankati": "Traditional ragi ball",
    "Natukodi Vepudu": "Country chicken dry fry",
    "Pomegranate Juice": "Fresh pomegranate juice", "Pineapple Juice": "Fresh pineapple juice",
    "Orange Juice": "Fresh orange juice", "Dryfruits Juice": "Rich dry fruits juice blend",
    "Watermelon Juice.": "Refreshing watermelon juice", "Sweet Lime Juice": "Sweet lime juice",
    "Apple Juice.": "Fresh apple juice", "Banana Juice": "Creamy banana juice",
    "Mixed Fruit Juice": "Mixed fruit juice blend", "Seasonal Fruit Juice": "Seasonal fresh fruit juice",
    "Chocolate Milkshake": "Rich chocolate milkshake", "Mango Milkshake": "Fresh mango milkshake",
    "Strawberry Milkshake": "Strawberry milkshake", "Vanilla Milkshake": "Classic vanilla milkshake",
    "Banana Milkshake": "Banana milkshake", "Apple Milkshake": "Apple milkshake",
    "Butterscotch Milkshakes": "Butterscotch milkshake", "Litchi Milkshake": "Litchi milkshake",
    "Oreo Milkshake": "Oreo cookie milkshake", "Pista Milkshakes": "Pistachio milkshake",
    "Virgin Mojito": "Refreshing virgin mojito", "Blue Lagoon": "Blue lagoon mocktail",
    "Strawberry Lemonade": "Tangy strawberry lemonade", "Fruit Crush": "Fresh fruit crush",
    "Blue Hawaiian": "Tropical blue Hawaiian mocktail", "Cinderella": "Cinderella mocktail",
    "Downstreet": "Downstreet signature mocktail", "Fruit Smarty": "Fruity smarty mocktail",
    "Pink Lady": "Elegant pink lady mocktail", "Rainbow Mocktail": "Colorful rainbow mocktail",
    "Shirley Temple": "Classic Shirley Temple", "Spicy Twist": "Spicy twist mocktail",
    "Virgin Pina Colada": "Creamy virgin piña colada",
    "Lassi Sweet": "Sweet Punjabi lassi", "Butter Milk": "Refreshing spiced buttermilk",
    "Lassi Salt": "Salted lassi",
    "Fruit Salad": "Fresh mixed fruit salad", "Arabian Delight": "Arabian-inspired dessert delight",
    "Fruit Salad With Icecream": "Fruit salad topped with ice cream",
    "Apricot Delite": "Apricot-flavored dessert", "Bengali Sweets": "Assorted Bengali sweets",
    "Masala Papad": "Crispy papad with masala topping", "Roasted Papad": "Plain roasted papad",
    "Cup Curd": "Fresh cup of curd", "Pineapple Raita": "Pineapple raita",
    "Veg Raita": "Mixed vegetable raita",
    "Boiled Vegetable Salad": "Healthy boiled vegetable salad", "Cucumber Salad": "Fresh cucumber salad",
    "Green Salad": "Fresh green salad", "Smoked Vegetable Salad": "Smoky grilled vegetable salad",
}

# ── Unsplash image pools ─────────────────────────────────────────────────────────
_POOLS = {
    "chinese_veg": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1601050690597-df0568f70950","1546069901-ba9599a7e63c","1540189549336-e6e99c3679fe",
        "1567337710282-00832b415979","1585032226651-759b368d7246","1572715376701-98568319fd0b",
        "1559847844-5315695dadae","1606574977732-e8e5f1f46c23","1564834724105-918b73d1b8e0",
        "1512621776951-a57141f2eefd","1574484284002-952d92456975","1567620832903-9fc6debc209f",
        "1606728035253-49e8a23146de","1599487488170-d11ec9c172f0","1565299624946-b28f40a0ae38",
        "1573080496219-bb080dd4f877","1585937421612-70a008356fbe","1580217593608-61931cefc821",
        "1614398751058-bca239de00ca","1603360946369-dc9bb6258143","1565557623262-b51c2513a641",
        "1606755962773-d324e0a13086","1612874742237-6526221588e3","1625937286520-3ef7955f3813",
        "1576402187878-974f70c890a5","1455619452474-d2be8b1e70cd","1596797038530-2c107229654b",
        "1604908176997-125f25cc6f3d","1612929633738-8fe44f7ec841","1631452180519-c014fe946bc7",
        "1559339352-11d035aa65de","1553621042-f6e147245754","1565680018093-ebb6e3062e7b",
    ]],
    "chinese_nonveg": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1626082927389-6cd097cdc6ec","1562967914-608f82629710","1608039755401-742074f0548d",
        "1567620832903-9fc6debc209f","1610057099431-d73a1c9d2f2f","1598515214211-89d3c73ae83b",
        "1587593810167-a84920ea0781","1619221882220-947b3d3c8861","1606728035253-49e8a23146de",
        "1599487488170-d11ec9c172f0","1632778149955-e80f8ceca2e8","1624726175512-19b9baf9fbd1",
        "1585937421612-70a008356fbe","1580217593608-61931cefc821","1614398751058-bca239de00ca",
        "1603360946369-dc9bb6258143","1565299624946-b28f40a0ae38","1606755962773-d324e0a13086",
        "1612874742237-6526221588e3","1625937286520-3ef7955f3813","1573080496219-bb080dd4f877",
    ]],
    "seafood": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1615141982883-c7ad0e69fd62","1504674900247-0877df9cc836","1535140728325-a4d3707eee61",
        "1510130113581-4ae76c0f6e7f","1559039448-9b03d2e3c18e","1519708227418-c8fd9a32b7a2",
        "1498654896293-37aacf113fd9","1580476262798-bddd9f4b7369","1551504734-5ee1c4a1479b",
        "1569058242567-93de6f36f8e6","1606731219412-213c1e68ca63","1617196035154-1e7e6e28b0db",
        "1610540881815-6e81b56d8ea5","1615361200098-9e630ec29b4e",
    ]],
    "soup": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1547592166-23ac45744acd","1603105037880-880cd4f5b2e6","1476718406336-bb5a9690ee2a",
        "1588566565463-180a5b2090d2","1613844237701-8f3664fc2eff","1594756202469-9ff9799b2e4e",
        "1604152135912-04a022e23696","1607330289024-1535c6b4e1c1","1583608205776-bfd35f0d9f83",
        "1509358271058-acd22cc93898","1617093727343-374698b1b08d","1597227129956-93bad7e18d08",
        "1603894584373-5ac82b2ae328",
    ]],
    "curry": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1565557623262-b51c2513a641","1585937421612-70a008356fbe","1631515243349-e0cb75fb8d3a",
        "1574484284002-952d92456975","1455619452474-d2be8b1e70cd","1631452180519-c014fe946bc7",
        "1596797038530-2c107229654b","1604908176997-125f25cc6f3d","1567337710282-00832b415979",
        "1612929633738-8fe44f7ec841","1576402187878-974f70c890a5","1574653853027-5382a3d23a15",
        "1628294895950-9805252327bc","1551881192-5e377f1b2142","1573080496219-bb080dd4f877",
        "1603894584373-5ac82b2ae328","1505253758473-96b7015fcd40","1607116667573-1c7d73636ba0",
        "1599043513900-ed6fe01d3833","1633321702518-7fecdafb94d5","1563379091339-03b21ab4a4f4",
        "1580217593608-61931cefc821","1606755962773-d324e0a13086","1612874742237-6526221588e3",
        "1625937286520-3ef7955f3813","1627662168223-7df99068099a","1617093727343-374698b1b08d",
        "1606728035253-49e8a23146de","1574484284002-952d92456975","1567188040759-fb8a883dc6d8",
        "1545247181-516773cae754","1606491956689-2ea866880049","1633321702518-7fecdafb94d5",
        "1609501676725-7186f017a4b7","1618449840665-9ed506d73a34","1617692855027-33b14f061079",
        "1642821373181-16a5bc9f5801","1574653853027-5382a3d23a15",
    ]],
    "tandoori": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1599487488170-d11ec9c172f0","1601050690117-94f5f6fa8bd7","1565557623262-b51c2513a641",
        "1628294895950-9805252327bc","1567188040759-fb8a883dc6d8","1574653853027-5382a3d23a15",
        "1551881192-5e377f1b2142","1573080496219-bb080dd4f877","1610057099431-d73a1c9d2f2f",
        "1598515214211-89d3c73ae83b","1603894584373-5ac82b2ae328","1606491956689-2ea866880049",
        "1545247181-516773cae754","1574484284002-952d92456975","1609501676725-7186f017a4b7",
        "1631452180519-c014fe946bc7","1617692855027-33b14f061079","1642821373181-16a5bc9f5801",
        "1618449840665-9ed506d73a34","1633321702518-7fecdafb94d5","1574653853027-5382a3d23a15",
        "1551881192-5e377f1b2142","1573080496219-bb080dd4f877",
    ]],
    "bread": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1565557623262-b51c2513a641","1600326145359-3a44909d1a39","1574071318508-1cdbab80d002",
        "1586444248879-bc604bc77212","1567620832903-9fc6debc209f","1555939594-58d7cb561ad1",
        "1601050690597-df0568f70950","1555507036-ab1f4038024a","1573675542321-f51b18e6e759",
        "1515024014929-c2ba2c4da1d0","1519864600395-3404e40a0eda","1509722747041-616f39b57569",
        "1528736235302-52922df5c122","1605888969139-42cca4308aa2","1586444248879-bc604bc77212",
        "1574071318508-1cdbab80d002",
    ]],
    "rice": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1596560548464-f010549b84d7","1516714435131-44d6b64dc6a2","1536304993881-460587633ee1",
        "1512058564366-18510be2db19","1645696301019-35adcc552067","1603133872878-684f208fb84b",
        "1604908176997-125f25cc6f3d","1563379091339-03b21ab4a4f4","1589302168068-964664d93dc0",
        "1631515243349-e0cb75fb8d3a","1596797038530-2c107229654b","1585937421612-70a008356fbe",
        "1574484284002-952d92456975","1612874742237-6526221588e3","1625937286520-3ef7955f3813",
        "1599043513900-ed6fe01d3833","1628294895950-9805252327bc","1607116667573-1c7d73636ba0",
        "1574653853027-5382a3d23a15","1603894584373-5ac82b2ae328","1551881192-5e377f1b2142",
        "1505253758473-96b7015fcd40","1627662168223-7df99068099a","1617093727343-374698b1b08d",
        "1606728035253-49e8a23146de",
    ]],
    "biryani": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1563379091339-03b21ab4a4f4","1589302168068-964664d93dc0","1631515243349-e0cb75fb8d3a",
        "1642821373181-16a5bc9f5801","1633321702518-7fecdafb94d5","1596797038530-2c107229654b",
        "1604908176997-125f25cc6f3d","1574484284002-952d92456975","1585937421612-70a008356fbe",
        "1565557623262-b51c2513a641","1606728035253-49e8a23146de","1612874742237-6526221588e3",
        "1625937286520-3ef7955f3813","1599043513900-ed6fe01d3833","1628294895950-9805252327bc",
        "1607116667573-1c7d73636ba0","1574653853027-5382a3d23a15","1603894584373-5ac82b2ae328",
        "1551881192-5e377f1b2142","1505253758473-96b7015fcd40","1627662168223-7df99068099a",
        "1617093727343-374698b1b08d","1545247181-516773cae754","1606491956689-2ea866880049",
        "1609501676725-7186f017a4b7","1618449840665-9ed506d73a34","1617692855027-33b14f061079",
        "1631452180519-c014fe946bc7","1567188040759-fb8a883dc6d8","1573080496219-bb080dd4f877",
        "1610057099431-d73a1c9d2f2f","1598515214211-89d3c73ae83b","1601050690117-94f5f6fa8bd7",
        "1599487488170-d11ec9c172f0","1574071318508-1cdbab80d002","1586444248879-bc604bc77212",
        "1555939594-58d7cb561ad1","1601050690597-df0568f70950","1555507036-ab1f4038024a",
        "1573675542321-f51b18e6e759","1515024014929-c2ba2c4da1d0","1519864600395-3404e40a0eda",
        "1509722747041-616f39b57569","1528736235302-52922df5c122","1605888969139-42cca4308aa2",
        "1576402187878-974f70c890a5","1455619452474-d2be8b1e70cd","1512058564366-18510be2db19",
        "1645696301019-35adcc552067","1603133872878-684f208fb84b",
    ]],
    "drinks": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1534353473418-4cfa6c56fd38","1600271886742-f049cd451bba","1572490122747-3968b75cc699",
        "1541658016709-82535e94bc69","1553787499-6f9133860278","1568901839119-631418a3910d",
        "1619158401201-8fa932695178","1579954115545-a95591f28bfc","1497034825429-c343d7c6a68f",
        "1501443762994-82bd5dace89a","1560008581-09826d1de69e","1570197571499-166b36435e9f",
        "1629385701021-fcd568a743e8","1576506295286-5cda18df43e7","1633933358116-a27b902fad35",
        "1514849302-984523450cf4","1563805042-7684c019e1cb","1551024506-0bccd828d307",
        "1580915411954-282cb1b0d780","1557142046-c704a3adf364","1505394033641-40c6ad1178d7",
        "1516559828984-fb3b99548b21","1621303837174-89787a7d4729","1587563871167-1ee9c731aefb",
        "1612203985729-70726954388c","1615478503562-ec2d8aa0e24e","1567206563064-6f60f40a2b57",
        "1543255006-d6395b6f1171","1595348020949-87cdfbb44174","1600002423562-975eabb78d5a",
        "1632170684742-9c8b38c1aeab","1628607189631-96e9e8a3cedc","1622483767028-3f66f32aef97",
        "1659432873335-3b5a6d7f1a4f","1625869767142-1fb8faf7e8d9",
    ]],
    "sides": [f"https://images.unsplash.com/photo-{pid}?w=800&q=80" for pid in [
        "1601050690597-df0568f70950","1585032226651-759b368d7246","1599487488170-d11ec9c172f0",
        "1567337710282-00832b415979","1573080496219-bb080dd4f877","1546069901-ba9599a7e63c",
        "1540189549336-e6e99c3679fe","1559847844-5315695dadae","1564834724105-918b73d1b8e0",
    ]],
}

_used: set[str] = set()
def _pick(cat: str, item: str) -> str:
    cl = cat.lower(); il = item.lower()
    if "soup" in cl: pool = _POOLS["soup"]
    elif "seafood" in cl or "prawn" in il or "fish" in il: pool = _POOLS["seafood"]
    elif "tandoori" in cl or "kebab" in il or "kabab" in il or "tikka" in il: pool = _POOLS["tandoori"]
    elif "grill" in cl: pool = _POOLS["tandoori"]
    elif "bread" in cl or "roti" in il or "naan" in il or "paratha" in il or "kulcha" in il: pool = _POOLS["bread"]
    elif "biryani" in cl: pool = _POOLS["biryani"]
    elif "noodle" in cl or "rice" in cl or "chopsu" in cl or "khichdi" in il or "pulao" in il: pool = _POOLS["rice"]
    elif "juice" in il or "milkshake" in il or "mojito" in il or "mocktail" in il or "lassi" in il or "butter milk" in il or "lemonade" in il or "crush" in il or "hawaiian" in il or "cinderella" in il or "downstreet" in il or "smarty" in il or "lady" in il or "temple" in il or "twist" in il or "colada" in il or "lagoon" in il or "smoothie" in il: pool = _POOLS["drinks"]
    elif "sweet" in il or "salad" in il or "dessert" in cl or "fruit" in il or "delight" in il or "bengali" in il: pool = _POOLS["drinks"]
    elif "papad" in il or "raita" in il or "curd" in il or "accomplishment" in cl: pool = _POOLS["sides"]
    elif "main course" in cl or "curry" in il or "masala" in il or "kadai" in il or "dal" in il or "palak" in il or "sabji" in il or "handi" in il or "rogan" in il or "nihari" in il or "angara" in il or "kurma" in il: pool = _POOLS["curry"]
    elif "chinese" in cl and ("egg" in il or "chicken" in il or "mutton" in il): pool = _POOLS["chinese_nonveg"]
    elif "chinese" in cl: pool = _POOLS["chinese_veg"]
    elif "rayalaseema" in cl: pool = _POOLS["curry"]
    else: pool = _POOLS["chinese_veg"]
    for url in pool:
        if url not in _used:
            _used.add(url)
            return url
    return pool[hash(item) % len(pool)]


def enrich():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active
    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for ci in range(1, 6):
        c = ws.cell(row=1, column=ci)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for ci, h in [(6, "Description"), (8, "Veg/Non-Veg"), (9, "Image URL")]:
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(wrap_text=True, vertical="center")
    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=2).value
        cat = ws.cell(row=r, column=1).value
        if not item: continue
        item_s = str(item).strip()
        cat_s = str(cat).strip() if cat else ""
        ws.cell(row=r, column=6, value=DESC.get(item_s, f"Delicious {item_s}"))
        ws.cell(row=r, column=8, value="Veg" if item_s in VEG_ITEMS else "Non-Veg")
        c = ws.cell(row=r, column=9, value=_pick(cat_s, item_s))
        c.font = Font(color="0563C1", underline="single", size=9)
    for col in range(1, 10):
        mx = max((len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, min(ws.max_row+1, 400))), default=8)
        ws.column_dimensions[get_column_letter(col)].width = min(mx + 3, 55)
    wb.save(XLSX_PATH)
    print(f"Enriched {XLSX_PATH.name}")


def build_payloads() -> list[dict]:
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    payloads = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        cat, item, price, hike, subcat = row[0], row[1], row[2], row[3], row[4]
        desc, veg_raw, img = row[5], row[7], row[8]
        if not item or price is None: continue
        hike_val = 0
        if hike is not None:
            try:
                h = float(hike)
                hike_val = round(h * 100, 1) if h < 1 else round(h, 1)
            except (ValueError, TypeError):
                hike_val = 0
        payloads.append({
            "name": str(item).strip(),
            "restaurantPrice": float(price),
            "hikePercentage": hike_val,
            "category": str(cat).strip() if cat else None,
            "subCategory": str(subcat).strip() if subcat else None,
            "isVeg": str(veg_raw).strip().lower() == "veg" if veg_raw else None,
            "isAvailable": True,
            "description": str(desc).strip() if desc else None,
            "image": [str(img).strip()] if img else [],
        })
    wb.close()
    return payloads


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--apply", action="store_true")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--bearer-token", default=os.environ.get("HONESTEATS_BEARER_TOKEN"))
    parser.add_argument("--delay", type=float, default=0.3)
    args = parser.parse_args()

    enrich()
    payloads = build_payloads()
    print(f"\n{len(payloads)} items | Restaurant: {RESTAURANT_ID}\n")
    menu_url = f"{args.api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"
    for i, p in enumerate(payloads, 1):
        vl = "Veg" if p["isVeg"] else "Non-Veg"
        print(f"[{i:3d}] {p['name']}")
        print(f"      ₹{p['restaurantPrice']} + {p['hikePercentage']}% | {p['category']} > {p['subCategory']} | {vl}")
        if not args.apply:
            print(f"      {json.dumps(p, indent=None, ensure_ascii=False)}")
        print()
    if not args.apply:
        print("=" * 60)
        print("DRY RUN — no HTTP calls.")
        print(f"To insert: python3 {sys.argv[0]} --apply")
        print("=" * 60)
        return 0
    s = requests.Session()
    s.headers["Content-Type"] = "application/json"
    if args.bearer_token:
        s.headers["Authorization"] = f"Bearer {args.bearer_token}"
    else:
        s.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS
    ok = err = 0
    for i, p in enumerate(payloads, 1):
        if args.delay > 0: time.sleep(args.delay)
        try:
            r = s.post(menu_url, json=p, timeout=60)
        except requests.RequestException as ex:
            err += 1; print(f"  ✗ [{i}] {p['name']}: {ex}", file=sys.stderr); continue
        if r.status_code == 201:
            ok += 1; print(f"  ✓ [{i}] {p['name']} created")
        else:
            err += 1; print(f"  ✗ [{i}] {p['name']}: HTTP {r.status_code} {r.text[:200]}", file=sys.stderr)
            if r.status_code == 401: print("\nUnauthorized.", file=sys.stderr); return 1
    print(f"\nDone: {ok} created, {err} failed.")
    return 0 if err == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
