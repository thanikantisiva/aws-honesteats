#!/usr/bin/env python3
"""
Import menu items from C3xYumDude-style Excel via HTTP API (same path as partner app).

POST {base}/api/v1/restaurants/{restaurantId}/menu

Auth (same as scripts/insert_restaurants_geohash.py):
  - Default: header x-retool-header = RETOOL_BYPASS_VALUE (env HONESTEATS_RETOOL_BYPASS)
  - Or: Authorization: Bearer <token> (--bearer-token / HONESTEATS_BEARER_TOKEN), optionally with --no-retool-bypass

Excel sheet (e.g. C3):
  A: S.No, B: Item name, C: Description, E: Price, F: Category, G: Subcategory, H: Veg or Non-Veg
  I (optional): Image URL — if empty, a matching stock image is chosen from item name + category.

Usage:
  python3 scripts/import_menu_from_xlsx.py --xlsx ~/Downloads/C3xYumDude.xlsx \\
    --restaurant-id RES-1775832301170-9755 --api-url https://api.yumdude.com

  python3 scripts/import_menu_from_xlsx.py ... --apply

Env:
  HONESTEATS_API_URL   — API host (no trailing /api/v1 required)
  HONESTEATS_RETOOL_BYPASS — x-retool-header value (omit with --no-retool-bypass)
  HONESTEATS_BEARER_TOKEN — JWT if not using retool bypass
"""
from __future__ import annotations

import argparse
import os
import sys
import time
from pathlib import Path

import openpyxl
import requests

_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))
from menu_item_images import resolve_menu_item_image

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get(
    "HONESTEATS_RETOOL_BYPASS",
    "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc",
)
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")


def _normalize_api_base(url: str) -> str:
    u = url.rstrip("/")
    if u.endswith("/api/v1"):
        u = u[: -len("/api/v1")]
    return u


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Import menu items from Excel via YumDude API")
    p.add_argument(
        "--xlsx",
        type=Path,
        default=Path.home() / "Downloads" / "C3xYumDude.xlsx",
        help="Path to workbook",
    )
    p.add_argument("--sheet", default="C3", help="Worksheet name")
    p.add_argument(
        "--restaurant-id",
        default="RES-1775832301170-9755",
        help="Restaurant id in URL path",
    )
    p.add_argument(
        "--api-url",
        default=DEFAULT_API_URL,
        help="API base URL (default: HONESTEATS_API_URL or https://api.yumdude.com)",
    )
    p.add_argument(
        "--no-retool-bypass",
        action="store_true",
        help="Do not send x-retool-header (use --bearer-token)",
    )
    p.add_argument(
        "--retool-bypass",
        default="",
        help="Override x-retool-header value (default: env or built-in dev default)",
    )
    p.add_argument(
        "--bearer-token",
        default=os.environ.get("HONESTEATS_BEARER_TOKEN", ""),
        help="Authorization Bearer token (restaurant/admin JWT)",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        default=True,
        help="Parse and print only (default)",
    )
    p.add_argument(
        "--apply",
        action="store_true",
        help="POST each row to the API",
    )
    p.add_argument("--hike-percentage", type=float, default=0.0, help="hikePercentage on all items")
    p.add_argument(
        "--delay-sec",
        type=float,
        default=float(os.environ.get("HONESTEATS_REQUEST_DELAY_SEC", "0.05")),
        help="Sleep between POSTs",
    )
    p.add_argument(
        "--no-images",
        action="store_true",
        help="Do not send image field (skip stock / column I)",
    )
    return p.parse_args()


def is_menu_data_row(a_val, b_val, f_val) -> bool:
    if a_val is None or b_val is None:
        return False
    if str(a_val).strip().upper() == "S.NO":
        return False
    try:
        int(a_val)
    except (TypeError, ValueError):
        return False
    name = str(b_val).strip()
    if not name or name.upper() == "DESCRIPTION":
        return False
    if not f_val or not str(f_val).strip():
        return False
    if str(f_val).strip().lower() == "category":
        return False
    return True


def parse_price(cell_val) -> float | None:
    if cell_val is None:
        return None
    if isinstance(cell_val, (int, float)):
        return round(float(cell_val), 2)
    s = str(cell_val).strip().replace("₹", "").replace(",", "")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def parse_is_veg(h_val) -> bool | None:
    if h_val is None or not str(h_val).strip():
        return None
    t = str(h_val).strip().lower()
    if "non" in t:
        return False
    if "veg" in t:
        return True
    return None


def build_session(args: argparse.Namespace) -> requests.Session:
    s = requests.Session()
    s.headers["Content-Type"] = "application/json"
    if args.no_retool_bypass:
        bypass = None
    elif (args.retool_bypass or "").strip():
        bypass = args.retool_bypass.strip()
    else:
        bypass = DEFAULT_RETOOL_BYPASS.strip() or None
    if bypass:
        s.headers[RETOOL_BYPASS_HEADER] = bypass
    token = (args.bearer_token or "").strip()
    if token:
        s.headers["Authorization"] = f"Bearer {token}"
    return s


def main() -> int:
    args = parse_args()
    if args.apply:
        args.dry_run = False

    if not args.xlsx.is_file():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        return 1

    base = _normalize_api_base(args.api_url)
    menu_url = f"{base}/api/v1/restaurants/{args.restaurant_id}/menu"

    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)
    if args.sheet not in wb.sheetnames:
        print(f"Sheet {args.sheet!r} not in {wb.sheetnames}", file=sys.stderr)
        return 1
    ws = wb[args.sheet]

    rows_out: list[tuple[str, str, str, float, bool | None, str | None, str | None]] = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row or 2000, values_only=True):
        cells = (list(row) + [None] * 9)[:9]
        a, b, c, _d, e, f, g, h, col_i = cells
        if not is_menu_data_row(a, b, f):
            continue
        price = parse_price(e)
        if price is None:
            print(f"SKIP (no price): {b!r}", file=sys.stderr)
            continue
        cat = str(f).strip()
        sub = str(g).strip() if g else None
        desc = str(c).strip() if c else None
        is_veg = parse_is_veg(h)
        excel_img = str(col_i).strip() if col_i is not None and str(col_i).strip() else None
        rows_out.append((cat, sub or "", str(b).strip(), price, is_veg, desc, excel_img))

    wb.close()

    print(f"API: POST {menu_url}")
    print(f"Restaurant: {args.restaurant_id} | Rows: {len(rows_out)}")
    print(f"Mode: {'DRY-RUN' if args.dry_run else 'APPLY'}")
    for i, (cat, sub, name, price, is_veg, desc, ximg) in enumerate(rows_out[:5], 1):
        img_preview = (
            ""
            if args.no_images
            else f" | img={resolve_menu_item_image(name, cat, sub, ximg)[:56]}…"
        )
        print(f"  {i}. {name} | {cat} / {sub} | ₹{price} | isVeg={is_veg}{img_preview}")
    if len(rows_out) > 5:
        print(f"  ... +{len(rows_out) - 5} more")

    if args.dry_run:
        print("No HTTP calls (pass --apply to insert via API).")
        return 0

    session = build_session(args)
    has_bypass = RETOOL_BYPASS_HEADER in session.headers
    has_bearer = "Authorization" in session.headers
    if not has_bypass and not has_bearer:
        print(
            "Error: no auth. Set x-retool-header (default) or pass --bearer-token / "
            "HONESTEATS_BEARER_TOKEN, or use default HONESTEATS_RETOOL_BYPASS.",
            file=sys.stderr,
        )
        return 1

    ok = 0
    err = 0
    for cat, sub, name, price, is_veg, desc, ximg in rows_out:
        payload: dict = {
            "name": name,
            "restaurantPrice": price,
            "hikePercentage": args.hike_percentage,
            "category": cat,
            "subCategory": sub or None,
            "isAvailable": True,
        }
        if desc:
            payload["description"] = desc
        if is_veg is not None:
            payload["isVeg"] = is_veg
        if not args.no_images:
            payload["image"] = resolve_menu_item_image(name, cat, sub, ximg)

        if args.delay_sec > 0:
            time.sleep(args.delay_sec)

        try:
            r = session.post(menu_url, json=payload, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"ERROR {name!r}: {ex}", file=sys.stderr)
            continue

        if r.status_code == 201:
            ok += 1
        else:
            err += 1
            print(f"ERROR {name!r}: HTTP {r.status_code} {r.text[:300]}", file=sys.stderr)
            if r.status_code == 401:
                print(
                    "\nUnauthorized: fix x-retool-header (prod uses secret from AWS) or use a valid "
                    "restaurant JWT with --bearer-token.",
                    file=sys.stderr,
                )
                return 1

    print(f"Done: {ok} created via API, {err} failed.")
    return 0 if err == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
