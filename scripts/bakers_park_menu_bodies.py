#!/usr/bin/env python3
"""
Enrich Bakers Park.xlsx and build menu POST bodies (no HTTP).

Expected columns (Sheet1):
  A: Category
  B: Subcategory
  C: Item Name
  D: Price (₹)
  E: hike (optional; defaults to 35 in JSON if missing)

Enrichment (column H only when blank):
  H: isVeg

JSON per row:
  {
    "name": C, "restaurantPrice": D, "hikePercentage": <E or 35>,
    "category": A, "subCategory": B, "isVeg": H, "isAvailable": true
  }

Usage:
  python3 scripts/bakers_park_menu_bodies.py
  python3 scripts/bakers_park_menu_bodies.py --xlsx "/path/Bakers Park.xlsx"
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_XLSX = Path.home() / "Downloads" / "Bakers Park.xlsx"
DEFAULT_RESTAURANT_ID = "RES-1777200686550-7882"
DEFAULT_OUT_JSON = Path.home() / "Downloads" / f"bakers_park_menu_bodies_{DEFAULT_RESTAURANT_ID}.json"
DEFAULT_HIKE = 35.0


def _clean(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _title_words(s: str) -> str:
    t = _clean(s)
    if not t:
        return ""
    return " ".join(w[:1].upper() + w[1:].lower() for w in t.split())


def _parse_price(v: Any) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    s = _clean(v).replace("₹", "").replace(",", "")
    return float(s)


def _parse_hike(v: Any, default: float) -> float:
    if v is None or _clean(v) == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(_clean(v).replace("%", ""))
    except ValueError:
        return default


def _parse_is_veg_cell(v: Any) -> bool | None:
    if v is None or _clean(v) == "":
        return None
    if isinstance(v, bool):
        return v
    s = _clean(v).lower()
    if s in {"true", "t", "yes", "1", "veg", "v", "vegetarian"}:
        return True
    if s in {"false", "f", "no", "0", "non-veg", "nonveg", "nv"}:
        return False
    return None


def _infer_is_veg(category: str, subcategory: str, name: str) -> bool:
    b = subcategory.lower()
    c = category.lower()
    n = name.lower()
    if "chicken" in b:
        return False
    if "chicken" in n and "veg" not in b:
        return False
    if "egg" in n or "omelette" in n:
        return False
    if "veg" in b:
        return True
    return True


def enrich_workbook(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    if _clean(ws.cell(1, 8).value) == "":
        ws.cell(1, 8, "isVeg")

    for row in range(2, (ws.max_row or 1) + 1):
        name = _clean(ws.cell(row, 3).value)
        if not name:
            continue
        if _parse_is_veg_cell(ws.cell(row, 8).value) is None:
            cat = _clean(ws.cell(row, 1).value)
            sub = _clean(ws.cell(row, 2).value)
            ws.cell(row, 8, _infer_is_veg(cat, sub, name))

    wb.save(path)
    wb.close()


def build_bodies(path: Path, *, hike_default: float) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    bodies: list[dict[str, Any]] = []

    for row in range(2, (ws.max_row or 1) + 1):
        name = _clean(ws.cell(row, 3).value)
        if not name:
            continue
        cat = _clean(ws.cell(row, 1).value)
        sub = _clean(ws.cell(row, 2).value)
        price = _parse_price(ws.cell(row, 4).value)
        hike = _parse_hike(ws.cell(row, 5).value, hike_default)
        is_veg = _parse_is_veg_cell(ws.cell(row, 8).value)
        if is_veg is None:
            is_veg = _infer_is_veg(cat, sub, name)

        bodies.append(
            {
                "name": _title_words(name),
                "restaurantPrice": price,
                "hikePercentage": hike,
                "category": _title_words(cat),
                "subCategory": _title_words(sub),
                "isVeg": is_veg,
                "isAvailable": True,
            }
        )

    wb.close()
    return bodies


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Bakers Park menu → JSON bodies (no insert).")
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--out-json", type=Path, default=DEFAULT_OUT_JSON)
    p.add_argument("--restaurant-id", default=DEFAULT_RESTAURANT_ID)
    p.add_argument("--hike-default", type=float, default=DEFAULT_HIKE)
    p.add_argument("--array-only", action="store_true")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    if not args.xlsx.is_file():
        print(f"Not found: {args.xlsx}", file=sys.stderr)
        return 1

    enrich_workbook(args.xlsx)
    bodies = build_bodies(args.xlsx, hike_default=args.hike_default)

    out_obj: Any = bodies if args.array_only else {"restaurantId": args.restaurant_id, "items": bodies}
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(json.dumps(out_obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    veg = sum(1 for b in bodies if b["isVeg"])
    print(f"Enriched: {args.xlsx}")
    print(f"Restaurant: {args.restaurant_id}")
    print(f"Bodies: {len(bodies)} (Veg: {veg}, Non-Veg: {len(bodies) - veg})")
    print(f"Wrote: {args.out_json}")
    print("No HTTP calls.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
