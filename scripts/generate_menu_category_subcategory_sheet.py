#!/usr/bin/env python3
"""Write Category / Subcategory / Item / Price menu to ~/Downloads from embedded CSV."""
from __future__ import annotations

import csv
import io
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path.home() / "Downloads" / "Menu_Category_Subcategory_Item_Price.xlsx"

CSV_BODY = """Category,Subcategory,Item,Price
Starters,,Masala Pappad (Full),60
Starters,,Gobi Manchuria (Single),50
Starters,,Gobi Manchuria (Full),100
Starters,,Onion Pakoda (Full),100
Starters,Gobi,65 / Ginger / Chilli / Finger Millet (Full),120
Starters,Vegetable,Manchuria / Chilli / Bullets / Ginger / Pakoda (Full),120
Starters,Aloo,Manchuria / Chilli / 65 / Ginger / French Fries (Full),120
Starters,Babycorn,Manchuria / 65 / Chilli / Ginger / Golden Fry (Full),180
Starters,Mushroom,Manchuria / 65 / Chilli / Ginger / Chettinadu (Full),180
Starters,Paneer,Manchuria / 65 / Chilli / Ginger / 555 / Majestic (Full),180
Starters,Paneer,Paneer Tikka (Full),220
Noodles,,Maggi,100
Noodles,,Veg Noodles,90
Noodles,,Lemon Noodles,90
Noodles,,Zeera Noodles,90
Noodles,,Gobi Noodles,100
Noodles,,Chilli Garlic Noodles,100
Noodles,,Schezwan Noodles,100
Noodles,,Macroni,120
Noodles,,Pasta,120
Noodles,,Bagara Noodles,100
Noodles,,Paneer Noodles,150
Noodles,,Kaju Noodles,150
Noodles,,Mushroom Noodles,150
Roti (Without Curry),,Pulka 1 No.,15
Roti (Without Curry),,Butter Pulka 1 No.,20
Roti (Without Curry),,Chapathi 1 No.,25
Roti (Without Curry),,Butter Chapathi 1 No.,30
Meals (Afternoon Only),,Line Meals,90
Meals (Afternoon Only),,Parcel Meals,110
North Indian Curries,,Tomato Curry (Single),80
North Indian Curries,,Tomato Curry (Full),140
North Indian Curries,,Plain Palak (Single),80
North Indian Curries,,Plain Palak (Full),140
North Indian Curries,,Mix Veg Curry (Single),80
North Indian Curries,,Mix Veg Curry (Full),140
North Indian Curries,,Gobi Masala (Single),80
North Indian Curries,,Gobi Masala (Full),140
North Indian Curries,,Green Peas Masala (Single),80
North Indian Curries,,Green Peas Masala (Full),140
North Indian Curries,,Chenna Masala (Single),80
North Indian Curries,,Chenna Masala (Full),140
North Indian Curries,,Aloo Gobi (Single),80
North Indian Curries,,Aloo Gobi (Full),140
North Indian Curries,,Aloo Palak (Single),80
North Indian Curries,,Aloo Palak (Full),140
North Indian Curries,,Aloo Mutter (Single),80
North Indian Curries,,Aloo Mutter (Full),140
North Indian Curries,,Gobi Palak (Single),80
North Indian Curries,,Gobi Palak (Full),140
North Indian Curries,,Capsicum Masala (Single),80
North Indian Curries,,Capsicum Masala (Full),140
North Indian Curries,,Malai Kofta (Full),150
North Indian Curries,,Paneer Mutter (Single),100
North Indian Curries,,Paneer Mutter (Full),180
North Indian Curries,,Baby Corn Masala (Single),100
North Indian Curries,,Baby Corn Masala (Full),180
North Indian Curries,,Palak Paneer (Single),100
North Indian Curries,,Palak Paneer (Full),180
North Indian Curries,,Methi Chaman (Single),100
North Indian Curries,,Methi Chaman (Full),180
North Indian Curries,,Butter Paneer Masala (Single),100
North Indian Curries,,Butter Paneer Masala (Full),180
North Indian Curries,,Mushroom Masala (Single),100
North Indian Curries,,Mushroom Masala (Full),180
North Indian Curries,,Kaju Paneer Masala (Single),100
North Indian Curries,,Kaju Paneer Masala (Full),200
North Indian Curries,,Kaju Mushroom Curry (Single),100
North Indian Curries,,Kaju Mushroom Curry (Full),200
North Indian Curries,,Kaju Masala (Single),100
North Indian Curries,,Kaju Masala (Full),200
North Indian Curries,,Mushroom Paneer (Single),100
North Indian Curries,,Mushroom Paneer (Full),200
North Indian Curries,,Paneer Burji (Single),100
North Indian Curries,,Paneer Burji (Full),220
Rice,,Curd Rice (Single),30
Rice,,Curd Rice (Full),50
Rice,,Lemon Rice,90
Rice,,Tomato Rice,90
Rice,,Onion Chilli Rice,90
Rice,,Veg Fried Rice,90
Rice,,Jeera Rice,90
Rice,,Azwain Rice,90
Rice,,Pudina Rice,100
Rice,,Kothimeera Fried Rice,100
Rice,,Gobi Rice,100
Rice,,Palak Rice,100
Rice,,Capsicum Fried Rice,100
Rice,,Aloo Fried Rice,100
Rice,,Garlic Fried Rice,100
Rice,,Schezwan Fried Rice,100
Rice,,Paneer Fried Rice,150
Rice,,Ghee Rice,150
Rice,,Mushroom Rice,150
Rice,,Kaju Rice,150
Rice,,Hong Kong Fried Rice,160
Rice,,Kaju Paneer Rice,160"""


def main() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Menu"

    reader = csv.reader(io.StringIO(CSV_BODY.strip()))
    for i, row in enumerate(reader):
        if i == 0:
            ws.append(row)
            continue
        cat, sub, item, price_s = row[0], row[1], row[2], row[3]
        price = int(price_s) if str(price_s).strip().isdigit() else price_s
        ws.append([cat, sub or None, item, price])

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
        row[3].alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate([28, 16, 52, 10], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    n_data = ws.max_row - 1
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.close()
    print(OUT)
    print(f"rows (excl. header): {n_data}")


if __name__ == "__main__":
    main()
