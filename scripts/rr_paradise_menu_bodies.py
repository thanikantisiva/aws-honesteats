#!/usr/bin/env python3
"""
Enrich RR Paradise Hotel menu workbook and build menu POST bodies (no HTTP).

Reads:  ~/Downloads/RR_Paradise_Hotel_Menu.xlsx
        Cols: A=Category, B=Subcategory, C=Item Name, D=Price

Enrichment (only fills blanks):
  H: isVeg   — determined from subcategory / item name

JSON body per row:
  {
    "name": C, "restaurantPrice": D, "hikePercentage": E,
    "category": A, "subCategory": B, "isVeg": H, "isAvailable": true
  }

Usage:
  python3 scripts/rr_paradise_menu_bodies.py
  python3 scripts/rr_paradise_menu_bodies.py --xlsx ... --out-json ...
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_XLSX = Path.home() / "Downloads" / "RR_Paradise_Hotel_Menu.xlsx"
DEFAULT_RESTAURANT_ID = "RES-1776847302757-5106"
DEFAULT_OUT_JSON = Path.home() / "Downloads" / f"rr_paradise_menu_bodies_{DEFAULT_RESTAURANT_ID}.json"
DEFAULT_HIKE = 35.0

NON_VEG_SUBCATEGORIES = frozenset({
    "chicken", "mutton", "seafood", "fish", "prawns", "egg",
})

NON_VEG_KEYWORDS = [
    "chicken", "mutton", "fish", "prawns", "prawn", "egg",
    "kamju", "tandoori chicken", "grill chicken",
    "lollipop", "drumstick", "tangidi", "bhatti",
    "liver",
]

ALWAYS_VEG_NAMES = frozenset({
    "biryani rice",
})


def _clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _title_case(value: Any) -> str:
    text = _clean(value)
    return " ".join(part[:1].upper() + part[1:] for part in text.split(" ")) if text else ""


def _parse_price(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    s = _clean(value).replace("₹", "").replace(",", "")
    return float(s)


def _parse_hike(value: Any, default: float) -> float:
    if value is None or _clean(value) == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    return float(_clean(value).replace("%", ""))


def _parse_is_veg(value: Any) -> bool | None:
    if value is None or _clean(value) == "":
        return None
    if isinstance(value, bool):
        return value
    s = _clean(value).lower()
    if s in {"true", "veg", "v", "yes", "1", "vegetarian"}:
        return True
    if s in {"false", "non-veg", "nonveg", "nv", "no", "0"}:
        return False
    return None


def _infer_is_veg(category: str, subcategory: str, name: str) -> bool:
    name_lower = name.lower()
    if name_lower in ALWAYS_VEG_NAMES:
        return True
    sub_lower = subcategory.lower()
    if sub_lower in NON_VEG_SUBCATEGORIES:
        return False
    for kw in NON_VEG_KEYWORDS:
        if kw in name_lower:
            return False
    cat_lower = category.lower()
    if cat_lower in {"seafood", "tandoor", "arabian"}:
        return False
    return True


def enrich_workbook(path: Path, *, hike_default: float) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    if _clean(ws.cell(1, 5).value) == "":
        ws.cell(1, 5, "hikePercentage")
    if _clean(ws.cell(1, 8).value) == "":
        ws.cell(1, 8, "isVeg")

    for row in range(2, (ws.max_row or 1) + 1):
        category = _clean(ws.cell(row, 1).value)
        subcategory = _clean(ws.cell(row, 2).value)
        name = _clean(ws.cell(row, 3).value)
        if not name:
            continue

        if _clean(ws.cell(row, 5).value) == "":
            ws.cell(row, 5, hike_default)

        if _parse_is_veg(ws.cell(row, 8).value) is None:
            ws.cell(row, 8, _infer_is_veg(category, subcategory, name))

    wb.save(path)
    wb.close()


def build_bodies(path: Path, *, hike_default: float) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    bodies: list[dict[str, Any]] = []

    for row in range(2, (ws.max_row or 1) + 1):
        category = _clean(ws.cell(row, 1).value)
        subcategory = _clean(ws.cell(row, 2).value)
        name = _clean(ws.cell(row, 3).value)
        if not name:
            continue

        is_veg = _parse_is_veg(ws.cell(row, 8).value)
        if is_veg is None:
            is_veg = _infer_is_veg(category, subcategory, name)

        body: dict[str, Any] = {
            "name": _title_case(name),
            "restaurantPrice": _parse_price(ws.cell(row, 4).value),
            "hikePercentage": _parse_hike(ws.cell(row, 5).value, hike_default),
            "category": _title_case(category),
            "subCategory": _title_case(subcategory),
            "isVeg": is_veg,
            "isAvailable": True,
        }
        bodies.append(body)

    wb.close()
    return bodies


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build RR Paradise menu JSON bodies without inserting.")
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--out-json", type=Path, default=DEFAULT_OUT_JSON)
    p.add_argument("--restaurant-id", default=DEFAULT_RESTAURANT_ID)
    p.add_argument("--hike-default", type=float, default=DEFAULT_HIKE)
    p.add_argument("--array-only", action="store_true",
                    help='Write bare array instead of {"restaurantId","items"}')
    return p.parse_args()


def main() -> int:
    args = parse_args()
    if not args.xlsx.is_file():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        return 1

    enrich_workbook(args.xlsx, hike_default=args.hike_default)
    bodies = build_bodies(args.xlsx, hike_default=args.hike_default)

    out_obj: Any
    if args.array_only:
        out_obj = bodies
    else:
        out_obj = {"restaurantId": args.restaurant_id, "items": bodies}

    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(json.dumps(out_obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    veg = sum(1 for b in bodies if b["isVeg"])
    non_veg = len(bodies) - veg
    print(f"Enriched workbook: {args.xlsx}")
    print(f"Restaurant: {args.restaurant_id}")
    print(f"Bodies: {len(bodies)} (Veg: {veg}, Non-Veg: {non_veg})")
    print(f"Wrote JSON: {args.out_json}")
    print("No HTTP calls were made.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
