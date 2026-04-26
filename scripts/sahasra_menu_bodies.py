#!/usr/bin/env python3
"""
Enrich Sahasra Food Court menu workbook and build menu POST bodies.

Does NOT insert anything. It only updates the workbook columns and writes JSON.

Expected workbook layout:
  A: category
  B: item name
  C: restaurantPrice
  D: hikePercentage

Enrichment:
  E: subCategory  (filled only when blank; default: same as category)
  H: isVeg        (filled only when blank; default: True)

JSON body per row:
  {
    "name": B,
    "restaurantPrice": C,
    "hikePercentage": D,
    "category": A,
    "subCategory": E,
    "isVeg": H,
    "isAvailable": true,
    "addOnOptions": [...]
  }

Usage:
  python3 scripts/sahasra_menu_bodies.py

  python3 scripts/sahasra_menu_bodies.py \\
    --xlsx /Users/user/Downloads/sahasra_food_court_menu.xlsx \\
    --out-json /Users/user/Downloads/sahasra_menu_bodies_RES-1777087265179-4590.json
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_XLSX = Path.home() / "Downloads" / "sahasra_food_court_menu.xlsx"
DEFAULT_RESTAURANT_ID = "RES-1777087265179-4590"
DEFAULT_OUT_JSON = Path.home() / "Downloads" / f"sahasra_menu_bodies_{DEFAULT_RESTAURANT_ID}.json"

def _clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _title_case(value: Any) -> str:
    text = _clean(value)
    return " ".join(part[:1].upper() + part[1:].lower() for part in text.split(" "))


def _parse_price(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    s = _clean(value).replace("₹", "").replace(",", "")
    return float(s)


def _parse_hike(value: Any) -> float:
    if value is None or _clean(value) == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return float(_clean(value).replace("%", ""))


def _parse_is_veg(value: Any) -> bool | None:
    if value is None or _clean(value) == "":
        return None
    if isinstance(value, bool):
        return value
    s = _clean(value).lower()
    if "non" in s and "veg" in s:
        return False
    if s in {"veg", "v", "vegetarian", "true", "yes", "1"}:
        return True
    if s in {"nonveg", "non-veg", "nv", "false", "no", "0"}:
        return False
    return None


def enrich_workbook(path: Path, *, default_is_veg: bool, subcategory: str | None) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    ws.cell(1, 1, ws.cell(1, 1).value or "Category")
    ws.cell(1, 2, ws.cell(1, 2).value or "Item Name")
    ws.cell(1, 3, ws.cell(1, 3).value or "restaurantPrice")
    ws.cell(1, 4, ws.cell(1, 4).value or "hikePercentage")
    ws.cell(1, 5, "subCategory")
    ws.cell(1, 8, "isVeg")

    for row in range(2, (ws.max_row or 1) + 1):
        category = _clean(ws.cell(row, 1).value)
        name = _clean(ws.cell(row, 2).value)
        if not category and not name:
            continue
        if category:
            category = _title_case(category)
            ws.cell(row, 1, category)
        if name:
            ws.cell(row, 2, _title_case(name))
        if _clean(ws.cell(row, 5).value) == "":
            ws.cell(row, 5, subcategory or category)
        ws.cell(row, 5, _title_case(ws.cell(row, 5).value))
        if _parse_is_veg(ws.cell(row, 8).value) is None:
            ws.cell(row, 8, default_is_veg)

    wb.save(path)
    wb.close()


def build_bodies(path: Path, *, default_is_veg: bool) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    bodies: list[dict[str, Any]] = []

    for row in range(2, (ws.max_row or 1) + 1):
        category = _title_case(ws.cell(row, 1).value)
        name = _title_case(ws.cell(row, 2).value)
        if not category and not name:
            continue
        if not name:
            print(f"Skipping row {row}: missing item name", file=sys.stderr)
            continue

        is_veg = _parse_is_veg(ws.cell(row, 8).value)
        body: dict[str, Any] = {
            "name": name,
            "restaurantPrice": _parse_price(ws.cell(row, 3).value),
            "hikePercentage": _parse_hike(ws.cell(row, 4).value),
            "category": category,
            "subCategory": _title_case(ws.cell(row, 5).value) or category,
            "isVeg": default_is_veg if is_veg is None else is_veg,
            "isAvailable": True,
        }
        bodies.append(body)

    wb.close()
    return bodies


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Sahasra menu JSON bodies without inserting.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out-json", type=Path, default=DEFAULT_OUT_JSON)
    parser.add_argument("--restaurant-id", default=DEFAULT_RESTAURANT_ID)
    parser.add_argument(
        "--subcategory",
        default=None,
        help="Value for blank column E. Default: copy the row's category.",
    )
    parser.add_argument(
        "--default-non-veg",
        action="store_true",
        help="Fill blank isVeg values as False instead of True.",
    )
    parser.add_argument(
        "--array-only",
        action="store_true",
        help='Write a bare array instead of {"restaurantId","items"}.',
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.xlsx.is_file():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        return 1

    default_is_veg = not args.default_non_veg
    enrich_workbook(args.xlsx, default_is_veg=default_is_veg, subcategory=args.subcategory)
    bodies = build_bodies(args.xlsx, default_is_veg=default_is_veg)

    out_obj: Any
    if args.array_only:
        out_obj = bodies
    else:
        out_obj = {"restaurantId": args.restaurant_id, "items": bodies}

    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(json.dumps(out_obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Enriched workbook: {args.xlsx}")
    print(f"Restaurant: {args.restaurant_id}")
    print(f"Bodies: {len(bodies)}")
    print(f"Wrote JSON: {args.out_json}")
    print("No HTTP calls were made.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
