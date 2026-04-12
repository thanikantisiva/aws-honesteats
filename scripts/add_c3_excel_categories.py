#!/usr/bin/env python3
"""Add Category, Subcategory, Veg or Non-Veg, and Image URL columns to C3xYumDude.xlsx sheet C3."""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

_SCRIPTS_DIR = Path(__file__).resolve().parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))
from menu_item_images import resolve_menu_item_image

# Item-name hints when section subcategory is not Veg/Non-Veg (e.g. Snacks, Beverages, Chaat)
_NONVEG_NAME = re.compile(
    r"\b(egg|chicken|mutton|fish|prawn|shrimp|lamb|beef|pork|meat|keema|salami|pepperoni|sausage)\b",
    re.I,
)

# Path to user's Excel (override with argv)
DEFAULT_XLSX = Path.home() / "Downloads" / "C3xYumDude.xlsx"

# Normalized section title (after emoji strip + whitespace) -> (category, subCategory)
SECTION_TO_CAT = {
    "SNACKS": ("Snacks", "General"),
    "CHEF SPECIAL": ("Chef Special", "General"),
    "STARTERS – VEG": ("Starters", "Veg"),
    "STARTERS - VEG": ("Starters", "Veg"),
    "STARTERS – NON-VEG": ("Starters", "Non-Veg"),
    "STARTERS - NON-VEG": ("Starters", "Non-Veg"),
    "PIZZA – VEG (8 INCHES ROUND)": ("Pizza", "Veg"),
    "PIZZA - VEG (8 INCHES ROUND)": ("Pizza", "Veg"),
    "PIZZA – NON VEG (8 INCHES ROUND)": ("Pizza", "Non-Veg"),
    "PIZZA - NON VEG (8 INCHES ROUND)": ("Pizza", "Non-Veg"),
    "BURGERS – VEG (NO CHIPS)": ("Burgers", "Veg"),
    "BURGERS - VEG (NO CHIPS)": ("Burgers", "Veg"),
    "BURGERS – NON VEG (NO CHIPS)": ("Burgers", "Non-Veg"),
    "BURGERS - NON VEG (NO CHIPS)": ("Burgers", "Non-Veg"),
    "SANDWICHES – VEG": ("Sandwiches", "Veg"),
    "SANDWICHES - VEG": ("Sandwiches", "Veg"),
    "SANDWICHES – NON-VEG": ("Sandwiches", "Non-Veg"),
    "SANDWICHES - NON-VEG": ("Sandwiches", "Non-Veg"),
    "FRENCH FRIES": ("Sides", "French Fries"),
    "COMBO'S – VEG": ("Combos", "Veg"),
    "COMBO'S - VEG": ("Combos", "Veg"),
    "COMBO'S – NON VEG": ("Combos", "Non-Veg"),
    "COMBO'S - NON VEG": ("Combos", "Non-Veg"),
    "MOCKTAILS": ("Beverages", "Mocktails"),
    "MILKSHAKES": ("Beverages", "Milkshakes"),
    "LASSI": ("Beverages", "Lassi"),
    "COOL CAKES": ("Cakes", "Cool Cakes"),
    "NORMAL CAKES MENU": ("Cakes", "Normal Cakes"),
    "PASTRIES": ("Bakery", "Pastries"),
    "BREADS & BAKERY": ("Bakery", "Breads"),
    "CHAAT": ("Chaat", "General"),
}


def normalize_section_title(raw: str) -> str:
    if not raw:
        return ""
    s = str(raw)
    # Remove emoji / pictographs (includes colored squares, food emoji, etc.)
    s = re.sub(
        r"[\U0001F300-\U0001FAFF\U00002600-\U000027BF\U00002300-\U000023FF\U0000200D]",
        "",
        s,
    )
    s = s.replace("\ufe0f", "").replace("\u200d", "")
    s = re.sub(r"\s+", " ", s).strip()
    # Normalize en-dash to hyphen for map lookup
    s = s.replace("\u2013", "-").replace("–", "-")
    # Leftovers after emoji strip (e.g. variation selector) before section name
    s = re.sub(r"^[\s\W_]+", "", s)
    return s.strip()


def is_section_header(a_val, b_val) -> bool:
    if a_val is None or b_val is None:
        return False
    a = str(a_val).strip().upper()
    if a != "S.NO":
        return False
    b = str(b_val)
    return bool(re.search(r"[\U0001F300-\U0001FAFF\U00002600-\U000027BF]", b)) or any(
        x in normalize_section_title(b).upper() for x in ("SNACKS", "PIZZA", "BURGER", "STARTERS", "SANDWICH", "COMBO", "MOCKTAIL", "MILKSHAKE", "LASSI", "CAKE", "PASTRY", "BREAD", "CHAAT", "FRIES", "CHEF")
    )


def is_data_row(a_val, b_val) -> bool:
    if a_val is None or b_val is None:
        return False
    try:
        int(a_val)
    except (TypeError, ValueError):
        return False
    if str(a_val).strip().upper() == "S.NO":
        return False
    return bool(str(b_val).strip())


def veg_or_nonveg(subcategory: str, item_name: str) -> str:
    """Return 'Veg' or 'Non-Veg' for display in Excel."""
    s = (subcategory or "").strip()
    if s == "Veg":
        return "Veg"
    if s == "Non-Veg":
        return "Non-Veg"
    name = (item_name or "").strip()
    if _NONVEG_NAME.search(name):
        return "Non-Veg"
    return "Veg"


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not path.is_file():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(path)
    if "C3" not in wb.sheetnames:
        print("Sheet 'C3' not found", file=sys.stderr)
        return 1
    ws = wb["C3"]

    current: tuple[str, str] = ("", "")
    filled = 0
    warnings: list[str] = []

    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value

        if is_section_header(a, b):
            norm = normalize_section_title(str(b))
            if norm in SECTION_TO_CAT:
                current = SECTION_TO_CAT[norm]
            else:
                # Try loose match: uppercase key
                key = norm.upper()
                found = None
                for k, v in SECTION_TO_CAT.items():
                    if k.upper() == key:
                        found = v
                        break
                if found:
                    current = found
                else:
                    current = (norm or "Unknown", "General")
                    warnings.append(f"Row {r}: unmapped section {norm!r}")
            ws.cell(r, 6).value = None
            ws.cell(r, 7).value = None
            ws.cell(r, 8).value = None
            ws.cell(r, 9).value = None
            continue

        if is_data_row(a, b):
            ws.cell(r, 6).value = current[0]
            ws.cell(r, 7).value = current[1]
            ws.cell(r, 8).value = veg_or_nonveg(current[1], str(b))
            filled += 1
        else:
            # leave F/G unchanged for blank rows
            pass

    images_filled = 0
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        f = ws.cell(r, 6).value
        g = ws.cell(r, 7).value
        if is_section_header(a, b):
            ws.cell(r, 9).value = None
            continue
        if not is_data_row(a, b):
            continue
        fv = str(f).strip() if f else ""
        if not fv or fv.lower() == "category":
            continue
        name = str(b).strip()
        sub = str(g).strip() if g else ""
        ws.cell(r, 9).value = resolve_menu_item_image(name, fv, sub, None)
        images_filled += 1

    # Column titles (after image pass — row 1 is a section row and would clear I1 otherwise)
    ws.cell(1, 6).value = "Category"
    ws.cell(1, 7).value = "Subcategory"
    ws.cell(1, 8).value = "Veg or Non-Veg"
    ws.cell(1, 9).value = "Image URL"

    out_path = path
    wb.save(out_path)
    wb.close()
    print(
        f"Saved {out_path} — filled Category, Subcategory, Veg or Non-Veg on {filled} item rows; "
        f"Image URL on {images_filled} rows."
    )
    for w in warnings[:20]:
        print("WARN:", w)
    if len(warnings) > 20:
        print(f"... and {len(warnings) - 20} more warnings")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
