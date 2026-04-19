#!/usr/bin/env python3
"""
Enrich ClassieJail menu Excel and emit menu-item JSON; optional POST to API.

Reads:  /Users/user/Downloads/ClassieJail_DineIn_vs_Zomato.xlsx
        Cols A–E: Category | Sub-Category | Item Name | Dine-In Price (₹) |
                  Our Hike on Dine-in %

Adds / refreshes columns:
  F — Description
  G — (blank spacer)
  H — isVeg (TRUE/FALSE)
  I — image (single Unsplash URL per row; URLs are distinct across the sheet)

Writes JSON array of objects (one per row), restaurantId trailing:
  name, restaurantPrice, hikePercentage, category, subCategory,
  isVeg, isAvailable, description, image, restaurantId

Usage:
  python3 scripts/enrich_classiejail_menu.py                    # enrich + JSON only
  python3 scripts/enrich_classiejail_menu.py --apply             # enrich + POST each item
  python3 scripts/enrich_classiejail_menu.py --from-json --apply # POST from JSON only

Prod: HONESTEATS_API_URL (default https://api.yumdude.com) + either
  HONESTEATS_RETOOL_BYPASS (x-retool-header) or HONESTEATS_BEARER_TOKEN.
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

XLSX_PATH = Path("/Users/user/Downloads/ClassieJail_DineIn_vs_Zomato.xlsx")
JSON_PATH = Path("/Users/user/Downloads/ClassieJail_DineIn_vs_Zomato.menu_items.json")
RESTAURANT_ID = "RES-1776575459110-2460"

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get(
    "HONESTEATS_RETOOL_BYPASS",
    "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc",
)
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

COL_CAT, COL_SUB, COL_NAME, COL_PRICE, COL_HIKE = 1, 2, 3, 4, 5
COL_DESC, COL_GAP, COL_VEG, COL_IMG = 6, 7, 8, 9


# ── Unsplash photo IDs grouped by food theme ─────────────────────────────────
# All IDs sourced from existing sister scripts (kitchengarden / kohinoor /
# maharaja / prakruthi). Themed pools help match content; a global de-dup keeps
# every assigned URL unique within the sheet. If themed pools run out for a
# given row, the next unused URL from the global pool is used. If even the
# global pool is exhausted, a "?sig=<row>" suffix is appended to a recycled
# image so the URL itself stays unique.
_POOLS: dict[str, list[str]] = {
    "soup": [
        "1547592166-23ac45744acd", "1603105037880-880cd4f5b2e6",
        "1476718406336-bb5a9690ee2a", "1588566565463-180a5b2090d2",
        "1613844237701-8f3664fc2eff", "1594756202469-9ff9799b2e4e",
        "1604152135912-04a022e23696", "1607330289024-1535c6b4e1c1",
        "1583608205776-bfd35f0d9f83", "1509358271058-acd22cc93898",
        "1617093727343-374698b1b08d", "1597227129956-93bad7e18d08",
        "1603894584373-5ac82b2ae328",
    ],
    "starters_veg": [
        "1601050690597-df0568f70950", "1546069901-ba9599a7e63c",
        "1540189549336-e6e99c3679fe", "1567337710282-00832b415979",
        "1585032226651-759b368d7246", "1572715376701-98568319fd0b",
        "1559847844-5315695dadae", "1606574977732-e8e5f1f46c23",
        "1564834724105-918b73d1b8e0", "1512621776951-a57141f2eefd",
        "1574484284002-952d92456975", "1567620832903-9fc6debc209f",
        "1606728035253-49e8a23146de", "1599487488170-d11ec9c172f0",
        "1565299624946-b28f40a0ae38", "1573080496219-bb080dd4f877",
        "1585937421612-70a008356fbe", "1580217593608-61931cefc821",
        "1614398751058-bca239de00ca", "1603360946369-dc9bb6258143",
    ],
    "starters_nonveg": [
        "1626082927389-6cd097cdc6ec", "1562967914-608f82629710",
        "1608039755401-742074f0548d", "1610057099431-d73a1c9d2f2f",
        "1598515214211-89d3c73ae83b", "1587593810167-a84920ea0781",
        "1619221882220-947b3d3c8861", "1632778149955-e80f8ceca2e8",
        "1624726175512-19b9baf9fbd1",
    ],
    "seafood": [
        "1615141982883-c7ad0e69fd62", "1504674900247-0877df9cc836",
        "1535140728325-a4d3707eee61", "1510130113581-4ae76c0f6e7f",
        "1559039448-9b03d2e3c18e", "1519708227418-c8fd9a32b7a2",
        "1498654896293-37aacf113fd9", "1580476262798-bddd9f4b7369",
        "1551504734-5ee1c4a1479b", "1569058242567-93de6f36f8e6",
        "1606731219412-213c1e68ca63", "1617196035154-1e7e6e28b0db",
        "1610540881815-6e81b56d8ea5", "1615361200098-9e630ec29b4e",
    ],
    "egg": [
        "1488477181946-6428a0291777", "1607103058027-4c5b87a85a55",
        "1551504734-5ee1c4a1479b",
    ],
    "tandoori": [
        "1601050690117-94f5f6fa8bd7", "1628294895950-9805252327bc",
        "1567188040759-fb8a883dc6d8", "1574653853027-5382a3d23a15",
        "1551881192-5e377f1b2142", "1545247181-516773cae754",
        "1606491956689-2ea866880049", "1609501676725-7186f017a4b7",
        "1617692855027-33b14f061079", "1633321702518-7fecdafb94d5",
        "1631452180519-c014fe946bc7", "1618449840665-9ed506d73a34",
        "1642821373181-16a5bc9f5801",
    ],
    "bread": [
        "1600326145359-3a44909d1a39", "1574071318508-1cdbab80d002",
        "1586444248879-bc604bc77212", "1555939594-58d7cb561ad1",
        "1555507036-ab1f4038024a", "1573675542321-f51b18e6e759",
        "1515024014929-c2ba2c4da1d0", "1519864600395-3404e40a0eda",
        "1509722747041-616f39b57569", "1528736235302-52922df5c122",
        "1605888969139-42cca4308aa2",
    ],
    "curry_veg": [
        "1565557623262-b51c2513a641", "1631515243349-e0cb75fb8d3a",
        "1455619452474-d2be8b1e70cd", "1596797038530-2c107229654b",
        "1604908176997-125f25cc6f3d", "1612929633738-8fe44f7ec841",
        "1576402187878-974f70c890a5", "1505253758473-96b7015fcd40",
        "1607116667573-1c7d73636ba0", "1599043513900-ed6fe01d3833",
        "1563379091339-03b21ab4a4f4", "1606755962773-d324e0a13086",
        "1612874742237-6526221588e3", "1625937286520-3ef7955f3813",
        "1627662168223-7df99068099a", "1545247181-516773cae754",
        "1606491956689-2ea866880049",
    ],
    "curry_nonveg": [
        "1559339352-11d035aa65de", "1553621042-f6e147245754",
        "1565680018093-ebb6e3062e7b", "1610057099431-d73a1c9d2f2f",
        "1598515214211-89d3c73ae83b", "1587593810167-a84920ea0781",
        "1619221882220-947b3d3c8861", "1632778149955-e80f8ceca2e8",
        "1624726175512-19b9baf9fbd1",
    ],
    "fried_rice": [
        "1596560548464-f010549b84d7", "1516714435131-44d6b64dc6a2",
        "1536304993881-460587633ee1", "1512058564366-18510be2db19",
        "1645696301019-35adcc552067", "1603133872878-684f208fb84b",
    ],
    "biryani": [
        "1589302168068-964664d93dc0", "1633321702518-7fecdafb94d5",
        "1565557623262-b51c2513a641", "1612874742237-6526221588e3",
        "1625937286520-3ef7955f3813", "1599043513900-ed6fe01d3833",
        "1628294895950-9805252327bc", "1607116667573-1c7d73636ba0",
        "1574653853027-5382a3d23a15", "1603894584373-5ac82b2ae328",
        "1551881192-5e377f1b2142", "1505253758473-96b7015fcd40",
        "1627662168223-7df99068099a", "1545247181-516773cae754",
        "1609501676725-7186f017a4b7", "1618449840665-9ed506d73a34",
        "1617692855027-33b14f061079",
    ],
    "drinks": [
        "1534353473418-4cfa6c56fd38", "1600271886742-f049cd451bba",
        "1572490122747-3968b75cc699", "1541658016709-82535e94bc69",
        "1553787499-6f9133860278", "1568901839119-631418a3910d",
        "1619158401201-8fa932695178", "1579954115545-a95591f28bfc",
        "1497034825429-c343d7c6a68f", "1501443762994-82bd5dace89a",
        "1560008581-09826d1de69e", "1570197571499-166b36435e9f",
        "1629385701021-fcd568a743e8", "1576506295286-5cda18df43e7",
        "1633933358116-a27b902fad35", "1514849302-984523450cf4",
        "1563805042-7684c019e1cb", "1551024506-0bccd828d307",
        "1580915411954-282cb1b0d780", "1557142046-c704a3adf364",
        "1505394033641-40c6ad1178d7", "1516559828984-fb3b99548b21",
        "1621303837174-89787a7d4729", "1587563871167-1ee9c731aefb",
        "1612203985729-70726954388c", "1615478503562-ec2d8aa0e24e",
        "1567206563064-6f60f40a2b57", "1543255006-d6395b6f1171",
        "1595348020949-87cdfbb44174", "1600002423562-975eabb78d5a",
        "1632170684742-9c8b38c1aeab", "1628607189631-96e9e8a3cedc",
        "1622483767028-3f66f32aef97", "1659432873335-3b5a6d7f1a4f",
        "1625869767142-1fb8faf7e8d9",
    ],
    "salad": [
        "1546069901-ba9599a7e63c", "1567620832903-9fc6debc209f",
        "1540189549336-e6e99c3679fe", "1559847844-5315695dadae",
        "1606728035253-49e8a23146de",
    ],
    "dessert": [
        "1571091718767-18b5b1457add", "1567620832903-9fc6debc209f",
        "1606312619070-d48b4c652a52", "1551024506-0bccd828d307",
    ],
    "raita_curd": [
        "1568901346375-23c9450c58cd", "1550547660-d9450f859349",
        "1571091718767-18b5b1457add", "1586190848861-99aa4a171e90",
    ],
    "mandi": [
        "1599487488170-d11ec9c172f0", "1601050690117-94f5f6fa8bd7",
        "1628294895950-9805252327bc", "1567188040759-fb8a883dc6d8",
        "1574653853027-5382a3d23a15", "1551881192-5e377f1b2142",
    ],
}


def _photo_url(pid: str, sig: int | None = None) -> str:
    base = f"https://images.unsplash.com/photo-{pid}?auto=format&fit=crop&w=800&q=80"
    return base if sig is None else f"{base}&sig={sig}"


# Validate IDs (defensive: drop anything that doesn't look like a real Unsplash photo id)
_RX_ID = re.compile(r"^\d{6,}-[0-9a-z]+$")
for k, lst in list(_POOLS.items()):
    _POOLS[k] = [pid for pid in lst if _RX_ID.match(pid)]


# Flatten + dedupe global pool, preserving themed order so themed picks come first
_seen: set[str] = set()
_GLOBAL_POOL: list[str] = []
for _pool_name in [
    "soup", "starters_veg", "starters_nonveg", "seafood", "egg", "tandoori",
    "bread", "curry_veg", "curry_nonveg", "fried_rice", "biryani", "drinks",
    "salad", "dessert", "raita_curd", "mandi",
]:
    for _pid in _POOLS[_pool_name]:
        if _pid not in _seen:
            _seen.add(_pid)
            _GLOBAL_POOL.append(_pid)


# Per-pool dedup view of unused photo ids (for themed picking)
class _Picker:
    def __init__(self):
        self.used: set[str] = set()
        self.recycle_idx = 0
        self.recycle_sig = 0

    def pick_themed(self, pool_key: str) -> str | None:
        for pid in _POOLS.get(pool_key, []):
            if pid not in self.used:
                self.used.add(pid)
                return _photo_url(pid)
        return None

    def pick_global(self) -> str | None:
        for pid in _GLOBAL_POOL:
            if pid not in self.used:
                self.used.add(pid)
                return _photo_url(pid)
        return None

    def pick_recycled(self) -> str:
        # All photo ids used; recycle with a unique sig to keep URL unique.
        pid = _GLOBAL_POOL[self.recycle_idx % len(_GLOBAL_POOL)]
        self.recycle_idx += 1
        self.recycle_sig += 1
        return _photo_url(pid, sig=self.recycle_sig)

    def pick(self, pool_key: str) -> str:
        return (
            self.pick_themed(pool_key)
            or self.pick_global()
            or self.pick_recycled()
        )


# ── Pool selection from category / subcategory / item name ────────────────────
def select_pool(cat: str, sub: str, name: str) -> str:
    cl = (cat or "").lower()
    sl = (sub or "").lower()
    nl = (name or "").lower()

    if "soup" in cl:
        return "soup"
    if "biryani" in cl or "biryani" in nl or "thali" in nl:
        return "biryani"
    if "mandi" in cl or "mandi" in sl or "mandi" in nl:
        return "mandi"
    if "fried rice" in cl or "fried rice" in nl or "noodle" in nl or "rice" in cl:
        return "fried_rice"
    if "bread" in cl or any(k in nl for k in ("naan", "roti", "paratha", "kulcha", "rumali")):
        return "bread"
    if "tandoori" in cl or "tandoori" in sl or any(k in nl for k in ("tandoori", "tikka", "kebab", "kabab")):
        return "tandoori"
    if "raita" in cl or "raita" in nl or "curd" in nl:
        return "raita_curd"
    if "salad" in cl or "salad" in nl or "papad" in nl:
        return "salad"
    if "dessert" in cl or any(k in nl for k in ("kheer", "meeta", "delight", "halwa", "ice cream", "icecream")):
        return "dessert"
    if "beverage" in cl or any(k in nl for k in ("juice", "lassi", "mocktail", "water", "soda", "milk shake", "milkshake", "smoothie")):
        return "drinks"
    if "platter" in cl:
        return "starters_nonveg" if any(k in nl for k in ("non-veg", "nonveg", "chicken", "mutton")) else "starters_veg"
    if "starters" in cl:
        if "sea food" in sl or "seafood" in sl or any(k in nl for k in ("fish", "prawn", "crab", "pomfret")):
            return "seafood"
        if "non-veg" in sl or "nonveg" in sl or any(k in nl for k in ("chicken", "mutton", "lamb", "beef", "egg")):
            return "starters_nonveg"
        return "starters_veg"
    if "egg" in cl or "egg" in sl:
        return "egg"
    if "rayalaseema" in cl:
        return "starters_nonveg"
    if "main course" in cl or "curry" in cl or "curry" in sl:
        if "non-veg" in sl or any(k in nl for k in ("chicken", "mutton", "fish", "prawn", "crab", "lamb", "beef", "egg")):
            return "curry_nonveg"
        return "curry_veg"
    return "starters_veg"


# ── isVeg classification ──────────────────────────────────────────────────────
_NONVEG_KEYWORDS = (
    "non-veg", "non veg", "nonveg", "chicken", "mutton", "fish", "prawn",
    "crab", "lamb", "beef", "egg", "bbq", "kebab", "kabab", "keema",
    "sea food", "seafood", "mandi",
)


def classify_veg(category: str, sub: str, name: str) -> bool:
    text = f"{(sub or '').lower()} {(name or '').lower()}"
    if any(k in text for k in _NONVEG_KEYWORDS):
        return False
    return True


# ── Description generation ────────────────────────────────────────────────────
def clean_display_name(raw: str | None) -> str:
    if not raw:
        return ""
    s = str(raw)
    s = re.sub(r"\s*\([^)]*\)\s*", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def build_description(category: str, sub: str, name: str, is_veg: bool) -> str:
    cat = (category or "").strip()
    sub = (sub or "").strip()
    base = clean_display_name(name)
    veg_word = "vegetarian" if is_veg else "non-vegetarian"
    house = "ClassieJail Restaurant"

    cl = cat.lower()
    sl = sub.lower()
    nl = (name or "").lower()

    if "soup" in cl:
        return f"{base} — house-style {veg_word} soup, simmered with aromatics. Served hot at {house}."
    if "biryani" in cl:
        return (
            f"{base} — long-grain basmati layered with spices and slow-dum cooked. "
            f"{'Vegetarian' if is_veg else 'Non-vegetarian'} biryani served with raita and salan at {house}."
        )
    if "mandi" in cl or "mandi" in sl:
        return f"{base} — Arabian-style mandi rice with slow-cooked, aromatic spices. {house} signature."
    if "tandoori" in cl or "tandoori" in sl or any(k in nl for k in ("tandoori", "tikka", "kebab", "kabab")):
        return f"{base} — char-grilled in the tandoor with house masalas; smoky and tender. {house}."
    if "bread" in cl or any(k in nl for k in ("naan", "roti", "paratha", "kulcha", "rumali")):
        return f"{base} — fresh from the tandoor, soft and pliable; pairs with curries and kebabs at {house}."
    if "fried rice" in cl or "rice" in cl:
        return f"{base} — wok-tossed {veg_word} rice with sauces and aromatics, Indo-Chinese style. {house}."
    if "starters" in cl or "platter" in cl or "rayalaseema" in cl:
        return f"{base} — {veg_word} starter with bold spices and crisp finish; great to share. {house}."
    if "egg" in cl:
        return f"{base} — egg-based dish prepared in {house}-style with a generous spice base."
    if "main course" in cl or "curry" in cl:
        return f"{base} — {veg_word} curry from {house}'s main course, slow-cooked in a rich masala."
    if "raita" in cl:
        return f"{base} — chilled curd-based accompaniment, perfect with biryanis and tandoori. {house}."
    if "salad" in cl:
        return f"{base} — fresh accompaniment to balance the spices on your plate. {house}."
    if "dessert" in cl:
        return f"{base} — sweet finish to your meal at {house}."
    if "beverage" in cl:
        return f"{base} — chilled refreshment from {house}'s beverages list."
    return f"{base} — a {veg_word} dish from {house}'s {cat or 'menu'}."


# ── Number parsing ────────────────────────────────────────────────────────────
def parse_num(val) -> float | int:
    if val is None or val == "":
        return 0
    if isinstance(val, (int, float)):
        return int(val) if float(val) == int(val) else float(val)
    try:
        x = float(str(val).strip().rstrip("%"))
        return int(x) if x == int(x) else x
    except ValueError:
        return 0


def enrich_workbook() -> list[dict]:
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    hdr_fill = PatternFill("solid", fgColor="2E86AB")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col, label in ((COL_DESC, "Description"), (COL_VEG, "isVeg"), (COL_IMG, "image")):
        c = ws.cell(1, col, label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    picker = _Picker()
    payloads: list[dict] = []
    seen_img_urls: set[str] = set()
    recycle_count = 0

    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, COL_NAME).value
        if name is None or str(name).strip() == "":
            continue
        cat = ws.cell(r, COL_CAT).value
        sub = ws.cell(r, COL_SUB).value
        price = ws.cell(r, COL_PRICE).value
        hike = ws.cell(r, COL_HIKE).value

        cat_s = str(cat).strip() if cat else ""
        sub_s = str(sub).strip() if sub else ""
        name_s = str(name).strip()

        is_veg = classify_veg(cat_s, sub_s, name_s)
        desc = build_description(cat_s, sub_s, name_s, is_veg)

        pool_key = select_pool(cat_s, sub_s, name_s)
        img_url = picker.pick(pool_key)
        if img_url in seen_img_urls:
            # Defensive — should not happen, but guard against any edge case.
            img_url = picker.pick_recycled()
        seen_img_urls.add(img_url)
        if "&sig=" in img_url:
            recycle_count += 1

        ws.cell(r, COL_DESC, desc).alignment = wrap
        ws.cell(r, COL_GAP, None)
        ws.cell(r, COL_VEG, bool(is_veg)).alignment = center
        ws.cell(r, COL_IMG, img_url).alignment = wrap

        payloads.append(
            {
                "name": name_s,
                "restaurantPrice": parse_num(price),
                "hikePercentage": parse_num(hike),
                "category": cat_s,
                "subCategory": sub_s,
                "isVeg": bool(is_veg),
                "isAvailable": True,
                "description": desc,
                "image": [img_url],
                "restaurantId": RESTAURANT_ID,
            }
        )

    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 4
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 64

    wb.save(XLSX_PATH)
    JSON_PATH.write_text(json.dumps(payloads, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    distinct_urls = len(seen_img_urls)
    print(f"Enriched rows: {len(payloads)}")
    print(f"Saved workbook: {XLSX_PATH}")
    print(f"Saved JSON: {JSON_PATH}")
    print(
        f"Image URLs: {distinct_urls} unique"
        f" (unique photo IDs: {len(_GLOBAL_POOL)},"
        f" recycled with &sig: {recycle_count})"
    )
    return payloads


def menu_item_api_body(row: dict) -> dict:
    """POST body for /api/v1/restaurants/{id}/menu (no restaurantId field)."""
    cat = (row.get("category") or "").strip() or None
    sub = (row.get("subCategory") or "").strip() or None
    desc = row.get("description")
    if isinstance(desc, str):
        desc = desc.strip() or None
    return {
        "name": row["name"],
        "restaurantPrice": float(row["restaurantPrice"]),
        "hikePercentage": float(row["hikePercentage"]),
        "category": cat,
        "subCategory": sub,
        "isVeg": row.get("isVeg", True),
        "isAvailable": row.get("isAvailable", True),
        "description": desc,
        "image": row.get("image") or [],
    }


def post_menu_items(
    payloads: list[dict],
    *,
    api_url: str,
    bearer_token: str | None,
    delay_sec: float,
) -> int:
    menu_url = f"{api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"
    session = requests.Session()
    session.headers["Content-Type"] = "application/json"
    if bearer_token:
        session.headers["Authorization"] = f"Bearer {bearer_token}"
    else:
        session.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS

    ok = err = 0
    for i, raw in enumerate(payloads, 1):
        body = menu_item_api_body(raw)
        if delay_sec > 0:
            time.sleep(delay_sec)
        try:
            r = session.post(menu_url, json=body, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"  ✗ [{i}] {body['name']}: {ex}", file=sys.stderr)
            continue
        if r.status_code == 201:
            ok += 1
            print(f"  ✓ [{i}] {body['name']}")
        else:
            err += 1
            print(
                f"  ✗ [{i}] {body['name']}: HTTP {r.status_code} {r.text[:300]}",
                file=sys.stderr,
            )
            if r.status_code == 401:
                print(
                    "\nUnauthorized: set HONESTEATS_RETOOL_BYPASS or HONESTEATS_BEARER_TOKEN.",
                    file=sys.stderr,
                )
                return 1

    print(f"\nPOST done: {ok} created, {err} failed. URL base: {api_url.rstrip('/')}")
    return 0 if err == 0 else 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich ClassieJail menu Excel + optional API import")
    parser.add_argument("--apply", action="store_true", help="POST each item to the API")
    parser.add_argument(
        "--from-json",
        action="store_true",
        help="Load payloads from JSON_PATH instead of re-enriching Excel",
    )
    parser.add_argument("--api-url", default=DEFAULT_API_URL, help="API base URL")
    parser.add_argument(
        "--bearer-token",
        default=os.environ.get("HONESTEATS_BEARER_TOKEN"),
        help="JWT (optional; else x-retool-header)",
    )
    parser.add_argument("--delay", type=float, default=0.3, help="Seconds between POSTs")
    args = parser.parse_args()

    if args.from_json:
        if not JSON_PATH.is_file():
            print(f"Missing {JSON_PATH}", file=sys.stderr)
            return 1
        payloads = json.loads(JSON_PATH.read_text(encoding="utf-8"))
        print(f"Loaded {len(payloads)} items from {JSON_PATH.name}")
    else:
        payloads = enrich_workbook()

    if not args.apply:
        print("No POST (pass --apply to import).")
        return 0

    return post_menu_items(
        payloads,
        api_url=args.api_url,
        bearer_token=args.bearer_token,
        delay_sec=args.delay,
    )


if __name__ == "__main__":
    raise SystemExit(main())
