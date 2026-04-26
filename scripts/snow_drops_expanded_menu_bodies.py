#!/usr/bin/env python3
"""
Enrich Snow_Drops_Menu_expanded.xlsx and build menu POST bodies (no HTTP).

Source columns:
  A: Category
  B: SubCategory
  C: Item Name
  D: Price (₹)
  E: addOnOptions (JSON string, optional)

Enrichment (only when column H is blank):
  H: isVeg  (boolean)

JSON per row:
  {
    "name": C, "restaurantPrice": D, "hikePercentage": 35,
    "category": A, "subCategory": B, "isVeg": H,
    "isAvailable": true,
    "addOnOptions": <parsed from E, omitted if empty>
  }

Wrap output: {"restaurantId", "items"} unless --array-only.

Usage:
  python3 scripts/snow_drops_expanded_menu_bodies.py
  python3 scripts/snow_drops_expanded_menu_bodies.py --xlsx ~/Downloads/Snow_Drops_Menu_expanded.xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_XLSX = Path.home() / "Downloads" / "Snow_Drops_Menu_expanded.xlsx"
DEFAULT_RESTAURANT_ID = "RES-1777198830257-9278"
DEFAULT_OUT_JSON = Path.home() / "Downloads" / f"snow_drops_menu_bodies_{DEFAULT_RESTAURANT_ID}.json"
DEFAULT_HIKE = 35.0


def _clean(v: Any) -> str:
    return str(v).strip() if v is not None else ""


def _title_words(s: str) -> str:
    t = _clean(s)
    if not t:
        return ""
    return " ".join(w[:1].upper() + w[1:].lower() for w in t.split())


def _parse_price(v: Any) -> float:
    if isinstance(v, (int, float)):
        return float(v)
    s = _clean(v).replace("₹", "").replace(",", "")
    return float(s)


def _parse_is_veg_cell(v: Any) -> bool | None:
    if v is None or _clean(v) == "":
        return None
    if isinstance(v, bool):
        return v
    s = _clean(v).lower()
    if s in {"true", "t", "yes", "1", "veg", "v", "vegetarian"}:
        return True
    if s in {"false", "f", "no", "0", "non-veg", "nonveg", "nv"}:
        return False
    return None


def _infer_is_veg(category: str, subcategory: str, name: str) -> bool:
    c = category.lower()
    b = subcategory.lower()
    s = name.lower()

    if c == "fried chicken":
        return False
    if c == "combos":
        if s.startswith("any veg ") or re.match(r"^any\s+veg\s+", s):
            return True
        if s.startswith("any chicken ") or re.match(r"^any\s+chicken\s+", s):
            return False
        return False
    if c == "burgers":
        return "chicken" not in b
    if c == "sandwiches":
        return "chicken" not in b
    if c == "pizza":
        return "chicken" not in b
    if c == "snacks":
        if "chicken" in s:
            return False
        return True
    return True


def _parse_add_ons(cell_val: Any) -> list[dict[str, Any]] | None:
    if cell_val is None:
        return None
    if isinstance(cell_val, list):
        return cell_val
    t = _clean(cell_val)
    if not t or t.lower() == "null":
        return None
    try:
        data = json.loads(t)
    except json.JSONDecodeError:
        return None
    if isinstance(data, list):
        return data
    return None


def enrich_workbook(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    ws.cell(1, 8, "isVeg")

    for row in range(2, (ws.max_row or 1) + 1):
        name = _clean(ws.cell(row, 3).value)
        if not name:
            continue
        if _parse_is_veg_cell(ws.cell(row, 8).value) is None:
            cat = _clean(ws.cell(row, 1).value)
            sub = _clean(ws.cell(row, 2).value)
            ws.cell(row, 8, _infer_is_veg(cat, sub, name))

    wb.save(path)
    wb.close()


def build_bodies(path: Path, *, hike: float) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    bodies: list[dict[str, Any]] = []

    for row in range(2, (ws.max_row or 1) + 1):
        name = _clean(ws.cell(row, 3).value)
        if not name:
            continue
        cat = _clean(ws.cell(row, 1).value)
        sub = _clean(ws.cell(row, 2).value)
        price = _parse_price(ws.cell(row, 4).value)
        addons = _parse_add_ons(ws.cell(row, 5).value)
        is_veg = _parse_is_veg_cell(ws.cell(row, 8).value)
        if is_veg is None:
            is_veg = _infer_is_veg(cat, sub, name)

        body: dict[str, Any] = {
            "name": _title_words(name),
            "restaurantPrice": price,
            "hikePercentage": hike,
            "category": _title_words(cat),
            "subCategory": _title_words(sub),
            "isVeg": is_veg,
            "isAvailable": True,
        }
        if addons:
            body["addOnOptions"] = addons
        bodies.append(body)

    wb.close()
    return bodies


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Snow Drops expanded menu → JSON bodies (no insert).")
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--out-json", type=Path, default=DEFAULT_OUT_JSON)
    p.add_argument("--restaurant-id", default=DEFAULT_RESTAURANT_ID)
    p.add_argument("--hike", type=float, default=DEFAULT_HIKE, help="hikePercentage for all items")
    p.add_argument("--array-only", action="store_true")
    return p.parse_args()


def main() -> int:
    args = parse_args()

    if not args.xlsx.is_file():
        print(f"Not found: {args.xlsx}", file=sys.stderr)
        return 1

    enrich_workbook(args.xlsx)
    bodies = build_bodies(args.xlsx, hike=args.hike)

    out_obj: Any = bodies if args.array_only else {"restaurantId": args.restaurant_id, "items": bodies}
    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(json.dumps(out_obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    with_addons = sum(1 for b in bodies if b.get("addOnOptions"))
    print(f"Enriched: {args.xlsx}")
    print(f"Restaurant: {args.restaurant_id}")
    print(f"Bodies: {len(bodies)} | with addOnOptions: {with_addons}")
    print(f"Wrote: {args.out_json}")
    print("No HTTP calls.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
