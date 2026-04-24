#!/usr/bin/env python3
"""
Build POST /menu JSON bodies from MakersOfMilkshakes split workbook (no HTTP).

Supported source layouts (first sheet):

1) **price_hike** (current Makers split file):
   A: Category, B: Item Name, C: Dine-In Price (₹), D: Hike

2) **size_price** (older split):
   A: Category, B: Item Name, C: Size, D: Dine-In Price (₹)

3) **enriched** (output of --write-enriched):
   A: Category, B: Item Name, C: restaurantPrice, D: hikePercentage,
   E: subCategory, F: Size, G: reserved, H: isVeg

Enrichment (written when --write-enriched is set):
  A: Category
  B: Item Name
  C: restaurantPrice
  D: hikePercentage
  E: subCategory  (default "Shakes")
  F: Size (preserved from original C)
  G: (reserved / empty)
  H: isVeg  (filled only when missing: default True for this menu)

Add-ons (extraPrice = --addon-price, default 30) for every row EXCEPT when
category matches (case-insensitive): Cheesy Shakes, Bubble Tea(s), Cold Coffee.

Does NOT call the API. Writes JSON array (or { restaurantId, items } with --wrap).

Usage:
  python3 scripts/makers_milkshakes_menu_bodies.py \\
    --xlsx ~/Downloads/MakersOfMilkshakes_DineIn_vs_Zomato_split_sizes.xlsx \\
    --out-json ~/Downloads/makers_menu_bodies_RES-1776965639587-4589.json

  python3 scripts/makers_milkshakes_menu_bodies.py ... --write-enriched ~/Downloads/makers_enriched.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl

DEFAULT_RESTAURANT_ID = "RES-1776965639587-4589"
DEFAULT_SUBCATEGORY = "Shakes"
ADD_ON_NAMES = [
    ("Banana", "ao_1"),
    ("Chocochip", "ao_2"),
    ("KitKat", "ao_3"),
    ("Munch", "ao_4"),
    ("5star", "ao_5"),
    ("Icecream", "ao_6"),
    ("Brownie", "ao_7"),
]

# Categories that must NOT get addOnOptions (match is normalized: lower, strip)
ADD_ON_EXCLUDE_CATEGORIES = frozenset(
    {
        "cheesy shakes",
        "bubble tea",
        "bubble teas",
        "cold coffee",
    }
)


def _norm_cat(value: object | None) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _category_excludes_addons(category: object | None) -> bool:
    key = _norm_cat(category)
    if not key:
        return True
    if key in ADD_ON_EXCLUDE_CATEGORIES:
        return True
    # "Bubble Tea" singular already in set
    return False


def _parse_bool_veg(cell_val: object | None) -> bool | None:
    if cell_val is None or (isinstance(cell_val, str) and not str(cell_val).strip()):
        return None
    if isinstance(cell_val, bool):
        return cell_val
    t = str(cell_val).strip().lower()
    if "non" in t and "veg" in t:
        return False
    if t in ("v", "veg", "vegetarian", "true", "yes", "1"):
        return True
    if t in ("nv", "non-veg", "nonveg", "false", "no", "0"):
        return False
    return None


def _parse_price(cell_val: object | None) -> float | None:
    if cell_val is None:
        return None
    if isinstance(cell_val, (int, float)):
        return round(float(cell_val), 2)
    s = str(cell_val).strip().replace("₹", "").replace(",", "")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


def _build_add_on_options(extra_price: float) -> list[dict]:
    return [
        {"name": name, "optionId": oid, "extraPrice": extra_price}
        for name, oid in ADD_ON_NAMES
    ]


def _parse_hike(cell_val: object | None) -> float | None:
    if cell_val is None or (isinstance(cell_val, str) and not str(cell_val).strip()):
        return None
    if isinstance(cell_val, (int, float)):
        return float(cell_val)
    s = str(cell_val).strip().replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None


def _read_rows_size_price(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Header: Category, Item Name, Size, Dine-In Price."""
    rows: list[dict] = []
    for r in range(2, (ws.max_row or 0) + 1):
        cat = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        size = ws.cell(r, 3).value
        price = _parse_price(ws.cell(r, 4).value)
        if not name or not str(name).strip():
            continue
        if price is None:
            continue
        rows.append(
            {
                "row": r,
                "category": str(cat).strip() if cat else "",
                "name": str(name).strip(),
                "size": str(size).strip() if size is not None and str(size).strip() else None,
                "restaurantPrice": price,
                "hikePercentage": None,
            }
        )
    return rows


def _read_rows_price_hike(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Header: Category, Item Name, Dine-In Price (₹), Hike."""
    rows: list[dict] = []
    for r in range(2, (ws.max_row or 0) + 1):
        cat = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        price = _parse_price(ws.cell(r, 3).value)
        hike = _parse_hike(ws.cell(r, 4).value)
        if not name or not str(name).strip():
            continue
        if price is None:
            continue
        rows.append(
            {
                "row": r,
                "category": str(cat).strip() if cat else "",
                "name": str(name).strip(),
                "size": None,
                "restaurantPrice": price,
                "hikePercentage": hike,
            }
        )
    return rows


def _read_rows_enriched_layout(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict]:
    """Enriched: A cat, B name, C price, D hike, E subcat, F size, G _, H isVeg."""
    rows: list[dict] = []
    for r in range(2, (ws.max_row or 0) + 1):
        cat = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        price = _parse_price(ws.cell(r, 3).value)
        hike = ws.cell(r, 4).value
        sub = ws.cell(r, 5).value
        size = ws.cell(r, 6).value
        is_veg_cell = ws.cell(r, 8).value
        if not name or not str(name).strip():
            continue
        if price is None:
            continue
        try:
            hike_f = float(hike) if hike is not None and str(hike).strip() != "" else 0.0
        except (TypeError, ValueError):
            hike_f = 0.0
        is_veg = _parse_bool_veg(is_veg_cell)
        rows.append(
            {
                "row": r,
                "category": str(cat).strip() if cat else "",
                "name": str(name).strip(),
                "restaurantPrice": price,
                "hikePercentage": hike_f,
                "subCategory": str(sub).strip() if sub else DEFAULT_SUBCATEGORY,
                "size": str(size).strip() if size is not None and str(size).strip() else None,
                "isVeg": is_veg,
            }
        )
    return rows


def _detect_layout(ws: openpyxl.worksheet.worksheet.Worksheet) -> str:
    h1 = ws.cell(1, 1).value
    h2 = ws.cell(1, 2).value
    h3 = ws.cell(1, 3).value
    h4 = ws.cell(1, 4).value
    if not (h1 and "category" in str(h1).lower() and h2 and "item" in str(h2).lower()):
        return "price_hike"
    t3 = str(h3 or "").lower()
    t4 = str(h4 or "").lower()
    if "size" in t3:
        return "size_price"
    if "restaurantprice" in t3.replace(" ", "") or (
        "restaurant" in t3 and "price" in t3
    ):
        return "enriched"
    if "hike" in t4 or "hike" in t3:
        return "price_hike"
    if "dine" in t3 or "price" in t3:
        return "price_hike"
    return "price_hike"


def write_enriched_workbook(
    in_path: Path,
    out_path: Path,
    *,
    subcategory: str,
    hike_default: float,
    default_is_veg: bool,
) -> None:
    wb = openpyxl.load_workbook(in_path)
    ws = wb[wb.sheetnames[0]]
    layout = _detect_layout(ws)
    if layout == "enriched":
        wb.close()
        raise SystemExit("Workbook already looks enriched; use a raw split file as --xlsx.")

    max_r = ws.max_row or 1
    data_rows: list[tuple[object, object, object, object]] = []
    for r in range(2, max_r + 1):
        cat = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        d = ws.cell(r, 4).value
        if not name and not cat:
            continue
        data_rows.append((cat, name, c, d))

    ws.cell(1, 1, "Category")
    ws.cell(1, 2, "Item Name")
    ws.cell(1, 3, "restaurantPrice")
    ws.cell(1, 4, "hikePercentage")
    ws.cell(1, 5, "subCategory")
    ws.cell(1, 6, "Size")
    ws.cell(1, 7, "")
    ws.cell(1, 8, "isVeg")

    out_r = 2
    for cat, name, c, d in data_rows:
        if layout == "size_price":
            price = _parse_price(d)
            hike_f = hike_default
            size_old = c
        else:
            price = _parse_price(c)
            hike_parsed = _parse_hike(d)
            hike_f = hike_default if hike_parsed is None else hike_parsed
            size_old = None
        if price is None:
            continue
        ws.cell(out_r, 1, cat)
        ws.cell(out_r, 2, name)
        ws.cell(out_r, 3, price)
        ws.cell(out_r, 4, hike_f)
        ws.cell(out_r, 5, subcategory)
        ws.cell(out_r, 6, size_old)
        ws.cell(out_r, 7, None)
        ws.cell(out_r, 8, default_is_veg)
        out_r += 1

    if out_r <= max_r:
        ws.delete_rows(out_r, max_r - out_r + 1)

    wb.save(out_path)
    wb.close()


def row_to_payload(
    row: dict,
    *,
    addon_price: float,
    default_hike: float,
    default_subcategory: str,
    default_is_veg: bool,
) -> dict:
    cat = row["category"]
    name = row["name"]
    price = row["restaurantPrice"]
    hp = row.get("hikePercentage")
    hike = default_hike if hp is None else float(hp)
    sub = row.get("subCategory") or default_subcategory
    is_veg = row.get("isVeg")
    if is_veg is None:
        is_veg = default_is_veg

    payload: dict = {
        "name": name,
        "restaurantPrice": price,
        "hikePercentage": hike,
        "category": cat,
        "subCategory": sub,
        "isVeg": is_veg,
        "isAvailable": True,
    }
    if not _category_excludes_addons(cat):
        payload["addOnOptions"] = _build_add_on_options(addon_price)
    return payload


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build menu POST bodies for MakersOfMilkshakes (no API).")
    p.add_argument(
        "--xlsx",
        type=Path,
        default=Path.home() / "Downloads" / "MakersOfMilkshakes_DineIn_vs_Zomato_split_sizes.xlsx",
    )
    p.add_argument(
        "--out-json",
        type=Path,
        default=Path.home() / "Downloads" / f"makers_menu_bodies_{DEFAULT_RESTAURANT_ID}.json",
    )
    p.add_argument("--restaurant-id", default=DEFAULT_RESTAURANT_ID)
    p.add_argument("--subcategory", default=DEFAULT_SUBCATEGORY)
    p.add_argument("--hike-default", type=float, default=0.0)
    p.add_argument("--addon-price", type=float, default=30.0)
    p.add_argument(
        "--default-non-veg",
        action="store_true",
        help="When column H is empty, set isVeg to False (default: True when H empty)",
    )
    p.add_argument(
        "--write-enriched",
        type=Path,
        default=None,
        help="Optional path to write enriched .xlsx (then read it for bodies unless --xlsx unchanged)",
    )
    p.add_argument(
        "--wrap",
        action="store_true",
        help='Write {"restaurantId","items":[...]} instead of a bare array',
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()
    if not args.xlsx.is_file():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        return 1

    read_path = args.xlsx
    if args.write_enriched:
        write_enriched_workbook(
            args.xlsx,
            args.write_enriched,
            subcategory=args.subcategory,
            hike_default=args.hike_default,
            default_is_veg=not args.default_non_veg,
        )
        read_path = args.write_enriched
        print(f"Wrote enriched workbook: {read_path}")

    wb = openpyxl.load_workbook(read_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    layout = _detect_layout(ws)
    if layout == "enriched":
        rows = _read_rows_enriched_layout(ws)
    elif layout == "size_price":
        rows = _read_rows_size_price(ws)
    else:
        rows = _read_rows_price_hike(ws)
    wb.close()

    bodies = [
        row_to_payload(
            r,
            addon_price=args.addon_price,
            default_hike=args.hike_default,
            default_subcategory=args.subcategory,
            default_is_veg=not args.default_non_veg,
        )
        for r in rows
    ]

    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    if args.wrap:
        out_obj = {"restaurantId": args.restaurant_id, "items": bodies}
    else:
        out_obj = bodies

    args.out_json.write_text(json.dumps(out_obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    with_addons = sum(1 for b in bodies if b.get("addOnOptions"))
    print(f"Restaurant (for your reference): {args.restaurant_id}")
    print(f"Rows read: {len(rows)} | Bodies: {len(bodies)} | With addOnOptions: {with_addons}")
    print(f"Wrote JSON: {args.out_json}")
    print("No HTTP calls were made.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
