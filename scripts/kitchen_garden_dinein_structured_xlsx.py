#!/usr/bin/env python3
"""Write The Kitchen Garden dine-in menu to a single Excel sheet."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

OUT_PATH = Path("/Users/user/Downloads/The_Kitchen_Garden_DineIn_Menu.xlsx")

ROWS: list[tuple[str, str, str, int | float]] = [
    # Veg Soup
    ("Veg Soup", "Soup", "Cream Tomato Soup", 130),
    ("Veg Soup", "Soup", "Cream Mushroom Soup", 140),
    ("Veg Soup", "Soup", "Spicy Corn With Almond", 150),
    # Non-Veg Soup
    ("Non-Veg Soup", "Soup", "Chicken Manchow Soup", 170),
    ("Non-Veg Soup", "Soup", "Chicken Royal Soup", 180),
    ("Non-Veg Soup", "Soup", "Mutton Soup", 200),
    ("Non-Veg Soup", "Soup", "Lumpum Soup", 170),
    # Rolls & Fries
    ("Rolls & Fries", "Starters", "Veg Spring Rolls", 210),
    ("Rolls & Fries", "Starters", "Veg Shanghai Rolls", 250),
    ("Rolls & Fries", "Starters", "Chicken Cheese Rolls", 320),
    ("Rolls & Fries", "Starters", "Chicken Spring Rolls", 290),
    ("Rolls & Fries", "Starters", "French Fries", 160),
    ("Rolls & Fries", "Starters", "TKG Special Fries", 180),
    # Veg Chinese
    ("Veg Chinese", "Starters", "Gobi 65", 220),
    ("Veg Chinese", "Starters", "Narmada Crispy Paneer", 370),
    ("Veg Chinese", "Starters", "Mushroom 65", 300),
    ("Veg Chinese", "Starters", "Dragon Paneer", 320),
    ("Veg Chinese", "Starters", "Chilli Babycorn", 300),
    ("Veg Chinese", "Starters", "Veg Cheese Stick", 290),
    ("Veg Chinese", "Starters", "Crispy Corn", 250),
    # Non-Veg Chinese
    ("Non-Veg Chinese", "Starters", "Chicken Lollipop", 250),
    ("Non-Veg Chinese", "Starters", "Chicken Lollipop", 410),
    ("Non-Veg Chinese", "Starters", "Supreme Chicken", 390),
    ("Non-Veg Chinese", "Starters", "Chicken 65", 320),
    ("Non-Veg Chinese", "Starters", "Chicken 555", 390),
    ("Non-Veg Chinese", "Starters", "Devil Wings", 380),
    ("Non-Veg Chinese", "Starters", "Chicken Majestic", 360),
    # Tandoori Veg
    ("Tandoori Veg", "Tandoori", "Tandoori Babycorn", 290),
    ("Tandoori Veg", "Tandoori", "Afghani Paneer Tikka", 370),
    ("Tandoori Veg", "Tandoori", "Paneer Tikka Kebab", 370),
    ("Tandoori Veg", "Tandoori", "Veg Seekh Kebab", 340),
    # Tandoori Non-Veg
    ("Tandoori Non-Veg", "Tandoori", "Tandoori Chicken", 320),
    ("Tandoori Non-Veg", "Tandoori", "Tandoori Chicken", 540),
    ("Tandoori Non-Veg", "Tandoori", "Tangadi Kebab", 280),
    ("Tandoori Non-Veg", "Tandoori", "Tangadi Kebab", 490),
    ("Tandoori Non-Veg", "Tandoori", "Chicken Tikka", 260),
    ("Tandoori Non-Veg", "Tandoori", "Chicken Tikka", 440),
    ("Tandoori Non-Veg", "Tandoori", "Afghani Chicken Tikka", 270),
    ("Tandoori Non-Veg", "Tandoori", "Afghani Chicken Tikka", 450),
    ("Tandoori Non-Veg", "Tandoori", "Dilruba Chicken Kebab", 450),
    ("Tandoori Non-Veg", "Tandoori", "Murgh Bati Bone Kebab", 320),
    ("Tandoori Non-Veg", "Tandoori", "Mutton King Kebab", 540),
    ("Tandoori Non-Veg", "Tandoori", "Mutton Seekh Kebab", 520),
    ("Tandoori Non-Veg", "Tandoori", "Fish Tikka", 490),
    ("Tandoori Non-Veg", "Tandoori", "Tandoori Prawns", 510),
    # Veg Chef Special
    ("Veg Chef Special", "Main", "Supreme Paneer", 340),
    ("Veg Chef Special", "Main", "Butter Garlic Mushroom", 300),
    ("Veg Chef Special", "Main", "Veg 99", 270),
    ("Veg Chef Special", "Main", "Veg Octopus", 290),
    # Non-Veg Chef Special
    ("Non-Veg Chef Special", "Main", "Kaju Chicken Pakoda", 390),
    ("Non-Veg Chef Special", "Main", "Dragon Chicken", 370),
    ("Non-Veg Chef Special", "Main", "Peri Peri Fried Chicken", 370),
    ("Non-Veg Chef Special", "Main", "Chicken Dumpling", 350),
    # Seafood Starter
    ("Seafood Starter", "Starters", "Loose Prawns", 420),
    ("Seafood Starter", "Starters", "Apollo Fish", 410),
    ("Seafood Starter", "Starters", "Prawns Vepudu", 420),
    ("Seafood Starter", "Starters", "Fish Fingers", 410),
    ("Seafood Starter", "Starters", "Russian Pan Fried Prawns", 430),
    ("Seafood Starter", "Starters", "Russian Pan Fried Fish", 420),
    # Grill
    ("Grill", "Grill", "Grill Chicken", 340),
    ("Grill", "Grill", "Grill Chicken", 580),
    # Noodles
    ("Noodles", "Chinese", "Veg Soft Noodles", 220),
    ("Noodles", "Chinese", "Veg Chilli Garlic Noodles", 240),
    ("Noodles", "Chinese", "Chicken Garlic Hakka Noodles", 260),
    ("Noodles", "Chinese", "Chicken Soft Noodles", 260),
    # Chinese Rice
    ("Chinese Rice", "Rice", "Veg Fried Rice", 230),
    ("Chinese Rice", "Rice", "Non-Veg Fried Rice", 250),
    ("Chinese Rice", "Rice", "Chicken Fried Rice", 280),
    ("Chinese Rice", "Rice", "Mushroom Fried Rice", 260),
    ("Chinese Rice", "Rice", "Veg Fried Rice", 210),
    ("Chinese Rice", "Rice", "Kaju Fried Rice", 280),
    ("Chinese Rice", "Rice", "Jeera Rice", 210),
    ("Chinese Rice", "Rice", "Mixed Veg Fried Rice", 250),
    ("Chinese Rice", "Rice", "Bati Fried Rice Veg", 240),
    ("Chinese Rice", "Rice", "Bati Fried Rice Non-Veg", 240),
    # Veg Curries
    ("Veg Curries", "Curry", "Dal Fry", 190),
    ("Veg Curries", "Curry", "Dal Thadka", 200),
    ("Veg Curries", "Curry", "Paneer Butter Masala", 330),
    ("Veg Curries", "Curry", "Kaju Paneer Masala", 350),
    ("Veg Curries", "Curry", "Paneer Kali Mirchi", 340),
    ("Veg Curries", "Curry", "Mushroom Tikka Masala", 310),
    ("Veg Curries", "Curry", "Kadai Veg Curry", 300),
    ("Veg Curries", "Curry", "Panchratan Veg Curry", 310),
    ("Veg Curries", "Curry", "Mokkajonna Palak Iguru", 390),
    ("Veg Curries", "Curry", "Palak Paneer", 310),
    # Egg Curries
    ("Egg Curries", "Curry", "Egg Masala", 250),
    ("Egg Curries", "Curry", "Kodi Guddu Kura", 260),
    ("Egg Curries", "Curry", "Kodi Guddu Vepudu", 250),
    # Indian Curries Non-Veg
    ("Indian Curries Non-Veg", "Curry", "Sholey Chicken Curry", 380),
    ("Indian Curries Non-Veg", "Curry", "Butter Chicken", 370),
    ("Indian Curries Non-Veg", "Curry", "Tawa Chicken", 410),
    ("Indian Curries Non-Veg", "Curry", "Munakai Chicken Curry", 360),
    ("Indian Curries Non-Veg", "Curry", "Matka Chicken Curry", 370),
    ("Indian Curries Non-Veg", "Curry", "Shabnam Chicken Curry", 380),
    # Mutton Curries
    ("Mutton Curries", "Curry", "Mutton Rogan Josh", 520),
    ("Mutton Curries", "Curry", "Methi Butter Mutton Curry", 530),
    ("Mutton Curries", "Curry", "Mutton Dalcha", 510),
    ("Mutton Curries", "Curry", "Gongura Mutton Curry", 540),
    # Seafood Curries
    ("Seafood Curries", "Curry", "Fish Masala", 440),
    ("Seafood Curries", "Curry", "Prawns Masala", 460),
    ("Seafood Curries", "Curry", "Gongura Prawns Curry", 470),
    ("Seafood Curries", "Curry", "Fish Tikka Masala", 460),
    # Indian Breads
    ("Indian Breads", "Bread", "Tandoori Roti", 35),
    ("Indian Breads", "Bread", "Butter Roti", 40),
    ("Indian Breads", "Bread", "Plain Naan", 50),
    ("Indian Breads", "Bread", "Butter Naan", 60),
    ("Indian Breads", "Bread", "Kulcha", 50),
    ("Indian Breads", "Bread", "Butter Kulcha", 60),
    ("Indian Breads", "Bread", "Rumali Roti", 35),
    ("Indian Breads", "Bread", "Pulka", 25),
    ("Indian Breads", "Bread", "Chapathi", 30),
    ("Indian Breads", "Bread", "Folding Butter Naan", 65),
    ("Indian Breads", "Bread", "Kashmir Naan", 115),
    ("Indian Breads", "Bread", "Masala Kulcha", 60),
    ("Indian Breads", "Bread", "Laccha Paratha", 60),
    ("Indian Breads", "Bread", "Plain Paratha", 50),
    ("Indian Breads", "Bread", "Aloo Paratha", 60),
    ("Indian Breads", "Bread", "Pudina Paratha", 60),
    ("Indian Breads", "Bread", "Kothimbir Paratha", 60),
    ("Indian Breads", "Bread", "Mutton Keema Naan", 185),
    ("Indian Breads", "Bread", "Chicken Keema Naan", 155),
    # Veg Dum Biryani
    ("Veg Dum Biryani", "Biryani", "Biryani Rice", 170),
    ("Veg Dum Biryani", "Biryani", "Veg Dum Biryani", 260),
    ("Veg Dum Biryani", "Biryani", "Veg Handi Biryani", 320),
    ("Veg Dum Biryani", "Biryani", "Veg Family Pack Biryani", 490),
    ("Veg Dum Biryani", "Biryani", "Veg Jumbo Biryani", 700),
    ("Veg Dum Biryani", "Biryani", "Paneer Biryani", 340),
    ("Veg Dum Biryani", "Biryani", "Mushroom Biryani", 320),
    ("Veg Dum Biryani", "Biryani", "Kaju Biryani", 340),
    ("Veg Dum Biryani", "Biryani", "Ulavacharu Biryani", 320),
    # Non-Veg Biryani
    ("Non-Veg Biryani", "Biryani", "Egg Biryani", 260),
    ("Non-Veg Biryani", "Biryani", "Mini Chicken Biryani", 220),
    ("Non-Veg Biryani", "Biryani", "Chicken Dum Biryani", 220),
    ("Non-Veg Biryani", "Biryani", "Chicken Handi Biryani", 360),
    ("Non-Veg Biryani", "Biryani", "Chicken Roast Biryani", 320),
    ("Non-Veg Biryani", "Biryani", "Chicken Family Pack Biryani", 640),
    ("Non-Veg Biryani", "Biryani", "Chicken Jumbo Pack Biryani", 930),
    ("Non-Veg Biryani", "Biryani", "Spl Chicken Biryani", 340),
    ("Non-Veg Biryani", "Biryani", "TKG Chicken Biryani", 320),
    # Mutton Biryani
    ("Mutton Biryani", "Biryani", "Mini Mutton Biryani", 280),
    ("Mutton Biryani", "Biryani", "Mutton Button Biryani", 390),
    ("Mutton Biryani", "Biryani", "Mutton Roast Biryani", 410),
    ("Mutton Biryani", "Biryani", "Mutton Family Pack Biryani", 780),
    ("Mutton Biryani", "Biryani", "Mutton Jumbo Pack Biryani", 1250),
    ("Mutton Biryani", "Biryani", "Mutton Kheema Biryani", 430),
    ("Mutton Biryani", "Biryani", "Nalli Ghost Biryani", 440),
    ("Mutton Biryani", "Biryani", "TKG Mutton Biryani", 410),
    ("Mutton Biryani", "Biryani", "Fish Biryani", 410),
    ("Mutton Biryani", "Biryani", "Prawns Biryani", 410),
    # Chef's Special Biryani
    ("Chef's Special Biryani", "Biryani", "Tawa Chicken Biryani", 380),
    ("Chef's Special Biryani", "Biryani", "Tangadi Biryani", 380),
    ("Chef's Special Biryani", "Biryani", "Lollipop Biryani", 380),
    ("Chef's Special Biryani", "Biryani", "Murgh Musalam Biryani", 400),
    ("Chef's Special Biryani", "Biryani", "Ulavacharu Chicken Biryani", 370),
    ("Chef's Special Biryani", "Biryani", "Ulavacharu Mutton Biryani", 440),
    ("Chef's Special Biryani", "Biryani", "Potlam Biryani", 380),
    # TKG Chef Special Rayalaseema
    ("TKG Chef Special Rayalaseema", "Main", "Mutton Kheema Ragi Mudda", 260),
    ("TKG Chef Special Rayalaseema", "Main", "Chicken Kheema Ragi Mudda", 200),
    ("TKG Chef Special Rayalaseema", "Main", "Natukodi Pulusu", 440),
    ("TKG Chef Special Rayalaseema", "Main", "Munnakai Pulusu with Ragi Mudda", 120),
    # Kichidi Items
    ("Kichidi Items", "Main", "Paneer Kheema Kichidi", 260),
    ("Kichidi Items", "Main", "Egg Kheema Kichidi", 230),
    ("Kichidi Items", "Main", "Mutton Kheema Kichidi", 280),
    ("Kichidi Items", "Main", "Chicken Kheema Kichidi", 250),
    # Mocktails
    ("Mocktails", "Beverages", "Rainbow", 170),
    ("Mocktails", "Beverages", "Dragon", 170),
    ("Mocktails", "Beverages", "Blue Lagoon", 160),
    ("Mocktails", "Beverages", "Dream Kiss", 170),
    ("Mocktails", "Beverages", "Virgin Mojito", 160),
    # Smoothies & Shakes
    ("Smoothies & Shakes", "Beverages", "Fruits Smoothie", 160),
    ("Smoothies & Shakes", "Beverages", "Vanilla Milkshake", 150),
    ("Smoothies & Shakes", "Beverages", "Banana Milkshake", 140),
    ("Smoothies & Shakes", "Beverages", "Strawberry Milkshake", 150),
    ("Smoothies & Shakes", "Beverages", "Chocolate Milkshake", 160),
    ("Smoothies & Shakes", "Beverages", "Oreo Milkshake", 160),
    ("Smoothies & Shakes", "Beverages", "Apple Milkshake", 160),
    ("Smoothies & Shakes", "Beverages", "Caramel Milkshake", 200),
    # Fresh Juice
    ("Fresh Juice", "Beverages", "Apple Juice", 100),
    ("Fresh Juice", "Beverages", "Pineapple Juice", 100),
    ("Fresh Juice", "Beverages", "Orange Juice", 100),
    ("Fresh Juice", "Beverages", "Pomegranate Juice", 110),
    ("Fresh Juice", "Beverages", "Watermelon Juice", 100),
    ("Fresh Juice", "Beverages", "Mix Fruit Juice", 120),
    # Beverages
    ("Beverages", "Beverages", "Fresh Lime Soda Sweet", 70),
    ("Beverages", "Beverages", "Fresh Lime Soda Salt", 70),
    ("Beverages", "Beverages", "Butter Milk", 60),
    ("Beverages", "Beverages", "Lassi Sweet", 80),
    ("Beverages", "Beverages", "Lassi Salt", 80),
    # TKG Special Sweets
    ("TKG Special Sweets", "Desserts", "Arabian Delight", 180),
    ("TKG Special Sweets", "Desserts", "Apricot Delight", 220),
    # Desserts
    ("Desserts", "Desserts", "Banana Split Sunday", 250),
    ("Desserts", "Desserts", "Gadbad Ice Cream", 250),
    ("Desserts", "Desserts", "Honey Moon Delight", 250),
    ("Desserts", "Desserts", "Triple Sunday", 250),
    ("Desserts", "Desserts", "Special Chocolate Sunday", 250),
    ("Desserts", "Desserts", "Fruits Salad", 140),
    ("Desserts", "Desserts", "Fruit Salad with Ice Cream", 220),
]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dine-In Menu"

    headers = ["Category", "Subcategory", "Item Name", "Price"]
    ws.append(headers)

    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="2E86AB")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    for col in range(1, 5):
        c = ws.cell(row=1, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center
        c.border = border

    for cat, sub, item, price in ROWS:
        ws.append([cat, sub, item, price])
        r = ws.max_row
        for col in range(1, 5):
            ws.cell(row=r, column=col).border = border
        ws.cell(row=r, column=4).alignment = right

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 10
    ws.freeze_panes = "A2"

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {len(ROWS)} rows to {OUT_PATH}")


if __name__ == "__main__":
    main()
