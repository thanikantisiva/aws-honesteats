#!/usr/bin/env python3
"""
Enrich WhatsApp menu workbook and build menu POST bodies (no HTTP).

Reads:  ~/Downloads/whatsapp_menu_items.xlsx
        Cols: A=Category, B=Item Name, C=Price, D=Hike

Enrichment (only fills blanks):
  E: subCategory
  H: isVeg

JSON body per row:
  {
    "name": B, "restaurantPrice": C, "hikePercentage": D,
    "category": A, "subCategory": E, "isVeg": H, "isAvailable": true
  }

Usage:
  python3 scripts/whatsapp_menu_bodies.py
  python3 scripts/whatsapp_menu_bodies.py --xlsx ... --out-json ...
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_XLSX = Path.home() / "Downloads" / "whatsapp_menu_items.xlsx"
DEFAULT_RESTAURANT_ID = "RES-1777091893663-3064"
DEFAULT_OUT_JSON = Path.home() / "Downloads" / f"whatsapp_menu_bodies_{DEFAULT_RESTAURANT_ID}.json"

# ── Per-item overrides: (category, subCategory, isVeg) ────────────────────────
# isVeg: True = veg, False = non-veg
ITEM_META: dict[str, tuple[str, str, bool]] = {
    # Burgers
    "Veg Burger":                               ("Burgers", "Veg Burgers", True),
    "Chicken Burger":                           ("Burgers", "Chicken Burgers", False),
    "Chicken Cheese Burger":                    ("Burgers", "Chicken Burgers", False),
    "Double Slice Burger":                      ("Burgers", "Premium Burgers", False),
    "Royal Burger":                             ("Burgers", "Premium Burgers", False),
    # Pizza — Veg
    "Corn Pizza (Medium) [9 inches]":           ("Pizza", "Veg Pizza", True),
    "Corn Pizza (Large) [10 inches]":           ("Pizza", "Veg Pizza", True),
    "Veg Pizza (Medium) [9 inches]":            ("Pizza", "Veg Pizza", True),
    "Veg Pizza (Large) [10 inches]":            ("Pizza", "Veg Pizza", True),
    "Paneer Pizza (Medium) [9 inches]":         ("Pizza", "Veg Pizza", True),
    "Paneer Pizza (Large) [10 inches]":         ("Pizza", "Veg Pizza", True),
    "Mushroom Pizza (Medium) [9 inches]":       ("Pizza", "Veg Pizza", True),
    "Mushroom Pizza (Large) [10 inches]":       ("Pizza", "Veg Pizza", True),
    # Pizza — Non-Veg
    "Chicken Pizza (Medium) [9 inches]":        ("Pizza", "Non-Veg Pizza", False),
    "Chicken Pizza (Large) [10 inches]":        ("Pizza", "Non-Veg Pizza", False),
    "Tandoori Pizza (Medium) [9 inches]":       ("Pizza", "Non-Veg Pizza", False),
    "Tandoori Pizza (Large) [10 inches]":       ("Pizza", "Non-Veg Pizza", False),
    "Fried Chicken Pizza (Medium) [9 inches]":  ("Pizza", "Non-Veg Pizza", False),
    "Fried Chicken Pizza (Large) [10 inches]":  ("Pizza", "Non-Veg Pizza", False),
    # Fried Chicken
    "1 Crispy Pc":                              ("Fried Chicken", "Crispy Chicken", False),
    "3 Crispy Pcs":                             ("Fried Chicken", "Crispy Chicken", False),
    "5 Crispy Pcs":                             ("Fried Chicken", "Crispy Chicken", False),
    "10 Crispy Pcs":                            ("Fried Chicken", "Crispy Chicken", False),
    "8 Wings":                                  ("Fried Chicken", "Wings & Lollipops", False),
    "5 Lollipops":                              ("Fried Chicken", "Wings & Lollipops", False),
    "6 Boneless Strips":                        ("Fried Chicken", "Strips & Popcorn", False),
    "Chicken Popcorn":                          ("Fried Chicken", "Strips & Popcorn", False),
    # Sandwich
    "Veg Sandwich":                             ("Sandwich", "Veg Sandwich", True),
    "Veg Cheese Sandwich":                      ("Sandwich", "Veg Sandwich", True),
    "Chicken Sandwich":                         ("Sandwich", "Non-Veg Sandwich", False),
    "Paneer Sandwich":                          ("Sandwich", "Veg Sandwich", True),
}


def _clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _title_case(value: Any) -> str:
    text = _clean(value)
    return " ".join(part[:1].upper() + part[1:] for part in text.split(" ")) if text else ""


def _parse_price(value: Any) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    s = _clean(value).replace("₹", "").replace(",", "")
    return float(s)


def _parse_hike(value: Any) -> float:
    if value is None or _clean(value) == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return float(_clean(value).replace("%", ""))


def _parse_is_veg(value: Any) -> bool | None:
    if value is None or _clean(value) == "":
        return None
    if isinstance(value, bool):
        return value
    s = _clean(value).lower()
    if s in {"true", "veg", "v", "yes", "1", "vegetarian"}:
        return True
    if s in {"false", "non-veg", "nonveg", "nv", "no", "0"}:
        return False
    return None


def enrich_workbook(path: Path) -> None:
    wb = openpyxl.load_workbook(path)
    ws = wb[wb.sheetnames[0]]

    ws.cell(1, 5, "subCategory")
    ws.cell(1, 8, "isVeg")

    for row in range(2, (ws.max_row or 1) + 1):
        name = _clean(ws.cell(row, 2).value)
        if not name:
            continue

        meta = ITEM_META.get(name)
        if meta:
            cat, sub, is_veg = meta
            ws.cell(row, 1, cat)
            if _clean(ws.cell(row, 5).value) == "":
                ws.cell(row, 5, sub)
            if _parse_is_veg(ws.cell(row, 8).value) is None:
                ws.cell(row, 8, is_veg)
        else:
            category = _clean(ws.cell(row, 1).value)
            if _clean(ws.cell(row, 5).value) == "":
                ws.cell(row, 5, category)
            if _parse_is_veg(ws.cell(row, 8).value) is None:
                ws.cell(row, 8, True)

    wb.save(path)
    wb.close()


def build_bodies(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    bodies: list[dict[str, Any]] = []

    for row in range(2, (ws.max_row or 1) + 1):
        name = _clean(ws.cell(row, 2).value)
        category = _clean(ws.cell(row, 1).value)
        if not name:
            continue

        is_veg = _parse_is_veg(ws.cell(row, 8).value)
        body: dict[str, Any] = {
            "name": _title_case(name),
            "restaurantPrice": _parse_price(ws.cell(row, 3).value),
            "hikePercentage": _parse_hike(ws.cell(row, 4).value),
            "category": _title_case(category),
            "subCategory": _title_case(ws.cell(row, 5).value) or _title_case(category),
            "isVeg": is_veg if is_veg is not None else True,
            "isAvailable": True,
        }
        bodies.append(body)

    wb.close()
    return bodies


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build WhatsApp menu JSON bodies without inserting.")
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--out-json", type=Path, default=DEFAULT_OUT_JSON)
    p.add_argument("--restaurant-id", default=DEFAULT_RESTAURANT_ID)
    p.add_argument("--array-only", action="store_true",
                    help='Write bare array instead of {"restaurantId","items"}')
    return p.parse_args()


def main() -> int:
    args = parse_args()
    if not args.xlsx.is_file():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        return 1

    enrich_workbook(args.xlsx)
    bodies = build_bodies(args.xlsx)

    out_obj: Any
    if args.array_only:
        out_obj = bodies
    else:
        out_obj = {"restaurantId": args.restaurant_id, "items": bodies}

    args.out_json.parent.mkdir(parents=True, exist_ok=True)
    args.out_json.write_text(json.dumps(out_obj, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    veg = sum(1 for b in bodies if b["isVeg"])
    non_veg = len(bodies) - veg
    print(f"Enriched workbook: {args.xlsx}")
    print(f"Restaurant: {args.restaurant_id}")
    print(f"Bodies: {len(bodies)} (Veg: {veg}, Non-Veg: {non_veg})")
    print(f"Wrote JSON: {args.out_json}")
    print("No HTTP calls were made.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
