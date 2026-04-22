"""Add Description column to ~/Downloads/menu_category_subcategory_item_price_inr.xlsx.

Re-run after editing DESCRIPTIONS: python3 scripts/menu_inr_add_descriptions.py
"""
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Keyed by (Category, Subcategory, Item) for disambiguation (e.g. Oreo thickshake vs milkshake).
DESCRIPTIONS = {
    ("Fried Chicken", "-", "Crispy Chicken (1 PC)"): "Single piece of golden, extra-crispy fried chicken.",
    ("Fried Chicken", "-", "Crispy Chicken (3 PC)"): "Three pieces of signature crispy fried chicken.",
    ("Fried Chicken", "-", "Crispy Chicken (5 PC)"): "Five pieces of crispy fried chicken—great for sharing.",
    ("Fried Chicken", "-", "Crispy Wings (5 PC)"): "Five crispy chicken wings, fried until crunchy outside and juicy inside.",
    ("Fried Chicken", "-", "Boneless Strips (5)"): "Five boneless chicken strips with a crisp coating.",
    ("Fried Chicken", "-", "Popcorn (Regular)"): "Bite-sized crispy chicken popcorn in a regular portion.",
    ("Fried Chicken", "-", "Popcorn (Large)"): "Bite-sized crispy chicken popcorn in a larger portion.",
    ("Indo American Chicken", "-", "Indo American Chicken (1 PC)"): "One piece of Indo–American style seasoned fried chicken.",
    ("Indo American Chicken", "-", "Indo American Chicken (2 PC)"): "Two pieces of Indo–American style fried chicken.",
    ("Indo American Chicken", "-", "Indo American Chicken (5 PC)"): "Five pieces of Indo–American style fried chicken.",
    ("Indo American Chicken", "-", "Indo American Wings (5 PC)"): "Five wings tossed in Indo–American style seasoning.",
    ("Indo American Chicken", "-", "Indo American Strips (4 PC)"): "Four boneless strips with Indo–American flavour profile.",
    ("Nashville Hot Chicken", "-", "Nashville Hot Chicken (1 PC)"): "Single piece of Nashville-style hot, spicy fried chicken.",
    ("Nashville Hot Chicken", "-", "Nashville Hot Chicken (2 PC)"): "Two pieces of Nashville hot fried chicken.",
    ("Nashville Hot Chicken", "-", "Nashville Hot Chicken (5 PC)"): "Five pieces of Nashville hot fried chicken.",
    ("Nashville Hot Chicken", "-", "Nashville Wings (5 PC)"): "Five spicy Nashville-style hot wings.",
    ("Nashville Hot Chicken", "-", "Nashville Strips (5 PC)"): "Five Nashville hot boneless chicken strips.",
    ("Nashville Hot Chicken", "-", "Juicy Wings (5 PC)"): "Five wings—juicy inside with a seasoned, spicy crust.",
    ("Nashville Hot Chicken", "-", "Juicy Strips (4 PC)"): "Four juicy boneless strips with bold seasoning.",
    ("Veg Burger", "With cheese", "Classic Veg Burger"): "Classic veg patty with cheese, veggies, and house sauces in a soft bun.",
    ("Veg Burger", "Without cheese", "Classic Veg Burger"): "Classic veg patty with fresh veggies and sauces—no cheese.",
    ("Veg Burger", "With cheese", "Bro Veg Burger"): "Loaded veg burger with cheese, premium toppings, and signature sauces.",
    ("Veg Burger", "Without cheese", "Bro Veg Burger"): "Loaded veg burger with premium toppings and sauces—no cheese.",
    ("Veg Burger", "With cheese", "Paneer Burger"): "Grilled or crumbed paneer with cheese and fresh fixings in a bun.",
    ("Non-Veg Sandwich", "With cheese", "Classic Chicken Sandwich"): "Classic grilled or fried chicken sandwich with melted cheese.",
    ("Non-Veg Sandwich", "Without cheese", "Classic Chicken Sandwich"): "Classic chicken sandwich with veggies and sauces—no cheese.",
    ("Non-Veg Sandwich", "With cheese", "Bro Chicken Sandwich"): "Hearty chicken sandwich with cheese and extra fillings.",
    ("Non-Veg Sandwich", "Without cheese", "Bro Chicken Sandwich"): "Hearty chicken sandwich packed with toppings—no cheese.",
    ("Non-Veg Sandwich", "With cheese", "Madmax Chicken Sandwich"): "Maxed-out chicken sandwich with cheese and bold flavours.",
    ("Non-Veg Sandwich", "Without cheese", "Madmax Chicken Sandwich"): "Maxed-out chicken sandwich without cheese.",
    ("Non-Veg Sandwich", "-", "Spice Hot Sauce"): "Extra side of spicy hot sauce for dipping or drizzling.",
    ("Veg Sandwich", "With cheese", "Classic Veg Sandwich"): "Classic veg fillings with cheese between toasted bread.",
    ("Veg Sandwich", "Without cheese", "Classic Veg Sandwich"): "Classic veg sandwich with fresh vegetables—no cheese.",
    ("Veg Sandwich", "With cheese", "Paneer Sandwich"): "Paneer, veggies, and cheese in a grilled or cold sandwich.",
    ("Veg Sandwich", "Without cheese", "Paneer Sandwich"): "Paneer and vegetable sandwich without cheese.",
    ("Veg Sandwich", "With cheese", "Bro Veg Sandwich"): "Generous veg sandwich with cheese and signature sauces.",
    ("Veg Sandwich", "Without cheese", "Bro Veg Sandwich"): "Generous veg sandwich with sauces—no cheese.",
    ("Wraps", "-", "Classic Chicken Wrap"): "Classic chicken, veggies, and sauce wrapped in a soft tortilla.",
    ("Wraps", "-", "Nashville Chicken Wrap"): "Nashville hot chicken with cool crunch wrapped for on-the-go eating.",
    ("Wraps", "-", "Indo American Chicken Wrap"): "Indo–American seasoned chicken with fresh fillings in a wrap.",
    ("Wraps", "-", "Veg Wrap"): "Mixed veg patty or grilled veggies with sauces in a tortilla wrap.",
    ("Mojitos", "-", "Deep Blue Sea"): "Refreshing blue curaçao–style mojito mocktail with lime and mint.",
    ("Mojitos", "-", "Virgin"): "Classic virgin mojito—lime, mint, and soda, no alcohol.",
    ("Mojitos", "-", "Melon"): "Cool melon-flavoured mojito with mint and citrus.",
    ("Mojitos", "-", "Raspberry"): "Sweet-tart raspberry mojito mocktail.",
    ("Mojitos", "-", "Blackberry"): "Berry-forward blackberry mojito with mint.",
    ("Mojitos", "-", "Mango"): "Tropical mango mojito with lime and mint.",
    ("Mojitos", "-", "Orange"): "Citrusy orange mojito mocktail.",
    ("Combos", "-", "Student Combo"): "Value combo curated for a filling meal on a budget.",
    ("Combos", "-", "Friendship Combo"): "Larger combo ideal for two or more to share.",
    ("Combos", "-", "BIG BRO Meal"): "Signature combo meal with mains, sides, and a drink-style pairing.",
    ("Combos", "-", "BRO Veg Meal"): "Veg combo with burger/sandwich-style main, side, and drink.",
    ("Combos", "-", "BRO CKN Meal"): "Chicken combo meal with crispy or saucy chicken, side, and beverage.",
    ("Waffles", "-", "Kitkat Waffle"): "Crisp waffle topped with KitKat pieces and chocolate drizzle.",
    ("Waffles", "-", "Dark & White"): "Waffle with dark and white chocolate sauces.",
    ("Waffles", "-", "Triple Chocolate"): "Waffle loaded with three chocolate elements—rich and indulgent.",
    ("Waffles", "-", "Naked Nutella"): "Warm waffle with generous Nutella—simple and decadent.",
    ("Waffles", "-", "Chocolate Overload"): "Chocolate-on-chocolate waffle for dessert lovers.",
    ("Fries", "-", "Regular (Salted)"): "Classic salted French fries—regular size.",
    ("Fries", "-", "Regular (Salted) Large"): "Classic salted French fries in a large portion.",
    ("Fries", "-", "Peri Peri Fries Small"): "Spicy peri peri seasoned fries—small.",
    ("Fries", "-", "Peri Peri Fries Large"): "Spicy peri peri seasoned fries—large.",
    ("Fries", "-", "Ckn Loaded Fries"): "Fries topped with chicken, sauces, and cheese-style drizzle.",
    ("Dips", "-", "Garlic Mayo Dip"): "Creamy garlic mayonnaise dip.",
    ("Dips", "-", "Sweet & Spicy Mayo"): "Mayo balanced with sweet heat—great for fries and strips.",
    ("Dips", "-", "Extra Cheese"): "Additional warm cheese sauce for burgers or fries.",
    ("Dips", "-", "Tandoori"): "Smoky tandoori-style dip for chicken and fries.",
    ("Dips", "-", "Island Dip"): "Tangy island-style dip with a hint of sweetness.",
    ("Thickshakes", "-", "Oreo"): "Extra-thick shake blended with Oreo cookies and ice cream.",
    ("Thickshakes", "-", "Kitkat"): "Dense shake with KitKat chunks blended in.",
    ("Thickshakes", "-", "Fivestar"): "Thick shake with Five Star chocolate bar pieces.",
    ("Thickshakes", "-", "Snickers"): "Thick shake with Snickers—caramel, nougat, and nuts in every sip.",
    ("Thickshakes", "-", "Chocochip Nutella"): "Thick Nutella-based shake with chocolate chips.",
    ("Thickshakes", "-", "Oreo Nutella"): "Oreo and Nutella combined in an ultra-thick shake.",
    ("Thickshakes", "-", "Peanut Butter"): "Rich peanut butter thickshake.",
    ("Thickshakes", "-", "P-Butter Nutella"): "Peanut butter and Nutella swirled into a thick shake.",
    ("Thickshakes", "-", "Caramel Nuts"): "Caramel, nuts, and ice cream in a thick blended shake.",
    ("Thickshakes", "-", "Strawberry"): "Strawberry thickshake—creamy and fruity.",
    ("Thickshakes", "-", "Blackcurrent"): "Blackcurrant-flavoured thick, creamy shake.",
    ("Thickshakes", "-", "Mango"): "Mango thickshake with tropical sweetness.",
    ("Milkshakes", "-", "Oreo"): "Classic milkshake blended with Oreo cookies.",
    ("Milkshakes", "-", "Kitkat"): "Milkshake with KitKat blended for crunch and chocolate.",
    ("Milkshakes", "-", "Fivestar"): "Milkshake with Five Star chocolate blended in.",
    ("Milkshakes", "-", "Snickers"): "Milkshake with Snickers pieces—caramel and nut notes.",
    ("Milkshakes", "-", "Chocochip Nutella"): "Nutella milkshake with chocolate chips.",
    ("Milkshakes", "-", "Oreo Nutella"): "Oreo and Nutella milkshake—lighter than thickshake.",
    ("Milkshakes", "-", "Peanut Butter"): "Creamy peanut butter milkshake.",
    ("Milkshakes", "-", "P-Butter Nutella"): "Peanut butter and Nutella milkshake.",
    ("Milkshakes", "-", "Caramel Nuts"): "Caramel and nut milkshake.",
    ("Milkshakes", "-", "Strawberry"): "Strawberry ice cream milkshake.",
    ("Milkshakes", "-", "Blackcurrent"): "Blackcurrant milkshake.",
    ("Milkshakes", "-", "Mango"): "Mango milkshake—smooth and refreshing.",
}


def main() -> None:
    path = Path.home() / "Downloads" / "menu_category_subcategory_item_price_inr.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Shift Price column from D to E
    max_r = ws.max_row
    ws.insert_cols(4, 1)  # new empty column D for Description

    ws.cell(row=1, column=4, value="Description")
    hdr = ws.cell(row=1, column=4)
    hdr.font = Font(bold=True, color="FFFFFF")
    hdr.fill = PatternFill("solid", fgColor="4472C4")
    hdr.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    missing = []
    for r in range(2, max_r + 1):
        cat = ws.cell(r, 1).value
        sub = ws.cell(r, 2).value
        item = ws.cell(r, 3).value
        key = (cat, sub, item)
        desc = DESCRIPTIONS.get(key)
        if desc is None:
            missing.append(key)
            desc = ""
        ws.cell(row=r, column=4, value=desc)

    if missing:
        raise SystemExit(f"Missing descriptions for {len(missing)} rows: {missing[:5]}...")

    for col in range(1, 6):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 22 if col != 4 else 52
        if col == 5:
            ws.column_dimensions[letter].width = 14

    for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(path)
    print(f"Updated {path} (Category | Subcategory | Item | Description | Price)")


if __name__ == "__main__":
    main()
