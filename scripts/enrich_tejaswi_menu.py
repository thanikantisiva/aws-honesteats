#!/usr/bin/env python3
"""
Enrich Tejaswi menu Excel and emit menu-item JSON; optional POST to API.

Reads:  ~/Downloads/Tejaswi_DineIn_vs_Zomato.xlsx
        Current cols A..N (from pricing script):
          A=Category, B=Subcategory, C=Item,
          D=Dine-In (\u20b9), E=Zomato (\u20b9),
          F=Zomato Hike %, G=R_z,
          H=Our Recommended Price P, I=Our Hike on Dine-in %,
          J=Rest from Us, K=Our Revenue,
          L=Customer saves vs Zomato (\u20b9), M=Customer saves vs Zomato %,
          N=Restaurant gains vs Zomato (\u20b9)

After enrichment, the workbook is rewritten so that:
  A=Category, B=Subcategory, C=Item (name),
  D=restaurantPrice (Our Recommended Price P),
  E=hikePercentage (Our Hike on Dine-in %),
  F=Dine-In (\u20b9), G=Zomato (\u20b9),
  H=isVeg, I=image,
  J=Zomato Hike %, K=R_z,
  L=description,
  M=Restaurant gets from Us, N=Our Revenue,
  O=Customer saves vs Zomato (\u20b9), P=Customer saves vs Zomato %,
  Q=Restaurant gains vs Zomato (\u20b9).

The JSON body matches what the user asked for:
  {
    "name": C, "restaurantPrice": D, "hikePercentage": E,
    "category": A, "subCategory": B, "isVeg": H,
    "isAvailable": true, "description": L, "image": [I]
  }

Writes JSON \u2192 ~/Downloads/Tejaswi_Menu.menu_items.json
Restaurant ID: RES-1776687331858-4639

Usage:
  python3 scripts/enrich_tejaswi_menu.py                     # enrich + JSON only (dry-run, default)
  python3 scripts/enrich_tejaswi_menu.py --apply             # enrich + POST to API
  python3 scripts/enrich_tejaswi_menu.py --from-json --apply # POST from existing JSON
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_PATH = Path.home() / "Downloads" / "Tejaswi_DineIn_vs_Zomato.xlsx"
JSON_PATH = Path.home() / "Downloads" / "Tejaswi_Menu.menu_items.json"
RESTAURANT_ID = "RES-1776687331858-4639"

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get(
    "HONESTEATS_RETOOL_BYPASS",
    "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc",
)
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

# ── Source column indices (current file layout) ───────────────────────────────
SRC_CAT, SRC_SUB, SRC_NAME = 1, 2, 3
SRC_DINE, SRC_ZOM = 4, 5
SRC_Z_HIKE, SRC_RZ, SRC_P, SRC_OUR_HIKE = 6, 7, 8, 9
SRC_REST_US, SRC_REV, SRC_SAVE, SRC_SAVE_PCT, SRC_GAIN = 10, 11, 12, 13, 14

# ── Target (rewritten) column indices ─────────────────────────────────────────
T_CAT, T_SUB, T_NAME = 1, 2, 3
T_PRICE, T_HIKE = 4, 5
T_DINE, T_ZOM = 6, 7
T_VEG, T_IMG = 8, 9
T_Z_HIKE, T_RZ = 10, 11
T_DESC = 12
T_REST_US, T_REV = 13, 14
T_SAVE, T_SAVE_PCT, T_GAIN = 15, 16, 17

TARGET_HEADERS = [
    "Category",                                # A
    "Subcategory",                             # B
    "Item",                                    # C
    "restaurantPrice (\u20b9)",                # D  ← Our Recommended Price P
    "hikePercentage (%)",                      # E  ← Our Hike on Dine-in %
    "Dine-In (\u20b9)",                        # F
    "Zomato (\u20b9)",                         # G
    "isVeg",                                   # H
    "image",                                   # I
    "Zomato Hike %",                           # J
    "R_z (\u20b9)",                            # K
    "description",                             # L
    "Restaurant gets from Us (\u20b9)",        # M
    "Our Revenue (per order) (\u20b9)",        # N
    "Customer saves vs Zomato (\u20b9)",       # O
    "Customer saves vs Zomato %",              # P
    "Restaurant gains vs Zomato (\u20b9)",     # Q
]

_BRAND = "Tejaswi Multicuisine"


# ── Unsplash photo pools (real IDs) ───────────────────────────────────────────
# Each value is an Unsplash photo slug (the part after "photo-" in the URL).
_POOLS: dict[str, list[str]] = {
    # Soups / starters — broths, bowls
    "soup_veg": [
        "1547592180-85f173990554", "1603105037880-880cd4edfb0d",
        "1583515470650-35cbbdeef9b3", "1476718406336-bb5a9690ee2a",
        "1604152135912-04a022e23696",
    ],
    "soup_nonveg": [
        "1547308283-b0efd5ff1be6", "1583515470650-35cbbdeef9b3",
        "1608500218890-c4f9077f79a1", "1515003197210-e0cd71810b5f",
    ],
    # Chicken starters / tandoori / dry
    "starter_chicken": [
        "1626082927389-6cd097cdc6ec", "1562967914-608f82629710",
        "1587593810167-a84920ea0781", "1604908176997-125f25cc6f3d",
        "1606755962773-d324e0a13086", "1598515214211-89d3c73ae83b",
        "1513185041617-8ab03f83d6c5", "1565557623262-b51c2513a641",
        "1585937421612-70a008356fbe", "1628294895950-9805252327bc",
    ],
    "starter_mutton": [
        "1599487488170-d11ec9c172f0", "1633237308525-cd587cf71926",
        "1631452180519-c014fe946bc7", "1589302168068-964664d93dc0",
        "1574484284002-952d92456975",
    ],
    "starter_fish": [
        "1535400875775-96ea15902b76", "1519708227418-c8fd9a32b7a2",
        "1604909052743-94e838986d24", "1467003909585-2f8a72700288",
        "1580959375944-abd7e991f971",
    ],
    "starter_prawns": [
        "1565680018434-b513d5e5fd47", "1625944228741-81ec20f8e34b",
        "1599021456807-25db0f974333", "1582845512747-e42001c95638",
    ],
    "starter_veg": [
        "1606491956689-2ea866880c84", "1585937421612-70a008356fbe",
        "1565299624946-b28f40a0ae38", "1606755962773-d324e0a13086",
        "1604152135912-04a022e23696", "1626700051175-f5d1f5e9c07a",
    ],
    "starter_paneer": [
        "1567337711067-af5f22bf3f0f", "1606491956689-2ea866880c84",
        "1601050690597-df0568f70950", "1630383249896-424e482df921",
        "1565557623262-b51c2513a641",
    ],
    "starter_gobi": [
        "1604908176997-125f25cc6f3d", "1600891964092-4316c288032e",
        "1626700051175-f5d1f5e9c07a",
    ],
    "starter_mushroom": [
        "1589302168068-964664d93dc0", "1519996529931-28324d5a630e",
        "1565299507177-b0ac66763828",
    ],
    "starter_babycorn": [
        "1573080496219-bb080dd4f877", "1630384060421-cb20d0e0649d",
    ],
    "starter_egg": [
        "1604908176997-125f25cc6f3d", "1513104890138-7c749659a591",
        "1467003909585-2f8a72700288",
    ],
    # Kebabs
    "kebab_chicken": [
        "1599487488170-d11ec9c172f0", "1585937421612-70a008356fbe",
        "1628294895950-9805252327bc", "1565299543923-37dd37887442",
        "1574484284002-952d92456975",
    ],
    # Curries
    "curry_veg": [
        "1565557623262-b51c2513a641", "1585937421612-70a008356fbe",
        "1596797038530-2c107229654b", "1601050690597-df0568f70950",
        "1589302168068-964664d93dc0",
    ],
    "curry_paneer": [
        "1567337711067-af5f22bf3f0f", "1601050690597-df0568f70950",
        "1565557623262-b51c2513a641", "1596797038530-2c107229654b",
    ],
    "curry_chicken": [
        "1565557623262-b51c2513a641", "1603105037303-8ca2d8a0acad",
        "1631452180519-c014fe946bc7", "1596797038530-2c107229654b",
        "1585937421612-70a008356fbe",
    ],
    "curry_mutton": [
        "1633237308525-cd587cf71926", "1631452180519-c014fe946bc7",
        "1589302168068-964664d93dc0", "1603105037303-8ca2d8a0acad",
    ],
    "curry_fish": [
        "1580959375944-abd7e991f971", "1535400875775-96ea15902b76",
        "1604909052743-94e838986d24",
    ],
    "curry_prawn": [
        "1625944228741-81ec20f8e34b", "1582845512747-e42001c95638",
    ],
    "curry_egg": [
        "1604908176997-125f25cc6f3d", "1565557623262-b51c2513a641",
    ],
    # Breads
    "bread_naan": [
        "1601050690597-df0568f70950", "1589302168068-964664d93dc0",
        "1565557623262-b51c2513a641", "1606491956689-2ea866880c84",
    ],
    "bread_roti": [
        "1601050690597-df0568f70950", "1589302168068-964664d93dc0",
    ],
    "bread_kulcha": [
        "1589302168068-964664d93dc0", "1630383249896-424e482df921",
    ],
    "bread_paratha": [
        "1601050690597-df0568f70950", "1630383249896-424e482df921",
    ],
    # Rice / Biryani / Pulao
    "rice_fried_veg": [
        "1563245372-f21724e3856d", "1603133872878-684f208fb84b",
        "1596797038530-2c107229654b",
    ],
    "rice_fried_chicken": [
        "1603133872878-684f208fb84b", "1565557623262-b51c2513a641",
        "1631452180519-c014fe946bc7",
    ],
    "rice_fried_egg": [
        "1603133872878-684f208fb84b", "1563245372-f21724e3856d",
    ],
    "rice_plain": [
        "1596797038530-2c107229654b", "1601050690597-df0568f70950",
    ],
    "biryani_veg": [
        "1604908176997-125f25cc6f3d", "1589302168068-964664d93dc0",
        "1565557623262-b51c2513a641", "1601050690597-df0568f70950",
    ],
    "biryani_chicken": [
        "1563379091339-03b21ab4a4f4", "1631452180519-c014fe946bc7",
        "1603105037303-8ca2d8a0acad", "1585937421612-70a008356fbe",
        "1565557623262-b51c2513a641", "1596797038530-2c107229654b",
    ],
    "biryani_mutton": [
        "1633237308525-cd587cf71926", "1631452180519-c014fe946bc7",
        "1603105037303-8ca2d8a0acad",
    ],
    "biryani_fish": [
        "1580959375944-abd7e991f971", "1535400875775-96ea15902b76",
    ],
    "biryani_prawn": [
        "1625944228741-81ec20f8e34b", "1582845512747-e42001c95638",
    ],
    "biryani_egg": [
        "1604908176997-125f25cc6f3d", "1565557623262-b51c2513a641",
    ],
    "pulao_veg": [
        "1596797038530-2c107229654b", "1563245372-f21724e3856d",
    ],
    "pulao_chicken": [
        "1603133872878-684f208fb84b", "1631452180519-c014fe946bc7",
    ],
}

_NOT_FOUND_POOL = "starter_chicken"  # safe default fallback


def _photo_url(pid: str, sig: int | None = None) -> str:
    base = f"https://images.unsplash.com/photo-{pid}?auto=format&fit=crop&w=800&q=80"
    return base if sig is None else f"{base}&sig={sig}"


# Build a global unique list so we can fall back when a themed pool is exhausted.
_GLOBAL: list[str] = []
_seen: set[str] = set()
for _pool, _ids in _POOLS.items():
    for _pid in _ids:
        if _pid not in _seen:
            _seen.add(_pid)
            _GLOBAL.append(_pid)


class _Picker:
    """Dedupes photo IDs across rows; falls back to signed variants if exhausted."""

    def __init__(self) -> None:
        self.used: set[str] = set()
        self._sig = 0

    def _themed(self, key: str) -> str | None:
        for pid in _POOLS.get(key, []):
            if pid not in self.used:
                self.used.add(pid)
                return _photo_url(pid)
        return None

    def _global(self) -> str | None:
        for pid in _GLOBAL:
            if pid not in self.used:
                self.used.add(pid)
                return _photo_url(pid)
        return None

    def _recycled(self) -> str:
        self._sig += 1
        pid = _GLOBAL[self._sig % len(_GLOBAL)]
        return _photo_url(pid, sig=self._sig)

    def pick(self, key: str) -> str:
        return self._themed(key) or self._themed(_NOT_FOUND_POOL) or self._global() or self._recycled()


# ── Veg / non-veg classification ──────────────────────────────────────────────
_NV_NAME_KW = (
    "chicken", "mutton", "fish", "prawn", "prawns", "egg", "beef", "pork",
    "kamju", "kumju",          # shark
    "natu kodi", "kodi",       # country chicken
    "murg", "tandoori",        # tandoori chicken / murg malai
    "kebab", "kabab",
    "kheema", "keema",
)
_VEG_HINTS_IN_NAME = ("veg ", "paneer", "gobi", "mushroom", "baby corn", "kaju", "palak")


def _is_nonveg(cat: str, sub: str, name: str) -> bool:
    cl = (cat or "").lower()
    sl = (sub or "").lower()
    nl = (name or "").lower()

    # Subcategory signals.
    if any(tag in sl for tag in ("non-veg", "non veg", "nonveg")):
        return True
    if sl.strip() == "veg":
        # ...unless name contains a non-veg keyword (safety).
        if any(k in nl for k in _NV_NAME_KW):
            return True
        return False
    if any(tag in sl for tag in ("chicken", "mutton", "fish", "prawn", "kamju")):
        return True

    # Category signals.
    if cl == "egg":
        return True
    if cl == "kebabs":
        return True

    # Name-based checks (biryanis, rice, breads, pulaos, noodles etc.).
    if any(k in nl for k in _NV_NAME_KW):
        # ...but an explicit veg hint wins (e.g. "Veg Biryani").
        if any(v in nl for v in _VEG_HINTS_IN_NAME) and not any(
            k in nl for k in ("chicken", "mutton", "fish", "prawn", "egg", "kamju", "kumju", "murg", "kheema", "keema")
        ):
            return False
        return True

    return False


# ── Pool selection ────────────────────────────────────────────────────────────
def select_pool(cat: str, sub: str, name: str) -> str:
    cl = (cat or "").lower()
    sl = (sub or "").lower()
    nl = (name or "").lower()

    if cl == "soups":
        return "soup_nonveg" if ("non" in sl or any(k in nl for k in ("chicken", "mutton", "prawn", "fish", "egg"))) else "soup_veg"

    if cl == "egg":
        return "starter_egg"

    if cl == "kebabs":
        return "kebab_chicken"

    if cl == "starters":
        if "chicken" in sl or "chicken" in nl or "tandoori" in nl or "murg" in nl:
            return "starter_chicken"
        if "mutton" in sl or "mutton" in nl:
            return "starter_mutton"
        if "fish" in sl or "fish" in nl:
            return "starter_fish"
        if "prawn" in sl or "prawn" in nl:
            return "starter_prawns"
        if "kamju" in sl or "kamju" in nl or "kumju" in nl:
            return "starter_fish"
        if "paneer" in nl:
            return "starter_paneer"
        if "gobi" in nl:
            return "starter_gobi"
        if "mushroom" in nl:
            return "starter_mushroom"
        if "baby corn" in nl:
            return "starter_babycorn"
        return "starter_veg"

    if cl == "curries":
        if "veg" == sl:
            if "paneer" in nl:
                return "curry_paneer"
            return "curry_veg"
        if "chicken" in sl:
            return "curry_chicken"
        if "mutton" in sl:
            return "curry_mutton"
        if "fish" in sl and "prawn" in sl:
            return "curry_prawn" if "prawn" in nl else "curry_fish"
        if "fish" in sl:
            return "curry_fish"
        if "prawn" in sl:
            return "curry_prawn"
        return "curry_veg"

    if cl.startswith("breads"):
        if "naan" in nl:
            return "bread_naan"
        if "roti" in nl:
            return "bread_roti"
        if "kulcha" in nl:
            return "bread_kulcha"
        if "paratha" in nl or "parota" in nl or "parata" in nl or "parotha" in nl:
            return "bread_paratha"
        return "bread_naan"

    if cl == "rice":
        if "chicken" in nl or "mutton" in nl or "prawn" in nl:
            return "rice_fried_chicken"
        if "egg" in nl:
            return "rice_fried_egg"
        if "fried" in nl:
            return "rice_fried_veg"
        return "rice_plain"

    if cl == "biryanis":
        if "veg" in nl and "chicken" not in nl and "mutton" not in nl and "prawn" not in nl and "fish" not in nl and "egg" not in nl and "kamju" not in nl:
            return "biryani_veg"
        if "paneer" in nl or "mushroom" in nl or "kaju" in nl:
            return "biryani_veg"
        if "mutton" in nl:
            return "biryani_mutton"
        if "fish" in nl:
            return "biryani_fish"
        if "prawn" in nl:
            return "biryani_prawn"
        if "egg" in nl:
            return "biryani_egg"
        return "biryani_chicken"

    if cl == "pulaos":
        if "chicken" in nl or "mutton" in nl or "prawn" in nl or "egg" in nl:
            return "pulao_chicken"
        return "pulao_veg"

    return "starter_veg"


# ── Description generator ─────────────────────────────────────────────────────
def build_description(cat: str, sub: str, name: str, is_veg: bool) -> str:
    cl = (cat or "").lower()
    sl = (sub or "").lower()
    nl = (name or "").lower()
    veg = "vegetarian" if is_veg else "non-vegetarian"

    if cl == "soups":
        return f"{name} \u2014 a comforting {veg} soup, richly flavoured and served piping hot at {_BRAND}."

    if cl == "starters":
        if "kamju" in sl or "kamju" in nl or "kumju" in nl:
            return f"{name} \u2014 boneless shark fillet (Kamju) tossed in our signature Andhra-style spices. {_BRAND}."
        if "chicken" in sl or "chicken" in nl:
            if "65" in nl:
                return f"{name} \u2014 classic Chicken 65 marinated in chilli, ginger-garlic and curry leaves. {_BRAND}."
            if "manchurian" in nl:
                return f"{name} \u2014 Indo-Chinese chicken tossed in a tangy soya-garlic manchurian sauce. {_BRAND}."
            if "tandoori" in nl:
                return f"{name} \u2014 yoghurt-marinated chicken chargrilled in the tandoor with aromatic spices. {_BRAND}."
            if "lollipop" in nl:
                return f"{name} \u2014 crispy chicken drumettes coated in a spicy red masala. {_BRAND}."
            if "gulzar" in nl:
                return f"{name} \u2014 our house-special Gulzar-style chicken, lightly spiced and dry-tossed. {_BRAND}."
            return f"{name} \u2014 tender chicken starter prepared in an Andhra-Chinese fusion style. {_BRAND}."
        if "mutton" in sl or "mutton" in nl:
            return f"{name} \u2014 slow-cooked mutton tossed in rich masalas for a smoky, spicy bite. {_BRAND}."
        if "fish" in sl or "fish" in nl:
            return f"{name} \u2014 fresh fish marinated in spices and fried crisp. {_BRAND}."
        if "prawn" in sl or "prawn" in nl:
            return f"{name} \u2014 juicy prawns tossed in house-blend spices. {_BRAND}."
        if "paneer" in nl:
            return f"{name} \u2014 cottage-cheese cubes marinated and tossed with peppers and Indo-Chinese spices. {_BRAND}."
        if "gobi" in nl:
            return f"{name} \u2014 crispy cauliflower florets tossed in a spicy sauce. {_BRAND}."
        if "mushroom" in nl:
            return f"{name} \u2014 button mushrooms tossed in aromatic spices and soya sauce. {_BRAND}."
        if "baby corn" in nl:
            return f"{name} \u2014 crunchy baby corn tossed in a tangy, spicy glaze. {_BRAND}."
        return f"{name} \u2014 flavourful {veg} starter from the {_BRAND} kitchen."

    if cl == "egg":
        if "bhurji" in nl or "burji" in nl:
            return f"{name} \u2014 spicy scrambled eggs with onion, tomato and green chillies. {_BRAND}."
        if "curry" in nl:
            return f"{name} \u2014 boiled eggs simmered in an Andhra-style onion-tomato masala. {_BRAND}."
        if "manchurian" in nl:
            return f"{name} \u2014 boiled eggs tossed in Indo-Chinese manchurian sauce. {_BRAND}."
        if "65" in nl:
            return f"{name} \u2014 Egg 65, chilli-coated boiled eggs with ginger, garlic and curry leaves. {_BRAND}."
        return f"{name} \u2014 egg preparation with signature house spices. {_BRAND}."

    if cl == "kebabs":
        if "tandoori" in nl:
            return f"{name} \u2014 yoghurt-marinated chicken chargrilled in the tandoor. {_BRAND}."
        if "kalmi" in nl:
            return f"{name} \u2014 Kalmi-style chicken leg kebab marinated in creamy spices and grilled. {_BRAND}."
        if "malai" in nl or "murg" in nl:
            return f"{name} \u2014 creamy Murg Malai kebab, tender and mildly spiced. {_BRAND}."
        if "tikka" in nl:
            return f"{name} \u2014 chicken tikka marinated in yoghurt and spices, chargrilled skewers. {_BRAND}."
        return f"{name} \u2014 tandoor-grilled chicken kebab, smoky and succulent. {_BRAND}."

    if cl == "curries":
        if "veg" == sl:
            if "paneer" in nl:
                return f"{name} \u2014 rich paneer curry simmered in a creamy onion-tomato gravy. {_BRAND}."
            if "kaju" in nl:
                return f"{name} \u2014 cashew-based curry in a creamy, mildly sweet gravy. {_BRAND}."
            if "mushroom" in nl:
                return f"{name} \u2014 mushroom curry in aromatic spiced gravy. {_BRAND}."
            if "palak" in nl:
                return f"{name} \u2014 spinach-based curry finished with ghee and Indian spices. {_BRAND}."
            if "hyderabadi" in nl:
                return f"{name} \u2014 Hyderabadi-style veg curry with aromatic spices. {_BRAND}."
            if "shahi" in nl or "kurma" in nl:
                return f"{name} \u2014 rich, creamy Shahi Kurma of mixed vegetables. {_BRAND}."
            return f"{name} \u2014 classic North-Indian vegetarian curry, simmered in spiced gravy. {_BRAND}."
        if "chicken" in sl:
            style = sl.replace("chicken", "").strip(" ()")
            flavour = style or "Andhra-style"
            return f"{name} \u2014 {flavour} chicken curry slow-cooked in a rich masala. {_BRAND}."
        if "mutton" in sl:
            return f"{name} \u2014 mutton on the bone slow-cooked in an aromatic Andhra masala. {_BRAND}."
        if "fish" in sl or "prawn" in sl:
            if "prawn" in nl:
                return f"{name} \u2014 juicy prawns simmered in a tangy coastal curry. {_BRAND}."
            return f"{name} \u2014 fresh fish simmered in a spicy coastal curry. {_BRAND}."
        return f"{name} \u2014 hearty {veg} curry from the {_BRAND} kitchen."

    if cl.startswith("breads"):
        if "naan" in nl:
            return f"{name} \u2014 soft tandoor-baked Indian flatbread. {_BRAND}."
        if "roti" in nl:
            return f"{name} \u2014 whole-wheat flatbread from the tandoor. {_BRAND}."
        if "kulcha" in nl:
            return f"{name} \u2014 leavened flatbread, soft and pillowy, fresh from the oven. {_BRAND}."
        if "paratha" in nl or "parota" in nl or "parata" in nl or "parotha" in nl:
            return f"{name} \u2014 flaky layered paratha, crisp on the outside, soft inside. {_BRAND}."
        return f"{name} \u2014 fresh-baked Indian bread. {_BRAND}."

    if cl == "rice":
        if "biryani rice" in nl:
            return f"{name} \u2014 fragrant biryani-spiced long-grain rice. {_BRAND}."
        if "jeera" in nl:
            return f"{name} \u2014 cumin-tempered basmati rice, light and aromatic. {_BRAND}."
        if "ghee" in nl:
            return f"{name} \u2014 basmati rice cooked in fragrant pure ghee. {_BRAND}."
        if "curd" in nl:
            return f"{name} \u2014 cool curd rice tempered with mustard and curry leaves. {_BRAND}."
        if "schezwan" in nl:
            return f"{name} \u2014 Indo-Chinese Schezwan fried rice with a spicy kick. {_BRAND}."
        if "fried" in nl:
            return f"{name} \u2014 wok-tossed Indo-Chinese fried rice with assorted vegetables. {_BRAND}."
        return f"{name} \u2014 comforting rice preparation from the {_BRAND} kitchen."

    if cl == "biryanis":
        if "mutton" in nl:
            return f"{name} \u2014 slow-dum-cooked mutton biryani layered with fragrant basmati and spices. {_BRAND}."
        if "prawn" in nl:
            return f"{name} \u2014 prawns layered with saffron-scented biryani rice and whole spices. {_BRAND}."
        if "fish" in nl:
            return f"{name} \u2014 fish biryani with delicate spices and long-grain basmati. {_BRAND}."
        if "egg" in nl:
            return f"{name} \u2014 boiled eggs layered with biryani rice and aromatic masalas. {_BRAND}."
        if "veg" in nl or "paneer" in nl or "mushroom" in nl or "kaju" in nl:
            return f"{name} \u2014 dum-cooked vegetarian biryani fragrant with whole spices. {_BRAND}."
        if "kamju" in nl:
            return f"{name} \u2014 Kamju (shark) biryani with robust Andhra spices. {_BRAND}."
        return f"{name} \u2014 dum-cooked chicken biryani layered with saffron-scented basmati. {_BRAND}."

    if cl == "pulaos":
        if "chicken" in nl:
            return f"{name} \u2014 fragrant chicken pulao cooked with whole spices and basmati rice. {_BRAND}."
        if "mutton" in nl:
            return f"{name} \u2014 mutton pulao cooked with whole spices and basmati. {_BRAND}."
        if "paneer" in nl:
            return f"{name} \u2014 paneer pulao with whole spices and herb-scented basmati. {_BRAND}."
        if "kaju" in nl:
            return f"{name} \u2014 cashew pulao with mild spices and fragrant basmati. {_BRAND}."
        return f"{name} \u2014 lightly-spiced vegetable pulao with basmati rice. {_BRAND}."

    return f"{name} \u2014 a {veg} favourite from the {_BRAND} kitchen."


# ── Helpers ───────────────────────────────────────────────────────────────────
def _has_value(val) -> bool:
    if val is None:
        return False
    s = str(val).strip()
    return bool(s) and s != "-"


def _num(val) -> float | int | None:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    s = str(val).strip()
    if not s or s in {"-", "\u2014"}:
        return None
    try:
        x = float(s.replace(",", "").rstrip("%"))
        return int(x) if x == int(x) else round(x, 2)
    except ValueError:
        return None


def _num_or_none(val):
    n = _num(val)
    return n


def _rounded(n, default=0):
    if n is None:
        return default
    if isinstance(n, float) and n == int(n):
        return int(n)
    return n


# ── Layout detection + pricing (inline, self-contained) ──────────────────────
ZOMATO_RATIO = 0.64     # restaurant share on Zomato
OUR_RATIO = 0.83        # restaurant share on our app
NO_Z_HIKE_PCT = 20.0    # fixed hike when Zomato price absent


def _find_col(ws, *keywords: str) -> int | None:
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").lower()
        if all(k in h for k in keywords):
            return c
    return None


def _detect_cols(ws) -> dict[str, int]:
    """Return logical→col-index map; works on source OR target layout."""
    cols: dict[str, int] = {"cat": 1, "sub": 2, "name": 3}
    # Name might actually be under a header; prefer column 3 by convention.
    # Dine-in column (always has "dine" in header).
    c = _find_col(ws, "dine")
    if c:
        cols["dine"] = c
    # Zomato price column (header "Zomato" but NOT "zomato hike").
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").lower()
        if "zomato" in h and "hike" not in h and "r_z" not in h:
            cols["zom"] = c
            break
    return cols


def compute_pricing(D: float | None, Z: float | None) -> dict:
    """Return all pricing columns for a single item."""
    out = {
        "price": None, "our_hike": None,
        "z_hike": None, "rz": None,
        "rest_us": None, "rev": None,
        "save": None, "save_pct": None, "gain": None,
    }
    if D is None or D <= 0:
        return out
    if Z is not None and Z > 0:
        R_z = Z * ZOMATO_RATIO
        P = round((0.40 * Z + 0.60 * R_z) / 0.85)
        if not (P < Z and OUR_RATIO * P > R_z):
            for delta in (0, -1, 1, -2, 2, -3, 3, -4, 4, -5, 5):
                cand = P + delta
                if cand > 0 and cand < Z and OUR_RATIO * cand > R_z:
                    P = cand
                    break
        out["price"] = P
        out["our_hike"] = round((P - D) / D * 100, 2)
        out["z_hike"] = round((Z - D) / D * 100, 2)
        out["rz"] = round(R_z, 2)
        out["rest_us"] = round(OUR_RATIO * P, 2)
        out["rev"] = round((1 - OUR_RATIO) * P, 2)
        out["save"] = round(Z - P, 2)
        out["save_pct"] = round((Z - P) / Z * 100, 2)
        out["gain"] = round(OUR_RATIO * P - R_z, 2)
    else:
        P = round(D * (1 + NO_Z_HIKE_PCT / 100))
        out["price"] = P
        out["our_hike"] = NO_Z_HIKE_PCT
        out["rest_us"] = round(OUR_RATIO * P, 2)
        out["rev"] = round((1 - OUR_RATIO) * P, 2)
    return out


# ── Main enrichment ───────────────────────────────────────────────────────────
def enrich_workbook() -> list[dict]:
    if not XLSX_PATH.is_file():
        raise SystemExit(f"Missing input workbook: {XLSX_PATH}")

    src_wb = load_workbook(XLSX_PATH)
    src = src_wb.active
    cols = _detect_cols(src)
    if "dine" not in cols or "zom" not in cols:
        raise SystemExit(
            f"Could not locate Dine-In / Zomato columns in {XLSX_PATH}. "
            f"Detected: {cols}"
        )

    # Read all rows into memory so we can rewrite the sheet cleanly.
    src_rows: list[dict] = []
    for r in range(2, src.max_row + 1):
        name = src.cell(r, cols["name"]).value
        if name is None or str(name).strip() == "":
            continue
        cat_v = src.cell(r, cols["cat"]).value
        sub_v = src.cell(r, cols["sub"]).value
        D = _num(src.cell(r, cols["dine"]).value)
        Z = _num(src.cell(r, cols["zom"]).value)
        pricing = compute_pricing(D, Z)
        src_rows.append({
            "cat": cat_v.strip() if isinstance(cat_v, str) else cat_v,
            "sub": sub_v.strip() if isinstance(sub_v, str) else sub_v,
            "name": str(name).strip(),
            "dine": D,
            "zom": Z,
            **pricing,
        })

    # Build a fresh workbook with the target layout.
    out_wb = Workbook()
    out = out_wb.active
    out.title = src.title or "Menu"

    hdr_fill = PatternFill("solid", fgColor="2E86AB")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for i, label in enumerate(TARGET_HEADERS, 1):
        c = out.cell(1, i, label)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = center

    picker = _Picker()
    payloads: list[dict] = []

    out_row = 2
    for row in src_rows:
        cat = str(row["cat"] or "").strip()
        sub_raw = row["sub"]
        sub = str(sub_raw).strip() if sub_raw not in (None, "") else ""
        sub = "" if sub == "-" else sub
        name = row["name"]

        is_veg = not _is_nonveg(cat, sub, name)
        desc = build_description(cat, sub, name, is_veg)
        img_url = picker.pick(select_pool(cat, sub, name))

        price = row["price"] if row["price"] is not None else row["dine"]  # fallback
        hike = row["our_hike"] if row["our_hike"] is not None else 0

        # Write target columns.
        out.cell(out_row, T_CAT, cat)
        out.cell(out_row, T_SUB, sub or "-")
        out.cell(out_row, T_NAME, name)
        out.cell(out_row, T_PRICE, _rounded(price))
        out.cell(out_row, T_HIKE, _rounded(hike))
        out.cell(out_row, T_DINE, _rounded(row["dine"]))
        out.cell(out_row, T_ZOM, _rounded(row["zom"]) if row["zom"] is not None else "-")
        out.cell(out_row, T_VEG, bool(is_veg)).alignment = center
        out.cell(out_row, T_IMG, img_url).alignment = wrap
        out.cell(out_row, T_Z_HIKE, row["z_hike"] if row["z_hike"] is not None else "-")
        out.cell(out_row, T_RZ, row["rz"] if row["rz"] is not None else "-")
        out.cell(out_row, T_DESC, desc).alignment = wrap
        out.cell(out_row, T_REST_US, row["rest_us"] if row["rest_us"] is not None else "-")
        out.cell(out_row, T_REV, row["rev"] if row["rev"] is not None else "-")
        out.cell(out_row, T_SAVE, row["save"] if row["save"] is not None else "-")
        out.cell(out_row, T_SAVE_PCT, row["save_pct"] if row["save_pct"] is not None else "-")
        out.cell(out_row, T_GAIN, row["gain"] if row["gain"] is not None else "-")

        payloads.append(
            {
                "name": name,
                "restaurantPrice": float(price) if price is not None else 0.0,
                "hikePercentage": float(hike) if hike is not None else 0.0,
                "category": cat or None,
                "subCategory": sub or None,
                "isVeg": bool(is_veg),
                "isAvailable": True,
                "description": desc,
                "image": [img_url],
                "restaurantId": RESTAURANT_ID,
            }
        )

        out_row += 1

    # Column widths.
    widths = {
        "A": 18, "B": 22, "C": 36,
        "D": 16, "E": 16, "F": 14, "G": 14,
        "H": 8, "I": 64,
        "J": 14, "K": 14, "L": 60,
        "M": 20, "N": 20, "O": 20, "P": 20, "Q": 20,
    }
    for col, w in widths.items():
        out.column_dimensions[col].width = w

    out_wb.save(XLSX_PATH)
    JSON_PATH.write_text(
        json.dumps(payloads, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    print(f"Enriched rows: {len(payloads)}")
    print(f"Saved workbook: {XLSX_PATH}")
    print(f"Saved JSON:     {JSON_PATH}")
    print(f"Unique images:  {len({p['image'][0] for p in payloads})}")
    return payloads


# ── Optional API import ───────────────────────────────────────────────────────
def menu_item_api_body(row: dict) -> dict:
    cat = (row.get("category") or "").strip() or None
    sub = (row.get("subCategory") or "").strip() or None
    desc = row.get("description")
    if isinstance(desc, str):
        desc = desc.strip() or None
    return {
        "name": row["name"],
        "restaurantPrice": float(row["restaurantPrice"]),
        "hikePercentage": float(row["hikePercentage"]),
        "category": cat,
        "subCategory": sub,
        "isVeg": row.get("isVeg", True),
        "isAvailable": row.get("isAvailable", True),
        "description": desc,
        "image": row.get("image") or [],
    }


def post_menu_items(
    payloads: list[dict],
    *,
    api_url: str,
    bearer_token: str | None,
    delay_sec: float,
) -> int:
    menu_url = f"{api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"
    session = requests.Session()
    session.headers["Content-Type"] = "application/json"
    if bearer_token:
        session.headers["Authorization"] = f"Bearer {bearer_token}"
    else:
        session.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS

    ok = err = 0
    for i, raw in enumerate(payloads, 1):
        body = menu_item_api_body(raw)
        if delay_sec > 0:
            time.sleep(delay_sec)
        try:
            r = session.post(menu_url, json=body, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"  x [{i}] {body['name']}: {ex}", file=sys.stderr)
            continue
        if r.status_code == 201:
            ok += 1
            print(f"  ok [{i}] {body['name']}")
        else:
            err += 1
            print(
                f"  x [{i}] {body['name']}: HTTP {r.status_code} {r.text[:300]}",
                file=sys.stderr,
            )
            if r.status_code == 401:
                print(
                    "\nUnauthorized: set HONESTEATS_RETOOL_BYPASS or HONESTEATS_BEARER_TOKEN.",
                    file=sys.stderr,
                )
                return 1

    print(f"\nPOST done: {ok} created, {err} failed. URL base: {api_url.rstrip('/')}")
    return 0 if err == 0 else 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich Tejaswi menu + optional API import")
    parser.add_argument("--apply", action="store_true", help="POST each item to the API")
    parser.add_argument("--from-json", action="store_true", help="Load payloads from JSON instead of re-enriching")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--bearer-token", default=os.environ.get("HONESTEATS_BEARER_TOKEN"))
    parser.add_argument("--delay", type=float, default=0.3)
    args = parser.parse_args()

    if args.from_json:
        if not JSON_PATH.is_file():
            print(f"Missing {JSON_PATH}", file=sys.stderr)
            return 1
        payloads = json.loads(JSON_PATH.read_text(encoding="utf-8"))
        print(f"Loaded {len(payloads)} items from {JSON_PATH.name}")
    else:
        payloads = enrich_workbook()

    if not args.apply:
        print("No POST (pass --apply to import).")
        return 0

    return post_menu_items(
        payloads,
        api_url=args.api_url,
        bearer_token=args.bearer_token,
        delay_sec=args.delay,
    )


if __name__ == "__main__":
    raise SystemExit(main())
