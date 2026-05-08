#!/usr/bin/env python3
"""
Enrich Savitha Udupi pricing Excel and emit menu-item JSON; optional POST to API.

Reads:  /Users/user/Downloads/SavithaUdupi Pricing.xlsx
        Cols A,B,C: Category | Item | Dine-In (₹)

Adds:
  I — subCategory (Title Case, derived from item name + category)

Writes JSON → /Users/user/Downloads/SavithaUdupi.menu_bodies.json with the
exact body shape requested by the user:
  {
    "name": <Item, Title Case>,
    "restaurantPrice": <Dine-In ₹>,
    "hikePercentage": 35,
    "category": <Category, Title Case>,
    "subCategory": <generated, Title Case>,
    "isVeg": true,
    "isAvailable": true
  }

Restaurant id: RES-1778251406858-7017

Usage:
  python3 scripts/enrich_savitha_udupi_menu.py                    # enrich + JSON only
  python3 scripts/enrich_savitha_udupi_menu.py --apply             # enrich + POST
  python3 scripts/enrich_savitha_udupi_menu.py --from-json --apply # POST from existing JSON
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

XLSX_PATH = Path("/Users/user/Downloads/SavithaUdupi Pricing.xlsx")
JSON_PATH = Path("/Users/user/Downloads/SavithaUdupi.menu_bodies.json")
RESTAURANT_ID = "RES-1778251406858-7017"
HIKE_PERCENTAGE = 35

RETOOL_BYPASS_HEADER = "x-retool-header"
DEFAULT_RETOOL_BYPASS = os.environ.get(
    "HONESTEATS_RETOOL_BYPASS",
    "9f2b7c4a6d1e8f30b5a9c2e7d4f1a6bc",
)
DEFAULT_API_URL = os.environ.get("HONESTEATS_API_URL", "https://api.yumdude.com")

COL_CAT, COL_NAME, COL_PRICE = 1, 2, 3
COL_SUB = 9  # column I


# ── Title-case helpers ───────────────────────────────────────────────────────
_LOWER_FIX = {
    "spl.": "Spl.",
    "veg.": "Veg.",
    "mix.": "Mix.",
    "mashroom": "Mushroom",  # source typo
    "panjabi": "Punjabi",    # source typo
    "(plate)": "(Plate)",
    "(single)": "(Single)",
    "(dry)": "(Dry)",
    "(gravy)": "(Gravy)",
    "(lunch)": "(Lunch)",
    "(1 pc)": "(1 pc)",
    "(4 pcs)": "(4 pcs)",
    "(6 pcs)": "(6 pcs)",
    "(8 pcs)": "(8 pcs)",
}


def title_case(text: str) -> str:
    """Title-case a string while preserving abbreviations like 'Spl.'/'Veg.'.

    - Splits on whitespace, capitalizes first letter of each token.
    - Keeps existing punctuation.
    - Applies a small fix-map to normalize common abbreviations / typos.
    """
    if text is None:
        return ""
    s = str(text).strip()
    if not s:
        return ""
    # Apply fix map first (case-insensitive substring replacement on whole tokens)
    tokens = re.split(r"(\s+)", s)
    out = []
    for tok in tokens:
        if not tok or tok.isspace():
            out.append(tok)
            continue
        low = tok.lower()
        if low in _LOWER_FIX:
            out.append(_LOWER_FIX[low])
            continue
        # Preserve all-caps short tokens like '&'
        if tok in ("&", "+", "-", "/"):
            out.append(tok)
            continue
        # Capitalize first alphabetic char, keep the rest as the source had it,
        # but also Title-case 2nd letter onward only if all-lowercase.
        if tok.isupper() and len(tok) > 1:
            out.append(tok)  # keep ALLCAPS like "BBQ"
        elif tok.islower():
            out.append(tok[:1].upper() + tok[1:])
        else:
            out.append(tok[:1].upper() + tok[1:])
    return "".join(out)


# ── Subcategory rules ────────────────────────────────────────────────────────
def derive_subcategory(category: str, name: str) -> str:
    """Return a meaningful Title-Case subcategory for the row."""
    cl = category.lower()
    nl = name.lower()

    # Breakfast & Tiffins
    if "breakfast" in cl or "tiffin" in cl:
        if "uthappam" in nl:
            return "Uthappam"
        if "pesarattu" in nl:
            return "Pesarattu"
        if "dosa" in nl:
            return "Dosa"
        if "idly" in nl:
            return "Idly"
        if "vada" in nl:
            return "Vada"
        if "poori" in nl:
            return "Poori"
        if "chapathi" in nl or "chapati" in nl:
            return "Chapathi"
        if "pongal" in nl:
            return "Pongal"
        if "upma" in nl:
            return "Upma"
        if "curd rice" in nl:
            return "Curd Rice"
        if "pulav" in nl or "pulao" in nl:
            return "Pulav"
        return "Tiffins"

    # Rice & Biryani
    if "rice" in cl and "biryani" in cl:
        if "biryani" in nl:
            return "Biryani"
        if "fried rice" in nl:
            return "Fried Rice"
        if "pulao" in nl or "pulav" in nl:
            return "Pulao"
        if "schezwan rice" in nl:
            return "Fried Rice"
        return "Rice"

    # Noodles
    if "noodle" in cl:
        return "Noodles"

    # Soups
    if "soup" in cl:
        return "Soup"

    # Starters
    if "starter" in cl:
        if "manchurian" in nl:
            return "Manchurian"
        if "chilli" in nl:
            return "Chilli"
        if "pepper roast" in nl:
            return "Pepper Roast"
        if "65" in nl:
            return "65"
        if "lollipop" in nl:
            return "Lollipop"
        if "toast" in nl:
            return "Toast"
        if "satay" in nl:
            return "Satay"
        if "crispy corn" in nl:
            return "Crispy Corn"
        if "majestic" in nl:
            return "Majestic"
        return "Starters"

    # Vegetable Curries
    if "vegetable curr" in cl or cl == "vegetable curries":
        if "paneer" in nl:
            return "Paneer Curry"
        if "mushroom" in nl or "mashroom" in nl:
            return "Mushroom Curry"
        if "kaju" in nl:
            return "Kaju Curry"
        if "kofta" in nl:
            return "Kofta"
        if "tandoori" in nl:
            return "Tandoori Curry"
        if "aloo" in nl:
            return "Aloo Curry"
        if "gobi" in nl:
            return "Gobi Curry"
        if "palak" in nl:
            return "Palak Curry"
        if "kadai" in nl:
            return "Kadai Curry"
        return "Veg Curry"

    # Special Curries
    if "special curr" in cl:
        if "paneer" in nl:
            return "Paneer Special"
        return "Veg Special"

    # Indian Breads
    if "bread" in cl:
        if "naan" in nl:
            return "Naan"
        if "kulcha" in nl:
            return "Kulcha"
        if "paratha" in nl:
            return "Paratha"
        if "roti" in nl:
            return "Roti"
        return "Indian Bread"

    # Chats
    if "chat" in cl:
        if "pav bhaji" in nl:
            return "Pav Bhaji"
        if "samosa" in nl:
            return "Samosa"
        if "kachori" in nl:
            return "Kachori"
        if "ragada" in nl or "ragda" in nl:
            return "Ragada"
        if "pani puri" in nl or "puri" in nl or "bhel" in nl or "papdi" in nl:
            return "Puri & Bhel"
        if "vada pav" in nl:
            return "Vada Pav"
        if "cutlet" in nl:
            return "Cutlet"
        return "Chat"

    # Snacks
    if "snack" in cl:
        if "pakoda" in nl or "pakora" in nl:
            return "Pakoda"
        if "finger chips" in nl or "chips" in nl:
            return "Finger Chips"
        return "Snack"

    return title_case(category)


# ── Number parsing ───────────────────────────────────────────────────────────
def parse_num(val) -> float | int:
    if val is None or val == "" or val == "-":
        return 0
    if isinstance(val, (int, float)):
        return int(val) if float(val) == int(val) else round(float(val), 2)
    try:
        x = float(str(val).strip().rstrip("%"))
        return int(x) if x == int(x) else round(x, 2)
    except ValueError:
        return 0


# ── Enrichment ───────────────────────────────────────────────────────────────
def enrich_workbook() -> list[dict]:
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    hdr_fill = PatternFill("solid", fgColor="2E86AB")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sub_header = ws.cell(1, COL_SUB, "Subcategory")
    sub_header.font = hdr_font
    sub_header.fill = hdr_fill
    sub_header.alignment = center

    payloads: list[dict] = []

    for r in range(2, ws.max_row + 1):
        raw_name = ws.cell(r, COL_NAME).value
        if raw_name is None or str(raw_name).strip() == "":
            continue
        raw_cat = ws.cell(r, COL_CAT).value
        raw_price = ws.cell(r, COL_PRICE).value

        name = title_case(str(raw_name))
        cat = title_case(str(raw_cat or ""))
        sub = title_case(derive_subcategory(cat, name))

        ws.cell(r, COL_SUB, sub).alignment = center

        price = parse_num(raw_price)
        if price <= 0:
            print(f"  ! skipping row {r} ({name!r}): non-positive price={raw_price!r}", file=sys.stderr)
            continue

        payloads.append(
            {
                "name": name,
                "restaurantPrice": float(price),
                "hikePercentage": HIKE_PERCENTAGE,
                "category": cat,
                "subCategory": sub,
                "isVeg": True,
                "isAvailable": True,
            }
        )

    ws.column_dimensions["I"].width = 22
    wb.save(XLSX_PATH)
    JSON_PATH.write_text(json.dumps(payloads, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Enriched rows: {len(payloads)}")
    print(f"Saved workbook: {XLSX_PATH}")
    print(f"Saved JSON:     {JSON_PATH}")
    cats = {}
    for p in payloads:
        key = (p["category"], p["subCategory"])
        cats[key] = cats.get(key, 0) + 1
    print("Category / Subcategory breakdown:")
    for (c, s), n in sorted(cats.items()):
        print(f"  {c:30s} | {s:20s} | {n}")
    return payloads


# ── API insert ───────────────────────────────────────────────────────────────
def post_menu_items(
    payloads: list[dict],
    *,
    api_url: str,
    bearer_token: str | None,
    delay_sec: float,
) -> int:
    menu_url = f"{api_url.rstrip('/')}/api/v1/restaurants/{RESTAURANT_ID}/menu"
    session = requests.Session()
    session.headers["Content-Type"] = "application/json"
    if bearer_token:
        session.headers["Authorization"] = f"Bearer {bearer_token}"
    else:
        session.headers[RETOOL_BYPASS_HEADER] = DEFAULT_RETOOL_BYPASS

    print(f"\nPOST {menu_url}")
    print(f"Items: {len(payloads)}\n")

    ok = err = 0
    for i, body in enumerate(payloads, 1):
        if delay_sec > 0:
            time.sleep(delay_sec)
        try:
            r = session.post(menu_url, json=body, timeout=60)
        except requests.RequestException as ex:
            err += 1
            print(f"  x [{i}] {body['name']}: {ex}", file=sys.stderr)
            continue
        if r.status_code == 201:
            ok += 1
            print(f"  + [{i}] {body['name']}")
        else:
            err += 1
            print(
                f"  x [{i}] {body['name']}: HTTP {r.status_code} {r.text[:300]}",
                file=sys.stderr,
            )
            if r.status_code == 401:
                print("\nUnauthorized: set HONESTEATS_RETOOL_BYPASS or HONESTEATS_BEARER_TOKEN.", file=sys.stderr)
                return 1

    print(f"\nPOST done: {ok} created, {err} failed.")
    return 0 if err == 0 else 1


def main() -> int:
    parser = argparse.ArgumentParser(description="Enrich SavithaUdupi pricing + optional API import")
    parser.add_argument("--apply", action="store_true", help="POST each item to the API")
    parser.add_argument("--from-json", action="store_true", help="Load payloads from JSON instead of re-enriching")
    parser.add_argument("--api-url", default=DEFAULT_API_URL)
    parser.add_argument("--bearer-token", default=os.environ.get("HONESTEATS_BEARER_TOKEN"))
    parser.add_argument("--delay", type=float, default=0.2)
    args = parser.parse_args()

    if args.from_json:
        if not JSON_PATH.is_file():
            print(f"Missing {JSON_PATH}", file=sys.stderr)
            return 1
        payloads = json.loads(JSON_PATH.read_text(encoding="utf-8"))
        print(f"Loaded {len(payloads)} items from {JSON_PATH.name}")
    else:
        payloads = enrich_workbook()

    if not args.apply:
        print("\nNo POST (pass --apply to import).")
        return 0

    return post_menu_items(
        payloads,
        api_url=args.api_url,
        bearer_token=args.bearer_token,
        delay_sec=args.delay,
    )


if __name__ == "__main__":
    raise SystemExit(main())
