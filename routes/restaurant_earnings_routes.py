"""Restaurant settlement / earnings routes"""
import io
import base64
import time
from datetime import datetime, timezone

from aws_lambda_powertools import Logger, Tracer, Metrics

from models.order import Order
from services.restaurant_earnings_service import RestaurantEarningsService
from utils.dynamodb import dynamodb_client, TABLES

logger = Logger()
tracer = Tracer()
metrics = Metrics()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _batch_fetch_orders(order_ids: list) -> dict:
    """Batch-fetch orders from DynamoDB. Returns {orderId: Order} map.

    Handles UnprocessedKeys with exponential backoff so no order is
    silently dropped under throttling or capacity pressure.
    """
    result = {}
    for i in range(0, len(order_ids), 100):
        chunk = order_ids[i:i + 100]
        pending = {TABLES['ORDERS']: {'Keys': [{'orderId': {'S': oid}} for oid in chunk]}}

        delay = 0.1  # initial backoff seconds
        while pending:
            response = dynamodb_client.batch_get_item(RequestItems=pending)

            for raw in response.get('Responses', {}).get(TABLES['ORDERS'], []):
                order = Order.from_dynamodb_item(raw)
                result[order.order_id] = order

            unprocessed = response.get('UnprocessedKeys', {})
            if unprocessed:
                logger.warning(
                    f"UnprocessedKeys returned for {len(unprocessed.get(TABLES['ORDERS'], {}).get('Keys', []))} "
                    f"orders — retrying after {delay:.1f}s"
                )
                time.sleep(delay)
                delay = min(delay * 2, 5)  # cap at 5 s
                pending = unprocessed
            else:
                pending = {}

    return result


def _order_to_row(order: Order, earning_date: str) -> dict:
    """Extract settlement fields from an Order object."""
    rev = order.revenue or {}
    platform_rev = rev.get('platformRevenue') or {}
    restaurant_rev = rev.get('restaurantRevenue') or {}

    food_commission = float(platform_rev.get('foodCommission') or 0)
    coupon_deduction = (
        float(restaurant_rev.get('couponDiscount') or 0) +
        float(restaurant_rev.get('itemCouponDiscount') or 0)
    )
    net_payout = float(restaurant_rev.get('finalPayout') or 0)

    return {
        'orderId'        : order.order_id,
        'date'           : earning_date,
        'createdAt'      : order.created_at if isinstance(order.created_at, str) else '',
        'grandTotal'     : round(float(order.grand_total or 0), 2),
        'foodTotal'      : round(float(order.food_total or 0), 2),
        'deliveryFee'    : round(float(order.delivery_fee or 0), 2),
        'platformFee'    : round(float(order.platform_fee or 0), 2),
        'paymentMethod'  : order.payment_method or '-',
        'restaurantName' : order.restaurant_name or '',
        'foodCommission' : round(food_commission, 2),
        'couponDeduction': round(coupon_deduction, 2),
        'netPayout'      : round(net_payout, 2),
    }


def _build_summary(rows: list) -> dict:
    return {
        'totalOrders'         : len(rows),
        'totalGMV'            : round(sum(r['foodTotal'] for r in rows), 2),
        'totalCommission'     : round(sum(r['foodCommission'] for r in rows), 2),
        'totalCouponDeduction': round(sum(r['couponDeduction'] for r in rows), 2),
        'netPayable'          : round(sum(r['netPayout'] for r in rows), 2),
    }


def _generate_xlsx(restaurant_id: str, restaurant_name: str, settlement_id: str,
                   start_date: str, end_date: str, rows: list) -> bytes:
    """Generate an XLSX settlement report using openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise Exception("openpyxl not installed — add to requirements.txt")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Settlement Report"

    RED   = "E8352A"
    DARK  = "2E241B"
    GRAY  = "F5F3EF"
    MID   = "EEEBE4"
    WHITE = "FFFFFF"

    thin   = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def fill(c):
        return PatternFill("solid", fgColor=c)

    def font(bold=False, color=DARK, size=10):
        return Font(bold=bold, color=color, name="Calibri", size=size)

    # ── Row 1: Title bar ─────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    ws['A1'].value     = "YumDude  •  Restaurant Settlement Report"
    ws['A1'].font      = Font(bold=True, color=WHITE, name="Calibri", size=14)
    ws['A1'].fill      = fill(RED)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 34

    # ── Rows 3–7: Info block ─────────────────────────────────────────────────
    info_items = [
        ("Restaurant Name",   restaurant_name or restaurant_id),
        ("Restaurant ID",     restaurant_id),
        ("Settlement Period", f"{start_date}  —  {end_date}"),
        ("Settlement ID",     settlement_id),
        ("Generated On",      datetime.now(timezone.utc).strftime("%d %b %Y, %H:%M UTC")),
    ]
    for offset, (label, value) in enumerate(info_items, start=3):
        lc = ws.cell(row=offset, column=1, value=label)
        lc.font = font(bold=True); lc.fill = fill(MID)
        lc.border = border; lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        ws.merge_cells(f'B{offset}:D{offset}')
        vc = ws.cell(row=offset, column=2, value=value)
        vc.font = font(); vc.fill = fill(WHITE)
        vc.border = border; vc.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    # ── Summary section ──────────────────────────────────────────────────────
    summary = _build_summary(rows)
    sr = 3 + len(info_items) + 1  # one blank row after info

    ws.merge_cells(f'A{sr}:H{sr}')
    sh = ws.cell(row=sr, column=1, value="Summary")
    sh.font = Font(bold=True, color=WHITE, name="Calibri", size=11)
    sh.fill = fill(DARK)
    sh.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws.row_dimensions[sr].height = 24

    summary_rows = [
        ("Total Orders",                           str(summary['totalOrders'])),
        ("Total Gross Menu Value",                 f"\u20b9{summary['totalGMV']:,.2f}"),
        ("Platform Commission",                    f"\u20b9{summary['totalCommission']:,.2f}"),
        ("Your Coupon Discount (restaurant-issued)", f"\u20b9{summary['totalCouponDeduction']:,.2f}"),
        ("Your Net Settlement",                    f"\u20b9{summary['netPayable']:,.2f}"),
        ("Earnings Formula",                       "Gross Menu Value \u2212 Platform Commission \u2212 Your Coupon Discount = Net Settlement"),
    ]
    for offset, (label, value) in enumerate(summary_rows, start=sr + 1):
        is_net = label.startswith("Net")
        lc = ws.cell(row=offset, column=1, value=label)
        lc.font = font(bold=True); lc.fill = fill(GRAY)
        lc.border = border; lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        ws.merge_cells(f'B{offset}:D{offset}')
        vc = ws.cell(row=offset, column=2, value=value)
        vc.font = Font(bold=is_net, color=RED if is_net else DARK, name="Calibri", size=10)
        vc.fill = fill(GRAY); vc.border = border
        vc.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    # ── Orders table ─────────────────────────────────────────────────────────
    tr = sr + len(summary_rows) + 2  # one blank row before table

    headers    = ["#", "Order ID", "Date & Time", "Gross Menu Value (₹)",
                  "Platform Commission (₹)", "Your Coupon Discount (₹)", "Your Net Payout (₹)", "Payment"]
    col_widths = [5, 22, 22, 22, 24, 26, 22, 14]

    for ci, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=tr, column=ci, value=h)
        cell.font  = Font(bold=True, color=WHITE, name="Calibri", size=10)
        cell.fill  = fill(RED); cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[tr].height = 36  # taller header to accommodate wrapped text

    for row_num, row in enumerate(rows, start=1):
        r        = tr + row_num
        row_fill = fill(WHITE) if row_num % 2 == 1 else fill(GRAY)

        try:
            dt       = datetime.fromisoformat(row['createdAt'].replace('Z', '+00:00'))
            date_str = dt.strftime("%d %b %Y  %H:%M")
        except Exception:
            date_str = row['createdAt'][:16] if row['createdAt'] else '-'

        values = [
            row_num, row['orderId'], date_str,
            row['foodTotal'],
            row['foodCommission'], row['couponDeduction'],
            row['netPayout'], row['paymentMethod'],
        ]
        for ci, val in enumerate(values, start=1):
            cell      = ws.cell(row=r, column=ci, value=val)
            is_num    = ci in (4, 5, 6, 7)
            is_payout = ci == 7
            cell.font = Font(bold=is_payout, color=DARK, name="Calibri", size=10)
            cell.fill = row_fill; cell.border = border
            cell.alignment = Alignment(horizontal='right' if is_num else 'center', vertical='center')
            if is_num and isinstance(val, float):
                cell.number_format = '#,##0.00'

    # ── Footer ───────────────────────────────────────────────────────────────
    fr = tr + len(rows) + 2
    ws.merge_cells(f'A{fr}:H{fr}')
    fc = ws.cell(row=fr, column=1,
                 value="System-generated report by YumDude. For disputes, contact support@yumdude.com")
    fc.font      = Font(italic=True, color="999999", name="Calibri", size=9)
    fc.alignment = Alignment(horizontal='center')

    ws.freeze_panes = ws.cell(row=tr + 1, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Route registration
# ─────────────────────────────────────────────────────────────────────────────

def register_restaurant_earnings_routes(app):
    """Register restaurant earnings / settlement routes."""

    @app.get("/api/v1/restaurants/<restaurant_id>/earnings/settlement/preview")
    @tracer.capture_method
    def get_settlement_preview(restaurant_id: str):
        """
        Preview unsettled restaurant earnings for a date range.

        Query params:
          startDate  YYYY-MM-DD
          endDate    YYYY-MM-DD
        """
        try:
            qp         = app.current_event.query_string_parameters or {}
            start_date = qp.get('startDate')
            end_date   = qp.get('endDate')

            if not start_date or not end_date:
                return {"error": "startDate and endDate query params required"}, 400

            earnings_list = RestaurantEarningsService.get_earnings_for_date_range(
                restaurant_id, start_date, end_date
            )
            unsettled = [e for e in earnings_list if not e.settled]
            order_ids = [e.order_id for e in unsettled if e.order_id]

            if not order_ids:
                return {
                    "restaurantId": restaurant_id, "startDate": start_date, "endDate": end_date,
                    "totalOrders": 0, "totalGMV": 0, "totalCommission": 0,
                    "totalCouponDeduction": 0, "netPayable": 0, "orders": []
                }, 200

            orders_map      = _batch_fetch_orders(order_ids)
            restaurant_name = ""
            rows            = []

            for earning in unsettled:
                if not earning.order_id or earning.order_id not in orders_map:
                    continue
                row = _order_to_row(orders_map[earning.order_id], earning.date.split('#')[0])
                if not restaurant_name:
                    restaurant_name = row['restaurantName']
                rows.append(row)

            summary = _build_summary(rows)
            return {
                "restaurantId"  : restaurant_id,
                "restaurantName": restaurant_name,
                "startDate"     : start_date,
                "endDate"       : end_date,
                **summary,
                "orders"        : rows,
            }, 200

        except Exception as e:
            logger.error("Error fetching settlement preview", exc_info=True)
            return {"error": "Failed to fetch settlement preview", "message": str(e)}, 500

    # ─────────────────────────────────────────────────────────────────────────

    @app.post("/api/v1/restaurants/<restaurant_id>/earnings/settlement/confirm")
    @tracer.capture_method
    def confirm_settlement(restaurant_id: str):
        """
        Confirm settlement for a restaurant.

        Marks all unsettled earnings in the date range as settled,
        generates an XLSX report, and returns it as a base64 string.

        Body (JSON):
          {
            "startDate":      "YYYY-MM-DD",
            "endDate":        "YYYY-MM-DD",
            "restaurantName": "optional – used in report header"
          }

        Response:
          {
            "settlementId":    "STL-XXXXXX-YYYYMMDDHHMMSS",
            "reportBase64":    "<base64-encoded XLSX>",
            "filename":        "yumdude_settlement_<id>_<range>.xlsx",
            ... summary fields ...
          }
        """
        try:
            body            = app.current_event.json_body or {}
            start_date      = body.get('startDate')
            end_date        = body.get('endDate')
            restaurant_name = body.get('restaurantName', '')

            if not start_date or not end_date:
                return {"error": "startDate and endDate required in request body"}, 400

            earnings_list = RestaurantEarningsService.get_earnings_for_date_range(
                restaurant_id, start_date, end_date
            )
            unsettled = [e for e in earnings_list if not e.settled]
            order_ids = [e.order_id for e in unsettled if e.order_id]

            if not order_ids:
                return {"error": "No unsettled orders found in this date range"}, 400

            orders_map = _batch_fetch_orders(order_ids)
            rows       = []

            for earning in unsettled:
                if not earning.order_id or earning.order_id not in orders_map:
                    continue
                row = _order_to_row(orders_map[earning.order_id], earning.date.split('#')[0])
                if not restaurant_name:
                    restaurant_name = row['restaurantName']
                rows.append(row)

            if not rows:
                return {"error": "Could not load order details for settlement"}, 400

            settlement_id = (
                f"STL-{restaurant_id[:6].upper()}-"
                f"{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}"
            )

            # Mark settled in DynamoDB
            RestaurantEarningsService.settle_earnings_for_orders(
                restaurant_id=restaurant_id,
                order_ids=order_ids,
                start_date=start_date,
                end_date=end_date,
                settlement_id=settlement_id,
            )

            # Generate XLSX and base64-encode
            xlsx_bytes  = _generate_xlsx(
                restaurant_id, restaurant_name, settlement_id, start_date, end_date, rows
            )
            report_b64  = base64.b64encode(xlsx_bytes).decode('utf-8')
            filename    = f"yumdude_settlement_{restaurant_id}_{start_date}_to_{end_date}.xlsx"
            summary     = _build_summary(rows)

            metrics.add_metric(name="RestaurantSettlementConfirmed", unit="Count", value=1)

            return {
                "settlementId"   : settlement_id,
                "restaurantId"   : restaurant_id,
                "restaurantName" : restaurant_name,
                "startDate"      : start_date,
                "endDate"        : end_date,
                **summary,
                "settledOrderIds": order_ids,
                "reportBase64"   : report_b64,
                "filename"       : filename,
            }, 200

        except Exception as e:
            logger.error("Error confirming settlement", exc_info=True)
            return {"error": "Failed to confirm settlement", "message": str(e)}, 500
